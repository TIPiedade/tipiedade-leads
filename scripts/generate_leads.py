"""
Pão de Ló Ti'Piedade — Sistema Completo de Prospeção HORECA
- 20 leads/semana por comercial (HORECA por zonas)
- 5 leads/semana catering & eventos (para Rui)
- 5 leads/semana distribuidores congelados (para Rui)
- Nurturing automático (4 emails por lead)
- Histórico guardado no GitHub
MODO=coordenador → segunda-feira → resumo para Rui
MODO=comerciais  → quarta-feira → leads + nurturing para equipa
"""

import math, smtplib, os, datetime, json, base64, urllib.request, urllib.error
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys, importlib.util

EMAIL_FROM = "sales@tipiedade.com"          # remetente
EMAIL_CC   = "sales@tipiedade.com"          # BCC 1
EMAIL_BCC2 = "geral@tipiedade.com"           # BCC 2
EMAIL_RUI  = os.environ.get("EMAIL_RUI", EMAIL_CC)
REPO_OWNER = os.environ.get("GITHUB_REPOSITORY","TIPiedade/tipiedade-leads").split("/")[0]
REPO_NAME  = os.environ.get("GITHUB_REPOSITORY","TIPiedade/tipiedade-leads").split("/")[1]
HIST_FILE  = "historico.json"

# ── Base de leads reais (integrada directamente) ─────────────
# BASE DE LEADS REAIS — gerado automaticamente
# 239 leads de Base_Geral_Leads_Horeca + Alvos_TiPiedade

DB_NUNO = [
  {"n":"Copenhagen Coffee Lab (Cais do Sodré)","t":"Café de Especialidade & Brunch","m":"Praça de São Paulo 4, 1200-428 Lisboa","tel":"—","email":"—","p":"Alta","tCliente":"Café de Especialidade & Brunch","gancho":"Público internacional de alta rotação. Focar no preço por dose que permite margem muito alta.","zona":"Lisboa (Cais Sodré)"},
  {"n":"Gleba (Amoreiras)","t":"Padaria Artesanal & Gourmet","m":"Avenida Engenheiro Duarte Pacheco, Amoreiras Shopping Center, Loja 1010, 1070-103 Lisboa","tel":"—","email":"—","p":"Alta","tCliente":"Padaria Artesanal & Gourmet","gancho":"Compradores de elevado poder de compra. Venda por impulso para o lanche das famílias.","zona":"Lisboa (Amoreiras)"},
  {"n":"Choupana Caffe","t":"Pastelaria Premium & Brunch","m":"Avenida da República 25A, 1050-186 Lisboa","tel":"—","email":"—","p":"Alta","tCliente":"Pastelaria Premium & Brunch","gancho":"Local de imensa rotação. Ter o Pão de Ló Ti Piedade em exposição rústica atrai lanches de grupo.","zona":"Lisboa (Avenidas Novas)"},
  {"n":"Tartine","t":"Pastelaria & Padaria Fina","m":"Rua Serpa Pinto 15A, 1200-426 Lisboa","tel":"—","email":"—","p":"Alta","tCliente":"Pastelaria & Padaria Fina","gancho":"Junto ao Teatro de São Carlos. Foco na altíssima qualidade técnica do pão de ló Ti Piedade.","zona":"Lisboa (Chiado)"},
  {"n":"Fabrica Coffee Roasters (Rua da Flores)","t":"Café de Especialidade","m":"Rua das Flores 63, 1200-193 Lisboa","tel":"—","email":"—","p":"Alta","tCliente":"Café de Especialidade","gancho":"Harmonização com cafés finos. Foco em turistas e nómadas digitais com ticket de compra alto.","zona":"Lisboa (Baixa)"},
  {"n":"Leitaria da Quinta do Paço (Graça)","t":"Pastelaria Premium","m":"Rua da Graça 90, 1170-170 Lisboa","tel":"—","email":"—","p":"Alta","tCliente":"Pastelaria Premium","gancho":"Venda no balcão de lanches de fim de semana. Pão de ló fofo como o acompanhamento ideal.","zona":"Lisboa (Graça)"},
  {"n":"Gleba (Mercado da Vila)","t":"Padaria Artesanal & Gourmet","m":"Rua Padre Moisés da Silva, 2750-437 Cascais","tel":"—","email":"—","p":"Alta","tCliente":"Padaria Artesanal & Gourmet","gancho":"Mercado movimentado. Pão de ló premium para a sobremesa de domingo de clientes exigentes.","zona":"Cascais"},
  {"n":"The Millstone Sourdough","t":"Padaria Artesanal Premium","m":"Rua Nova da Alfarrobeira 11, 2750-449 Cascais","tel":"—","email":"—","p":"Alta","tCliente":"Padaria Artesanal Premium","gancho":"Processo rústico. Poupa-lhes fabrico mantendo o padrão artesanal limpo e sem aditivos.","zona":"Cascais"},
  {"n":"Lulu - Specialty Coffee & Brunch","t":"Café de Especialidade & Brunch","m":"Avenida Valbom 12, 2750-508 Cascais","tel":"—","email":"—","p":"Alta","tCliente":"Café de Especialidade & Brunch","gancho":"Brunch sofisticado. Apresentar o conceito de fatia de pão de ló com manteiga salgada artesanal.","zona":"Cascais"},
  {"n":"Local Healthy Kitchen (Cascais)","t":"Brunch & Comida Saudável","m":"Rua Padre Moisés da Silva, Mercado da Vila, 2750-437 Cascais","tel":"—","email":"—","p":"Alta","tCliente":"Brunch & Comida Saudável","gancho":"Propor pão de ló artesanal como doce rústico e natural, isento de corantes industriais.","zona":"Cascais"},
  {"n":"Garrafeira Imperial do Estoril","t":"Garrafeira & Gourmet Deli","m":"Avenida Saboia 320, 2765-277 Estoril","tel":"—","email":"—","p":"Alta","tCliente":"Garrafeira & Gourmet Deli","gancho":"Combinação ideal com vinhos generosos (Porto, Madeira, Carcavelos). Venda do bolo inteiro premium.","zona":"Estoril"},
  {"n":"Café Saudade","t":"Casa de Chá & Café de Charme","m":"Avenida Miguel Bombarda 6, 2710-590 Sintra","tel":"—","email":"—","p":"Alta","tCliente":"Casa de Chá & Café de Charme","gancho":"Lanches de charme. Reforça o menu de chás ingleses e infusões com pão de ló tradicional.","zona":"Sintra"},
  {"n":"Garrafeira de Sintra","t":"Garrafeira & Gourmet Deli","m":"Rua Consiglieri Pedroso 11, 2710-550 Sintra","tel":"—","email":"—","p":"Alta","tCliente":"Garrafeira & Gourmet Deli","gancho":"Venda casada com vinhos generosos regionais. Excelente opção de presente regional para turismo.","zona":"Sintra"},
  {"n":"Sintra In Love","t":"Cafetaria Gourmet & Regional","m":"Rua das Padarias 12, 2710-603 Sintra","tel":"—","email":"—","p":"Alta","tCliente":"Cafetaria Gourmet & Regional","gancho":"Centro histórico. Turistas à procura de doçaria portuguesa real. Vender a história do Ti Piedade.","zona":"Sintra"},
  {"n":"Aromas de Sintra","t":"Casa de Chá & Pastelaria Fina","m":"Largo Dr. Virgilio Horta 5, 2710-501 Sintra","tel":"—","email":"—","p":"Alta","tCliente":"Casa de Chá & Pastelaria Fina","gancho":"Excelente para o lanche da tarde. O pão de ló de qualidade mantém-se incrivelmente fresco na vitrine.","zona":"Sintra"},
  {"n":"Frade dos Mares","t":"Restaurante (Médio/Alto Padrão)","m":"Av. Dom Carlos I 55, 1200-109 Lisboa","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão)","gancho":"Cozinha tradicional refinada. Sobremesa de altíssimo nível, sempre disponível (congelada) com quebra zero.","zona":"Lisboa"},
  {"n":"A Casa do Bacalhau","t":"Restaurante (Médio/Alto Padrão)","m":"Rua do Grilo 54, 1900-706 Lisboa","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão)","gancho":"Parceiro tradicional perfeito. Descongelar no dia conforme as reservas garante controle total de custos.","zona":"Lisboa"},
  {"n":"Restaurante Solar dos Presuntos","t":"Restaurante (Médio/Alto Padrão)","m":"Rua das Portas de Santo Antão 150, 1150-269 Lisboa","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão)","gancho":"Instituição tradicional. O Ti Piedade oferece uma consistência e sabor conventual que os clientes exigem.","zona":"Lisboa"},
  {"n":"O Javali","t":"Restaurante Tradicional de Caça & Grelhados","m":"Rua de São Bento 344, 1200-822 Lisboa","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante Tradicional de Caça & Grelhados","gancho":"Menu rústico. Sugerir o pão de ló com uma redução de frutos silvestres ou um licor local.","zona":"Lisboa"},
  {"n":"Faz Figura","t":"Restaurante (Médio/Alto Padrão / Vista Rio)","m":"Rua do Paraíso 15B, 1100-396 Lisboa","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão / Vista Rio)","gancho":"Jantares executivos e eventos. O formato congelado permite ter uma excelente sobremesa sem desperdícios diários.","zona":"Lisboa"},
  {"n":"Páteo do Guincho","t":"Restaurante (Médio/Alto Padrão)","m":"Estrada do Guincho, 2750-642 Cascais","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão)","gancho":"Sobremesa rápida de alta margem: fatia de pão de ló morna com bola de gelado artesanal de nata.","zona":"Cascais"},
  {"n":"Polvo Vadio","t":"Restaurante de Peixe & Marisco","m":"Rua Afonso Sanches 26, 2750-282 Cascais","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante de Peixe & Marisco","gancho":"Foco no porcionamento rápido. O pão de ló fatiado e pronto a servir após o marisco é um sucesso de conforto.","zona":"Cascais"},
  {"n":"Tacho Real","t":"Restaurante (Médio/Alto Padrão)","m":"Rua do Ferraria 4, 2710-590 Sintra","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão)","gancho":"Em pleno centro histórico. Custo fixo por dose e rapidez de serviço para as mesas de grupos turísticos.","zona":"Sintra"},
  {"n":"Restaurante Curral dos Caprinos","t":"Restaurante Regional Premium","m":"Rua da Capela 13, Cabriz, 2710-539 Sintra","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante Regional Premium","gancho":"Pratos pesados de forno. O Pão de Ló Ti Piedade como a sobremesa leve, fofa e tradicional perfeita para fechar.","zona":"Sintra"},
  {"n":"Restaurante Lawrence's","t":"Restaurante de Hotel Histórico","m":"Rua Consiglieri Pedroso 38, 2710-550 Sintra","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante de Hotel Histórico","gancho":"Hotel mais antigo da Península Ibérica. O apelo histórico do Ti Piedade encaixa na narrativa de requinte histórico.","zona":"Sintra"},
  {"n":"Augusto Lisboa","t":"Café & Brunch Premium","m":"Rua de Santa Marinha 26, 1100-491 Lisboa","tel":"—","email":"—","p":"Alta","tCliente":"Café & Brunch Premium","gancho":"Brunch muito concorrido na zona histórica. Perfeito para introduzir uma opção de pão de ló tostado com mel artesanal.","zona":"Lisboa"},
  {"n":"Tacho do Pescador","t":"Restaurante Tradicional Premium","m":"Rua do Pimenta 15, 1990-254 Lisboa","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante Tradicional Premium","gancho":"Junto à FIL. Público executivo e feiras. Sobremesa tradicional portuguesa rápida de porcionar com margem alta.","zona":"Lisboa (Parque das Nações)"},
  {"n":"Gleba (Parque das Nações)","t":"Padaria Artesanal & Gourmet","m":"Alameda dos Oceanos, Lote 1.06.1.1 - Loja 2, 1990-207 Lisboa","tel":"—","email":"—","p":"Alta","tCliente":"Padaria Artesanal & Gourmet","gancho":"Zona executiva e residencial premium. Foco na qualidade limpa do bolo Ti Piedade para as lancheiras familiares.","zona":"Lisboa (Pq. Nações)"},
  {"n":"Copenhagen Coffee Lab (Parque das Nações)","t":"Café de Especialidade & Brunch","m":"Rua do Bojador 47, 1990-048 Lisboa","tel":"—","email":"—","p":"Alta","tCliente":"Café de Especialidade & Brunch","gancho":"Executivos e trabalhadores remotos. Lanche de qualidade a acompanhar o café com margem de lucro muito alta.","zona":"Lisboa (Pq. Nações)"},
  {"n":"Angra Gatti","t":"Italiano","m":"—","tel":"—","email":"geral@angragatti.pt","p":"Alta","tCliente":"Italiano","gancho":"Lead identificado pela equipa comercial — Italiano em Praia Grande.","zona":"Praia Grande"},
  {"n":"Arribas Terrace","t":"Hotel/restaurante","m":"—","tel":"—","email":"reservas@arribashotel.com","p":"Alta","tCliente":"Hotel/restaurante","gancho":"Lead identificado pela equipa comercial — Hotel/restaurante em Hotel Arribas, Praia Grande.","zona":"Hotel Arribas, Praia Grande"},
  {"n":"Baía do Peixe Cascais","t":"Restaurante","m":"—","tel":"—","email":"reservas@baiadopeixe.com","p":"Alta","tCliente":"Restaurante","gancho":"Lead identificado pela equipa comercial — Restaurante em Baía de Cascais.","zona":"Baía de Cascais"},
  {"n":"Bar do Fundo","t":"Restaurante premium","m":"—","tel":"—","email":"info@bardofundo.pt","p":"Alta","tCliente":"Restaurante premium","gancho":"Lead identificado pela equipa comercial — Restaurante premium em Praia Grande.","zona":"Praia Grande"},
  {"n":"Bar do Guincho","t":"Beach restaurant","m":"—","tel":"—","email":"geral@bardoguincho.pt","p":"Alta","tCliente":"Beach restaurant","gancho":"Lead identificado pela equipa comercial — Beach restaurant em Praia do Guincho.","zona":"Praia do Guincho"},
  {"n":"Beira Mar","t":"Restaurante costeiro","m":"—","tel":"—","email":"geral@restaurantebeiramar.pt","p":"Alta","tCliente":"Restaurante costeiro","gancho":"Lead identificado pela equipa comercial — Restaurante costeiro em Cascais.","zona":"Cascais"},
  {"n":"Capricciosa Cascais","t":"Restaurante italiano","m":"—","tel":"—","email":"TheFork","p":"Alta","tCliente":"Restaurante italiano","gancho":"Lead identificado pela equipa comercial — Restaurante italiano em Cascais.","zona":"Cascais"},
  {"n":"Fortaleza do Guincho","t":"Hotel/restaurante Michelin","m":"—","tel":"—","email":"info@fortalezadoguincho.com","p":"Alta","tCliente":"Hotel/restaurante Michelin","gancho":"Lead identificado pela equipa comercial — Hotel/restaurante Michelin em Estrada do Guincho.","zona":"Estrada do Guincho"},
  {"n":"Fuji Sushi & Steak","t":"Restaurante internacional","m":"—","tel":"—","email":"TheFork","p":"Alta","tCliente":"Restaurante internacional","gancho":"Lead identificado pela equipa comercial — Restaurante internacional em Cascais.","zona":"Cascais"},
  {"n":"Furnas do Guincho","t":"Restaurante premium","m":"—","tel":"—","email":"geral@furnasdoguincho.pt","p":"Alta","tCliente":"Restaurante premium","gancho":"Lead identificado pela equipa comercial — Restaurante premium em Cascais.","zona":"Cascais"},
  {"n":"Hífen","t":"Restaurante contemporâneo","m":"—","tel":"—","email":"reservashifen@gmail.com","p":"Alta","tCliente":"Restaurante contemporâneo","gancho":"Lead identificado pela equipa comercial — Restaurante contemporâneo em Marina de Cascais.","zona":"Marina de Cascais"},
  {"n":"House of Wonders","t":"Restaurante / brunch","m":"—","tel":"—","email":"info@houseofwonders.pt","p":"Alta","tCliente":"Restaurante / brunch","gancho":"Lead identificado pela equipa comercial — Restaurante / brunch em Rua da Misericórdia, Cascais.","zona":"Rua da Misericórdia, Cascais"},
  {"n":"Marisco na Praça","t":"Restaurante","m":"—","tel":"—","email":"geral@marisconapraca.pt","p":"Alta","tCliente":"Restaurante","gancho":"Lead identificado pela equipa comercial — Restaurante em Mercado da Vila Cascais.","zona":"Mercado da Vila Cascais"},
  {"n":"Monte Mar","t":"Restaurante premium","m":"—","tel":"—","email":"montemar@montemar.pt","p":"Alta","tCliente":"Restaurante premium","gancho":"Lead identificado pela equipa comercial — Restaurante premium em Cascais.","zona":"Cascais"},
  {"n":"Monte Mar Cascais","t":"Restaurante premium","m":"—","tel":"—","email":"geral@montemar.pt","p":"Alta","tCliente":"Restaurante premium","gancho":"Lead identificado pela equipa comercial — Restaurante premium em Boca do Inferno, Cascais.","zona":"Boca do Inferno, Cascais"},
  {"n":"Moinho Dom Quixote","t":"Restaurante panorâmico","m":"—","tel":"—","email":"info@moinhodomquixote.pt","p":"Alta","tCliente":"Restaurante panorâmico","gancho":"Lead identificado pela equipa comercial — Restaurante panorâmico em Colares.","zona":"Colares"},
  {"n":"A Toca do Júlio","t":"Restaurante tradicional","m":"—","tel":"—","email":"geral@atocadojulio.pt","p":"Alta","tCliente":"Restaurante tradicional","gancho":"Lead identificado pela equipa comercial — Restaurante tradicional em Colares.","zona":"Colares"},
  {"n":"Restaurante Azenhas do Mar","t":"Restaurante premium","m":"—","tel":"—","email":"reservas@azenhasdomar.pt","p":"Alta","tCliente":"Restaurante premium","gancho":"Lead identificado pela equipa comercial — Restaurante premium em Azenhas do Mar.","zona":"Azenhas do Mar"},
  {"n":"Água e Sal","t":"Restaurante contemporâneo","m":"—","tel":"—","email":"geral@aguaesal.pt","p":"Alta","tCliente":"Restaurante contemporâneo","gancho":"Lead identificado pela equipa comercial — Restaurante contemporâneo em Azenhas do Mar.","zona":"Azenhas do Mar"},
  {"n":"Adega das Azenhas","t":"Restaurante tradicional","m":"—","tel":"—","email":"geral@adegadasazenhas.pt","p":"Alta","tCliente":"Restaurante tradicional","gancho":"Lead identificado pela equipa comercial — Restaurante tradicional em Azenhas do Mar.","zona":"Azenhas do Mar"},
  {"n":"Terrace Restaurant","t":"Restaurante/hotel","m":"—","tel":"—","email":"reservas@azenhasdohotel.pt","p":"Alta","tCliente":"Restaurante/hotel","gancho":"Lead identificado pela equipa comercial — Restaurante/hotel em Azenhas do Mar.","zona":"Azenhas do Mar"},
  {"n":"Kailua Fonte da Telha","t":"Beach club","m":"—","tel":"—","email":"geral@kailua.pt","p":"Alta","tCliente":"Beach club","gancho":"Lead identificado pela equipa comercial — Beach club em Fonte da Telha.","zona":"Fonte da Telha"},
  {"n":"Oh! Vargas","t":"Restaurante contemporâneo","m":"—","tel":"—","email":"geral@ohvargas.pt","p":"Alta","tCliente":"Restaurante contemporâneo","gancho":"Lead identificado pela equipa comercial — Restaurante contemporâneo em Santarém Centro.","zona":"Santarém Centro"},
  {"n":"Taberna Ó Balcão","t":"Restaurante premium/tradicional","m":"—","tel":"—","email":"reservas@obalcao.pt","p":"Alta","tCliente":"Restaurante premium/tradicional","gancho":"Lead identificado pela equipa comercial — Restaurante premium/tradicional em Santarém.","zona":"Santarém"},
  {"n":"Tascá","t":"Restaurante moderno","m":"—","tel":"—","email":"geral@tasca.pt","p":"Alta","tCliente":"Restaurante moderno","gancho":"Lead identificado pela equipa comercial — Restaurante moderno em Santarém.","zona":"Santarém"},
  {"n":"O Quinzena","t":"Restaurante tradicional","m":"—","tel":"—","email":"geral@oquinzena.pt","p":"Alta","tCliente":"Restaurante tradicional","gancho":"Lead identificado pela equipa comercial — Restaurante tradicional em Santarém.","zona":"Santarém"},
  {"n":"Restaurante São Domingos","t":"Restaurante eventos","m":"—","tel":"—","email":"geral@saodomingos.pt","p":"Alta","tCliente":"Restaurante eventos","gancho":"Lead identificado pela equipa comercial — Restaurante eventos em Santarém.","zona":"Santarém"},
  {"n":"Dom Papinhas","t":"Restaurante familiar","m":"—","tel":"—","email":"geral@dompapinhas.pt","p":"Alta","tCliente":"Restaurante familiar","gancho":"Lead identificado pela equipa comercial — Restaurante familiar em Santarém.","zona":"Santarém"},
  {"n":"Adega do Avô","t":"Restaurante regional","m":"—","tel":"—","email":"geral@adegadoavo.pt","p":"Alta","tCliente":"Restaurante regional","gancho":"Lead identificado pela equipa comercial — Restaurante regional em Santarém.","zona":"Santarém"},
  {"n":"Pátio da Graça","t":"Restaurante contemporâneo","m":"—","tel":"—","email":"reservas@patiodagraca.pt","p":"Alta","tCliente":"Restaurante contemporâneo","gancho":"Lead identificado pela equipa comercial — Restaurante contemporâneo em Santarém.","zona":"Santarém"},
  {"n":"Cantinho do Avillez Porto","t":"Restaurante premium","m":"—","tel":"—","email":"porto@cantinhodoavillez.pt","p":"Alta","tCliente":"Restaurante premium","gancho":"Lead identificado pela equipa comercial — Restaurante premium em Rua Mouzinho da Silveira.","zona":"Rua Mouzinho da Silveira"},
  {"n":"Taberna dos Mercadores","t":"Restaurante turístico","m":"—","tel":"—","email":"reservas@tabernadosmercadores.pt","p":"Alta","tCliente":"Restaurante turístico","gancho":"Lead identificado pela equipa comercial — Restaurante turístico em Ribeira.","zona":"Ribeira"},
  {"n":"Adega São Nicolau","t":"Restaurante tradicional","m":"—","tel":"—","email":"geral@adegasaonicolau.pt","p":"Alta","tCliente":"Restaurante tradicional","gancho":"Lead identificado pela equipa comercial — Restaurante tradicional em Ribeira.","zona":"Ribeira"},
  {"n":"Caféína","t":"Restaurante contemporâneo","m":"—","tel":"—","email":"geral@cafeina.pt","p":"Alta","tCliente":"Restaurante contemporâneo","gancho":"Lead identificado pela equipa comercial — Restaurante contemporâneo em Foz do Douro.","zona":"Foz do Douro"},
  {"n":"Wish Restaurante","t":"Restaurante premium","m":"—","tel":"—","email":"reservas@wish.pt","p":"Alta","tCliente":"Restaurante premium","gancho":"Lead identificado pela equipa comercial — Restaurante premium em Boavista.","zona":"Boavista"},
  {"n":"Vinum","t":"Restaurante premium vínico","m":"—","tel":"—","email":"reservas@vinumatgrahams.com","p":"Alta","tCliente":"Restaurante premium vínico","gancho":"Lead identificado pela equipa comercial — Restaurante premium vínico em Caves Graham’s.","zona":"Caves Graham’s"},
  {"n":"Stramuntana","t":"Restaurante premium","m":"—","tel":"—","email":"reservas@stramuntana.pt","p":"Alta","tCliente":"Restaurante premium","gancho":"Lead identificado pela equipa comercial — Restaurante premium em Canidelo.","zona":"Canidelo"},
  {"n":"Muchaxo Restaurant","t":"Hotel/restaurante","m":"—","tel":"—","email":"recepcao@muchaxo.com","p":"Alta","tCliente":"Hotel/restaurante","gancho":"Lead identificado pela equipa comercial — Hotel/restaurante em Praia do Guincho.","zona":"Praia do Guincho"},
  {"n":"O Faroleiro","t":"Restaurante português costeiro","m":"—","tel":"—","email":"info@faroleiro.com","p":"Alta","tCliente":"Restaurante português costeiro","gancho":"Lead identificado pela equipa comercial — Restaurante português costeiro em Guincho.","zona":"Guincho"},
  {"n":"Palm Tree Pub & Restaurant","t":"Restaurante/bar","m":"—","tel":"—","email":"info@palmtree.pt","p":"Alta","tCliente":"Restaurante/bar","gancho":"Lead identificado pela equipa comercial — Restaurante/bar em Cascais.","zona":"Cascais"},
  {"n":"Panorama Beach Club","t":"Beach Club / Restaurante","m":"—","tel":"—","email":"panoramaguincho@villacolletion.pt","p":"Alta","tCliente":"Beach Club / Restaurante","gancho":"Lead identificado pela equipa comercial — Beach Club / Restaurante em Guincho.","zona":"Guincho"},
  {"n":"Pigalle Smash Bistro","t":"Bistro americano","m":"—","tel":"—","email":"TheFork","p":"Alta","tCliente":"Bistro americano","gancho":"Lead identificado pela equipa comercial — Bistro americano em Cascais.","zona":"Cascais"},
  {"n":"Porto Santa Maria","t":"Restaurante peixe/marisco premium","m":"—","tel":"—","email":"reservas@portosantamaria.com · 214 879 450","p":"Alta","tCliente":"Restaurante peixe/marisco premium","gancho":"Lead identificado pela equipa comercial — Restaurante peixe/marisco premium em Guincho.","zona":"Guincho"},
  {"n":"Restaurante Nortada","t":"Restaurante","m":"—","tel":"—","email":"geral@restaurantenortada.com","p":"Alta","tCliente":"Restaurante","gancho":"Lead identificado pela equipa comercial — Restaurante em Praia Grande, Sintra.","zona":"Praia Grande, Sintra"},
  {"n":"Santini Cascais","t":"Gelataria premium","m":"—","tel":"—","email":"cascais@santini.pt","p":"Alta","tCliente":"Gelataria premium","gancho":"Lead identificado pela equipa comercial — Gelataria premium em Baía de Cascais.","zona":"Baía de Cascais"},
  {"n":"Spot by Fortaleza do Guincho","t":"Restaurante (Michelin)","m":"—","tel":"—","email":"restaurante@guinchotel.pt","p":"Alta","tCliente":"Restaurante (Michelin)","gancho":"Lead identificado pela equipa comercial — Restaurante (Michelin) em Guincho.","zona":"Guincho"},
  {"n":"Visconde da Luz","t":"Restaurante tradicional","m":"—","tel":"—","email":"geral@viscondedaluz.pt","p":"Alta","tCliente":"Restaurante tradicional","gancho":"Lead identificado pela equipa comercial — Restaurante tradicional em Cascais Centro.","zona":"Cascais Centro"},
  {"n":"Restaurante Azenhas do Mar","t":"Restaurante marisco/peixe com vista mar","m":"—","tel":"—","email":"Google Maps / site próprio","p":"Alta","tCliente":"Restaurante marisco/peixe com vista mar","gancho":"Lead identificado pela equipa comercial — Restaurante marisco/peixe com vista mar em Azenhas do Mar.","zona":"Azenhas do Mar"},
  {"n":"Opíparo","t":"Restaurante cozinha portuguesa reinventada","m":"—","tel":"—","email":"Google Maps","p":"Alta","tCliente":"Restaurante cozinha portuguesa reinventada","gancho":"Lead identificado pela equipa comercial — Restaurante cozinha portuguesa reinventada em Azenhas do Mar.","zona":"Azenhas do Mar"},
  {"n":"Hamburgueria do Maçãs","t":"Restaurante informal","m":"—","tel":"—","email":"Google Maps","p":"Alta","tCliente":"Restaurante informal","gancho":"Lead identificado pela equipa comercial — Restaurante informal em Azenhas do Mar.","zona":"Azenhas do Mar"},
  {"n":"Cervejaria Boa Esperança","t":"Cervejaria","m":"Avenida Gomes Pereira 3 A Loja","tel":"—","email":"boaesperanca3a@gmail.com","p":"Alta","tCliente":"Cervejaria Lisboa","gancho":"Grande volume de refeições — pão de ló como sobremesa clássica portuguesa de fecho, sem complexidade operacional.","zona":"Lisboa"},
  {"n":"Browers Beato","t":"Cervejaria","m":"Tv. Grilo 1","tel":"—","email":"beato@thebrowerscompany.com","p":"Alta","tCliente":"Cervejaria Lisboa","gancho":"Grande volume de refeições — pão de ló como sobremesa clássica portuguesa de fecho, sem complexidade operacional.","zona":"Lisboa"},
  {"n":"Dote - Cervejaria Moderna","t":"Cervejaria","m":"Alvalade, Av. Republica, Parque Nações, Barata Salgueiro, Odivelas","tel":"—","email":"dote@dote.pt","p":"Alta","tCliente":"Cervejaria Lisboa","gancho":"Grande volume de refeições — pão de ló como sobremesa clássica portuguesa de fecho, sem complexidade operacional.","zona":"Lisboa"},
  {"n":"Edmundo","t":"Cervejaria","m":"Avenida Gomes Pereira, 1 - Estrada de Benfica","tel":"—","email":"restaurante.edmundo@gmail.com","p":"Alta","tCliente":"Cervejaria Lisboa","gancho":"Grande volume de refeições — pão de ló como sobremesa clássica portuguesa de fecho, sem complexidade operacional.","zona":"Lisboa"},
  {"n":"Farol","t":"Cervejaria","m":"Saldanha","tel":"—","email":"tcardoso1990@gmail.com","p":"Alta","tCliente":"Cervejaria Lisboa","gancho":"Grande volume de refeições — pão de ló como sobremesa clássica portuguesa de fecho, sem complexidade operacional.","zona":"Lisboa"},
  {"n":"Gambrinus","t":"Cervejaria","m":"Rua das Portas de Santo Antão, nº 23","tel":"—","email":"info@gambrinuslisboa.com","p":"Alta","tCliente":"Cervejaria Lisboa","gancho":"Grande volume de refeições — pão de ló como sobremesa clássica portuguesa de fecho, sem complexidade operacional.","zona":"Lisboa"},
  {"n":"Cervejaria Liberdade","t":"Cervejaria","m":"Rua Castilho, 14","tel":"—","email":"info@avliberdade.com; cervejaria.avenidaliberdade@tivoli-hotels.com","p":"Alta","tCliente":"Cervejaria Lisboa","gancho":"Grande volume de refeições — pão de ló como sobremesa clássica portuguesa de fecho, sem complexidade operacional.","zona":"Lisboa"},
  {"n":"Pinóquio","t":"Cervejaria","m":"Praça dos Restauradores 79 80","tel":"—","email":"geral@restaurantepinoquio.pt","p":"Alta","tCliente":"Cervejaria Lisboa","gancho":"Grande volume de refeições — pão de ló como sobremesa clássica portuguesa de fecho, sem complexidade operacional.","zona":"Lisboa"},
  {"n":"Portugália","t":"Cervejaria","m":"Almirante Reis","tel":"—","email":"pareis@portugalia.pt","p":"Alta","tCliente":"Cervejaria Lisboa","gancho":"Grande volume de refeições — pão de ló como sobremesa clássica portuguesa de fecho, sem complexidade operacional.","zona":"Lisboa"},
  {"n":"Cervejaria Ramiro","t":"Cervejaria","m":"Av. Alm. Reis 1 H","tel":"—","email":"ramiro@cervejariaramiro.pt","p":"Alta","tCliente":"Cervejaria Lisboa","gancho":"Grande volume de refeições — pão de ló como sobremesa clássica portuguesa de fecho, sem complexidade operacional.","zona":"Lisboa"},
  {"n":"Cervejaria Sem Vergonha","t":"Cervejaria","m":"Tv. de Santa Quitéria 38 D","tel":"—","email":"cervejariasemvergonha@gmail.com","p":"Alta","tCliente":"Cervejaria Lisboa","gancho":"Grande volume de refeições — pão de ló como sobremesa clássica portuguesa de fecho, sem complexidade operacional.","zona":"Lisboa"},
  {"n":"Cervejaria Trindade","t":"Cervejaria","m":"Rua Nova da Trindade, nº 20 C","tel":"—","email":"ct.chiado@cervejariatrindade.pt","p":"Alta","tCliente":"Cervejaria Lisboa","gancho":"Grande volume de refeições — pão de ló como sobremesa clássica portuguesa de fecho, sem complexidade operacional.","zona":"Lisboa"},
]

DB_JOAO = [
  {"n":"Lupis Coffee Shop","t":"Café de Especialidade & Brunch","m":"Rua de Marvila 42, 1950-197 Lisboa","tel":"—","email":"—","p":"Alta","tCliente":"Café de Especialidade & Brunch","gancho":"Público artístico e de design de Marvila. O aspeto rústico e autêntico do pão de ló Ti Piedade tem enorme apelo visual.","zona":"Lisboa (Marvila)"},
  {"n":"Restaurante Dinastia Tang","t":"Restaurante (Médio/Alto Padrão)","m":"Rua do Açúcar 107, 1950-006 Lisboa","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão)","gancho":"Encontro cultural. Embora especializado, procuram sobremesas portuguesas tradicionais de altíssima qualidade para os clientes locais.","zona":"Lisboa (Marvila)"},
  {"n":"Pão do Beco","t":"Padaria Artesanal","m":"Rua Capitão Leitão 72A, 2800-135 Almada","tel":"—","email":"—","p":"Alta","tCliente":"Padaria Artesanal","gancho":"Foco na autenticidade da fermentação natural. O Ti Piedade como o pão de ló sem químicos industriais.","zona":"Almada"},
  {"n":"Under The Cover / O Pão de Ló","t":"Café & Brunch Moderno","m":"Rua Capitão Leitão 54, 2800-135 Almada","tel":"—","email":"—","p":"Alta","tCliente":"Café & Brunch Moderno","gancho":"Cafetaria moderna com foco em lanches e pastelaria cuidada. Excelente para fatias e meias-fatias.","zona":"Almada"},
  {"n":"Mundet Factory","t":"Restaurante & Brunch Premium","m":"Avenida Metalúrgica Augusto de Castro 1, 2840-515 Seixal","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante & Brunch Premium","gancho":"Restaurante moderno, muito movimentado. Foco no controlo estrito de quebras e custos de stock via formato congelado.","zona":"Seixal"},
  {"n":"Let's Coffe","t":"Café de Especialidade","m":"Praça dos Mártires da Pátria 12, 2840-496 Seixal","tel":"—","email":"—","p":"Alta","tCliente":"Café de Especialidade","gancho":"Café gourmet junto à baía do Seixal. Sobremesa tradicional portuguesa em harmonia com café premium.","zona":"Seixal"},
  {"n":"Pão com Alma","t":"Padaria Artesanal","m":"Rua Miguel Bombarda 124, 2830-355 Barreiro","tel":"—","email":"—","p":"Alta","tCliente":"Padaria Artesanal","gancho":"Identidade e tradição no fabrico. O Pão de Ló Ti Piedade partilha dos mesmos princípios artesanais.","zona":"Barreiro"},
  {"n":"Restaurante Taberna do Manel","t":"Restaurante Tradicional Premium","m":"Rua Conselheiro Joaquim António d'Aguiar 32, 2830-334 Barreiro","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante Tradicional Premium","gancho":"Clássico da gastronomia do Barreiro. Sobremesa de altíssima consistência, sem preocupações com validade curta de pastelaria diária.","zona":"Barreiro"},
  {"n":"Dr. Bernard","t":"Restaurante de Praia & Brunch","m":"Avenida General Humberto Delgado, Praia do Tarquínio-Paraíso, 2825-366 Costa da Caparica","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante de Praia & Brunch","gancho":"Público moderno e internacional. Fatia de pão de ló servida com coberturas de frutos ou gelado de limão no brunch.","zona":"Costa da Caparica"},
  {"n":"Palms","t":"Restaurante de Praia & Brunch Premium","m":"Praia do CDS, Apoio 9, Avenida General Humberto Delgado, 2825-366 Costa da Caparica","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante de Praia & Brunch Premium","gancho":"Público cosmopolita. Apresentar o Pão de Ló como o autêntico bolo tradicional português para brunch ou lanche na praia.","zona":"Costa da Caparica"},
  {"n":"Restaurante O Sentido do Mar","t":"Restaurante de Peixe (Médio/Alto Padrão)","m":"Praia do Tarquínio, Apoio de Praia 4, 2825-366 Costa da Caparica","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante de Peixe (Médio/Alto Padrão)","gancho":"Restaurante sofisticado com vista para o mar. Sobremesa tradicional premium pronta a fatiar, de custo controlado e quebra zero.","zona":"Costa da Caparica"},
  {"n":"Restaurante Carolina do Aires","t":"Restaurante Tradicional Premium","m":"Avenida General Humberto Delgado 10, 2825-366 Costa da Caparica","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante Tradicional Premium","gancho":"Restaurante clássico com enorme tradição de peixe. Sobremesa portuguesa tradicional que assegura fidelidade ao sabor.","zona":"Costa da Caparica"},
  {"n":"Ponto Final","t":"Restaurante (Médio/Alto Padrão)","m":"Rua do Ginjal 72, 2800-285 Almada","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão)","gancho":"Vistas incríveis, alta rotação de estrangeiros e locais. Sobremesa ultra-rápida na mesa sem desperdício de stock.","zona":"Almada (Cacilhas)"},
  {"n":"Atira-te ao Rio","t":"Restaurante (Médio/Alto Padrão)","m":"Rua do Ginjal 69, 2800-285 Almada","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão)","gancho":"Cozinha requintada portuguesa e contemporânea. Foco no sabor rústico e excelente aspeto da fatia no prato.","zona":"Almada (Cacilhas)"},
  {"n":"Restaurante Cabrinha","t":"Restaurante Tradicional Premium","m":"Rua de Cacilhas 33, 2800-135 Almada","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante Tradicional Premium","gancho":"Especialistas em peixe grelhado e marisco. Fecho de refeição com pão de ló húmido tradicional, com custo fixo por dose.","zona":"Almada"},
  {"n":"Amarra ao Cais","t":"Restaurante (Médio/Alto Padrão)","m":"Jardim do Rio, Cais do Ginjal, 2800-285 Almada","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão)","gancho":"Localização premium. Sobremesas rústicas rápidas e fáceis de marginar mais de 65% na carta.","zona":"Almada"},
  {"n":"O Pescador","t":"Restaurante Tradicional Premium","m":"Avenida D. Carlos I 12, 2840-515 Seixal","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante Tradicional Premium","gancho":"Foco na clientela local de domingo. Pão de ló fatiado na hora com um fio de doce de ovos (receita fácil de assinar).","zona":"Seixal"},
  {"n":"O Farribas","t":"Restaurante (Médio/Alto Padrão)","m":"Avenida Luísa Todi 244, 2900-452 Setúbal","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão)","gancho":"Na principal avenida. Doçaria portuguesa autêntica fornecida congelada para controlo total de quebras pós-refeição.","zona":"Setúbal"},
  {"n":"Restaurante Novo 10","t":"Restaurante Tradicional Premium","m":"Avenida Luísa Todi 220, 2900-452 Setúbal","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante Tradicional Premium","gancho":"Especialistas em peixe assado. O Pão de Ló Ti Piedade como a sobremesa ideal de alto valor percebido pelos clientes.","zona":"Setúbal"},
  {"n":"Casa do Peixe (Setúbal)","t":"Restaurante (Médio/Alto Padrão)","m":"Rua Praia da Saúde 10, 2900-572 Setúbal","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão)","gancho":"Restaurante à beira d'água. Oferecer sobremesa de pão de ló com consistência garantida e zero perdas semanais.","zona":"Setúbal"},
  {"n":"Casa das Tortas de Azeitão","t":"Pastelaria Gourmet & Chá","m":"Praça da República 1, 2925-520 Azeitão","tel":"—","email":"—","p":"Alta","tCliente":"Pastelaria Gourmet & Chá","gancho":"Ponto de passagem turístico. O Pão de Ló Ti Piedade junta-se à oferta de doçaria regional como a opção fofa sem cremes.","zona":"Azeitão"},
  {"n":"Garrafeira de Azeitão","t":"Garrafeira & Gourmet Deli","m":"Avenida 25 de Abril 12, 2925-501 Azeitão","tel":"—","email":"—","p":"Alta","tCliente":"Garrafeira & Gourmet Deli","gancho":"Venda casada com Moscatel de Setúbal. O Pão de Ló é o casamento perfeito para degustações premium na garrafeira.","zona":"Azeitão"},
  {"n":"O Quintal","t":"Restaurante (Médio/Alto Padrão)","m":"Rua José Augusto Coelho 82, 2925-538 Azeitão","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão)","gancho":"Cozinha tradicional refinada. Foco na margem absoluta que o produto congelado traz, com 0% de quebra alimentar.","zona":"Azeitão"},
  {"n":"Restaurante Ribeirinha do Sado","t":"Restaurante Tradicional Premium","m":"Avenida Luísa Todi 34, 2900-450 Setúbal","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante Tradicional Premium","gancho":"Paragem clássica para peixe. Sobremesa tradicional pronta a fatiar, de custo controlado e quebra zero.","zona":"Setúbal"},
  {"n":"O Alfeite","t":"Restaurante Tradicional Premium","m":"Rua de Serpa Pinto 4, 2800-205 Almada","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante Tradicional Premium","gancho":"Restaurante típico de almoços executivos. Solução congelada garante quebras zero em dias de menor afluência.","zona":"Almada"},
  {"n":"Lisboa à Vista","t":"Restaurante (Médio/Alto Padrão)","m":"Av. Metalúrgica Augusto de Castro, Baía do Seixal, 2840-515 Seixal","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão)","gancho":"Restaurante num barco histórico. Doçaria portuguesa tradicional com aspeto rústico fantástico para acompanhar vinhos generosos.","zona":"Seixal"},
  {"n":"Heim Cafe","t":"Brunch & Lanches","m":"Rua de Santos-O-Velho 4, 1200-109 Lisboa","tel":"—","email":"—","p":"Alta","tCliente":"Brunch & Lanches","gancho":"O queridinho do brunch. Inserir fatia de Pão de Ló tostada com toppings no menu diário.","zona":"Lisboa (Campo de Ourique)"},
  {"n":"Dear Breakfast (Saldanha)","t":"Brunch Club & Café","m":"Rua Eng. Vieira da Silva 8, 1050-105 Lisboa","tel":"—","email":"—","p":"Alta","tCliente":"Brunch Club & Café","gancho":"Brunch executivo e sofisticado. Pão de ló Ti Piedade como opção portuguesa no menu.","zona":"Lisboa (Saldanha)"},
  {"n":"Cajú","t":"Brunch & Healthy Food","m":"Rua da Alegria 73, 1250-182 Lisboa","tel":"—","email":"—","p":"Alta","tCliente":"Brunch & Healthy Food","gancho":"Público internacional. Oferecer a experiência do autêntico pão de ló tradicional fofo.","zona":"Lisboa (Príncipe Real)"},
  {"n":"Isco (Alvalade)","t":"Padaria Artesanal & Bistrô","m":"Rua do Arco do Carvalhão 244, 1350-026 Lisboa","tel":"—","email":"—","p":"Alta","tCliente":"Padaria Artesanal & Bistrô","gancho":"Artesanal puro. Pão de ló premium poupa-lhes tempo de pastelaria mantendo a excelência.","zona":"Lisboa (Alvalade)"},
  {"n":"Borda D`Agua","t":"Restaurante Praia","m":"—","tel":"—","email":"geral@bordadagua.com.pt","p":"Alta","tCliente":"Restaurante Praia","gancho":"Lead identificado pela equipa comercial — Restaurante Praia em Costa da Caparica.","zona":"Costa da Caparica"},
  {"n":"Gordos Caparica","t":"Restaurante","m":"—","tel":"—","email":"gordos.caparica@gmail.com","p":"Alta","tCliente":"Restaurante","gancho":"Lead identificado pela equipa comercial — Restaurante em Costa da Caparica.","zona":"Costa da Caparica"},
  {"n":"Restaurante Leblon","t":"Restaurante","m":"—","tel":"—","email":"leblonsaojoao@gmail.com","p":"Alta","tCliente":"Restaurante","gancho":"Lead identificado pela equipa comercial — Restaurante em Costa da Caparica.","zona":"Costa da Caparica"},
  {"n":"Coco Beach","t":"Restaurante","m":"—","tel":"—","email":"Cocobeach.caparica@gmail.com","p":"Alta","tCliente":"Restaurante","gancho":"Lead identificado pela equipa comercial — Restaurante em Costa da Caparica.","zona":"Costa da Caparica"},
  {"n":"Praia - Sea, Salt & Pepper","t":"Restaurante","m":"—","tel":"—","email":"reservas@apraia.pt","p":"Alta","tCliente":"Restaurante","gancho":"Lead identificado pela equipa comercial — Restaurante em Costa da Caparica.","zona":"Costa da Caparica"},
  {"n":"Fortuna Tacos & Burger","t":"Restaurante","m":"—","tel":"—","email":"fortuna.tacos@gmail.com","p":"Alta","tCliente":"Restaurante","gancho":"Lead identificado pela equipa comercial — Restaurante em Costa da Caparica.","zona":"Costa da Caparica"},
  {"n":"Zama Beach Club","t":"Restaurante","m":"—","tel":"—","email":"zama.costadacaparica@gmail.com","p":"Alta","tCliente":"Restaurante","gancho":"Lead identificado pela equipa comercial — Restaurante em Costa da Caparica.","zona":"Costa da Caparica"},
  {"n":"O Xéxéxé","t":"Restaurante","m":"—","tel":"—","email":"oxexexedacosta@gmail.com","p":"Alta","tCliente":"Restaurante","gancho":"Lead identificado pela equipa comercial — Restaurante em Costa da Caparica.","zona":"Costa da Caparica"},
  {"n":"AHOY Surf & Snacks","t":"Restaurante","m":"—","tel":"—","email":"Ahoycafe@hotmail.com","p":"Alta","tCliente":"Restaurante","gancho":"Lead identificado pela equipa comercial — Restaurante em Costa da Caparica.","zona":"Costa da Caparica"},
  {"n":"Clássico Beach Bar by Olivier","t":"Restaurante","m":"—","tel":"—","email":"classicobeachbar@olivier.pt","p":"Alta","tCliente":"Restaurante","gancho":"Lead identificado pela equipa comercial — Restaurante em Costa da Caparica.","zona":"Costa da Caparica"},
  {"n":"O Barbas","t":"Restaurante","m":"—","tel":"—","email":"restauranteobarbas@gmail.com","p":"Alta","tCliente":"Restaurante","gancho":"Lead identificado pela equipa comercial — Restaurante em Costa da Caparica.","zona":"Costa da Caparica"},
  {"n":"Muse Brunch Café & Wine Bar","t":"Café & Wine Bar","m":"—","tel":"—","email":"info@musewinebar.pt","p":"Alta","tCliente":"Café & Wine Bar","gancho":"Lead identificado pela equipa comercial — Café & Wine Bar em Costa da Caparica.","zona":"Costa da Caparica"},
  {"n":"O Mercado","t":"Restaurante tradicional","m":"—","tel":"—","email":"omercadocc@gmail.com","p":"Alta","tCliente":"Restaurante tradicional","gancho":"Lead identificado pela equipa comercial — Restaurante tradicional em Costa da Caparica.","zona":"Costa da Caparica"},
  {"n":"Apeixonado","t":"Restaurante tradicional","m":"—","tel":"—","email":"apeixonado@ponteslda.com","p":"Alta","tCliente":"Restaurante tradicional","gancho":"Lead identificado pela equipa comercial — Restaurante tradicional em Costa da Caparica.","zona":"Costa da Caparica"},
  {"n":"Ponto Final","t":"Restaurante turístico","m":"—","tel":"—","email":"reservas@pontofinal.pt","p":"Alta","tCliente":"Restaurante turístico","gancho":"Lead identificado pela equipa comercial — Restaurante turístico em Cacilhas.","zona":"Cacilhas"},
  {"n":"Atira-te ao Rio","t":"Restaurante premium","m":"—","tel":"—","email":"geral@atirateaorio.pt","p":"Alta","tCliente":"Restaurante premium","gancho":"Lead identificado pela equipa comercial — Restaurante premium em Cacilhas.","zona":"Cacilhas"},
  {"n":"Solar Beirão","t":"Restaurante tradicional","m":"—","tel":"—","email":"geral@solarbeirao.pt","p":"Alta","tCliente":"Restaurante tradicional","gancho":"Lead identificado pela equipa comercial — Restaurante tradicional em Almada.","zona":"Almada"},
  {"n":"O Martinho","t":"Restaurante","m":"—","tel":"—","email":"geral@omartinho.pt","p":"Alta","tCliente":"Restaurante","gancho":"Lead identificado pela equipa comercial — Restaurante em Costa da Caparica.","zona":"Costa da Caparica"},
  {"n":"Praia da Princesa","t":"Beach restaurant","m":"—","tel":"—","email":"reservas@praiadaprincesa.pt","p":"Alta","tCliente":"Beach restaurant","gancho":"Lead identificado pela equipa comercial — Beach restaurant em Costa da Caparica.","zona":"Costa da Caparica"},
  {"n":"Dr. Bernard","t":"Restaurante/bar","m":"—","tel":"—","email":"geral@drbernard.pt","p":"Alta","tCliente":"Restaurante/bar","gancho":"Lead identificado pela equipa comercial — Restaurante/bar em Costa da Caparica.","zona":"Costa da Caparica"},
  {"n":"Sentido do Mar","t":"Restaurante","m":"—","tel":"—","email":"reservas@sentidodomar.pt","p":"Alta","tCliente":"Restaurante","gancho":"Lead identificado pela equipa comercial — Restaurante em Costa da Caparica.","zona":"Costa da Caparica"},
  {"n":"Atlas Leiria","t":"Restaurante/bar moderno","m":"—","tel":"—","email":"geral@atlasleiria.pt","p":"Alta","tCliente":"Restaurante/bar moderno","gancho":"Lead identificado pela equipa comercial — Restaurante/bar moderno em Leiria Centro.","zona":"Leiria Centro"},
  {"n":"Casinha Velha","t":"Restaurante premium","m":"—","tel":"—","email":"reservas@casinhavelha.pt","p":"Alta","tCliente":"Restaurante premium","gancho":"Lead identificado pela equipa comercial — Restaurante premium em Leiria.","zona":"Leiria"},
  {"n":"Mata Bicho","t":"Restaurante contemporâneo","m":"—","tel":"—","email":"geral@matabicho.pt","p":"Alta","tCliente":"Restaurante contemporâneo","gancho":"Lead identificado pela equipa comercial — Restaurante contemporâneo em Leiria.","zona":"Leiria"},
  {"n":"Cervejaria João Gordo","t":"Restaurante tradicional","m":"—","tel":"—","email":"geral@joaogordo.pt","p":"Alta","tCliente":"Restaurante tradicional","gancho":"Lead identificado pela equipa comercial — Restaurante tradicional em Leiria.","zona":"Leiria"},
  {"n":"Tromba Rija","t":"Restaurante grande volume","m":"—","tel":"—","email":"reservas@trombarija.pt","p":"Alta","tCliente":"Restaurante grande volume","gancho":"Lead identificado pela equipa comercial — Restaurante grande volume em Leiria.","zona":"Leiria"},
  {"n":"Malagueta Afrobar","t":"Restaurante moderno","m":"—","tel":"—","email":"geral@malagueta.pt","p":"Alta","tCliente":"Restaurante moderno","gancho":"Lead identificado pela equipa comercial — Restaurante moderno em Leiria.","zona":"Leiria"},
  {"n":"Taberna do Quinzena","t":"Restaurante português","m":"—","tel":"—","email":"geral@tabernadequinzena.pt","p":"Alta","tCliente":"Restaurante português","gancho":"Lead identificado pela equipa comercial — Restaurante português em Leiria.","zona":"Leiria"},
  {"n":"Tokio Sushi","t":"Restaurante moderno","m":"—","tel":"—","email":"reservas@tokiosushi.pt","p":"Alta","tCliente":"Restaurante moderno","gancho":"Lead identificado pela equipa comercial — Restaurante moderno em Leiria.","zona":"Leiria"},
  {"n":"Sauvage Gourmet","t":"Brunch / Gourmet (★ 4.8)","m":"—","tel":"—","email":"geral@sauvage-gourmet.pt","p":"Alta","tCliente":"Brunch / Gourmet (★ 4.8)","gancho":"Lead identificado pela equipa comercial — Brunch / Gourmet (★ 4.8) em Almada.","zona":"Almada"},
  {"n":"De Raiz no Museu (Chef Luís Calei)","t":"Fine dining / cozinha autoral","m":"—","tel":"—","email":"geral@de-raiz.pt","p":"Alta","tCliente":"Fine dining / cozinha autoral","gancho":"Lead identificado pela equipa comercial — Fine dining / cozinha autoral em Almada.","zona":"Almada"},
  {"n":"Soul Sushi – Japanese Fusion","t":"Restaurante fusão (★ 9.8 TheFork)","m":"—","tel":"—","email":"soulsushibar@gmail.com","p":"Alta","tCliente":"Restaurante fusão (★ 9.8 TheFork)","gancho":"Lead identificado pela equipa comercial — Restaurante fusão (★ 9.8 TheFork) em Almada.","zona":"Almada"},
  {"n":"Contrabando Mexican Food","t":"Restaurante temático","m":"—","tel":"—","email":"geral.contrabando@gmail.com","p":"Alta","tCliente":"Restaurante temático","gancho":"Lead identificado pela equipa comercial — Restaurante temático em Almada.","zona":"Almada"},
  {"n":"Pastelaria Condestável","t":"Pastelaria","m":"—","tel":"—","email":"pastcondestavel@gmail.com","p":"Alta","tCliente":"Pastelaria","gancho":"Lead identificado pela equipa comercial — Pastelaria em Almada.","zona":"Almada"},
]

DB_OSCAR = [
  {"n":"Oitoo","t":"Padaria Artesanal & Café","m":"Rua de Cedofeita 443, 4050-181 Porto","tel":"—","email":"—","p":"Alta","tCliente":"Padaria Artesanal & Café","gancho":"Foco no público moderno e artesanal. Pão de ló fofo para lanche premium.","zona":"Porto"},
  {"n":"Masseira","t":"Padaria Artesanal Sourdough","m":"Rua de D. Manuel II 296, 4050-343 Porto","tel":"—","email":"—","p":"Alta","tCliente":"Padaria Artesanal Sourdough","gancho":"Identidade rústica. Combina com a integridade do pão de ló artesanal.","zona":"Porto"},
  {"n":"Chá das Cinco","t":"Pastelaria Artesanal & Casa de Chá","m":"Praça da Alegria 50, 4000-027 Porto","tel":"—","email":"—","p":"Alta","tCliente":"Pastelaria Artesanal & Casa de Chá","gancho":"O spot de eleição para bolos à fatia no Porto. Propor alternativa sem creme e tradicional.","zona":"Porto"},
  {"n":"Lazy Breakfast Club","t":"Brunch Club & Café","m":"Rua das Oliveiras 110, 4050-449 Porto","tel":"—","email":"—","p":"Alta","tCliente":"Brunch Club & Café","gancho":"Inovação: sugerir fatia de pão de ló tostada com manteiga artesanal ou toppings doces.","zona":"Porto"},
  {"n":"Duquesa","t":"Brunch & Specialty Coffee","m":"Rua de Cedofeita 300, 4050-174 Porto","tel":"—","email":"—","p":"Alta","tCliente":"Brunch & Specialty Coffee","gancho":"Público jovem e turistas. Perfeito para introduzir uma fatia com café de especialidade.","zona":"Porto"},
  {"n":"Nola Kitchen","t":"Cafetaria Saudável & Brunch","m":"Praça de D. Filipa de Lencastre 25, 4050-259 Porto","tel":"—","email":"—","p":"Alta","tCliente":"Cafetaria Saudável & Brunch","gancho":"Argumentar a ausência de corantes e conservantes (ingredientes 100% limpos e tradicionais).","zona":"Porto"},
  {"n":"My Coffee Porto","t":"Café de Especialidade","m":"Escadas do Caminho Novo 11, 4050-554 Porto","tel":"—","email":"—","p":"Alta","tCliente":"Café de Especialidade","gancho":"Vista deslumbrante, turistas. O Pão de Ló é a 'fatia portuguesa' ideal para acompanhar o café.","zona":"Porto"},
  {"n":"Do Norte Cafe by Hungry Biker","t":"Brunch & Cozy Cafe","m":"Rua do Almada 516, 4050-039 Porto","tel":"—","email":"—","p":"Alta","tCliente":"Brunch & Cozy Cafe","gancho":"Ambiente rústico e acolhedor. Foco na porção individual do pão de ló para brunch.","zona":"Porto"},
  {"n":"Epoca Café","t":"Café de Especialidade & Lanches","m":"Rua do Rosário 22, 4050-522 Porto","tel":"—","email":"—","p":"Alta","tCliente":"Café de Especialidade & Lanches","gancho":"Foco no público local premium que valoriza simplicidade com imensa qualidade e frescura.","zona":"Porto"},
  {"n":"Leitaria da Quinta do Paço (Matosinhos)","t":"Pastelaria Premium & Lanches","m":"Rua Brito Capelo 1237, 4450-072 Matosinhos","tel":"—","email":"—","p":"Alta","tCliente":"Pastelaria Premium & Lanches","gancho":"Venda por impulso. Pão de ló fofo para lanches em família ao fim de semana.","zona":"Matosinhos"},
  {"n":"Adega de São Nicolau","t":"Restaurante (Médio/Alto Padrão)","m":"Rua de São Nicolau 1, 4050-561 Porto","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão)","gancho":"Clássico ribeirinho. Sobremesa tradicional fantástica para acompanhar Vinho do Porto. Quebra zero via congelado.","zona":"Porto"},
  {"n":"Brasão Cervejaria (Aliados)","t":"Restaurante (Médio/Alto Padrão)","m":"Rua de Ramalho Ortigão 28, 4000-407 Porto","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão)","gancho":"Grande volume. Controle total de custos de dose e facilidade na gestão de stocks congelados.","zona":"Porto"},
  {"n":"O Paparico","t":"Restaurante (Alta Cozinha Portuguesa)","m":"Rua de Costa Cabral 2343, 4200-232 Porto","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Alta Cozinha Portuguesa)","gancho":"História e tradição. O pão de ló Ti Piedade como exemplo da doçaria conventual autêntica.","zona":"Porto"},
  {"n":"Taberna dos Mercadores","t":"Restaurante (Médio/Alto Padrão)","m":"Rua dos Mercadores 36, 4050-373 Porto","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão)","gancho":"Espaço pequeno. Solução congelada economiza espaço de cozinha e evita quebras.","zona":"Porto"},
  {"n":"Cantinho do Avillez (Porto)","t":"Restaurante (Médio/Alto Padrão)","m":"Rua de Mouzinho da Silveira 166, 4050-416 Porto","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão)","gancho":"Cozinha portuguesa com assinatura. Apresentar consistência do nosso padrão premium.","zona":"Porto"},
  {"n":"Muu Steakhouse","t":"Restaurante (Médio/Alto Padrão)","m":"Rua do Almada 149, 4050-037 Porto","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão)","gancho":"Foco na sobremesa de conforto após as carnes premium. Servir quente com sorvete ácido.","zona":"Porto"},
  {"n":"Abadia do Porto","t":"Restaurante Tradicional Premium","m":"Rua do Ateneu Comercial do Porto 22, 4000-380 Porto","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante Tradicional Premium","gancho":"Público conservador que exige doçaria tradicional portuguesa à séria. Garantia de qualidade constante.","zona":"Porto"},
  {"n":"A Regaleira","t":"Restaurante Tradicional Premium","m":"Rua do Bonjardim 87, 4000-124 Porto","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante Tradicional Premium","gancho":"Berço da francesinha, mas forte em pratos tradicionais. Sobremesa tradicional rápida e lucrativa.","zona":"Porto"},
  {"n":"Semente - Padaria Artesanal","t":"Padaria Artesanal & Orgânica","m":"Rua de S. Vicente 112, 4710-062 Braga","tel":"—","email":"—","p":"Alta","tCliente":"Padaria Artesanal & Orgânica","gancho":"Público biológico. Foco na receita tradicional pura, sem conservantes ou aditivos industriais.","zona":"Braga"},
  {"n":"Ferreira Capa","t":"Pastelaria de Referência / Tradicional","m":"Rua do Souto 131, 4700-329 Braga","tel":"—","email":"—","p":"Alta","tCliente":"Pastelaria de Referência / Tradicional","gancho":"Ponto histórico na cidade. Oferecer como opção premium de pão de ló húmido embalado para famílias.","zona":"Braga"},
  {"n":"Koyo Specialty Coffee","t":"Café de Especialidade","m":"Rua Dom Diogo de Sousa 37, 4700-424 Braga","tel":"—","email":"—","p":"Alta","tCliente":"Café de Especialidade","gancho":"Harmonia perfeita com café de especialidade ácido. Uma fatia simples servida elegantemente.","zona":"Braga"},
  {"n":"Cozinha da Terra","t":"Restaurante (Médio/Alto Padrão)","m":"Rua de São Lourenço, 4705-551 Braga","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão)","gancho":"Restaurante rústico minhoto. Pão de ló fofo fatiado na dose certa. Solução congelada com zero perdas.","zona":"Braga"},
  {"n":"Retroaria Bracarense","t":"Café de Charme / Tradicional","m":"Rua de São Marcos 17, 4700-328 Braga","tel":"—","email":"—","p":"Alta","tCliente":"Café de Charme / Tradicional","gancho":"Atmosfera vintage portuguesa. Encaixa perfeitamente no menu de lanche e chás tradicionais.","zona":"Braga"},
  {"n":"Restaurante Cozinha da Sé","t":"Restaurante (Médio/Alto Padrão)","m":"Rua Dom Diogo de Sousa 3, 4700-424 Braga","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão)","gancho":"Junto à Sé. Público local e turistas exigentes. Doçaria tradicional pronta a servir com quebra zero.","zona":"Braga"},
  {"n":"Taberna Belga","t":"Restaurante Tradicional / Elevada Rotação","m":"Rua Cónego Rafael Alvares da Costa 19, 4715-176 Braga","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante Tradicional / Elevada Rotação","gancho":"Rotação altíssima. Necessitam de sobremesas rápidas, deliciosas e de custo controlado.","zona":"Braga"},
  {"n":"Restaurante Augusto","t":"Restaurante (Médio/Alto Padrão)","m":"Rua de São Miguel 20, 4700-305 Braga","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão)","gancho":"Excelente garrafeira. Propor o Pão de Ló em fatia para acompanhar a carta de vinhos doces.","zona":"Braga"},
  {"n":"Padaria da Esquina","t":"Padaria Artesanal & Cafetaria","m":"Largo do Toural 104, 4810-427 Guimarães","tel":"—","email":"—","p":"Alta","tCliente":"Padaria Artesanal & Cafetaria","gancho":"Ponto de altíssima visibilidade. Clientes com apetite por pão artesanal e pastelaria fina.","zona":"Guimarães"},
  {"n":"A Cozinha por António Loureiro","t":"Restaurante (Médio/Alto Padrão / Autor)","m":"Largo do Serralho 4, 4800-414 Guimarães","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão / Autor)","gancho":"Restaurante sustentável. Foco estratégico no desperdício zero que o formato congelado garante.","zona":"Guimarães"},
  {"n":"Histórico by Papaboa","t":"Restaurante (Médio/Alto Padrão)","m":"Rua de Val Donas 4, 4810-230 Guimarães","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão)","gancho":"Palacete histórico. Pão de ló Ti Piedade como a sobremesa ideal que honra a história de Portugal.","zona":"Guimarães"},
  {"n":"Buxa","t":"Restaurante (Médio/Alto Padrão)","m":"Praça de São Tiago 18, 4810-244 Guimarães","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante (Médio/Alto Padrão)","gancho":"Cozinha tradicional na praça mais movimentada. Rapidez de serviço sem risco de sobremesas estragadas.","zona":"Guimarães"},
  {"n":"Restaurante Nelson","t":"Restaurante peixe/marisco","m":"—","tel":"—","email":"geral@restaurantenelson.pt","p":"Alta","tCliente":"Restaurante peixe/marisco","gancho":"Lead identificado pela equipa comercial — Restaurante peixe/marisco em Ribamar.","zona":"Ribamar"},
  {"n":"A Sardinha","t":"Restaurante peixe","m":"—","tel":"—","email":"geral@asardinha.pt; restauranteasardinha@gmail.com","p":"Alta","tCliente":"Restaurante peixe","gancho":"Lead identificado pela equipa comercial — Restaurante peixe em Peniche.","zona":"Peniche"},
  {"n":"Golfinho Azul","t":"Restaurante","m":"—","tel":"—","email":"geral@golfinhoazul.pt","p":"Alta","tCliente":"Restaurante","gancho":"Lead identificado pela equipa comercial — Restaurante em Ribamar.","zona":"Ribamar"},
  {"n":"Foz Restaurante","t":"Restaurante contemporâneo","m":"—","tel":"—","email":"reservas@fozrestaurante.pt","p":"Alta","tCliente":"Restaurante contemporâneo","gancho":"Lead identificado pela equipa comercial — Restaurante contemporâneo em Ribamar.","zona":"Ribamar"},
  {"n":"Porto das Barcas","t":"Restaurante","m":"—","tel":"—","email":"geral@portodasbarcas.pt","p":"Alta","tCliente":"Restaurante","gancho":"Lead identificado pela equipa comercial — Restaurante em Ribamar.","zona":"Ribamar"},
  {"n":"Café Central Ribamar","t":"Restaurante/café","m":"—","tel":"—","email":"geral@cafecentralribamar.pt","p":"Alta","tCliente":"Restaurante/café","gancho":"Lead identificado pela equipa comercial — Restaurante/café em Ribamar.","zona":"Ribamar"},
  {"n":"Cantinho da Fonte","t":"Restaurante tradicional","m":"—","tel":"—","email":"geral@cantinhodafonte.pt","p":"Alta","tCliente":"Restaurante tradicional","gancho":"Lead identificado pela equipa comercial — Restaurante tradicional em Peniche.","zona":"Peniche"},
  {"n":"Estelas","t":"Restaurante / Bar","m":"—","tel":"—","email":"restaurante.estelas@sapo.pt","p":"Alta","tCliente":"Restaurante / Bar","gancho":"Lead identificado pela equipa comercial — Restaurante / Bar em Peniche.","zona":"Peniche"},
  {"n":"Pateo da Saudade","t":"Restaurante tradicional","m":"—","tel":"—","email":"pateodasaudade@gmail.com","p":"Alta","tCliente":"Restaurante tradicional","gancho":"Lead identificado pela equipa comercial — Restaurante tradicional em Peniche.","zona":"Peniche"},
  {"n":"Profresco","t":"Restaurante","m":"—","tel":"—","email":"profresco@profresco.pt; geral@profresco.pt","p":"Alta","tCliente":"Restaurante","gancho":"Lead identificado pela equipa comercial — Restaurante em Peniche.","zona":"Peniche"},
  {"n":"Mar d’Areia","t":"Restaurante","m":"—","tel":"—","email":"geral@mardareia.pt","p":"Alta","tCliente":"Restaurante","gancho":"Lead identificado pela equipa comercial — Restaurante em Ericeira.","zona":"Ericeira"},
  {"n":"Tik Tapas","t":"Tapas/wine bar","m":"—","tel":"—","email":"reservas@tiktapas.pt","p":"Alta","tCliente":"Tapas/wine bar","gancho":"Lead identificado pela equipa comercial — Tapas/wine bar em Ericeira.","zona":"Ericeira"},
  {"n":"Adega da Vila","t":"Restaurante tradicional","m":"—","tel":"—","email":"geral@adegadavila.pt","p":"Alta","tCliente":"Restaurante tradicional","gancho":"Lead identificado pela equipa comercial — Restaurante tradicional em Mafra.","zona":"Mafra"},
  {"n":"Uni Sushi","t":"Restaurante moderno","m":"—","tel":"—","email":"geral@unisushi.pt","p":"Alta","tCliente":"Restaurante moderno","gancho":"Lead identificado pela equipa comercial — Restaurante moderno em Ericeira.","zona":"Ericeira"},
  {"n":"Furnas Ericeira","t":"Restaurante premium","m":"—","tel":"—","email":"reservas@furnasericeira.pt","p":"Alta","tCliente":"Restaurante premium","gancho":"Lead identificado pela equipa comercial — Restaurante premium em Ericeira.","zona":"Ericeira"},
  {"n":"Pedra Dura","t":"Restaurante grande volume","m":"—","tel":"—","email":"geral@pedradura.pt","p":"Alta","tCliente":"Restaurante grande volume","gancho":"Lead identificado pela equipa comercial — Restaurante grande volume em Ericeira.","zona":"Ericeira"},
  {"n":"Restaurante Prim","t":"Restaurante contemporâneo","m":"—","tel":"—","email":"geral@prim.pt","p":"Alta","tCliente":"Restaurante contemporâneo","gancho":"Lead identificado pela equipa comercial — Restaurante contemporâneo em Ericeira.","zona":"Ericeira"},
  {"n":"Jangada","t":"Restaurante marisco","m":"—","tel":"—","email":"reservas@jangada.pt","p":"Alta","tCliente":"Restaurante marisco","gancho":"Lead identificado pela equipa comercial — Restaurante marisco em Ericeira.","zona":"Ericeira"},
  {"n":"Cozinha da Sé","t":"Restaurante premium","m":"—","tel":"—","email":"reservas@cozinhadase.pt","p":"Alta","tCliente":"Restaurante premium","gancho":"Lead identificado pela equipa comercial — Restaurante premium em Braga Centro.","zona":"Braga Centro"},
  {"n":"Dona Júlia","t":"Restaurante tradicional","m":"—","tel":"—","email":"geral@donajulia.pt","p":"Alta","tCliente":"Restaurante tradicional","gancho":"Lead identificado pela equipa comercial — Restaurante tradicional em Braga.","zona":"Braga"},
  {"n":"Taberna Belga","t":"Restaurante grande volume","m":"—","tel":"—","email":"geral@tabernabelga.pt","p":"Alta","tCliente":"Restaurante grande volume","gancho":"Lead identificado pela equipa comercial — Restaurante grande volume em Braga.","zona":"Braga"},
  {"n":"Retrokitchen","t":"Restaurante contemporâneo","m":"—","tel":"—","email":"geral@retrokitchen.pt","p":"Alta","tCliente":"Restaurante contemporâneo","gancho":"Lead identificado pela equipa comercial — Restaurante contemporâneo em Braga.","zona":"Braga"},
  {"n":"Restaurante Tia Isabel","t":"Restaurante tradicional","m":"—","tel":"—","email":"reservas@tiaisabel.pt","p":"Alta","tCliente":"Restaurante tradicional","gancho":"Lead identificado pela equipa comercial — Restaurante tradicional em Braga.","zona":"Braga"},
  {"n":"Bacalhau na Vila","t":"Restaurante português","m":"—","tel":"—","email":"geral@bacalhaunavila.pt","p":"Alta","tCliente":"Restaurante português","gancho":"Lead identificado pela equipa comercial — Restaurante português em Braga.","zona":"Braga"},
  {"n":"Michizaki","t":"Restaurante moderno","m":"—","tel":"—","email":"geral@michizaki.pt","p":"Alta","tCliente":"Restaurante moderno","gancho":"Lead identificado pela equipa comercial — Restaurante moderno em Braga.","zona":"Braga"},
  {"n":"Restaurante Arcoense","t":"Restaurante premium","m":"—","tel":"—","email":"geral@arcoense.pt","p":"Alta","tCliente":"Restaurante premium","gancho":"Lead identificado pela equipa comercial — Restaurante premium em Braga.","zona":"Braga"},
  {"n":"Brasão Coliseu","t":"Restaurante grande volume","m":"—","tel":"—","email":"geral@brasao.pt","p":"Alta","tCliente":"Restaurante grande volume","gancho":"Lead identificado pela equipa comercial — Restaurante grande volume em Porto Centro.","zona":"Porto Centro"},
  {"n":"TerraPlana Café","t":"Restaurante/brunch","m":"—","tel":"—","email":"geral@terraplana.pt","p":"Alta","tCliente":"Restaurante/brunch","gancho":"Lead identificado pela equipa comercial — Restaurante/brunch em Porto.","zona":"Porto"},
  {"n":"Flow Restaurant & Bar","t":"Restaurante premium","m":"—","tel":"—","email":"reservas@flow.pt","p":"Alta","tCliente":"Restaurante premium","gancho":"Lead identificado pela equipa comercial — Restaurante premium em Porto.","zona":"Porto"},
  {"n":"DOP","t":"Restaurante fine dining","m":"—","tel":"—","email":"geral@doprestaurante.pt","p":"Alta","tCliente":"Restaurante fine dining","gancho":"Lead identificado pela equipa comercial — Restaurante fine dining em Porto.","zona":"Porto"},
  {"n":"Tapabento","t":"Restaurante turístico","m":"—","tel":"—","email":"geral@tapabento.pt","p":"Alta","tCliente":"Restaurante turístico","gancho":"Lead identificado pela equipa comercial — Restaurante turístico em Porto Centro.","zona":"Porto Centro"},
  {"n":"Terminal 4450","t":"Restaurante premium","m":"—","tel":"—","email":"reservas@terminal4450.pt","p":"Alta","tCliente":"Restaurante premium","gancho":"Lead identificado pela equipa comercial — Restaurante premium em Matosinhos.","zona":"Matosinhos"},
  {"n":"Gaveto","t":"Restaurante referência","m":"—","tel":"—","email":"geral@gaveto.pt","p":"Alta","tCliente":"Restaurante referência","gancho":"Lead identificado pela equipa comercial — Restaurante referência em Matosinhos.","zona":"Matosinhos"},
  {"n":"O Gaveto Marisqueira","t":"Marisqueira","m":"—","tel":"—","email":"reservas@gaveto.pt","p":"Alta","tCliente":"Marisqueira","gancho":"Lead identificado pela equipa comercial — Marisqueira em Matosinhos.","zona":"Matosinhos"},
  {"n":"Salta o Muro","t":"Restaurante contemporâneo","m":"—","tel":"—","email":"geral@saltaomuro.pt","p":"Alta","tCliente":"Restaurante contemporâneo","gancho":"Lead identificado pela equipa comercial — Restaurante contemporâneo em Matosinhos.","zona":"Matosinhos"},
  {"n":"Meia-Nau","t":"Restaurante moderno","m":"—","tel":"—","email":"geral@meianau.pt","p":"Alta","tCliente":"Restaurante moderno","gancho":"Lead identificado pela equipa comercial — Restaurante moderno em Matosinhos.","zona":"Matosinhos"},
  {"n":"Casa Serrão","t":"Restaurante tradicional","m":"—","tel":"—","email":"reservas@casaserrao.pt","p":"Alta","tCliente":"Restaurante tradicional","gancho":"Lead identificado pela equipa comercial — Restaurante tradicional em Matosinhos.","zona":"Matosinhos"},
  {"n":"Tito II","t":"Restaurante peixe/marisco","m":"—","tel":"—","email":"geral@tito.pt","p":"Alta","tCliente":"Restaurante peixe/marisco","gancho":"Lead identificado pela equipa comercial — Restaurante peixe/marisco em Matosinhos.","zona":"Matosinhos"},
  {"n":"Restaurante Mauritânia","t":"Restaurante grande volume","m":"—","tel":"—","email":"geral@mauritania.pt","p":"Alta","tCliente":"Restaurante grande volume","gancho":"Lead identificado pela equipa comercial — Restaurante grande volume em Matosinhos.","zona":"Matosinhos"},
  {"n":"DeCastro Gaia","t":"Restaurante contemporâneo","m":"—","tel":"—","email":"geral@decastro.pt","p":"Alta","tCliente":"Restaurante contemporâneo","gancho":"Lead identificado pela equipa comercial — Restaurante contemporâneo em Gaia.","zona":"Gaia"},
  {"n":"Ar de Rio","t":"Restaurante turístico","m":"—","tel":"—","email":"reservas@arderio.pt","p":"Alta","tCliente":"Restaurante turístico","gancho":"Lead identificado pela equipa comercial — Restaurante turístico em Cais de Gaia.","zona":"Cais de Gaia"},
  {"n":"Casa Adão","t":"Restaurante tradicional","m":"—","tel":"—","email":"geral@casaadao.pt","p":"Alta","tCliente":"Restaurante tradicional","gancho":"Lead identificado pela equipa comercial — Restaurante tradicional em Gaia.","zona":"Gaia"},
  {"n":"7 Grottos","t":"Restaurante moderno","m":"—","tel":"—","email":"geral@7grottos.pt","p":"Alta","tCliente":"Restaurante moderno","gancho":"Lead identificado pela equipa comercial — Restaurante moderno em Gaia.","zona":"Gaia"},
  {"n":"Esplanada do Teleférico","t":"Restaurante turístico","m":"—","tel":"—","email":"geral@esplanadadoteleferico.pt","p":"Alta","tCliente":"Restaurante turístico","gancho":"Lead identificado pela equipa comercial — Restaurante turístico em Gaia.","zona":"Gaia"},
  {"n":"Mar na Brasa","t":"Restaurante peixe","m":"—","tel":"—","email":"geral@marnabrasa.pt","p":"Alta","tCliente":"Restaurante peixe","gancho":"Lead identificado pela equipa comercial — Restaurante peixe em Gaia.","zona":"Gaia"},
  {"n":"Marisqueira de Ribamar","t":"Marisqueira (desde 1972)","m":"—","tel":"—","email":"marisqueiraribamar.com","p":"Alta","tCliente":"Marisqueira (desde 1972)","gancho":"Lead identificado pela equipa comercial — Marisqueira (desde 1972) em Ribamar.","zona":"Ribamar"},
  {"n":"Cervejaria O Pescador","t":"Cervejaria/Marisqueira","m":"—","tel":"—","email":"opescadorribamar@hotmail.com","p":"Alta","tCliente":"Cervejaria/Marisqueira","gancho":"Lead identificado pela equipa comercial — Cervejaria/Marisqueira em Ribamar.","zona":"Ribamar"},
  {"n":"Casa Rodrigues","t":"Restaurante familiar","m":"—","tel":"—","email":"casarodriguesribamar@gmail.com","p":"Alta","tCliente":"Restaurante familiar","gancho":"Lead identificado pela equipa comercial — Restaurante familiar em Ribamar.","zona":"Ribamar"},
  {"n":"Do Mar À Mesa","t":"Restaurante peixe/marisco","m":"—","tel":"—","email":"domaramesa.geral@gmail.com","p":"Alta","tCliente":"Restaurante peixe/marisco","gancho":"Lead identificado pela equipa comercial — Restaurante peixe/marisco em Ribamar.","zona":"Ribamar"},
  {"n":"Restaurante do Parque","t":"Restaurante tradicional","m":"—","tel":"—","email":"restaurantedoparque.peniche@gmail.com","p":"Alta","tCliente":"Restaurante tradicional","gancho":"Lead identificado pela equipa comercial — Restaurante tradicional em Peniche.","zona":"Peniche"},
  {"n":"Sushi Fish","t":"Restaurante moderno","m":"—","tel":"—","email":"geral@sushifish.pt","p":"Alta","tCliente":"Restaurante moderno","gancho":"Lead identificado pela equipa comercial — Restaurante moderno em Peniche.","zona":"Peniche"},
  {"n":"Taberna do Ganhão","t":"Restaurante","m":"—","tel":"—","email":"geral@tabernadoganhao.pt","p":"Alta","tCliente":"Restaurante","gancho":"Lead identificado pela equipa comercial — Restaurante em Peniche.","zona":"Peniche"},
  {"n":"Tasca do Joel","t":"Restaurante referência","m":"—","tel":"—","email":"reservas@tascadojoel.pt; reservas.tascadojoel@gmail.com","p":"Alta","tCliente":"Restaurante referência","gancho":"Lead identificado pela equipa comercial — Restaurante referência em Peniche.","zona":"Peniche"},
  {"n":"Tribeca Restaurante-Brasserie","t":"Brasserie","m":"—","tel":"—","email":"​tribeca-peniche@hotmail.com; tribeca@tribeca-restaurante.com","p":"Alta","tCliente":"Brasserie","gancho":"Lead identificado pela equipa comercial — Brasserie em Peniche.","zona":"Peniche"},
]


# ── Função de acesso ──────────────────────────────────────────
def get_db_horeca():
    """Retorna a base de leads reais por comercial."""
    return {
        "nuno":  DB_NUNO,
        "joao":  DB_JOAO,
        "oscar": DB_OSCAR,
    }

DB_HORECA_REAL = get_db_horeca()

# ── Base extra (catering + distribuidores) — integrada directamente ──
"""
Base de dados de leads para catering/eventos e distribuidores de congelados.
Importado pelo generate_leads.py
"""

# ════════════════════════════════════════════════════════════════
# CATERING & EVENTOS (Portugal inteiro)
# ════════════════════════════════════════════════════════════════
DB_CATERING = [
  {"n":"Doce Evento","t":"Catering & Eventos","m":"R. Tomás Ribeiro 34, Lisboa","tel":"213 540 210","email":"geral@doceevento.pt","p":"Alta","tCliente":"Catering premium casamentos e eventos corporativos","gancho":"Volume elevado por evento — unidose 85g é sobremesa elegante sem logística complexa.","zona":"Lisboa"},
  {"n":"Saveurs Catering","t":"Catering & Eventos","m":"Av. da República 50, Lisboa","tel":"217 960 400","email":"info@saveurs.pt","p":"Alta","tCliente":"Catering corporativo e eventos internacionais","gancho":"Clientela exigente — produto artesanal português com 40 anos diferencia a proposta.","zona":"Lisboa"},
  {"n":"Essência de Sabor","t":"Catering & Eventos","m":"R. Actor Vale 8, Lisboa","tel":"214 105 020","email":"geral@essenciadesabor.pt","p":"Alta","tCliente":"Casamentos e eventos sociais","gancho":"Sobremesa individual elegante — sem necessidade de pasteleiro no evento.","zona":"Lisboa"},
  {"n":"Eurest Portugal","t":"Catering & Eventos","m":"Av. Fontes Pereira de Melo 16, Lisboa","tel":"213 186 000","email":"geral@eurest.pt","p":"Alta","tCliente":"Catering colectivo / grandes volumes","gancho":"Volume de refeições diário — produto congelado em dose individual garante qualidade constante.","zona":"Lisboa"},
  {"n":"Gertal Companhia Geral","t":"Catering & Eventos","m":"R. Actor Tasso 12, Lisboa","tel":"213 619 200","email":"info@gertal.pt","p":"Alta","tCliente":"Catering colectivo / empresas e hospitais","gancho":"Grande volume com necessidade de sobremesa diferenciada e de fácil execução.","zona":"Lisboa"},
  {"n":"Quinta de Sant'Ana Eventos","t":"Catering & Eventos","m":"Gradil, Mafra","tel":"261 963 480","email":"eventos@quintadesantana.pt","p":"Alta","tCliente":"Eventos de luxo / casamentos premium","gancho":"Casamentos em quinta histórica — pão de ló artesanal é a sobremesa com mais narrativa.","zona":"Oeste"},
  {"n":"Solar de Mil Reis Eventos","t":"Catering & Eventos","m":"Av. Eng. Duarte Pacheco, Leiria","tel":"244 820 000","email":"eventos@solarmilreis.pt","p":"Alta","tCliente":"Eventos e banquetes","gancho":"Região com crescimento de eventos — produto artesanal diferencia a oferta de sobremesas.","zona":"Centro"},
  {"n":"Monte da Ravasqueira","t":"Catering & Eventos","m":"Estrada Monte da Ravasqueira, Arraiolos","tel":"266 498 280","email":"eventos@ravasqueira.com","p":"Alta","tCliente":"Eventos premium / enoturismo","gancho":"Enoturismo de luxo com sobremesas de autor — pão de ló Ti'Piedade como proposta artesanal.","zona":"Alentejo"},
  {"n":"Herdade do Esporão Eventos","t":"Catering & Eventos","m":"Herdade do Esporão, Reguengos de Monsaraz","tel":"266 509 280","email":"turismo@esporao.com","p":"Alta","tCliente":"Enoturismo / eventos internacionais","gancho":"Eventos com clientela internacional premium — produto português com receita secular.","zona":"Alentejo"},
  {"n":"Catering Delícias do Norte","t":"Catering & Eventos","m":"R. do Bonjardim 312, Porto","tel":"222 054 780","email":"info@deliciasdoporto.pt","p":"Alta","tCliente":"Catering casamentos / norte","gancho":"Norte com grande tradição de casamentos — dose individual em congelado facilita logística.","zona":"Porto"},
  {"n":"Quinta da Aveleda Eventos","t":"Catering & Eventos","m":"Quinta da Aveleda, Penafiel","tel":"255 718 200","email":"enoturismo@aveleda.pt","p":"Alta","tCliente":"Enoturismo e eventos premium","gancho":"Quinta histórica com eventos de alto valor — produto artesanal português complementa a experiência.","zona":"Norte"},
  {"n":"Vatel Portugal","t":"Catering & Eventos","m":"R. Gonçalo Cristóvão 195, Porto","tel":"222 073 900","email":"porto@vatel.pt","p":"Média","tCliente":"Escola de hotelaria / catering","gancho":"Formação e eventos — produto de referência para demonstrações e menus de escola.","zona":"Porto"},
  {"n":"CateringLab","t":"Catering & Eventos","m":"Av. de Braga 210, Guimarães","tel":"253 400 100","email":"info@cateringlab.pt","p":"Média","tCliente":"Catering eventos empresariais","gancho":"Região com indústria têxtil — eventos empresariais com necessidade de produto premium acessível.","zona":"Norte"},
  {"n":"Nobre Catering","t":"Catering & Eventos","m":"R. Alexandre Herculano 12, Coimbra","tel":"239 701 200","email":"geral@nobrecatering.pt","p":"Média","tCliente":"Catering académico e social","gancho":"Coimbra com muitos eventos académicos — sobremesa artesanal diferencia o menu.","zona":"Centro"},
  {"n":"Sabor a Festa Catering","t":"Catering & Eventos","m":"R. dos Correeiros 4, Faro","tel":"289 890 120","email":"geral@saborafesta.pt","p":"Média","tCliente":"Catering Algarve / turismo","gancho":"Algarve com turismo de alto valor — produto artesanal português autêntico.","zona":"Algarve"},
  {"n":"Quinta dos Vales Eventos","t":"Catering & Eventos","m":"Sítio dos Vales, Lagoa, Algarve","tel":"282 431 036","email":"eventos@quintadosvales.pt","p":"Alta","tCliente":"Enoturismo / casamentos Algarve","gancho":"Destino de casamentos internacionais — produto artesanal português é proposta de valor forte.","zona":"Algarve"},
  {"n":"Alma Catering","t":"Catering & Eventos","m":"R. Prior do Crato 30, Évora","tel":"266 705 360","email":"info@almacatering.pt","p":"Média","tCliente":"Catering Alentejo / eventos culturais","gancho":"Évora Património Mundial — eventos com clientela que valoriza produto nacional autêntico.","zona":"Alentejo"},
  {"n":"Banquetes Royal","t":"Catering & Eventos","m":"R. do Campo Alegre 1070, Porto","tel":"226 074 500","email":"geral@banquetesroyal.pt","p":"Média","tCliente":"Casamentos e banquetes","gancho":"Volume por evento — dose individual elimina desperdício em refeições de grande grupo.","zona":"Porto"},
  {"n":"Sabores com História","t":"Catering & Eventos","m":"Av. Infante Santo 42, Setúbal","tel":"265 522 800","email":"geral@saborescomhistoria.pt","p":"Média","tCliente":"Catering eventos sul","gancho":"Margem sul em crescimento — produto artesanal para eventos de nível médio-alto.","zona":"Sul"},
  {"n":"Quinta do Rol Eventos","t":"Catering & Eventos","m":"Torres Vedras","tel":"261 967 040","email":"eventos@quintadorol.com","p":"Alta","tCliente":"Enoturismo / casamentos Oeste","gancho":"Casamentos em adega — pão de ló artesanal marida perfeitamente com vinho e tradição.","zona":"Oeste"},
]

# ════════════════════════════════════════════════════════════════
# DISTRIBUIDORES DE CONGELADOS (zonas não cobertas pela equipa)
# ════════════════════════════════════════════════════════════════
DB_DISTRIBUIDORES = [
  {"n":"Frigoríficos do Algarve","t":"Distribuidor Congelados","m":"Zona Industrial, Loulé","tel":"289 416 200","email":"geral@frigorificos-algarve.pt","p":"Alta","tCliente":"Distribuidor regional congelados Algarve","gancho":"Algarve sem cobertura — canal HORECA forte com turismo de alto valor durante todo o ano.","zona":"Algarve"},
  {"n":"Distrinor Alimentar","t":"Distribuidor Congelados","m":"Zona Industrial de Braga","tel":"253 607 500","email":"comercial@distrinor.pt","p":"Alta","tCliente":"Distribuidor alimentar Norte","gancho":"Cobertura norte — complementa carteira de congelados com produto artesanal de alto valor percebido.","zona":"Norte"},
  {"n":"Irmãos Antunes Distribuição","t":"Distribuidor Congelados","m":"Zona Industrial, Évora","tel":"266 748 300","email":"geral@irmaosantunes.pt","p":"Alta","tCliente":"Distribuidor Alentejo","gancho":"Alentejo sem representação — região em crescimento turístico e gastronómico.","zona":"Alentejo"},
  {"n":"Frioeste","t":"Distribuidor Congelados","m":"R. Industrial, Viseu","tel":"232 420 800","email":"comercial@frioeste.pt","p":"Alta","tCliente":"Distribuidor congelados Viseu / Dão-Lafões","gancho":"Região interior sem cobertura — base de restauração e hotelaria a crescer.","zona":"Interior Norte"},
  {"n":"Caloricedos","t":"Distribuidor Congelados","m":"Zona Industrial, Castelo Branco","tel":"272 330 500","email":"geral@caloricedos.pt","p":"Alta","tCliente":"Distribuidor Beira Interior","gancho":"Beira Interior sem cobertura — território com hotelaria rural e turismo de natureza.","zona":"Interior"},
  {"n":"Setasul Distribuição","t":"Distribuidor Congelados","m":"Zona Industrial, Setúbal","tel":"265 700 400","email":"comercial@setasul.pt","p":"Alta","tCliente":"Distribuidor Setúbal / Alentejo Litoral","gancho":"Zona costeira com restauração forte — produto diferenciador na carteira de congelados.","zona":"Sul"},
  {"n":"Transmontana Alimentar","t":"Distribuidor Congelados","m":"Zona Industrial, Bragança","tel":"273 331 200","email":"geral@transmontanaalimentar.pt","p":"Alta","tCliente":"Distribuidor Trás-os-Montes","gancho":"Região com gastronomia forte e turismo rural crescente — produto com identidade nacional.","zona":"Nordeste"},
  {"n":"Frioraia","t":"Distribuidor Congelados","m":"Zona Industrial, Santarém","tel":"243 300 200","email":"comercial@frioraia.pt","p":"Alta","tCliente":"Distribuidor Ribatejo / Médio Tejo","gancho":"Zona limítrofe às áreas cobertas — pode complementar a distribuição com maior penetração.","zona":"Ribatejo"},
  {"n":"Atlântico Frio","t":"Distribuidor Congelados","m":"Zona Industrial, Viana do Castelo","tel":"258 800 300","email":"geral@atlanticofrio.pt","p":"Alta","tCliente":"Distribuidor Minho / Alto Lima","gancho":"Norte com tradição de festas e eventos — produto artesanal com grande aceitação local.","zona":"Norte"},
  {"n":"Giroalimentar","t":"Distribuidor Congelados","m":"Zona Industrial, Figueira da Foz","tel":"233 400 500","email":"comercial@giroalimentar.pt","p":"Alta","tCliente":"Distribuidor Pinhal Litoral / Figueira","gancho":"Costa com restauração turística — Ti'Piedade como sobremesa de referência no linear.","zona":"Centro Litoral"},
  {"n":"Friponente","t":"Distribuidor Congelados","m":"Zona Industrial, Portalegre","tel":"245 200 400","email":"geral@friponente.pt","p":"Média","tCliente":"Distribuidor Alto Alentejo","gancho":"Região com potencial não explorado — pousadas, hotéis rurais e restauração de qualidade.","zona":"Alentejo"},
  {"n":"Norte Frio Distribuição","t":"Distribuidor Congelados","m":"Zona Industrial, Vila Real","tel":"259 370 300","email":"comercial@nortefrio.pt","p":"Média","tCliente":"Distribuidor Douro / Trás-os-Montes Sul","gancho":"Enoturismo do Douro em crescimento — produto artesanal português premium.","zona":"Douro"},
  {"n":"Mediterrânico Alimentar","t":"Distribuidor Congelados","m":"Zona Industrial, Tavira","tel":"281 325 600","email":"geral@mediterranico-alimentar.pt","p":"Alta","tCliente":"Distribuidor Sotavento Algarvio","gancho":"Algarve Oriental com turismo internacional — produto artesanal de alto valor percebido.","zona":"Algarve"},
  {"n":"Expofrio","t":"Distribuidor Congelados","m":"Zona Industrial, Aveiro","tel":"234 380 700","email":"comercial@expofrio.pt","p":"Alta","tCliente":"Distribuidor Aveiro / Baixo Vouga","gancho":"Região industrial com restauração em crescimento — Ti'Piedade na carteira de congelados premium.","zona":"Centro Norte"},
  {"n":"Frioguarda","t":"Distribuidor Congelados","m":"Zona Industrial, Guarda","tel":"271 210 500","email":"geral@frioguarda.pt","p":"Média","tCliente":"Distribuidor Serra da Estrela / Beira Alta","gancho":"Turismo de montanha e neve — produto de doçaria artesanal premium em contexto de acolhimento.","zona":"Interior Norte"},
]


# ════════════════════════════════════════════════════════════════
# BASE DE LEADS HORECA (comerciais)
# ════════════════════════════════════════════════════════════════

DB_HORECA = {
"Lisboa": [
  {"n":"Tasca do Chico","t":"Restaurante","m":"R. do Diário de Notícias 39, Lisboa","tel":"965 059 670","email":"info@tascadochico.pt","p":"Alta","tCliente":"Tasca contemporânea","gancho":"Rotatividade alta de turistas — sobremesa pronta a servir com zero desperdício."},
  {"n":"Solar dos Presuntos","t":"Restaurante","m":"R. das Portas de Santo Antão 150, Lisboa","tel":"213 424 253","email":"geral@solardospresuntos.com","p":"Alta","tCliente":"Restaurante tradicional / turismo","gancho":"Volume de turistas — pão de ló é produto de eleição para terminar uma refeição típica."},
  {"n":"Martinho da Arcada","t":"Restaurante","m":"Pr. do Comércio 3, Lisboa","tel":"218 879 259","email":"geral@martinhodaarcada.pt","p":"Alta","tCliente":"Histórico / turismo","gancho":"O produto mais português para o restaurante mais antigo de Lisboa."},
  {"n":"Pastéis de Belém","t":"Pastelaria","m":"R. de Belém 84, Lisboa","tel":"213 637 423","email":"geral@pasteisdebelem.pt","p":"Alta","tCliente":"Pastelaria icónica / turismo","gancho":"Duas referências da doçaria artesanal portuguesa — produto complementar."},
  {"n":"Landeau Chocolate","t":"Café","m":"R. das Flores 70, Lisboa","tel":"214 792 178","email":"hello@landeau.pt","p":"Alta","tCliente":"Café de nicho / produto","gancho":"Valoriza produtos com história — dose individual em congelado resolve logística."},
  {"n":"Taberna Rua das Flores","t":"Restaurante","m":"R. das Flores 103, Lisboa","tel":"213 479 418","email":"info@tabernaruadasflores.pt","p":"Média","tCliente":"Tasca contemporânea","gancho":"Rotatividade alta — produto pronto a servir elimina desperdício."},
  {"n":"Pharmácia","t":"Restaurante","m":"R. Marechal Saldanha 1, Lisboa","tel":"213 465 146","email":"geral@museudafarmacia.pt","p":"Média","tCliente":"Restaurante temático / cultura","gancho":"Clientela atenta à origem — pão de ló como sobremesa de autor."},
  {"n":"Decadente","t":"Restaurante","m":"R. de São Pedro de Alcântara 45, Lisboa","tel":"213 957 936","email":"info@odecadente.pt","p":"Média","tCliente":"Bistrô / residentes","gancho":"Carta rotativa — dose individual encaixa na filosofia anti-desperdício."},
  {"n":"Copenhagen Coffee Lab","t":"Café","m":"R. Nova da Piedade 10, Lisboa","tel":"—","email":"hello@copenhagencoffeelab.com","p":"Média","tCliente":"Café de especialidade","gancho":"Pairing pão de ló + café de especialidade."},
  {"n":"Mercearia do Bairro","t":"Mercearia Gourmet","m":"R. do Açúcar 83, Lisboa","tel":"—","email":"geral@merceariadobairro.pt","p":"Média","tCliente":"Mercearia gourmet","gancho":"Produto artesanal com 40 anos — diferenciador face ao industrial."},
  {"n":"Clube de Jornalistas","t":"Restaurante","m":"R. das Trinas 129, Lisboa","tel":"213 977 138","email":"geral@clubedejornalistas.com","p":"Média","tCliente":"Restaurante clássico","gancho":"Clientela fiel — sobremesa clássica como âncora de carta."},
  {"n":"Tasca do Lagarto","t":"Restaurante","m":"R. dos Bacalhoeiros 34, Lisboa","tel":"—","email":"info@tascadolagarto.pt","p":"Média","tCliente":"Tasca / turismo","gancho":"Alfama — turistas com apetência por sobremesas genuinamente portuguesas."},
  {"n":"Café de São Bento","t":"Café","m":"R. de São Bento 212, Lisboa","tel":"213 952 911","email":"geral@cafesaobento.pt","p":"Média","tCliente":"Café clássico","gancho":"Clientela estabelecida — pão de ló como sobremesa de balcão premium."},
  {"n":"O Pitéu da Graça","t":"Restaurante","m":"Pr. da Graça 96, Lisboa","tel":"218 870 565","email":"—","p":"Baixa","tCliente":"Restaurante de bairro","gancho":"Dose individual congelada controla custo e elimina desperdício."},
  {"n":"Mini Bar Teatro","t":"Restaurante","m":"R. António Maria Cardoso 58, Lisboa","tel":"211 305 393","email":"info@minibar.pt","p":"Média","tCliente":"Restaurante criativo","gancho":"Público jovem — pão de ló chocolate ou canela como sobremesa diferenciada."},
],
"Santarém": [
  {"n":"Restaurante O Salazares","t":"Restaurante","m":"R. de São Martinho 2, Santarém","tel":"243 322 384","email":"info@osalazares.pt","p":"Alta","tCliente":"Restaurante tradicional","gancho":"Capital gastronómica regional — receita secular tem aceitação natural."},
  {"n":"Restaurante Portas do Sol","t":"Restaurante","m":"Jardim das Portas do Sol, Santarém","tel":"243 309 520","email":"info@portasdosol.pt","p":"Alta","tCliente":"Restaurante com vista / turismo","gancho":"Turismo de alto valor — sobremesa em volume sem perda de qualidade."},
  {"n":"Hotel Cristal Santarém","t":"Hotel","m":"R. Francisco Moreira 7, Santarém","tel":"243 377 575","email":"reservas@hotelcristal.pt","p":"Alta","tCliente":"Hotel 3* / negócios","gancho":"Carta sólida sem pasteleiro próprio."},
  {"n":"Pastelaria Bijou","t":"Pastelaria","m":"Av. Bernardo Santareno, Santarém","tel":"243 322 507","email":"—","p":"Alta","tCliente":"Pastelaria de referência local","gancho":"Dose individual — impulso ao balcão."},
  {"n":"Pastelaria Paraíso","t":"Pastelaria","m":"Av. Marquês de Sá da Bandeira, Santarém","tel":"243 322 441","email":"—","p":"Média","tCliente":"Pastelaria de bairro","gancho":"Diferenciação com produto artesanal."},
  {"n":"Tasca do Escondidinho","t":"Restaurante","m":"R. Capelo e Ivens, Santarém","tel":"243 323 991","email":"—","p":"Média","tCliente":"Tasca tradicional","gancho":"Pão de ló como sobremesa caseira."},
  {"n":"Mercearia Tradicional do Ribatejo","t":"Mercearia Gourmet","m":"Lg. da Feira, Santarém","tel":"—","email":"—","p":"Média","tCliente":"Mercearia gourmet","gancho":"Identidade ribatejana — produto regional artesanal."},
],
"Linha Sintra–Cascais": [
  {"n":"Bar do Fundo","t":"Restaurante","m":"Av. Alfredo Coelho, Praia Grande, Colares","tel":"219 282 092","email":"info@bardofundo.pt","p":"Alta","tCliente":"Restaurante premium / vista mar","gancho":"Clientela de poder de compra elevado — sobremesa com narrativa."},
  {"n":"Restaurante Azenhas do Mar","t":"Restaurante","m":"Lugar das Piscinas, 2705-098 Colares","tel":"219 280 739","email":"info@azenhasdomar.com","p":"Alta","tCliente":"Restaurante icónico / turismo","gancho":"Mais fotografado de Portugal — sobremesa artesanal reforça o posicionamento."},
  {"n":"Taberna Clandestina","t":"Restaurante","m":"R. Afonso Sanches 36, Cascais","tel":"916 229 630","email":"info@tabernaclandestina.pt","p":"Alta","tCliente":"Gastropub / residentes premium","gancho":"Carta criativa portuguesa — Ti'Piedade como sobremesa de referência."},
  {"n":"Hífen","t":"Restaurante","m":"Av. Dom Carlos I 48, Cascais","tel":"915 546 537","email":"info@hifenrestaurant.com","p":"Alta","tCliente":"Restaurante de referência / Cascais","gancho":"Produto com história diferencia a experiência."},
  {"n":"Hotel Palácio Estoril","t":"Hotel","m":"R. Particular, Estoril","tel":"214 648 000","email":"info@palacioestoril.com","p":"Alta","tCliente":"Hotel 5* / luxo","gancho":"Qualidade constante em grande volume sem pasteleiro."},
  {"n":"Lawrence's Hotel (Sintra)","t":"Hotel","m":"R. Consiglieri Pedroso 38, Sintra","tel":"219 105 500","email":"info@lawrenceshotel.com","p":"Alta","tCliente":"Boutique hotel / turismo cultural","gancho":"Hotel histórico — produto artesanal complementa a narrativa de autenticidade."},
  {"n":"Gourmet Italiano (Cascais)","t":"Mercearia Gourmet","m":"Av. Infante Dom Henrique 1027 D, Cascais","tel":"214 842 127","email":"info@gourmetitaliano.pt","p":"Alta","tCliente":"Deli gourmet / expatriados","gancho":"Clientela internacional com elevado poder de compra."},
  {"n":"Emporium Gourmet","t":"Mercearia Gourmet","m":"Av. Nossa Senhora do Cabo 101, Cascais","tel":"211 541 588","email":"info@emporiumgourmet.pt","p":"Alta","tCliente":"Mercearia gourmet","gancho":"História de 40 anos — receita da D. Piedade vende-se sozinha."},
  {"n":"Taberna Económica de Cascais","t":"Restaurante","m":"R. Sebastião José de Carvalho e Melo 35, Cascais","tel":"214 832 214","email":"info@tabernaeconomicadecascais.com","p":"Alta","tCliente":"Taberna / turismo","gancho":"Volume de turistas — produto congelado garante consistência."},
  {"n":"Angra Gatti","t":"Restaurante","m":"Av. Alfredo Coelho 57, Praia Grande, Colares","tel":"965 770 247","email":"info@angragatti.com","p":"Alta","tCliente":"Restaurante italiano / destino","gancho":"Sobremesa portuguesa como proposta de fecho de refeição."},
  {"n":"Mana","t":"Restaurante","m":"Tv. Navegantes 13, Cascais","tel":"915 669 206","email":"info@manacascais.pt","p":"Média","tCliente":"Restaurante & bar / trendy","gancho":"Chocolate ou canela na carta."},
  {"n":"Café Paris (Sintra)","t":"Café","m":"Pr. da República 32, Sintra","tel":"219 232 375","email":"—","p":"Média","tCliente":"Café turístico","gancho":"Produto icónico para turistas internacionais em Sintra."},
  {"n":"Casa da Galé","t":"Restaurante","m":"Av. Alfredo Coelho 61, Praia Grande, Colares","tel":"219 291 218","email":"—","p":"Média","tCliente":"Restaurante peixe / local","gancho":"Final natural de uma refeição de peixe."},
],
"SuperIndep_Nuno": [
  {"n":"Mercado Municipal de Campo de Ourique","t":"Supermercado Independente","m":"R. Coelho da Rocha, Lisboa","tel":"213 954 628","email":"mercado@cm-lisboa.pt","p":"Alta","tCliente":"Mercado alimentar / produto fresco","gancho":"Clientela premium — dose individual congelada com margem interessante."},
  {"n":"Honest Greens Market (Cascais)","t":"Supermercado Independente","m":"Av. Marginal, São João do Estoril","tel":"—","email":"—","p":"Média","tCliente":"Supermercado independente / saudável","gancho":"Classe média-alta — produto artesanal diferencia o linear."},
],
"Margem Sul": [
  {"n":"O Farol Design Hotel","t":"Hotel","m":"R. do Farol 1, Cacilhas, Almada","tel":"210 407 040","email":"info@farolhotel.com","p":"Alta","tCliente":"Boutique hotel / design","gancho":"Hotel de autor — produto artesanal de alto valor percebido."},
  {"n":"Hotel Sana Sesimbra","t":"Hotel","m":"Av. 25 de Abril, Sesimbra","tel":"212 289 000","email":"info@sesimbra.sanahotels.com","p":"Alta","tCliente":"Resort / lazer","gancho":"F&B de resort — qualidade constante sem pasteleiro."},
  {"n":"Restaurante Ribamar (Sesimbra)","t":"Restaurante","m":"Av. dos Náufragos 29, Sesimbra","tel":"212 233 853","email":"info@restauranteribamar.com","p":"Alta","tCliente":"Restaurante peixe / turismo","gancho":"Destino de verão — produto pronto a servir em período de pico."},
  {"n":"Tasca D'Avenida","t":"Restaurante","m":"Av. Dom Afonso Henriques 10C, Almada","tel":"968 348 036","email":"—","p":"Alta","tCliente":"Tasca contemporânea","gancho":"Almoços de negócios — sobremesa premium com margem confortável."},
  {"n":"The Baptist","t":"Restaurante","m":"R. Afonso Galo 56, Almada","tel":"212 750 996","email":"—","p":"Média","tCliente":"Restaurante casual","gancho":"Volume e rotatividade — sobremesa pronta a servir."},
  {"n":"Restaurante Paladar (Palmela)","t":"Restaurante","m":"R. João de Deus, Palmela","tel":"—","email":"—","p":"Média","tCliente":"Restaurante regional","gancho":"Turismo de interior com apetência por produto artesanal."},
  {"n":"Pastelaria A Floresta (Barreiro)","t":"Pastelaria","m":"Av. Bento Gonçalves, Barreiro","tel":"—","email":"—","p":"Média","tCliente":"Pastelaria de bairro","gancho":"Dose individual ao balcão com margem interessante."},
],
"SuperIndep_Joao": [
  {"n":"Supermercado Apolónia (Leiria)","t":"Supermercado Independente","m":"Av. Heróis de Angola, Leiria","tel":"244 859 900","email":"leiria@apolonia.pt","p":"Alta","tCliente":"Supermercado premium independente","gancho":"Foco em produto nacional — Ti'Piedade é referência perfeita."},
  {"n":"Mini Mercado Tradicional da Caparica","t":"Supermercado Independente","m":"R. da Liberdade, Costa da Caparica","tel":"—","email":"—","p":"Média","tCliente":"Mini mercado / bairro de praia","gancho":"Zona balnear — produto congelado no linear de sobremesas."},
],
"Costa Oeste (S. Martinho–Vieira)": [
  {"n":"Hotel Columbano (S. Martinho do Porto)","t":"Hotel","m":"Av. Marginal, S. Martinho do Porto","tel":"262 989 220","email":"info@hotelcolumbano.com","p":"Alta","tCliente":"Hotel 4* / praia","gancho":"Procura de verão intensa — produto congelado mantém qualidade."},
  {"n":"Tasca do Zé (Nazaré)","t":"Restaurante","m":"R. Mouzinho de Albuquerque 22, Nazaré","tel":"262 551 945","email":"—","p":"Alta","tCliente":"Tasca turística / Nazaré","gancho":"Turismo internacional — pão de ló como sobremesa típica portuguesa."},
  {"n":"Restaurante O Casalinho (Nazaré)","t":"Restaurante","m":"R. do Elevador 24, Nazaré","tel":"262 552 608","email":"—","p":"Alta","tCliente":"Restaurante peixe / turismo","gancho":"Turistas internacionais — produto de identidade nacional."},
  {"n":"Hotel Maré (Nazaré)","t":"Hotel","m":"R. Mouzinho de Albuquerque 8, Nazaré","tel":"262 550 000","email":"info@hotelmare.com","p":"Alta","tCliente":"Hotel / turismo surf","gancho":"Público de surf que valoriza produto artesanal local."},
  {"n":"Solar de Alcobaça","t":"Restaurante","m":"Pr. 25 de Abril, Alcobaça","tel":"262 598 312","email":"—","p":"Média","tCliente":"Restaurante turístico / mosteiro","gancho":"Destino patrimonial — produto com raízes medievais."},
  {"n":"Restaurante A Tasquinha (Alcobaça)","t":"Restaurante","m":"R. Frei António Brandão 2, Alcobaça","tel":"262 582 397","email":"—","p":"Média","tCliente":"Restaurante local","gancho":"Produto artesanal como âncora da carta de sobremesas."},
],
"Leiria": [
  {"n":"Tromba Rija","t":"Restaurante","m":"R. Professores Portelas, Marrazes, Leiria","tel":"244 855 072","email":"geral@trombarja.com","p":"Alta","tCliente":"Restaurante regional / referência","gancho":"Referência gastronómica — produto artesanal encaixa na valorização do território."},
  {"n":"Hotel Eurosol Leiria","t":"Hotel","m":"R. Comissão da Iniciativa, Leiria","tel":"244 838 201","email":"info@eurosolleiria.pt","p":"Alta","tCliente":"Hotel 4* / negócios","gancho":"Carta consistente — produto congelado resolve sobremesas sem pastelaria."},
  {"n":"Pastelaria Garrett","t":"Pastelaria","m":"Pr. Rodrigues Lobo, Leiria","tel":"244 812 370","email":"—","p":"Alta","tCliente":"Pastelaria de referência","gancho":"Complementa o catálogo e diferencia da concorrência."},
  {"n":"Tasca da Barrosinha","t":"Restaurante","m":"Pr. Rodrigues Lobo 2, Leiria","tel":"244 823 703","email":"—","p":"Alta","tCliente":"Tasca tradicional","gancho":"Clientela local — sobremesa clássica com margem confortável."},
  {"n":"O Funil (Leiria)","t":"Restaurante","m":"Av. Heróis de Angola 66, Leiria","tel":"244 832 522","email":"—","p":"Média","tCliente":"Restaurante casual","gancho":"Volume de almoços — produto pronto acelera rotatividade."},
  {"n":"Patisserie Almonda (Tomar)","t":"Pastelaria","m":"Av. Marquês de Tomar, Tomar","tel":"249 312 252","email":"—","p":"Média","tCliente":"Pastelaria de destino","gancho":"Destino turístico — produto português de referência."},
],
"Ericeira–Caldas da Rainha": [
  {"n":"Marginal (Peniche)","t":"Restaurante","m":"Estr. Marginal Norte, Peniche","tel":"968 907 248","email":"marginalrestaurante@gmail.com","p":"Alta","tCliente":"Restaurante premium / costa","gancho":"Vista mar — sobremesa artesanal fecha a experiência."},
  {"n":"Restaurante Sueste (Ericeira)","t":"Restaurante","m":"R. Eduardo Burnay 22, Ericeira","tel":"261 862 108","email":"info@sueste.pt","p":"Alta","tCliente":"Restaurante peixe / turismo surf","gancho":"Comunidade surf internacional — produto artesanal autêntico."},
  {"n":"Hotel Termas das Caldas","t":"Hotel","m":"Pr. 25 de Abril, Caldas da Rainha","tel":"262 830 200","email":"info@termascaldas.pt","p":"Alta","tCliente":"Hotel termal / saúde","gancho":"Produto sem conservantes, ingredientes simples e receita secular."},
  {"n":"Adega do Caseiro (Caldas)","t":"Restaurante","m":"R. Eng. Duarte Pacheco, Caldas da Rainha","tel":"262 831 291","email":"—","p":"Alta","tCliente":"Restaurante regional","gancho":"Referência local — identidade regional reforça o posicionamento."},
  {"n":"Chico Neto (Ribamar)","t":"Restaurante","m":"R. das Armaçõe 26, Ribamar","tel":"261 422 106","email":"—","p":"Alta","tCliente":"Restaurante peixe / local","gancho":"Clientela de fim-de-semana — sobremesa clássica para famílias."},
  {"n":"O Viveiro (Ribamar)","t":"Restaurante","m":"R. das Armaçõe 7, Ribamar","tel":"261 422 197","email":"—","p":"Alta","tCliente":"Restaurante peixe / vista mar","gancho":"Vista mar — produto artesanal eleva a carta sem complexidade."},
  {"n":"Cafetaria Puro Cake Lab","t":"Pastelaria","m":"Pr. Jacob Rodrigues Pereira 18, Peniche","tel":"916 950 480","email":"info@purocakelab.pt","p":"Alta","tCliente":"Pastelaria artesanal","gancho":"Valoriza produto artesanal — Ti'Piedade como oferta complementar."},
  {"n":"Pastelaria Princesa do Mar","t":"Pastelaria","m":"R. António Maria Oliveira 34, Peniche","tel":"262 782 929","email":"—","p":"Alta","tCliente":"Pastelaria local","gancho":"Dose individual com margem de revenda interessante."},
  {"n":"Restaurante Ó Baleal","t":"Restaurante","m":"Baleal, Peniche","tel":"—","email":"—","p":"Alta","tCliente":"Restaurante surf / natureza","gancho":"Baleal ícone do surf — produto artesanal autêntico."},
],
"SuperIndep_Oscar": [
  {"n":"Supermercado Apolónia (Porto)","t":"Supermercado Independente","m":"R. de Júlio Dinis 826, Porto","tel":"226 066 730","email":"porto@apolonia.pt","p":"Alta","tCliente":"Supermercado premium independente","gancho":"Foco em produto nacional — Ti'Piedade é referência natural."},
  {"n":"Mercado Bom Sucesso (Porto)","t":"Supermercado Independente","m":"Pr. do Bom Sucesso 74, Porto","tel":"226 088 800","email":"info@mercadobomsucesso.com","p":"Alta","tCliente":"Mercado gourmet / turismo","gancho":"Mercado de referência no Porto — produto artesanal com 40 anos encaixa naturalmente."},
],
"Coimbra": [
  {"n":"Fangas Mercearia Bar","t":"Mercearia Gourmet","m":"R. Fernandes Tomás 45, Coimbra","tel":"239 115 540","email":"info@fangas.pt","p":"Alta","tCliente":"Mercearia gourmet / bar","gancho":"Curadoria nacional — pão de ló com 40 anos é produto natural."},
  {"n":"Hotel Quinta das Lágrimas","t":"Hotel","m":"Santa Clara, Coimbra","tel":"239 802 380","email":"reservas@quintadaslagrimas.pt","p":"Alta","tCliente":"Hotel 5* / romance","gancho":"Narrativa histórica — produto artesanal reforça experiência portuguesa."},
  {"n":"Restaurante O Trovador","t":"Restaurante","m":"Lg. da Sé Velha 15, Coimbra","tel":"239 825 475","email":"info@otrovador.pt","p":"Alta","tCliente":"Restaurante histórico / fado","gancho":"Fado ao vivo — pão de ló como sobremesa de fim de noite."},
  {"n":"Café Santa Cruz","t":"Café","m":"Pr. 8 de Maio, Coimbra","tel":"239 833 617","email":"cafesantacruz@sapo.pt","p":"Alta","tCliente":"Café histórico / turismo","gancho":"Mais emblemático de Coimbra — produto artesanal com história."},
  {"n":"Pastelaria Briosa","t":"Pastelaria","m":"R. Direita, Coimbra","tel":"239 824 764","email":"—","p":"Alta","tCliente":"Pastelaria de referência","gancho":"Complementa o catálogo sem competir diretamente."},
  {"n":"Adega Paço do Conde","t":"Restaurante","m":"R. Paço do Conde 1, Coimbra","tel":"239 825 605","email":"—","p":"Média","tCliente":"Restaurante clássico","gancho":"Clientela académica — sobremesa clássica universitária."},
  {"n":"Tasca da Rua Nova","t":"Restaurante","m":"R. Nova 44, Coimbra","tel":"239 826 669","email":"—","p":"Média","tCliente":"Tasca contemporânea","gancho":"Carta curta — produto artesanal como âncora."},
],
"Porto": [
  {"n":"O Gaveto","t":"Restaurante","m":"R. Roberto Ivens 826, Matosinhos","tel":"229 381 879","email":"info@restauranteogaveto.com","p":"Alta","tCliente":"Restaurante peixe / referência","gancho":"Clientela exigente — pão de ló é a sobremesa natural de uma refeição portuguesa."},
  {"n":"Mercearia das Flores","t":"Mercearia Gourmet","m":"R. das Flores 110, Porto","tel":"222 013 290","email":"info@merceariadas flores.pt","p":"Alta","tCliente":"Mercearia gourmet / design","gancho":"Curadoria nacional — história de 40 anos e receita intacta."},
  {"n":"Hotel Infante de Sagres","t":"Hotel","m":"Pr. Filipa de Lencastre 62, Porto","tel":"223 398 500","email":"info@hotelinfantesagres.pt","p":"Alta","tCliente":"Hotel 5* histórico","gancho":"Hotel histórico — produto artesanal complementa experiência premium."},
  {"n":"Café Majestic","t":"Café","m":"R. de Santa Catarina 112, Porto","tel":"222 003 887","email":"geral@cafemajestic.com","p":"Alta","tCliente":"Café histórico / turismo","gancho":"Dos cafés mais visitados da Europa — doçaria nacional premium."},
  {"n":"Taberninha do Manel","t":"Restaurante","m":"Av. Gustavo Eiffel 274, Porto","tel":"222 086 389","email":"—","p":"Alta","tCliente":"Tasca histórica / turismo","gancho":"Pão de ló como âncora de cozinha portuguesa."},
  {"n":"Pastelaria Luca","t":"Pastelaria","m":"R. de Sá da Bandeira 118, Porto","tel":"222 084 010","email":"—","p":"Alta","tCliente":"Pastelaria de referência","gancho":"Produto artesanal de autor complementa o catálogo."},
  {"n":"Casa de Pasto da Palmeira","t":"Restaurante","m":"R. de Palmeira 2, Porto","tel":"222 005 753","email":"—","p":"Alta","tCliente":"Casa de pasto / turismo","gancho":"Cozinha portuguesa simples — pão de ló é a sobremesa perfeita."},
  {"n":"Aduela","t":"Restaurante","m":"R. do Oliveiras 38, Porto","tel":"222 008 757","email":"—","p":"Média","tCliente":"Restaurante casual / wine bar","gancho":"Jovem e urbano — chocolate ou canela como sobremesa diferenciada."},
],
"Braga": [
  {"n":"Bem Me Quer (Braga)","t":"Restaurante","m":"Pr. do Município, Braga","tel":"253 278 916","email":"geral@restaurantebemmequeer.pt","p":"Alta","tCliente":"Restaurante de referência","gancho":"Âncora premium da carta de sobremesas."},
  {"n":"Hotel Meliá Braga","t":"Hotel","m":"Av. General Carrilho da Silva Pinto 8, Braga","tel":"253 144 000","email":"melia.braga@melia.com","p":"Alta","tCliente":"Hotel 5* / congressos","gancho":"Volume de eventos — dose individual para grande escala com qualidade consistente."},
  {"n":"Restaurante Inácio","t":"Restaurante","m":"Campo das Hortas 4, Braga","tel":"253 613 235","email":"—","p":"Alta","tCliente":"Restaurante clássico","gancho":"Referência local — produto artesanal reforça qualidade."},
  {"n":"Pastelaria Riquexó","t":"Pastelaria","m":"Av. Central 69, Braga","tel":"253 215 055","email":"—","p":"Alta","tCliente":"Pastelaria clássica","gancho":"Complementa o catálogo sem concorrer diretamente."},
  {"n":"Pastelaria Oliveira","t":"Pastelaria","m":"R. do Souto 128, Braga","tel":"253 215 990","email":"—","p":"Alta","tCliente":"Pastelaria de referência","gancho":"Muito frequentada — diferenciação com margem elevada."},
  {"n":"Taberna Belga","t":"Restaurante","m":"R. de Maximinos 121, Braga","tel":"253 204 786","email":"—","p":"Média","tCliente":"Gastropub / cerveja artesanal","gancho":"Sobremesa artesanal portuguesa como fecho diferenciado."},
],
"Guimarães": [
  {"n":"Solar do Arco","t":"Restaurante","m":"R. de Santa Maria 48, Guimarães","tel":"253 513 072","email":"info@solardoarco.pt","p":"Alta","tCliente":"Restaurante histórico","gancho":"Centro histórico Património Mundial — receita secular encaixa perfeitamente."},
  {"n":"Pousada de Guimarães","t":"Hotel","m":"R. Conde de Margaride 153, Guimarães","tel":"253 511 249","email":"pousadaguimaraes@pousadas.pt","p":"Alta","tCliente":"Pousada histórica / turismo","gancho":"Mosteiro medieval — receita levada ao Japão no séc. XVI."},
  {"n":"El Rei","t":"Restaurante","m":"Pr. de São Tiago 20, Guimarães","tel":"253 419 096","email":"—","p":"Alta","tCliente":"Restaurante centro histórico","gancho":"Clientela turística internacional com apetência por produto típico."},
  {"n":"Pastelaria Clarinha","t":"Pastelaria","m":"R. de Santo António, Guimarães","tel":"253 512 552","email":"—","p":"Alta","tCliente":"Pastelaria de referência","gancho":"Complementa o catálogo com margem interessante."},
  {"n":"Mercearia Vimaranes","t":"Mercearia Gourmet","m":"Lg. do Toural, Guimarães","tel":"—","email":"—","p":"Média","tCliente":"Mercearia gourmet","gancho":"Identidade nacional forte para mercearia gourmet."},
  {"n":"Sabores do Minho","t":"Restaurante","m":"R. de Couros 24, Guimarães","tel":"—","email":"—","p":"Média","tCliente":"Restaurante regional","gancho":"Pão de ló como sobremesa de eleição em carta regional."},
],
}

COMERCIAIS = {
    "nuno":  {"nome":"Nuno",  "email":os.environ.get("EMAIL_NUNO",""),  "canal":"horeca", "zonas":["Lisboa","Santarém","Linha Sintra–Cascais","SuperIndep_Nuno"]},
    "joao":  {"nome":"João",  "email":os.environ.get("EMAIL_JOAO",""),  "canal":"horeca", "zonas":["Lisboa","Margem Sul","Costa Oeste (S. Martinho–Vieira)","Leiria","SuperIndep_Joao"]},
    "oscar": {"nome":"Óscar", "email":os.environ.get("EMAIL_OSCAR",""), "canal":"horeca", "zonas":["Ericeira–Caldas da Rainha","Coimbra","Porto","Braga","Guimarães","SuperIndep_Oscar"]},
    "rui_catering":     {"nome":"Rui", "email":EMAIL_RUI, "canal":"catering",     "zonas":[]},
    "rui_distribuidores":{"nome":"Rui","email":EMAIL_RUI, "canal":"distribuidores","zonas":[]},
}

TIPOS_HORECA = ["Restaurante","Pastelaria","Hotel","Mercearia Gourmet","Café","Supermercado Independente"]

# ════════════════════════════════════════════════════════════════
# UTILITÁRIOS
# ════════════════════════════════════════════════════════════════

def semana_num():
    return datetime.date.today().isocalendar()[1]

def semana_datas():
    d = datetime.date.today()
    seg = d - datetime.timedelta(days=d.weekday())
    sex = seg + datetime.timedelta(days=4)
    meses = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
    return f"{seg.day} {meses[seg.month-1]} – {sex.day} {meses[sex.month-1]} {sex.year}"

def hoje():
    return datetime.date.today().strftime("%Y-%m-%d")

def seeded_shuffle(lst, seed):
    result = list(lst)
    for i in range(len(result)-1, 0, -1):
        j = int(abs(math.sin(seed*(i+1)*9301+49297)*233280)) % (i+1)
        result[i], result[j] = result[j], result[i]
    return result

def gerar_leads_horeca(com_id, sem):
    """Gera 20 leads para o comercial, usando a base de dados real dos ficheiros Excel."""
    # Usar base real se disponível, senão fallback para DB_HORECA
    pool_real = DB_HORECA_REAL.get(com_id, [])
    pool_fallback = []
    if not pool_real:
        com = COMERCIAIS[com_id]
        for zona in com["zonas"]:
            for lead in DB_HORECA.get(zona, []):
                if lead["t"] in TIPOS_HORECA:
                    pool_fallback.append({**lead, "zona": zona, "canal":"HORECA"})
        pool = pool_fallback
    else:
        pool = [{**l, "canal":"HORECA"} for l in pool_real]
    
    shuffled = seeded_shuffle(pool, sem*1000 + list(COMERCIAIS.keys()).index(com_id)+1)
    alta  = [l for l in shuffled if l["p"]=="Alta"]
    resto = [l for l in shuffled if l["p"]!="Alta"]
    return (alta+resto)[:20]

def gerar_leads_canal(db, canal_id, sem, n=5):
    shuffled = seeded_shuffle(db, sem*2000 + hash(canal_id)%1000)
    return [{**l, "canal": canal_id} for l in shuffled[:n]]

def lead_key(canal_id, lead):
    return f"{canal_id}::{lead['n']}::{lead.get('zona','PT')}"

# ════════════════════════════════════════════════════════════════
# HISTÓRICO — GitHub API
# ════════════════════════════════════════════════════════════════

def github_api(method, path, data=None):
    token = os.environ.get("GH_PAT","")
    if not token:
        print("[AVISO] GH_PAT não configurado — histórico não será guardado.")
        return None
    url = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/{path}"
    headers = {"Authorization":f"Bearer {token}","Accept":"application/vnd.github+json","X-GitHub-Api-Version":"2022-11-28","Content-Type":"application/json"}
    body = json.dumps(data).encode() if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as r:
            return json.loads(r.read())
    except urllib.error.HTTPError as e:
        print(f"[GitHub API] {method} {path} → {e.code}: {e.read().decode()}")
        return None

def ler_historico():
    resp = github_api("GET", f"contents/{HIST_FILE}")
    if not resp:
        return {"leads":{},"ultima_atualizacao":None,"versao":"2.0"}, None
    content = base64.b64decode(resp["content"]).decode("utf-8")
    return json.loads(content), resp["sha"]

def escrever_historico(historico, sha):
    content_b64 = base64.b64encode(json.dumps(historico, ensure_ascii=False, indent=2).encode()).decode()
    data = {"message":f"histórico: {hoje()}","content":content_b64}
    if sha: data["sha"] = sha
    result = github_api("PUT", f"contents/{HIST_FILE}", data)
    print(f"{'✓' if result else '✗'} Histórico {'guardado' if result else 'ERRO ao guardar'}")

def registar_envios(historico, canal_id, leads, email_num, comercial_nome):
    d = hoje()
    for lead in leads:
        k = lead_key(canal_id, lead)
        if k not in historico["leads"]:
            historico["leads"][k] = {
                "nome": lead["n"], "tipo": lead["t"],
                "zona": lead.get("zona","Portugal"), "morada": lead.get("m",""),
                "email_lead": lead.get("email",""), "tel": lead.get("tel",""),
                "comercial": comercial_nome, "canal_id": canal_id,
                "canal": lead.get("canal",""), "prioridade": lead.get("p",""),
                "tipologia_cliente": lead.get("tCliente",""),
                "gancho": lead.get("gancho",""),
                "semana_entrada": semana_num(),
                "data_entrada": d,
                "estado": "Em aberto",
                "emails_enviados": [],
                "visitas": [], "notas": ""
            }
        entry = historico["leads"][k]
        if not any(e["num"]==email_num and e["data"]==d for e in entry["emails_enviados"]):
            entry["emails_enviados"].append({"num":email_num,"data":d,"assunto":f"Nurturing Email {email_num}"})
    historico["ultima_atualizacao"] = d
    return historico

def calcular_email_num(historico, canal_id, lead):
    k = lead_key(canal_id, lead)
    if k not in historico["leads"]: return 1
    enviados = [e["num"] for e in historico["leads"][k].get("emails_enviados",[])]
    for n in [1,2,3,4]:
        if n not in enviados: return n
    return None

# ════════════════════════════════════════════════════════════════
# EXCEL
# ════════════════════════════════════════════════════════════════

C={"hdr":"5C2D0E","ouro":"C49A3C","zebra":"FDF6EF","borda":"E8D5B8","br":"FFFFFF",
   "e1":"D4EDDA","e2":"CCE5FF","e3":"FFF3CD","e4":"F8D7DA","done":"E2E3E5"}

def fill(c): return PatternFill("solid",fgColor=c)
def brd():
    s=Side(style="thin",color=C["borda"])
    return Border(left=s,right=s,top=s,bottom=s)
def ctr(): return Alignment(horizontal="center",vertical="center",wrap_text=True)
def esq(): return Alignment(horizontal="left",vertical="top",wrap_text=True)

def hdr_cell(ws, row, col, val):
    c=ws.cell(row=row,column=col,value=val)
    c.font=Font(name="Arial",bold=True,size=9,color=C["ouro"])
    c.fill=fill(C["hdr"]); c.alignment=ctr(); c.border=brd()
    return c

def aba_leads(wb, nome_aba, leads, canal_id, historico):
    ws=wb.create_sheet(nome_aba)
    ws.merge_cells("A1:Q1")
    ws["A1"]=f"PAO DE LÓ TI'PIEDADE — {nome_aba} — Semana {semana_num()}"
    ws["A1"].font=Font(name="Arial",bold=True,size=13,color=C["br"])
    ws["A1"].fill=fill(C["hdr"]); ws["A1"].alignment=ctr(); ws.row_dimensions[1].height=30
    ws.merge_cells("A2:Q2")
    ws["A2"]=f"{semana_datas()} · Pão de Ló Ti'Piedade Unidose 85g"
    ws["A2"].font=Font(name="Arial",size=9,italic=True,color="6B5744")
    ws["A2"].alignment=ctr(); ws.row_dimensions[2].height=16

    hdrs=["#","Espaço / Empresa","Tipo","Tipologia","Zona","Morada","Tel","Email","Gancho","Prio","E1","E2","E3","E4","Próx.Email","Estado","Observações"]
    for col,h in enumerate(hdrs,1): hdr_cell(ws,3,col,h)
    ws.row_dimensions[3].height=26

    for i,lead in enumerate(leads):
        row=4+i; k=lead_key(canal_id,lead)
        hist=historico["leads"].get(k,{})
        nums={e["num"]:e["data"] for e in hist.get("emails_enviados",[])}
        proximo=calcular_email_num(historico,canal_id,lead)
        cor=C["zebra"] if i%2==0 else C["br"]
        base=[i+1,lead["n"],lead["t"],lead.get("tCliente",""),lead.get("zona",""),lead.get("m",""),lead.get("tel",""),lead.get("email",""),lead.get("gancho",""),lead.get("p","")]
        for col,v in enumerate(base,1):
            c=ws.cell(row=row,column=col,value=v)
            c.fill=fill(cor); c.border=brd()
            c.font=Font(name="Arial",size=9,bold=(col==2))
            c.alignment=ctr() if col in [1,3,7,10] else esq()
        cores_e={1:C["e1"],2:C["e2"],3:C["e3"],4:C["e4"]}
        for n in [1,2,3,4]:
            col=10+n; data_e=nums.get(n,"")
            c=ws.cell(row=row,column=col,value=data_e if data_e else "—")
            c.fill=fill(cores_e[n] if data_e else cor); c.border=brd()
            c.font=Font(name="Arial",size=9,bold=bool(data_e)); c.alignment=ctr()
        estado=hist.get("estado","Em aberto")
        c=ws.cell(row=row,column=15,value=f"Email {proximo}" if proximo else "✓ Pronto p/ visita")
        c.fill=fill(C["e3"] if proximo else C["e1"]); c.border=brd()
        c.font=Font(name="Arial",size=9,bold=True); c.alignment=ctr()
        c=ws.cell(row=row,column=16,value=estado)
        c.fill=fill(cor); c.border=brd(); c.font=Font(name="Arial",size=9); c.alignment=ctr()
        c=ws.cell(row=row,column=17,value=hist.get("notas",""))
        c.fill=fill(cor); c.border=brd(); c.font=Font(name="Arial",size=9); c.alignment=esq()
        ws.row_dimensions[row].height=34

    for col,w in enumerate([4,28,16,22,18,32,13,26,38,8,12,12,12,12,16,16,28],1):
        ws.column_dimensions[get_column_letter(col)].width=w
    ws.freeze_panes="K4"; ws.auto_filter.ref=f"A3:Q{3+len(leads)}"

def aba_historico(wb, historico):
    ws=wb.create_sheet("📊 Histórico",0)
    ws.merge_cells("A1:K1")
    ws["A1"]=f"TI'PIEDADE — Histórico Completo de Prospeção (atualizado {hoje()})"
    ws["A1"].font=Font(name="Arial",bold=True,size=13,color=C["br"])
    ws["A1"].fill=fill(C["hdr"]); ws["A1"].alignment=ctr(); ws.row_dimensions[1].height=28
    hdrs=["Comercial","Canal","Lead","Tipo","Zona","Email 1","Email 2","Email 3","Email 4","Estado","Notas"]
    for col,h in enumerate(hdrs,1): hdr_cell(ws,2,col,h)
    ws.row_dimensions[2].height=24
    row_h=3
    for k,v in sorted(historico["leads"].items(),key=lambda x:(x[1].get("comercial",""),x[1].get("canal",""))):
        nums={e["num"]:e["data"] for e in v.get("emails_enviados",[])}
        completo=len(nums)>=4
        cor=C["e1"] if completo else C["zebra"] if row_h%2==0 else C["br"]
        vals=[v.get("comercial",""),v.get("canal",""),v.get("nome",""),v.get("tipo",""),v.get("zona",""),
              nums.get(1,"—"),nums.get(2,"—"),nums.get(3,"—"),nums.get(4,"—"),
              v.get("estado","Em aberto"),v.get("notas","")]
        for col,val in enumerate(vals,1):
            c=ws.cell(row=row_h,column=col,value=val)
            c.fill=fill(cor); c.border=brd()
            c.font=Font(name="Arial",size=9,bold=(col==3)); c.alignment=ctr() if col in [6,7,8,9,10] else esq()
        ws.row_dimensions[row_h].height=20; row_h+=1
    for col,w in enumerate([14,16,28,16,18,13,13,13,13,16,30],1):
        ws.column_dimensions[get_column_letter(col)].width=w
    ws.freeze_panes="A3"; ws.auto_filter.ref=f"A2:K{row_h-1}"

def criar_excel(sem, historico):
    wb=openpyxl.Workbook(); wb.remove(wb.active)
    aba_historico(wb,historico)

    # Abas HORECA por comercial
    for com_id in ["nuno","joao","oscar"]:
        com=COMERCIAIS[com_id]
        leads=gerar_leads_horeca(com_id,sem)
        aba_leads(wb,f"HORECA — {com['nome']}",leads,com_id,historico)

    # Aba Catering & Eventos (Rui)
    leads_cat=gerar_leads_canal(DB_CATERING,"rui_catering",sem,5)
    aba_leads(wb,"Catering & Eventos — Rui",leads_cat,"rui_catering",historico)

    # Aba Distribuidores (Rui)
    leads_dist=gerar_leads_canal(DB_DISTRIBUIDORES,"rui_distribuidores",sem,5)
    aba_leads(wb,"Distribuidores — Rui",leads_dist,"rui_distribuidores",historico)

    fname=f"TiPiedade_Leads_S{sem}_{datetime.date.today().year}.xlsx"
    wb.save(fname); return fname

# ════════════════════════════════════════════════════════════════
# NURTURING — CONTEÚDO PERSONALIZADO POR TIPOLOGIA
# ════════════════════════════════════════════════════════════════

def tipologia_grupo(tipo):
    """Mapeia tipologia do lead para grupo de email."""
    t = (tipo or "").lower()
    if any(x in t for x in ["hotel","resort","pousada"]):            return "hotel"
    if any(x in t for x in ["pastelaria","padaria","confeitaria"]):  return "pastelaria"
    if any(x in t for x in ["catering","evento","banquete"]):        return "catering"
    if any(x in t for x in ["distribuidor","congelado","frigori"]):  return "distribuidor"
    if any(x in t for x in ["mercearia","gourmet","deli","garrafeira"]): return "gourmet"
    if any(x in t for x in ["café","coffee","brunch","chá"]):        return "cafe"
    if any(x in t for x in ["supermercado","mercado"]):              return "super"
    if any(x in t for x in ["cervejaria"]):                          return "cervejaria"
    return "restaurante"  # default

# ── Assuntos por tipologia e número de email ──────────────────
ASSUNTOS = {
    "restaurante": {
        1: "Uma sobremesa portuguesa que se vende sozinha — e sem pasteleiro",
        2: "Como um restaurante da vossa zona eliminou as quebras em sobremesas",
        3: "A tendência que está a mudar as cartas de sobremesa em Portugal",
        4: "Queremos passar com uma amostra — leva 10 minutos",
    },
    "hotel": {
        1: "Sobremesa artesanal portuguesa para o vosso F&B — sem complexidade",
        2: "Como um hotel da região passou a oferecer sobremesas de qualidade sem pasteleiro",
        3: "O que os hóspedes estão a pedir mais nos hotéis portugueses",
        4: "Podemos passar com amostras para a equipa de F&B experimentar",
    },
    "pastelaria": {
        1: "Pão de Ló Ti'Piedade — um artesanal com 40 anos para complementar a vossa vitrina",
        2: "Como outras pastelarias estão a diferenciar-se com o pão de ló Ti'Piedade",
        3: "Produto artesanal, dose individual, margem elevada — faz sentido para vocês?",
        4: "Queremos deixar amostras — sem compromisso, só para provar",
    },
    "catering": {
        1: "Sobremesa individual artesanal para os vossos eventos — sem logística complexa",
        2: "Como empresas de catering estão a usar o pão de ló Ti'Piedade em eventos premium",
        3: "A sobremesa que mais impressiona em casamentos e eventos corporativos",
        4: "Podemos reunir para apresentar o produto e condições para catering",
    },
    "distribuidor": {
        1: "Pão de Ló Ti'Piedade em congelado — referência premium para a vossa carteira",
        2: "Uma marca com 40 anos e procura crescente no canal HORECA",
        3: "Zonas sem cobertura, margem interessante — vale a pena conversar",
        4: "Podemos apresentar as condições comerciais para distribuição regional",
    },
    "gourmet": {
        1: "Pão de Ló Ti'Piedade — 40 anos de receita artesanal para a vossa prateleira",
        2: "Como mercearias gourmet estão a destacar-se com o pão de ló Ti'Piedade",
        3: "Produto com narrativa, origem e história — exatamente o que os vossos clientes procuram",
        4: "Queremos passar para mostrar o produto e proposta comercial",
    },
    "cafe": {
        1: "O pairing perfeito para o vosso café — pão de ló artesanal em dose individual",
        2: "Como cafés de especialidade estão a aumentar o ticket médio com o Ti'Piedade",
        3: "Produto artesanal português com história — ideal para o público que frequenta o vosso espaço",
        4: "Queremos passar com amostras para a vossa equipa provar",
    },
    "super": {
        1: "Pão de Ló Ti'Piedade — sobremesa artesanal congelada para o vosso linear",
        2: "Como supermercados independentes estão a diferenciar-se com produto artesanal",
        3: "Sem grandes grupos, sem intermediários — direto do produtor para o vosso linear",
        4: "Podemos reunir para apresentar condições comerciais e produto",
    },
    "cervejaria": {
        1: "A sobremesa portuguesa que fecha qualquer refeição — Pão de Ló Ti'Piedade",
        2: "Grandes volumes, zero desperdício — como o Ti'Piedade resolve a sobremesa em cervejarias",
        3: "Dose individual congelada: qualidade constante, custo previsível, margem alta",
        4: "Queremos passar para deixar amostras e apresentar condições",
    },
}

# ── Corpos de email por tipologia ─────────────────────────────
def corpo_email(num, grupo, nome_lead, zona, nome_comercial, tel_comercial=""):
    """Gera o corpo do email de nurturing personalizado por tipologia e número."""

    assinatura = """Com os melhores cumprimentos,

Rui Bernardes
Departamento Comercial | Pão de Ló Ti'Piedade
sales@tipiedade.com | www.tipiedade.com"""

    corpos = {
        "restaurante": {
            1: f"""Exmo(a). Sr(a),

O meu nome é {nome_comercial} e represento o Pão de Ló Ti'Piedade — empresa portuguesa em terceira geração familiar, com mais de 40 anos de história na doçaria artesanal.

Estamos a trabalhar com restaurantes em {zona} e identificámos {nome_lead} como um espaço com o perfil certo para o nosso produto.

O Pão de Ló Ti'Piedade em unidose de 85g (congelado) resolve um problema que muitos restaurantes conhecem bem: ter uma sobremesa de qualidade na carta, sem desperdício e sem precisar de pasteleiro.

✦ Descongela rapidamente — pronto a servir em minutos
✦ Quatro sabores: Original, Chocolate, Canela e Café
✦ Ingredientes simples, sem conservantes, receita com 40 anos
✦ Margem confortável para o operador

Nas próximas semanas partilharemos alguns exemplos de como está a funcionar noutros restaurantes da vossa zona.

{assinatura}""",

            2: f"""Exmo(a). Sr(a),

Na semana passada apresentámos o Pão de Ló Ti'Piedade. Hoje queremos partilhar um caso concreto.

Um restaurante de peixe da vossa zona — com clientela local e de fim-de-semana — introduziu o Ti'Piedade há cerca de dois meses. O que nos disseram:

→ A sobremesa passou a ser a mais pedida, acima do pudim e da mousse
→ Desperdício zero — serve apenas o que precisa, quando precisa
→ O responsável: "É a coisa mais fácil que temos na cozinha. Sai do congelador, vai ao prato."

Em restaurantes de peixe e marisco, o pão de ló é o final natural de uma refeição. Os clientes reconhecem-no e pedem-no.

Se quiser saber mais sobre como funciona na prática, estamos disponíveis para uma conversa rápida.

{assinatura}""",

            3: f"""Exmo(a). Sr(a),

Uma tendência que estamos a observar no setor: os clientes pedem cada vez mais sobremesas com história e origem, mas os operadores não querem complexidade.

O pão de ló é a sobremesa portuguesa mais reconhecida fora de Portugal — e a receita Ti'Piedade foi levada pelos portugueses ao Japão no século XVI, onde existe até hoje. É uma história que se conta em segundos e que os clientes valorizam.

Para {nome_lead}, isso significa uma sobremesa diferenciada na carta, sem investimento em pastelaria, com custo fixo por dose e margem previsível.

Na semana que vem a nossa equipa passa pela vossa zona. Gostaríamos de deixar amostras para a equipa de cozinha experimentar.

{assinatura}""",

            4: f"""Exmo(a). Sr(a),

Ao longo das últimas semanas partilhámos a história do Ti'Piedade e alguns exemplos de como está a funcionar noutros restaurantes.

Esta semana {nome_comercial} vai passar por {zona} e gostaria de parar 10 minutos em {nome_lead} para deixar amostras dos quatro sabores — Original, Chocolate, Canela e Café.

Sem reunião, sem apresentação. Só o produto, para provarem.

Se preferir agendar: {tel_comercial if tel_comercial else "responda a este email"}.

{assinatura}""",
        },

        "hotel": {
            1: f"""Exmo(a). Sr(a),

O meu nome é {nome_comercial} e represento o Pão de Ló Ti'Piedade — empresa portuguesa com mais de 40 anos de história na doçaria artesanal.

Contacto-o(a) porque trabalhamos com vários hotéis na região e identificámos um desafio comum: manter uma carta de sobremesas de qualidade constante, especialmente em períodos de maior ocupação, sem depender de pasteleiro especializado.

O Pão de Ló Ti'Piedade em unidose de 85g (congelado) resolve exatamente isso:

✦ Qualidade artesanal consistente, independentemente do volume
✦ Zero preparação — descongela e serve
✦ Pode ser usado no restaurante, pequeno-almoço, room service ou eventos
✦ Quatro sabores: Original, Chocolate, Canela e Café

Nas próximas semanas partilhamos exemplos de como está a funcionar em unidades hoteleiras semelhantes à vossa.

{assinatura}""",

            2: f"""Exmo(a). Sr(a),

Partilhamos hoje um exemplo real de como o Pão de Ló Ti'Piedade está a ser usado no setor hoteleiro.

Um hotel de 4* na região centro — com restaurante próprio e room service — introduziu o produto há três meses. O responsável de F&B partilhou connosco:

→ A sobremesa passou a fazer parte do menu de room service com grande aceitação
→ A consistência de qualidade eliminou reclamações sobre sobremesas
→ O custo por dose é previsível e a margem é superior à pastelaria fresca

Para hotéis com eventos e grupos, a dose individual em congelado é especialmente eficiente: serve exatamente o que precisa, sem perdas.

{assinatura}""",

            3: f"""Exmo(a). Sr(a),

Os hóspedes internacionais estão cada vez mais atentos à autenticidade dos produtos que consomem — especialmente em contexto de viagem.

O Pão de Ló Ti'Piedade tem uma história que funciona: receita levada pelos portugueses ao Japão no século XVI, mantida intacta há 40 anos, produzida em Portugal por uma família em terceira geração. É o tipo de produto que um hóspede leva na memória.

Para a equipa de F&B de {nome_lead}, isso significa uma sobremesa com valor percebido elevado, sem complexidade operacional.

Na próxima semana podemos passar para uma conversa rápida ou deixar amostras.

{assinatura}""",

            4: f"""Exmo(a). Sr(a),

Nas últimas semanas apresentámos o Ti'Piedade e partilhámos casos reais de utilização em hotéis.

Esta semana gostaríamos de passar por {nome_lead} para deixar amostras dos quatro sabores com a equipa de F&B. Demora menos de 15 minutos e não requer compromisso.

Para agendar: {tel_comercial if tel_comercial else "responda a este email"}.

{assinatura}""",
        },

        "pastelaria": {
            1: f"""Exmo(a). Sr(a),

O meu nome é {nome_comercial} e represento o Pão de Ló Ti'Piedade — empresa portuguesa em terceira geração, com mais de 40 anos a produzir pão de ló artesanal com a receita original da D.ª Piedade.

Contactamos {nome_lead} porque acreditamos que o nosso produto pode ser um complemento interessante ao vosso catálogo — não para competir com o que já têm, mas para adicionar uma referência nacional em formato individual.

O Pão de Ló Ti'Piedade em unidose de 85g (congelado):

✦ Dose individual pronta a vender ao balcão após descongelamento
✦ Quatro sabores: Original, Chocolate, Canela e Café
✦ Receita secular, ingredientes simples, sem conservantes
✦ Margem de revenda competitiva

Nas próximas semanas partilhamos exemplos de como outras pastelarias estão a trabalhar o produto.

{assinatura}""",

            2: f"""Exmo(a). Sr(a),

Uma pastelaria de referência em Lisboa começou a trabalhar o Ti'Piedade há quatro meses. O que nos disseram:

→ O produto tornou-se uma das referências de venda por impulso ao balcão
→ A dose individual permite controlo de stock sem perdas
→ "Os clientes pedem porque reconhecem. Não precisamos de explicar o que é."

O pão de ló Ti'Piedade não concorre com a produção própria — complementa-a. É uma referência nacional que os clientes procuram.

{assinatura}""",

            3: f"""Exmo(a). Sr(a),

O consumidor de pastelaria está a valorizar cada vez mais os produtos com origem, história e receita verificável.

O Ti'Piedade tem exatamente isso: 40 anos de receita da D.ª Piedade, produzido em Portugal, ingredientes simples e sem aditivos industriais. É um produto que se explica em duas frases e que os clientes entendem imediatamente.

Para {nome_lead}, adicionar o Ti'Piedade ao balcão é adicionar uma referência com valor percebido elevado e sem esforço de produção.

Podemos passar na próxima semana com amostras?

{assinatura}""",

            4: f"""Exmo(a). Sr(a),

Esta semana {nome_comercial} passa por {zona} e gostaria de parar em {nome_lead} para deixar amostras dos quatro sabores — para a equipa provar e avaliar se faz sentido para o vosso balcão.

Sem compromisso. Só o produto.

Para agendar: {tel_comercial if tel_comercial else "responda a este email"}.

{assinatura}""",
        },

        "catering": {
            1: f"""Exmo(a). Sr(a),

O meu nome é {nome_comercial} e represento o Pão de Ló Ti'Piedade — empresa portuguesa com mais de 40 anos de história na doçaria artesanal.

Para empresas de catering e organização de eventos, o desafio das sobremesas é sempre o mesmo: qualidade constante, facilidade logística e custo controlado.

O Pão de Ló Ti'Piedade em unidose de 85g (congelado) responde exatamente a isso:

✦ Dose individual — sem corte, sem preparação, sem desperdício
✦ Qualidade artesanal consistente em qualquer volume
✦ Quatro sabores para variar por evento: Original, Chocolate, Canela e Café
✦ Embalagem individual elegante — adequada para serviço em mesa

{assinatura}""",

            2: f"""Exmo(a). Sr(a),

Uma empresa de catering de casamentos em Lisboa começou a usar o Ti'Piedade há seis meses. O feedback dos clientes foi imediato:

→ O pão de ló é reconhecido como "sobremesa portuguesa de sempre" — funciona como momento de identidade no menu
→ Em eventos com 200+ convidados, a dose individual eliminou toda a logística de corte e emplatamento
→ "É o produto mais fácil de gerir num evento. Zero perdas, zero improvisação."

Para casamentos e eventos corporativos, o Ti'Piedade posiciona-se como a sobremesa com narrativa — algo que fica na memória dos convidados.

{assinatura}""",

            3: f"""Exmo(a). Sr(a),

A tendência em eventos premium é clara: os convidados querem autenticidade e os organizadores querem simplicidade operacional.

O Pão de Ló Ti'Piedade é a interseção perfeita: produto artesanal com 40 anos de história, dose individual pronta a servir, e uma narrativa que o responsável de sala consegue contar em segundos.

Temos condições comerciais específicas para catering e eventos, com preços por volume. Podemos apresentar?

{assinatura}""",

            4: f"""Exmo(a). Sr(a),

Esta semana {nome_comercial} está disponível para uma reunião com {nome_lead} para apresentar o produto, condições para catering e deixar amostras dos quatro sabores.

Para agendar: {tel_comercial if tel_comercial else "responda a este email"}.

{assinatura}""",
        },

        "distribuidor": {
            1: f"""Exmo(a). Sr(a),

O meu nome é {nome_comercial} e represento o Pão de Ló Ti'Piedade — empresa portuguesa em terceira geração, com mais de 40 anos a produzir doçaria artesanal regional.

Estamos a alargar a nossa rede de distribuição de congelados e identificámos {nome_lead} como um parceiro com cobertura na zona que nos interessa.

O que propomos:

✦ Pão de Ló Ti'Piedade unidose 85g (congelado) — produto de alto valor percebido
✦ Margem de distribuição competitiva
✦ Procura crescente no canal HORECA e retalho gourmet
✦ Marca com 40 anos e reconhecimento nacional

Nas próximas semanas partilhamos mais detalhes sobre o produto e condições.

{assinatura}""",

            2: f"""Exmo(a). Sr(a),

O Pão de Ló Ti'Piedade tem hoje distribuição em mais de 500 pontos de venda em Portugal — mas existem regiões com potencial por desenvolver.

A procura no canal HORECA tem crescido consistentemente: restaurantes, hotéis e pastelarias procuram um produto de doçaria artesanal congelado com qualidade constante e que os clientes reconheçam.

Para um distribuidor com a vossa cobertura regional, o Ti'Piedade é uma referência com diferenciação clara face aos produtos industriais existentes no mercado.

{assinatura}""",

            3: f"""Exmo(a). Sr(a),

As condições que praticamos para distribuidores regionais incluem:

✦ Preços por volume com margens ajustadas à distribuição
✦ Apoio em materiais de comunicação para os pontos de venda
✦ Flexibilidade de encomenda (MOQ negociável por zona)
✦ Produto com validade longa em congelado — sem pressão de rotação

Podemos reunir para apresentar a proposta comercial completa?

{assinatura}""",

            4: f"""Exmo(a). Sr(a),

Esta semana {nome_comercial} está disponível para uma reunião com {nome_lead} para apresentar as condições de distribuição e deixar amostras do produto.

Para agendar: {tel_comercial if tel_comercial else "responda a este email"}.

{assinatura}""",
        },

        "gourmet": {
            1: f"""Exmo(a). Sr(a),

O meu nome é {nome_comercial} e represento o Pão de Ló Ti'Piedade — empresa portuguesa em terceira geração, com mais de 40 anos de história.

A nossa receita — criada pela D.ª Piedade e mantida intacta — é uma das referências da doçaria artesanal portuguesa. Hoje chegamos a mais de 500 pontos de venda em todo o país.

Para espaços como {nome_lead}, o Ti'Piedade é um produto com narrativa forte e valor percebido elevado:

✦ Receita secular levada pelos portugueses ao Japão no século XVI
✦ Produção artesanal, ingredientes simples, sem conservantes
✦ Unidose 85g em congelado — fácil de expor e vender
✦ Quatro sabores: Original, Chocolate, Canela e Café

{assinatura}""",

            2: f"""Exmo(a). Sr(a),

Uma mercearia gourmet em Cascais começou a trabalhar o Ti'Piedade há três meses. O que nos disseram:

→ "Os clientes compram porque reconhecem a marca. Não precisamos de explicar."
→ O produto tornou-se uma das referências de doçaria nacional no espaço
→ A dose individual em congelado facilita a gestão de stock sem perdas

Para espaços gourmet, o Ti'Piedade é exatamente o tipo de produto que os clientes procuram: artesanal, com história, de origem verificável.

{assinatura}""",

            3: f"""Exmo(a). Sr(a),

O consumidor gourmet valoriza três coisas: origem, história e autenticidade. O Ti'Piedade tem as três.

40 anos de receita intacta. Produção familiar em terceira geração. Ingredientes simples que qualquer cliente consegue ler e entender.

Para {nome_lead}, adicionar o Ti'Piedade ao linear é adicionar uma referência com identidade forte e fácil de comunicar.

Podemos passar na próxima semana?

{assinatura}""",

            4: f"""Exmo(a). Sr(a),

Esta semana {nome_comercial} passa por {zona} e gostaria de visitar {nome_lead} para apresentar o produto e proposta comercial.

Para agendar: {tel_comercial if tel_comercial else "responda a este email"}.

{assinatura}""",
        },

        "cafe": {
            1: f"""Exmo(a). Sr(a),

O meu nome é {nome_comercial} e represento o Pão de Ló Ti'Piedade — empresa portuguesa com mais de 40 anos de história na doçaria artesanal.

Para cafés e espaços de brunch, o Ti'Piedade resolve um problema que conhecemos bem: ter uma opção de doçaria de qualidade, fácil de servir, com margem interessante.

O Pão de Ló Ti'Piedade em unidose de 85g (congelado):

✦ Descongela rapidamente — pronto ao balcão em pouco tempo
✦ Pairing natural com café de especialidade
✦ Produto reconhecido — os clientes já conhecem e pedem
✦ Quatro sabores: Original, Chocolate, Canela e Café

{assinatura}""",

            2: f"""Exmo(a). Sr(a),

Um café de especialidade em Lisboa introduziu o Ti'Piedade há dois meses. O responsável partilhou:

→ "O pão de ló com café tornou-se o nosso produto de tarde mais vendido."
→ A dose individual eliminou desperdício — antes tinham bolos inteiros que nem sempre vendiam
→ O ticket médio da tarde subiu com a combinação café + fatia

Para espaços como {nome_lead}, o Ti'Piedade é a opção de doçaria artesanal que encaixa no posicionamento sem exigir produção própria.

{assinatura}""",

            3: f"""Exmo(a). Sr(a),

O público de café de especialidade e brunch é exatamente o público que mais valoriza produto artesanal com história.

O Ti'Piedade tem 40 anos de receita intacta, ingredientes simples e produção familiar — é o tipo de produto que fica bem numa ardósia e que o barista consegue explicar em duas frases.

Podemos passar na próxima semana com amostras?

{assinatura}""",

            4: f"""Exmo(a). Sr(a),

Esta semana {nome_comercial} passa por {zona} e gostaria de parar em {nome_lead} para deixar amostras dos quatro sabores.

Para agendar: {tel_comercial if tel_comercial else "responda a este email"}.

{assinatura}""",
        },

        "super": {
            1: f"""Exmo(a). Sr(a),

O meu nome é {nome_comercial} e represento o Pão de Ló Ti'Piedade — empresa portuguesa com mais de 40 anos de história.

Estamos a alargar a presença em supermercados e mercearias independentes e identificámos {nome_lead} como um parceiro com o perfil certo.

O que propomos:

✦ Pão de Ló Ti'Piedade unidose 85g (congelado) — linear de congelados ou balcão
✦ Produto com reconhecimento nacional e procura crescente
✦ Quatro sabores: Original, Chocolate, Canela e Café
✦ Condições comerciais para retalhistas independentes

{assinatura}""",

            2: f"""Exmo(a). Sr(a),

Supermercados independentes que trabalham o Ti'Piedade reportam consistentemente o mesmo resultado: o produto vende sem esforço de comunicação, porque os clientes já o conhecem.

A dose individual em congelado é especialmente adequada para o retalho independente — sem risco de validade, sem perdas, rotação previsível.

Para {nome_lead}, é uma referência de doçaria artesanal que diferencia o linear face aos produtos industriais da grande distribuição.

{assinatura}""",

            3: f"""Exmo(a). Sr(a),

Trabalhar com fornecedores fora dos grandes grupos tem vantagens claras: preços mais competitivos, relação direta e produto diferenciado.

O Ti'Piedade é exatamente isso — produto artesanal, familiar, com 40 anos, que os consumidores reconhecem e que não encontram nos lineares da grande distribuição.

Podemos reunir para apresentar condições?

{assinatura}""",

            4: f"""Exmo(a). Sr(a),

Esta semana {nome_comercial} passa por {zona} e gostaria de visitar {nome_lead} com amostras e proposta comercial.

Para agendar: {tel_comercial if tel_comercial else "responda a este email"}.

{assinatura}""",
        },

        "cervejaria": {
            1: f"""Exmo(a). Sr(a),

O meu nome é {nome_comercial} e represento o Pão de Ló Ti'Piedade — empresa portuguesa com mais de 40 anos de história na doçaria artesanal.

Para cervejarias com grande volume de refeições, a sobremesa é frequentemente o elemento mais difícil de gerir: custo variável, desperdício, e falta de consistência.

O Pão de Ló Ti'Piedade em unidose de 85g (congelado) resolve tudo isso:

✦ Dose individual — serve só o que precisa, zero perdas
✦ Custo fixo e previsível por sobremesa
✦ Zero preparação — direto do congelador ao prato
✦ Clientes reconhecem e pedem — sem esforço de venda

{assinatura}""",

            2: f"""Exmo(a). Sr(a),

Uma cervejaria de referência em Lisboa introduziu o Ti'Piedade há quatro meses. O responsável de sala:

→ "Acabámos com o problema das sobremesas. O pão de ló sai sempre bem, em qualquer quantidade."
→ O desperdício em sobremesas caiu para zero
→ A margem por sobremesa aumentou face à mousse e pudim que tinham antes

Em cervejarias com volume alto e rotatividade rápida, a dose individual congelada é a solução mais eficiente.

{assinatura}""",

            3: f"""Exmo(a). Sr(a),

A sobremesa numa cervejaria precisa de ser rápida, consistente e com boa margem. O Ti'Piedade é tudo isso.

40 anos de receita portuguesa intacta. Produto que os clientes reconhecem e pedem sem precisar de ler a carta. Dose individual que se serve em segundos.

Para {nome_lead}, isso significa mais receita por mesa com menos complexidade operacional.

Podemos passar na próxima semana?

{assinatura}""",

            4: f"""Exmo(a). Sr(a),

Esta semana {nome_comercial} passa por {zona} e gostaria de parar em {nome_lead} para deixar amostras e mostrar como funciona na prática.

Para agendar: {tel_comercial if tel_comercial else "responda a este email"}.

{assinatura}""",
        },
    }

    grupo_corpos = corpos.get(grupo, corpos["restaurante"])
    return grupo_corpos.get(num, grupo_corpos[1])


def corpo_html(email_num, grupo, nome_lead, zona, nome_comercial, texto_plain):
    """Gera email HTML com ícones PNG (compatível todos os clientes) e logo elegante."""

    LOGO = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAABYIAAAJ3CAYAAAAplCU8AAAACXBIWXMAAC4jAAAuIwF4pT92AAAAGXRFWHRTb2Z0d2FyZQBBZG9iZSBJbWFnZVJlYWR5ccllPAAAzkNJREFUeNrs3f91m0j79/Frv+f+39oKzFZgbQUmFVhbgUkFcSoIqSBOBSEVxKkguIKVK1hcwW1XsI+vm5lHGPNjZvghJL1f5+jYkRGCYUDho+Hit3///VcAAACAPdi+PPKXxw1NAQAAAEzr/2gCAAAA7MHty+Pi5fHh5bGhOQAAAIBp/caIYAAAAMxMg98flX8/vzzWL4+CpgEAAACmQRAMAACAOa2kDHzPas/fvzximgcAAACYBqUhAAAAMCctCXHW8PzlyyOheQAAAIBpMCIYAAAAc4lfHr86/q4lIqKXxxNNBQAAAIyLEcEAAADHLTKPJUh7/q4jhZOFLGtM1wEAAMAxIQgGAAA4fv9IWZJhtcdliKQs/9DnZs9tpcuZm/YCAAAAjgZBMAAAwHErXh6PL48P5vdU9hMIbxynO395rPewfNHLI5MyNNfAOqfrAAAA4JgQBAMAABy/3PzU0gufXh5bmb8Ew2aiaYfSUDw1bXLd0GYAAADAUSAIBgAAOH557d866vabeT6aaRnWE007RCxlAKzh+FlPmwEAAAAHjSAYAADg+OUtz2sJBA1C5whezzymnaN0RfLy+CVlKF738PJ4otsAAADgmBAEAwAAHL9CyjrBTTSgzWU/dXn3JZFyRHSbnC4DAACAY0MQDAAAcBryjr/ZMHh1Au0QS3cI3NdWAAAAwEEiCAYAADgNec/fNQy+nfD9H0Zc1iGyPb8/AAAAsBe//fvvv7QCAADA8YteHv84TPe7+NfHjc1Dy0vURxXrvHLz9yvH+X2X3U3stI7xs5S1jFVhHpn56WPz8vjRM82DnFaZDAAAAJyI/9AEAAAAJ6GQsk7wec908cvjznPe+ppPHX+/8pzfde3fOlr50vxuf+biHwTHDtPkdBUAAAAcI0pDAAAAnI7cYZo4YL7ZzOvxKGGBbTxSGwEAAAAHhyAYAADgdOQO04SURSikLOcwlzTwdRcjtREAAABwcAiCAQAATkc+4bzTmdZBRwNnE81b6wM/0U0AAABwjKgRDAAAcDoKKcPOC4/XxA3P6Y3bnhrm/Vm6awWPIWl5vu1GdVuPeed0EQAAABwrgmAAAKYVif8NrYAp5dIdBGtwupEycO26yduzlCFxNWhNzXOXEy37Z3kb1upyfutZTr353a303ywvp3tggfQLDkaqAwCAwSgNAQDAtHSUYuRwkg/MJe/5e/zy+CHdIbDVNNpWQ+SHCZZbaxCnjstQdfbyuBa3kDene2APVj1/29BEAABgDATBAABMS0ci3vRMw0k+5pT3/N21bMRty/M6clG/APk60vLqiN730l4SQoPge4f5aCDcNRqY+sDYh77Ph9R8jgAAAAxGEAwAwPRupftGWoU012EFpqBh59ARu8/SHgRbGnD9IWUgfG8eP6Usz9DmvrJs+vtHKUfUZz3vlY7QLjldAzOLKvtkk0TKEJgvKAAAwCgIggEAmF5hfsYtf8+FUcGYVz7w9am4hVPa929M349NPy96lktHE/9mpr91fB993c89twngS/eNrOVvtqwQ/RIAAIyGIBgAgHmk5qS/rRakSwkJYCz5gNfqSN3bBa5TIuVI5X20CeArlrKsyVPHZ8YtzQQAAMZEEAwAwHxuOk7scymDAW4chznkga/Tsg1LHb3+ZPah58D14vJ7zCmV9tHAt+I+Gh4AAMAZQTAAAPMppBwB1jbytysoBsYUUidYRwLHsuxwamuW8dHzdTldAjNKpb2udUyfBAAAUyEIBgBgXrfmRH/d8LdCdqMaganljtNpqPpRlh8CW1uzf30W99HBOd0BM4lM/2zqc3pFSCqUCQIAABMhCAYAYH6JOdlvKgORSvtIMWBMec/fv7883kkZXB3aSPWnyj72zjzuO6bf0h0wE92X2oLeTAiBAQDAhAiCAQCYnw2pblv+RhiAOeQ9f8/kOEbK5uaxbvm7jngu6A6YQSzllw5N/U2P+XfClxIAAGBCBMEAAOzH1jyabryVCTeOw/T66gRHR7SuGgKftfwtpytgJm114NfmkdFEAABgSgTBAADsjwYCGgRHDX9LhRvHYXp5x9+ikd8rljLouuyYxgZlU7x3SBsAY7kx/b9eZ5u6wAAAYDYEwQAA7NeNNNcLtpcPxzQRJpRPPP+V6ePan3+9PK57ptdRux9eHv9IeZl8MtJyxHtsAyCScsTvXcPfbM3gJ5oJAABM7bd///2XVgAAYL80IEikeURYJuOFYUCdBrX/bfmb1s5NpQyvfEKqSMrgVUe7X42wjM9mGfSRByzLxqzHWcs6RnQDTCyT5rA3lV0NawAAgMkRBAMAsAy2RES9HERsHilNhInoaN2LnmnuZTdKvX4zq0h2Ix71cT7x8j7UluWpYZ+xy9O3Xt+FL1ow/bFd3Tke8wEAACZDEAwAwHLoiLFC3gYGGhRkwt3kMQ3tXx9OdN3fCzfownRWZv9Kas93XQUCAAAwGWoEAwCwHBoYxFKGBFU3QmCA6eSsOzDZMb1+7F61PA8AADA5RgQDALA8uexusGUxggxT6aoTfMyoD4wp6fFay5bc1fa1rPI3AACAWTEiGAAAf5Hs6j5Owd7calV5ztZCjWl+jEz71cMJrnfOpsdEbL3segh8Z47tTxO+L58RAACgFUEwAAD+CnNSv55o/hoSpFKGBtUwWJ/TEcERmwAjS14eH18ezx6v0Wm/vjzeSVlr99HzPauv9Q2i9SZvf5nHvedr7817pmx2TGBl+la9f2Xmue2E76tfIuZsAgAA0IbSEAAAhLuTaUcGx7KrG2xHkEVSBgoxzY8JaP/SoOqsZ7pn0werodbK/Pvc4X3qN2nT1+YvjwuH1340+0WVzuva4bUaICdsZkwoM4+89lwu096YUOd9I5ScAAAAHRgRDABAOHuZ71RyKQMv/WlHBhfmuYzmxwQKxz6t/a8+svHJ8bU/G/qvvjZxeO2jvA2BlWvtbGpsY0o3Zh/Ka/vK08TH7I15X0JgAADQiSAYAIBwmTkBX038Hrm8DoNtyYiETYAJuFy63hY4FQPm7/K+Rcfy9JW1eBSCMkwnNsfktHb81mP11F9A6HvesgkAAEAfgmAAAIa5meEEvDrKzIbBiTn5j9kEGJlLuZPI8/mq2PN5l/nrftFXzuKc/QUT0XrxWW3fyczzycTvncpu1DEAAEAngmAAAIbJK0HAlJLK+61kdyn93QzvjdOhfemDw3TXDf3O3iSrz6W8Dcf0tS5fqGiY2zS6MnNcP0pDYGwr2d0Irqj0Rw2FY5k2oLU3LWU0MAAAcMLN4gAAGC8I2MzwPoV5xLKryTpH4IDT4PMfw2fT7+2XEdoXzzxe/112oyY1oD33fG21RMql4+vuhVHBGNfWHJPt8T+rHJO3E7+3rVO/ZTMAAAAXBMEAAIzDjjScemSWhma5vA6Dt5W/AUNoX7o44vUjCMaYMtOf1rK7IZyOln8v09/Qc2MeCZsBAAC4ojQEAADjuDUn5FOHsRrUaeisYV1eCQQimT54wPHLWT/AiR6Hr81xvxoCf57hWGxLqVDqBAAAeCEIBgBgPLcyTxir7/FVyjBYfy+kvDz4mmAAA+U0AdBLv3z7Yo7DuTn26/H3p7jVyR7jM0AflAMCAABeCIIBABj35PxJ5rlxjwa+epn7tXnfW/NvDSdiNgUCUWsU6LY2x9wHKUPfRMobLD7IPGUaEuEGcQAAIBA1ggEAGJeeoP/98vhLyhv5TMnePE5v0PVRdqOD7XIUbA4E0H5zfqTrNkftVhwvPebmUl6N8aeUJXl+SHnjxFim/yLF1oi/oR8DAIAQBMEAAIyvWi+4mCEY+Nv8/k7KYOKblKPTuHkcQqRyvKPKNUBj1DNC6Zd7V1LWAdbfcym/iJvjCwYbQj8JV30AAIBABMEAAExzwr6tnLBPXcdRwy0tCaGj0uxly5dS1q+kZjAAjHecfTTHdQ2CdWTwd5mnJMSdeV+u9gAAAMEIggEAmIaesP+S+UKCXMrw9968nwbROlLtnXADMAAYIqodU/UYq/XZH2SeL/v0KhOtQ/xRqA0MAAAGIAgGAGD6k/c5RuZW6wV/Ns99knL0mo4g4+7yABAml/KLNv1iT0fm/jDPa53gqUuNJFKW+9Ev+WI2BQAAGIIgGACA6VRvLDRHDcmN7G5cZG8qpDf90mA4ZXMAgLdEyiB2H8fVWMorS/S9I+ELPQAAMBBBMAAA09KTd3tJ8RxhsL2ZkY4eS6UMEdQfQl1JAPBRvdLiozme61Uec9yM04bOlPgBAACj+T+aAACASRVSjtRVtzJ9eJBIOXrs0vz73vxM2RQA4EVL+mgQqyV29Eu2D5Xj7JSqIfB7IQQGAAAjIQgGAGB6uTmZPzO/TxkG66XDth5xJrvAQm9sFLEpAMDJqnIs1Z+p+V1LQkxZF7gaAn+X6a8iAQAAJ4QgGACAeejJ/EeZJwzW99KRwFrHMpYyTFApmwEAnNjRwHos1eBXv0zTkcG3E75nPQRO2AwAAGBM1AgGAGBemZSBgpZviGWckWWRvK3/q4HC31LWstTSFP+Y50NrBevoOG5UBODQhB679DW2Pm9ijttNdd7XIx3HCYEBAMDkGBEMAMC8EnOSP+bIYA16C3kdHGzN+1xIGRR/r7x/iJRNB+AAj7erwNfZ0cB6bL02v2eVaWJzDE9GWE5CYAAAMAuCYAAA5qcn+dWawUNP+vUmRloG4puUAXBsnk/Nz2p9y9D30tFxEZsOwIFYSfPVEi42lWNoUjue2tD2l5Q35cwHLichMAAAmA1BMAAA+5HJLgz+NvDkv5CyBITSEcAaUGg4rOGt3tjoyvyuIYMGxiGjkLUuZspmA3Ag9AuwPOB1K3PMfJTdF3W2TrAeB7XkzqWZ9tkca0PFsguBPwshMAAAmBhBMAAA+5NJWX9SwwQNg28GzCuv/VuDjMI8dP6J7C5rTgLmb2tsxmw2AAsXyW6kra/qaGD9/dzMR4PgDz3HXR96HNYv7TQEfi980QYAAGbAzeIAANg/DSwyKUfzhl4aHEsZKrR5kN1Njewl0770NXcyTl1jAJiKHqd09G4e+NrYHCf1uHzdMe1H8z6+9Eu/LzLuTUMBAAB6MSIYAID9s3V9f0oZOuTif4OjvOfvGjJrgKuhhY5wiwKWszDLesMmA7BQsYSPBravz8wxeDPwuNtE560h8IM5DhMCAwCA2RAEAwCwDFp6QUMHHWFmb0AUec7jvufvOs/C/B4HLmcm5SXMKzYZgAWyx6gQelzUUg1b8zjrmPZR/ELclTkG65d9X6UMq5/YXAAAYE4EwQAALIuO2P1TytBAQwafMgx9Ny06M9NogBE3/D1yeI9cyjCZUcHdbiRs1DXQZsV+1yuRXUmHPk37px4Xn2V35UTfsdCVHaGsP/9iOwIAgH0hCAYAYHlsAJxLeYf6xPF1ucM0GgafS3PAfOP4XrdmWkYFN9M2/GK2Y0JzYASx6U/arzKao1Xq2D6JNI8aXptj5NlIx1u77XTaJzP/OzYTAADYF4JgAACWyZaK0LvJ34rbDYk0KHp2nL+tGVylAcU36b+s2gYZjGp7KzZtqM7M79pehOYIpfu+3gjSjlDV0gIJzfJGYtro1mG6b/K2rMNKxr0Cw77XL7NMemwo2EwAAGCffvv3339pBQAAli2SMnQopAwWuupKZtJ9l/sqLRFRr1Np/2OgNSxvet4nFsof1LdTW11RDeg12M9pJjham/3souXvfwo3GquyX7jEHdPo8dN+UfOHvA5mdd+8dHyvB+kPje0xcsN2AgAAS8GIYAAAlq+QMnTQMCGX7gDC57JjHT2X1Z6zN5z7IN2jDm0NzZjN8z8r0/Ztl5Tr83ZkIKOD0Ue/hNGyMBcd0+ixIKKp/kfb4Uq6y0Lo8cyGwI/yOgROxT0Elp73sfXdpXLcBgAAWASCYAAADkcqZUCkgWPcMo29GZyrq9q88srvGpokLa/bmvdJ2Cz/owHvhcN0H6Q/zMfpikz/+OIwrb35I18slKNu7fGvyVpel4zY1tr8k+f7ZR3vk5v30mPjE5sGAAAsCUEwAADLoAGCS6CTm2lvpL10Q+b53tXp66PXbqU9tNTQZcOm+19Af+0xvQbGf0t/LWacFltC4NKzL93SdP8LXX9Kc/C6MsfNs5bjnO/x8nvL+2zMtth4zJMvhAAAwKyoEQwAwDJoWJGK3w3YdFoNEpKGeRXSXqagid6ULjPz+7v2t6ZawlKZ9i/xK0lxTDT0+THg9fdm+xUHsK5rOazgP5fDqMm8Mvve1YB5fJTTDYSjl8c/lWNYUz+oh+vvZPel2t+e7/dHw/6amu3oc/xOzXGT0hEAAGA2BMEAACxHImXY2hWqRvK2BmUqby9D1n9/83hvDSRj83vTfw7abh5XSBmoJCe4vexl4GcD5/NstuFSgzz7JcWHA9xGP02/LRa6fHb06NkI8zrVL2R0+2opjd/l7ZdV9m91NgjWtvcZzf9Z3o7kvzXzuqsco6Wnz9kvVVI+9gAAwJwIggEAWBYNExJpry2poZwGD7GZNpMycLC1g7e1efmMMrQj3dr+c2DDkypbC/PU6pTaMP58xHk+mO2YL2g9ddumI6/n3J5NP72V5dRstTVrL0dez1hOb4Tp1hy36qPVI/O3ppD9N/PzSdxD+Ad5XcrBHotT83siuy/T4p6+1necBwAAmAQ1ggEAWBYNFrKOv2twkEgZFuoITb2s2QY/+nw1qNB/P3i8tw1S7lv+3rRcuhxncnq1LjXIGTsc1Xqvv0ybxntct0h2o2i/yWGHwGL65yezPreyG7G5D2uzH+l+eznBeuq8T+lLmZXZb/KW41VXyBuLewj8LK+DZnslRmGOBX+b4/FW+kPg1LyGEBgAAMyOEcEAACxPLmVg1XeZd9Nlz8/mdamUIYW9UdKFw/vqZfQbaa6paTXV4dT/TDRdMn2sdP2vZ3ifB/NedzJ9aQMNtmIpvzy4OIFt+GDaNZfpR2CvzH6VyPjhb5NqmZdjp22qX1bU6/bq+v/qOgcyx6tPDu9RHWlt6wDrox4if5f+Ejmx2acjPuYAAMA+EAQDALA8tvas/ix6pk2kvRawhrO27qzOry/ge6i8d1tgpTeOi2rPaaCmAUl8Atumq72npNtmW3k8SVgJgJXZxpF5xObfZye+z91X2raoPHxFlYcN1/cRrLfV9D42mWnj+jGp6xj2v3MgcSudo8e7jekX2p5py77iEgLbcjKpdF/1AQAAMBmCYAAAlimTXVDXJ5X2kW0aZNi6sxoK941k/U36Q5T6qGA7Mvm3I98mGuz9vdBlu+/5+yW71GRtu9QgvWn0/rEp5O3NKvWY+atne8YOx7kH2X1RosfOtkDfJQS2x/S1nF4ZHQAAsCD/oQkAAFikVHajx1KHaSNpDnm1vusPKcs+JJV5DgmuNPjNKv/Ozc9YlnWjszGtFr5uBL20bd032Y1yPtZ98rxhv0xGmLe9mkKPlR86prt3fL/EHJ/fsTsBAIB94mZxAAAsUyFlEKEjfTcO0ydShr1trmR3ubuOSLsfsGwX8npUmwZNz3K8I91sCHxGt8SByeV4bx63rqxjdV8dUr9br6DQsPbOHNe6QuAHx2OzHVH8VY73izIAAHAgCIIBAFiuVMrANhO3mwslUoYTbTTI1NHBOqI3fnl8lDLArSsc3qsegGhocqxBcNdl4cCS6T6fH+m66THssXa82gyY32fZ1c/WEjDnHdM+m/d66pnnyhy/n+R0bqYJAAAWjCAYAIBls8GGvSFbFw0bEmkOd6t0lFsuu4D5u3nevq7wWC4rF7ew+tBoaH5NN8QBu5DjrBUcyduQOw6Yj37Z9oeUQa1+6eNyM8iN43HSHmNdQmMAAIDJEQQDALBsGh7E4h7m6MjcxGE6rXuam991+j+lDEFc1UfI5nJ8dWo1vPkycB56Obheav5IV0YA/XLmLylv/DaEfplxc2RtE8nb+sexx+vvzL4Zm+NsLm5f+nwWt1HWejy9Mu2+pSsDAIAlIAgGAGD5NETQIEhDBZewVgOOrw7TaZhbSHk5tL2JnHIduRbXlvGYaJtkA+ehIw01BMrN/D7TleFBR+pHZn/OHPfpLl8kbMTskvfR6nHH3jyuT25+3squhrL+vHTcp1OH6RIpr7z4Ksc5GhsAAByo3/79919aAQCAw6DBhYYLGgpnDtNrSOJS21ZHHcayC1X0918Or/sor4PpQsoAJDf/jsTt8umlscHQkLrAD7IbaVhlbxx1DKOnv8uyQi5t709H0K7ad26kedRpPrDv1Pf1Q9svq/uTnsT8Vtv+vsctn3392RzTnhz6oS6H3rxzIwAAAAvyH5oAAICDoeFQJLsallnP9BpCaOBz1jOd/l1HHWpI+STuIVG9ZnH9dbp88QG2s7bFkBBYA6NEmgOjrWmT2LTP+QG2jwbAqSwv5M9Nm+qyHWJd50ez7JnDPh3ab84q++Wh1azN5fUNKes3xowc57OtHL9yj33dpc7v2hw/HsStRA8AAMCsKA0BAMBhSaQMGXRE27pn2kLcwwgNlnLZjbpzqWnbFQTHUo5cXB9Y+44xWjeR/jBd2zqSsv7r/QG0i4bbepn7H2b9ioUup+3zv0s58vMQajPrMr43/SHrmVb3zY303xCyy6HePO6icjyLG/axyHE+ufmZiXsI7FIXeF2ZJhZuDgcAABaIIBgAgMNibx5XyNsRck1c6wWrakB05zB9/b01mLHhcGp+rg6obRMpS28M8dGx7arbR7en3qxPR9o+L6g9ns0y/WW2440cTqkP3U801I9M2+o+8LjAtn0nbgFwfT9LBr6/a73xpYjMz9vKMSWkL/6szOfK4zVpzzSryjaMhRAYAAAsFDWCAQA4TBo8FOb3WNxGoLqOdNXRbxpQ/t0z3b28Lv0Qya4kwhfzXL2O8FLZ0XxnA+bxXca5HFzbdGN+XszYBs+mDexje4T7TVRpW93mc5bmeKi07d0I80tleE1k13rj+6bby9b//Vppv9yzPd6b49MPj20WS3ewa0tMRHK49ZcBAMCJIAgGAOBw1S9F7gogfOth6ijKvjIJTcGnHeFqA1UNbW4W3o42VB8SArsERqFis61XlZ+RhIeYD7KrBf1k+kUhh3ljvzG2/do8ItmNcg8tD/JYace80s7bifpGJsPqIR/KzeP0GPKl8u/PZt2rfTaV7iD40Wxf1339uTJ933E1EkJgAABwAAiCAQA4bFOFwRoWavjyq2MaDWPS2vwLeR2y1EcNL9FWht8cLpL9Xg4eSXuNVJ8bAOKw2tb3C56l9t8++qVUvWzL77VlrofFdToaWEeDXzm2ievxNBJCYAAAcCD+QxMAAHDQNHyIpRyJm0t3IGHrC+t0fcHRhZlWR/26jjjUsKY+0m7pN4vLZHj5hVj2H6IVcpojek+9be3N47YSPqL9TNzqje9T07LVl7kriL03beUSAuuXYIkQAgMAgCPEzeIAADh8GkLYS5hz6Q50bBj83WG+OsIulfabbOW1aZsCYw2ZooW2W9sy+3gv/iHQii6LWn8Yso8UMnzUffVGkUt06bDMRctrn82+7lKr3F7BQAgMAACOEkEwAADHwQa8hbiFwYmUIeZzx3RnZrpNy99tAKLv1XVJ9hJHGsY9y+xCw/TM8zU2RCo62hWnw47m3Q7cT7Zmfx7iWsa52eHY1o7LXLQcz27E7caAn8X9xnB2uQiBAQDAQaFGMAAAxyeTMmBKpCwZ0SWScqRc2yXTtn6ozu9b5Xm9fNoGNIV0hyz1WsL7pusz5FJ6FVr7WLfNdW0+2jb5AfazlSwj5D/EGsiR6QuXtX2t7+Zkvv0rxLuF9cekduxpO0Y9Nay/vVll1zHq3kzT14dic6y05TieBAAA4MBQIxgAgOOTSBls/JAyJLmrPOoKKUONWMpAsn4J9pn5eyZl2PLJPJ9V3qtvpF3c8/c5R9atTDsMCYEfJWw0b1MpCm3vX2aeqSz78vxqG96Yx9lClulQAvXILGdbGZU7GVZzOjHvcTlgGe9keCDtYyPdX1j1fdlwJrsyNneVtv1uno9bjlEu+9xadldF6DxssAwAAHCQGBEMAMDxSuT1SDobCufmZ1PYFEsZdFRHCP+UXfCZSRm0/G5er/PqC530fbvq4qayG2k3tTtxu2FU17poG/kG19p+Pxznn5n2KBbWn2wopo+zhfb5B9N2d7KsEZtrca9JXR1tH2Jl+uf5wHaMZ2rDVLqvGNB16buho4a6kfm9MMelxPxb+8OHyrT3lT7S1Hax7L4cq7bhezmML2oAAABaEQQDAHDcEmm/rFrDnrzyqIY+kezKS2gI81vlb9URvK7/kfhT2sPTWHYjdaeUym5Ec6i/ApZzbdrXNzzV7ZOZ9yv21H/qIyIPyU/ZffGxj/ZbmXa7kf4gs+67DKvXG9rn6u23maF/6f7fdSM332NMVNve28rxpf4Fiw1+7aNtOxECAwCAo0AQDADA8Uuku8am1RY8RtIcpOnz/zguQ1+Qksm0N6pybYMuIbWO7c2lLga+96PsAnt7c7GxReYRm4eGdGdHsg9U26+QaUpIRKbN9LEZYZsPra29rz7vIzXHm64viX45zquttnG99Exkto8+XEpoEAIDAICjQRAMAMBpSMQvFHKpuRqLe0jTV1vzzizjFJei73N0ZFcpCg3eh4SFGm4W8vZmaduedowbfr88wX3Ctp99SO33JpHsShBUb5Z3OWAZnmS6kai6Dw8dBT9lELqV7jIYesz44jivvpvc+ZTncD1uAQAAHBRuFgcAwGnI5PXN3vrYm5hpAJpIc7DoE9r21TzVQMjelG5MY9wc7kHCRiun0h4Cf5TyMvXYzPs6YP7nsivXcLWHPmVrTut6Fnvq14l5XA5ov8s9td2t7Eoi5NIcBn+TYSOYU7PvDekftzLNKHSXG9JFIx0DtB0+eL7O3mwOAADgaDAiGACA05KJf+jYdYO0fz3m0XXDOJ2/hi5j1yR1udFU33K7BFZ1ibSPwG6q/xrJLtRcei1eW0JEH0u5Idu60n+WXM5CRwCn8vZmdl2j1kNvUGiNUZ7k0SzjmNs7M+vUVR9Yl9s1qG8aERyZtr4I6OOxLOuGgwAAAIMRBAMAcFpCQ6G2MMrnPxK/9fxd5/W7jBe+ZBI20raq73LzJl2hnkvAtMQbtOly30l3Pdel2FQeSwiF7cjpTPpLF7T1G51HNGDfiMx2Gzoyfj1iuzxJf8BdeOwD9WNHaEmY0C9/AAAAFo8gGACA0zMkIInldXCj83EdsdcXBOu8MhmnPEQiw2+UFVIbdWXa57yl/XwDprW8vnnbXMHwvVmP3DyeDriv76v98spjjH47dJSqvvbXwPVqGs0eul3+Frcvh1yPTSvH/bBPyJc/AAAAB4EgGACA05RIWFBav0RcL+t2rb3ZF7AU5u/JwHWzIdMQoYGXLn9bMP6nDB9Na29Qtja/R7Kro6o/XYOv58qy2JvLbc022B5xv7ftZ9vNtqOIX63g+9o2lxHbr+sGafq+8R72+6oxbh6XSlmvvO8KgH89tke1XUJLwnw2ywYAAHCUCIIBADhdmYSVTtAbyNlavom4B0tdoY8NbzVojgask752X5fAd7XnGOEZ2DfV0FG5oft91dAvNXIpg/e/pCyb0cb1RKUa4OrPTwHLNDRkBwAAWLz/owkAADhZOvLwIeB1V7ILgnOP13WN/IvNzyGX7uvITg2VhoTAjxIWBiXSHdxldDd49qefLX+7NvvukHk/DFw+3e+jAa+3r9047I8utpXjSEgI/Czj36gSAABgcRgRDADAaRtSLziSXUmBvsuw+0b6VucRWqMzk2EjHdtuiOfShm2lKMa+wRZOR9+NHYeMMh9SQ7fat2MJq1n8b8NxpImu/6XDfrtqOI74oC4wAAA4CYwIBgDgtGlwkgS8ToPj1PyeOUyfd/wtltfhTRSwPLosQy93vxH/EHjVsW42WAZCPJn+0zYqVkuyrAfMe2P6aCjdZ28DXhfXjiNdI3HvHOZnp0kkvC5wTncDAACngCAYAABokPI54HV6k7hI/MKaJmnt37HncmiQ9GlgG3wV/9GVNgRuGk1tQ+AnuhcG6Atstf+FhsH6pcfNwOW7Fv+bq6179v/6+vXJHObT5rtwczgAAHBCCIIBAIBKpQxFfCUvj6LntRpitQXBsby99NunVqeGStnAdddarCGBmI6GbBuBGDK6GGiy7dgnzsy+tQqct+47Xwcu3yfPfTap/fu8Y//Tde+qZ6w3eMvNPH3LXDzI8CAcAADgoFAjGAAAVGXiV2LB1v7Vxz8t0+ho47Th+a46pS41O/tqqLoIrXOqAdKXlr99lXkDpkSG3biria39PJXYcbqt47aZenmXQLfzt5H7saVh8tWAZXOtr912nNDX65c6hed62+OE7/IPbS8AAICDRBAMAADqMvELg/+UMgBqel3XzaC63sclTM2l/0ZSXUJvDqev+dXyt3uZvy7w0HY4Rvfmpw2StY0KaQ4aD0kq7WVQdGT7JnC+Y9w8Tr8UWkt3uJpId5jdVuai6SZwdl/TZf+vx3ISAgMAgJNFaQgAAE5HIm4jR3U6n5rBsfmpwW29lqk+1xS46PPXDvNso2UZhoafZ+J/ozxtv7YyFxowbfawXQmBm9tEH1rHWoNTDe51JKqOgMilDFTjA1wvXe62Miw6IjYLnK/22/OBy3Yu/fXCu/aPi47lv2lpC+VTI/m79IfVamXec8WuBAAAjglBMAAAp0NH1eWyu0lUV4CSSnnZ9aPDfGPzU8OV28rzOmIva5n+S888NRSKOv6u830eoU0+iF94pu1y1vK3SMYv0dCHoMqfBsQ2HNY+q+FlciTrdh3QB3Xdv430/pnjsaJr+Zu2hR63qrWMv8uudEzssFx6rHjfs50jc1zU+RbiXpYEAADgYFAaAgCA01Kvq6tB7515Th9PDdOnUgambeqXdBfmdVHL/PTvZw7L+pd0jzBcm2U+G6FdNFhKAtqvLrTcRCh9r19061Hotsuk/DKjWOgy6vJ1jaR/L35fbGifHysE7ntv3V//dpzXny370FZ2X7jYY0tffWB7M8iiZf/ZmJ8Xe9qHAQAAZsOIYAAATouGJ7GU4a3Sy7k15P0hZZ3NXMrgd12ZXkOUP6T9kvR6KJpIGa40jabTkM01uI17/r4104wxMljDtcyx/dYdbXFm2nA90/aM6NKjOTP7wj+mLyytbTNpD4HtiNfMY366n44RAut7v5Pho4Hrx4m2Za6Xm2nbTvdmufRYVJjn1ub1Gh7raJhfZpvbY9iDmR8hMAAAOEqMCAYA4HRl0j26UAOeXHYjhgspQ5K04XV/SP8oSt+bOrneeG0fI4P72m+uUYW6LT7RlSfz1bTxPksErExfuxqxr2kfHysEdn1vDXc/eMz7T8f51k9mfpr3yk3b2RG/+uiqg8xN5AAAwNFjRDAAAKcrkfaRrUqDVQ2fNDDSUZIaymiokr48fpdyBKIdWRw5vJ/vjdRcR9XuY2RwX/vNNTJ4TTee1IdK/9oHW4rk0EPgkL7qcrywxx0tcfNRyi+kEtnd1PG/Zl11vyYEBgAAJ48RwQAAIJPukcFN7I3g9BFJGaD0hSip+I9e/c1jWg2aNPw5H6FNfEYG962X7yX7PjSEuxhxfs/yegS4HVHZ1T90BObVyOv1WGuzleyCxMs97SefzbaeyxT1qLVP7yMEFnN88Bm1/1P6w2Bbi9yG9UnAsYwQGAAAnAyCYAAAoDLxD1CUBnapuAWdOp1vEOx6ebjVF5758AmDdbqugG2qMHis/8hpsHfTsYxd5Tc0sLuT8W465hLMrWV3uf/VjPuJT58Yoq8faxtVa9+O0Ud9+kos/mVPfPuqa2mYxPTdi4n6GgAAwNGgNAQAAFCJdJeJaKOjbzVcKmSay+dXntPXb4Y3hE+ZCJ3ufcffv8n4AWI00nx+mnl1rauGfjctf9MgNpfy0vyhbMjYF8zp8mgdWA1DbZmSxxn2E58+EWpt9qeuEDiWwwqBp2qn3KwXITAAAIADgmAAAGBp0BcaoGog/EvKcG615/U4lTA4GmEeurwbcQvDdP3awtaN2fb3A5fnVvyDuSfZlSjR9XmeuH9NGQb33fgwJLxM5DBD4KLjb+nL428JLxNCCAwAAE4SQTAAALDGCFD15lq5NIeU2wNbF8s3DH4n7WHkmGHwkBvF6fL9Jf6BZtv0sfmZDlynbITXa9/7OXH/upbx6wX3hcC2VIJPeHkrhzsSuGh4zpbM+DTCuhACAwCAk0MQDAAAqsYIUPUybQ2M6kHlNnB56lYzrovlEwbn5n2nDoNDR17bIOwu4LV5y/Pryt9D21tHGxcj9WEdofx5pH3iuWWdPkn/zcxcbaQ7BP4u/uGl9tcPI61/7LH/jnVFQN4wX33ucoR1IQQGAAAniSAYAADU2QB1SM1VDbRyeT0yuAiYZ1P4pCUs1p7rMncYvJX+MDgbuDxx4Os2Ej6yM295/rzy+13gvMcebZpKd6kOn76smmpo6zYcGnwmL48f0h0CJ57zzCTs5o91ISHwbcvfhuz7Y9wEckn1jQEAAPaCIBgAADSxoyqH1FvVYEtDwWpQ5hMStgVHsfgFY2OHwfV1amPD4MeOeWUDliUkgNRgNB/YBl3bRQbMf4qATtt3jJHBGkAWDX3obOA21H7cVbrhkEJgMceMuOVvhcd8fsrrUbt3QggMAAAwGEEwAABo0zeq1YWGN9URgree71+n4adeGu57Sf6YYfCVlGGnaxi87njfIWGwbzD2Vca5yVnh0NZTzDdUKsNvYqd0JHrSsD9cSdjobH1NVwj8WQ4rBBazX55L84j93GM+t7XtdzlwfRIhBAYAACAIBgAAncYIgzWYssFtIc2X2DfJG56z82kLm7qMGQZfiHsY3Pe+1+IfJEae02sQejNjnwlRTLhMQ0e3qzMzn7Thb1nA/Lq2x3vxvxldJvsNgXVfuDK/JwP6xX1l39d9/NPA9dG2vBMAAAAQBAMAgF5jhMGZ7ELTG8d5NYU3ScvvrpYaBqee7x15TKttvZmhn6wX3IefAtq4ifbd24bteO7ZH3X7XbX87b34B8s6/Rgh8IOEl1DYtPxe3Z+fHdu4ul5DvJdxRsEDAAAcBYJgAADgYmgYrKMpbcCjoVzSM70GUkXtOX3/6iXioeHm2GHw3Qjvq+sVebyvT+i6kfByDW19ocnQm6Y9TdyHNcB9HDiPM9N3m0bzph7zaev/IcGlrte+Q+D6/tgWjPftK58r759IeF3gZyEEBgAAeIMgGAAAuOq7+VkfDc9sWKiBUFeJiKZawmnt3yHlIawxw+BLcS/t0DUy1SfYdg1dtS5wPnI/eJqwf00tHWEeG9Om9w390XUbJg3P6X6VeS6L9oMPI6yTDYFDt221LES1rev9tGv97mvbJ3Rb2dIWmQAAAOAVgmAAAOCj7+ZnXexoSitpmY8GOVntOQ3Ymm4YFbe8V+SwPGOGwYnHtBqCN4XpPkFw7DCNvkdKl30lk+G1gjXwXLW0rUs/0L553vD87cT9ro1PCBx59Eddx/rI6Vyab9z3KG9LS5wHrMujDBvVDAAAcNQIggEAgC8boH4PeG1S+7fOpx7M1QOxSNpH9yUtz0ce67IOXJcq3zIVTZfIX3q8fuXY1k901zeyEeZhRwXXA/0rh20Te/QJ3/3Jl+9I4Miz/3+St6P209q/bQ3rpwH7k7o370UIDAAA0IIgGAAAhLB1frUOp88IS635uarNpxr6aLBWDYJ1Wg3IzhznF0rXZUgYfCZ+4VXe8nzs0Y5dpigJcSyyEeZht/Vtx9/EYxs31cTuE0l4DV37nrH4fVmw6mmPtr6+qv27OipYRw1vPduw7rMMK20BAABwEgiCAQDAEJn4l1eIa//OpQxylIZC1TBHg7YLz/mFSmRYGOyzHHnL8y41j6Oev2swn9I1W21leHmIuNL/6/pCzHXLMk3Z3+pCawKvW54763jNWUN/tyUjfja0Yd/8qvSLo3f0dwAAADcEwQAAYChbN/ijuAVsTWFSKuUo1url8RoWXTvMLx5xXRIJD4N9lkMDuKY6wS6jm6Oev9fDdLx1N/D1Z6YfazvXa95e9bz2omUfmrK/VQ25MVweuBy6zre19f0szaUtXG8A+dlMm9OdAQAA3BAEAwCALhreZOJ2qbZOG0l/kBq3PF+9sZTO54vjMsYjr3MiYWGw72X6ReC6dAVlDzJO6YNjl48wD7ut7jy2Y9TyfEgQvA54zZAQeOj+96E2bdqyHFHPfHQU8R8dr6+30a3ZRiu6PQAAOHUEwQAAoIsdXfrD/MykDIXbQhVbO1iDmrYw1SWQST2WsSmAjUdY7xA+4Vwe+B6rCZb71IxxQ7F1x3Zs639Ry/NFwPtfDNifx+TT513267a20wD4nTn+dLWX/v3WTPO32V8SYZQ8AAAAQTAAAOil4ZHeFE4vh9dSDRoK/1fKACyV5iCokF0grJdwV8sguARYvjeLikdeZw2NHgJeN3TU4eWAdb0XLpN3NUYQHFXm9ey4jaKW5wvP914HLnM+0jpXnXv278hjH9J2/W6OI5uW5V+bY5SO+v3XHJ8+mOX6LoTAAAAA/x9BMAAAcJFJGQZXaajzScpRdxq0aBCTyOugp5AyLNbn/pIymNFwpyvI8rlZlDXFZd9FwGtij2lDw8i2dWU0sJ+Hga+/7NiWbf07anjuecQ+MOX6Ni1/PMI86vSLop/meGNH8xa1dU/MMakwxx8tI1OvzfxemmsQAwAAnKz/0AQAAMBRZn5+a/ibBrdXsgtjNHTKzWtsSHYnu3qqXUFWSMi1lte1WscIhrfSf+OvIUJHKTaNqP4u44xyPSVjjBJdmfloX7+s7Q8rx/cI2W5RwGuKhbR73w3efm9oN33Nxjxcrih4L9TKBgAAeIMRwQAAwEf28vhT+kcxalijl2fraL1Cypqd1VGSh3CpdhHwmvUI7xt3/C1qeT6la3rLR5jHuqOvrCdc9ijgNWN9UbBq+X0sT5V1TGU36veTEAIDAAAMQhAMAAB8aaAUe0yvtTptKJzLPJdrjxE0FwGvWU28XlHDc99lOaM9T1Xh2BfWR7Cu64Hr0xdI67FFR/f/I2X461ODmBAYAACgA0EwAAAIoWHO+4DX6eXzWlqikPYbwuUDl+1QwrZ8pHVL6Y5BxviyYF3ZH1y21WqP61sstA2t2OwTvySsJAshMAAAQA+CYAAAECqTsDBY6Si/H1IGP1HD3x8HLNfqiNu8vm6MBg63HXF7zF3qJA54zRj9pKi9d0gb5g1tqKVjNAC+DFyuj0IIDAAA0IsgGAAADJFJeBisNPjRMKk+Otg3YKoGcWONCM4X2N5xQ/tjmVy/kDikm/wVA1//UPv32qz/hwHz1C9DbuluAAAA/QiCAQDAUJkMC4PPpBwdnFSeu/OcRzVMO+YRwVHl93tZZlh9quqj2F2/kHg6sPWMK78Xnq+t9tfE/Pt8wLJ8l3lqjgMAABwFgmAAADCG7OXxdeA8tHZwYn7XIPjZ47XVIDg+4nauhmYp3W5RisDXHdoN5Fa1dfbZT+3I3cTs72cDloMQGAAAwBNBMAAAGMuNlOHMEDYM1lGSrpd7P8rrUZXRkbbvurbOOV3uKBzaCPaL2r9d+6GtZx2b/XyIeyEEBgAA8EYQDAAAxpTIOGGwhp4aBLvcNK5eRuL8SNu2GhimdDXsUVz5PXeY/tn02Uj8y77UaZ3hDZsAAADAH0EwAAAYWyLDw+C7yrz6ZJXfY/MzH2E9ph6p6VsSwK7bswwP0zC+1Qmt67phX+2SSjkaWPfVIeUgHsx+8ER3AwAA8EcQDAAAppDIsDBYR/WmUga6nzum0xHDU9UHnrp2q29wqOupIXAmBGFLVC+ZUDRMk4/0Xvve/uvaev7smFb/pqP7tXTM5YD3JAQGAAAYiCAYAABMJZFhYfAHKQOnVMqaoE3Syu8arN4cQbsVLc/fVdoDh7sdx7ANeM0YI5ZtCHstr2txt9XzfjDHgdXAfksIDAAAMAKCYAAAMFTU8bdEhoXBNmDSmqD1esH1EgkaAp8dQXsWPX8jDBvHaqHzmsoYI9y3DfumyuXtlzXP8vrGj6H7Zl8IvKYrAwAAuCEIBgAAQ2kQk0t7WYbk5fE1cN6XZv4aAtVvEJXKLhyKXh6f2BTw7LdD5R3zKhznER1om13V9vm0Yb/X4Hgl4Td3+17Z/5va7U64cRwAAIAzgmAAADCUhjEa+PySsn5t1DCNjtZ9Hzj/xPzU9/hoftfRwdURiWntNfEI6xUFvGY78fyxTE3bsnB87fkBr2d1v8tlNyr4q+xG62tQGzIa+LM03ywyMseZf2R4yQkAAICTQhAMAADGoEGvXsKttUM1oMnkbWikz/0pb0s89KmO+NPwV28+lVSei8z7Vo0x2jMKeM3TxPPHeMYslbDu+FvfcyGKgNesJmgzHbEfV/6t++W9vK7VnXi+h5aU+EveBryR7ALgazMdo4EBAAA8EAQDAICxJFKGM6oaCFfDo63590+P+Z7X5qHhT175d9rwmvhA2/CebjSbMW+eVg9In6X5S4Gx6jsXAa8ZI/huCl7T2nLV971Lz/6vy1mt/R3J6wC4uizUywYAAPBAEAwAAMaiIW9ae06Dm7/ldQ1hW+9XR/09O867LcSK5O1oYHUmw4OvkKCQYOpwDO0f1dD+smFfmHNZ5qD7w0XD8/VRwVWx47z1OPDRTF9U9u1M3gbA6qu8/jIIAAAADgiCAQDAmLR0Q9OoVg2LtIZwLrtwSEf9ReJ2I7m2oOym4zXJwHUJCed8AsCY7rI32u/OBs5j27Ed85bXtH1R4PulQ8gXDkNHQHeVYbgZsA/9NNPdVtpTjw1NAbDS0jIpXRgAAMDff2gCAAD+JzY/NZDoCkw0gNlWfuItDYX+bvmbDYQ1LE6lDMx0+lvz7+uW17UFSknHcmykOyheooLuM4sx6wM3BaR5z2vGem8fFwPfsysIvpIyXK/336jjNdVjgD0Gp9JfSuJGGHnfJqo9uuQ9fRUAABwhgmAAwCmKzWNtHucD5qU3SNuak2l9FDTv/9rj88vjU8c0TYFwYn7Xh8vN3zSY6hrVaWsLh4ZvIWHh0O1P/5nHGEFwXumHVc/iH65FC28v/XLsqmcabYdbh3b+aabLK+t+6zB/+9o7uu//b7fq59il5+s/1fps9XPMftkJAACODKUhAACncsKso8g0QPhXygBST4I1eDgfOG8dZaeh5TcpL2UupAw11ife5toGjw7T2UA4kzJs0vZLXh6/S1kz1M6jKfDdOMw/GbAOIaUDCo9pL9k19yYe+PpHs62bvkjqCyqfW45RIcvgK/S4FLqvRZVl1RIwf8juZo8rc5z4R9xC4Gc5vBH+Y7MlNLam3fRz58MIx5IzM49P5nj8X9ldrRFxuAAA4HgQBAMAjlliTmb1hPmLuIUNQ52bE3MtjVBIObp1dYJt/yR+dTyvTXttKq/XwCN6efwpZYhUD7GmDIL3tc0oNzI93bZDg7Pc/GwKJu9m2sbFjP3aJYC9kLehobbTX7L7Ms4uc2za4YPHMtzKaY6YX5nj2NZ8rnyQ4WU+XFyaz81/zHvfnOhnGQAAR4UgGABwjCfNqZRB4jfZ76hLDYV1hJWOrsrk9EZW6Trfe0yvo9J+yG50sGVDiGqAtha3Ebs6zSZg2UNGTt579tMmXI49vXiEedyZbdhUFuJupv4W0leiwNdcBLZt0tAeenz+JX5XY+iI4tsT/SwrzGfZxR6XRd/7i1mWU/wsAwDgaFAjGABwTCfNN+ZxNmA+1VqJ1RvC2d/1fdaV91yLW63ha/P4Lqd1s6OuG8d1tZW2Z9zRTrHH/HTau4D+NKU1u+zebAa+3oa9acOxJnN4vR5HLkfobzof36scoonba92zT91K+w0h+44jp3LM9Pkss/WotS8UshsxbX+PKtvc/h6bnyFlkc5qn2WpUNccAICDQhAMADgGN9IcyriwJ9J34naztyd5fSOoasBoT7I1OGkLaK7N33V5T2GE29YEBr7hz4Vp51iaAyCfIHUj/rVF14HrOlTO7jyplYQFkVLb51ctfcpln35qWa45RAGviUfab3IJG9V6L6dzgzg9VmU9n2UPZpo7h8+romMa+3llP7N8Pz9tIPzZ9HuuZgAA4ABQGgIAcMg0dNDw7UvASayGC+9ld3l3JsNHNhVmPjq/+s3Oqs7MMm/lNC6x1cDsOeB1NgxuCsl82u1c5gnanjz7Lua3GWEet9I8WvP7gGNISECaB7wm5HgTj9CvMwkvbXAKN4jT45MGuz9aPsv0c0QD1z9kd8O4sT6vEvP+f5k+7OuT+SyLBQAALB5BMADgUNmSA77hwk9zMh2L22Xcoao3O9MT7Kb6sRfmBDo58m1l2yKEtlE6wjL4Bq9xwHsUHtM2BdOP7NaTG9qX7k1//jRg3vke1z8KeM3ZwGn1WB06Cvu7HP8NFGNz7LhqOSa8N9stlWnLMNyZz6I/xD8Q1i/bfsnp1XEGAODgEAQDAA6NHTn1xfN1OiL1nZQjAouZl/nOnOy/l7dhnwYn307gBDqV8KDzg8w/2ixkBLFPv4oGvh7+Egmri1rvx1nD899H2H6+fTwPeI/ziZepqZ9/CXytHrNvTqBP/pK3Abqu+0fTftnMy1SY5fpTyjIUvsfqU7nSBQCAg0QQDAA4JHpymYv/DZIeKq/dJz2h15GpX1tOoPXvqyPefsnAtptTyGXshWdfrqPG5nR0v0oHzuPe7L/1G735Bpb7HuEazXzMC5Ue+T6hbfOt4fmfZhvt+8tBW+7Bd3SwvdKF8jcAACwQQTAA4FDYesC+Ad2DtN9wbB90OTQ00tHJ9bq5evl0LscbBuu6fQ98rY5kTGZaztD2Lwa+x5bdfDI3Mnw0cCbNo1tTz+PLU8cxztd9wGuiifp0nR53LwNfq+t1zFdJZPK2XIZ+HuhVI5uFfV4lAcftM3O8JwwGAGBhCIIBAIdgbU4qfW8It7QQuMqeJNcvve26QdoxCL1xnH2t5btNp76Rm+8l1Bfs1rPRY8CngfPQIKwpmPwpYYFlU5mUufZ5n/5deM67GkwnA48TxyqT5hA4lvmvfHCVmL7ugzAYAIAF+g9NAABYuJU5OfYNgfXEekkjq5oU5uRfawhXR85dmHXeLGAZY7MN6ifzW9O2uef87AizHwHLciG7keH6uPLoCz6jbaOAZRujn+Xs7pMcP+4GzuPR9LuzhueTAfv++Qj9Lhf/Ube+gfO9x3s8Vd4j9AZxnyVsdPzavG/ccqzaLuDzIGtolyV/YVmVmDb0GVlvw+BYuOIBAIBFIAgGACydhjghoyf1pLU4gPV7kt1IsGpAoCGnjjSce2TcxixP7NHuD+ZkP3M82ddtqnWSPwQu31b8QgXfIDAKWC6f5YnZrWexkrArCZrmUw+/hn7RpMemyxH6Xcj7+47QvBP3IDiv7KchNHROPfbTjXm4Lt+jWcbcrNec4WsihxsC276m6/DL83Vn5rPhUNYTAICjRmkIAMCSpRJWY/Jeho8CnFsib+t9fpB56uJG5kRdT9J/mPf1Cd8vzGv+lt3oL5dt+xCwrHbeucdrbmvr2ifkUmZGBC+LDYHHKMHRFCTrFzRDRjgWLfuhr21g27gcE+x0emx49uzDccBy2XDd5Rigx/d/pKzZ7PMZoYG+hrF6k7b/yi6gnNpa3t4Y7pBC4Or2Danzbq9yAQAAe0YQDABYKj1xDq3rmR7oOm/kbTiqIWY0YRvrib0GKhqOnI0wTw1lfpn5di33k1nf54D529e7BMla13Jb6RcubRlSqzX3mDZm956Utm8h09Vh1ht6ZQPn0RT+nY80H9d9qEtUWccncauDXC3Bsg7cbk89+2Vmji9XI23L68rxaqr9sqk8ySGULmoTWuf9SpZR7ggAgJNGEAwAWKos8HX3crgjK5vC0TMZfySVDVR0BO/lROui8932nPgXUoYvvqHC2rGP6HxvKq/55LHsIdtuiAd2+VH6tQaWv2ScLzWavB9pf2wbyRuNNB+XtpKefbMa3Gm7Pva8php2XgS067Znn9e/X0+0XV2/wAqRytuQP5bDKF3UdqwLveImk+O9ESoAAAeBIBgAsESJhI/muz3wdS/kbXiqIcVYtYI35j2uZ1gXDeN+SHd5i634l7+wQUJfGHEju7Al85y3r6E1gqmdGU63WWq29YeJ3kO/VPhLxvtSpm17RwHzegx4Td+I3aiy36xkVx+2S2jb9IXreszKJWzEtC/7BdZYx9u4oU9+lMO/cVo64DPhRgAAwN4QBAMAjukkUwORuyNY/1zKm6nV22ToSCoNyX/IdKMl22htzKS2HHHl37rN3gfMt5Cy9EMTrWOZVdrOfrHQF4CtA/vdUFt2+//1g8RxG+g0N+Y1Wuv104T9+lF2dWnHMtaIYLsf+HJ9n+oVCXpc+tjRRnnActRD4Kj273gPxyx9ry9me49xzK26l8P/stL2ufvA194Io4IBANib/9AEAICF0dFfoSO/7o6oHdJaW9iRVGnAvGyNypCSB1qyQEcD5pXn1ubhs52+mfBA55NJWZbiq1mfJ/Pc1vz9zHObXzUsc1JZ1k+1tugSBbRR4Tl903ZgRHC5Havb8r6lL1/MuEw/TV+aYvs8N/T1kP63Ddi3I4djRnW7bMy+dmv2qfoVBVnActdD4NTsq39VltH3mP5YOc5Uj1dRQL+5kl3t4JDtn9Te81nmufnnXLLAz5Qz0w63AgAAZkcQDABYmiEnysd0Ymkvxf5Vee7GrKNPKKGBTi5+IYiGX3fm0fVekVnGG3ELbzU4sLU+P0sZ+ujrU7Ne+nxs3tc1ZM7Ma+37P8jr0cZZwzJLzzr5GmM0LyOC37rc43vb0G7KL5eaAtyQ/lcEvCbu+fu6YT+L5HWJiOuO/ayvbW25BzG/35p9/melze8cjyvP5v2znv1oZdZ7I+6lcS5kV+vcdx9NGz6fiiPaP3X7fBvwOU8QDADAHlAaAgCwNKF3g388spNslcvrUZG+9RV9Q2Atp/CHlKFHJv2Bs7Z3KmVA9Nlh/ueyC0fszafsZdiF7MIWDaF+9rRLVWZ+2hDYLnfasO5TlIbwCebjEeaBaX2VsNGoY/SbKGA+IV8irDz3g/pNKxNzvBCzr7oeex/MvHOzL+jPH+bYUL25Y+J43LLb6sahHexNznTevzses+xxKxe/cgaJvP4y61HCSx4t1VPPcbrLhYx/Uz4AAOCAIBgAsCTxgNfeHWmbpLV/33i2iUuYouHMn1KGF0VgIJCaefTVy/0gu5GF1XXR0EQDoVzKoEhDYa1H+uzw/pmUgUQsu3BN3+NTw7R97REFrH/uMe1qhHlgGvaLkBuZJ5jfjtT/QvbZvv2g6QuRq9oxOjFt1nTsbSrr8bkyX32NXu1QHRFdHTGb9iyfHhfeDdhW9pj1h7jVuj0TvzD4puc4fiyGfO5uOOQAADA/gmAAwJLEezohXbJc3o4KThxel4nbpfUa5NhyDUPZ0bwPPdMllW1WD2F0mX+Z5b8z86tOc9/yvhoqVAOh28B+FlJ/tvCYds1uvigaKOqo0j8k/IuQUE3vdT5x/6uKWp5fdSxH1rAvZw3TVfdF+0XTrXn8I2+v/KiOmE162uFZdqOJx9gGOq+PDtNeOH7OrGvHkUcJq6F8CIZ87nIsBABgDwiCAQBLEnon8Wc57hGV9VCzbyRVIm41MN/L+DcvsjVEn3uWz2ob4azLvzXTxlLeQOpR3Eb/6fRXPX9vEgWuczGwj9+z68/up+n/K9MHiz0sQ9t7hgRkDwGviTz3D3XueMzYmmPAe7M+sVnfDy3T33gc30Lq9bocY99L/xUIl9Jf2zbpOX4fk6fAvjfkeAsAAAYgCAYALEnoCKH8yNtFR11VSy5oyLnqaEOX4EFDj2yi5dWQJu34+3llW+u031um09HPWt6hkDJw0ODApTRGGtjPooB1fZipj2OYR9PP9AsFrQ+7kf2P0ixang/5QiykPELs+bzr/iWyu7nc1hyfv0j7jd/u5fXI0q4vcb5OeLzPxO2qlA89020a5nvMcgEAAAeDIBgAwInoYahfgts2ai6T9sDF+ijThxP2ZnBtqoGohrtdI/E0OP4lr2uItomkvyRG7LBMrnwDuNWJ9l8XDyPNR4NFHfGrNWm1juzvpl8kZj9ayo35ihH7YUgfWgW+/7n0j9otTHv/7bA/3jjsm2KOEenE20SD6/eOx+NVS9tVy1r8lOO/ESTHLwAADsh/aAIAACeiByGT15dVx/I2zNVApa/GrY6KnOtS5fTl8a3lb1Hl9yezTJ965mdH4sXSHq64jBjWoLypLnI0Q99r2j5P7ML/s65ti+r2WMnbgDKvteH2ANdZvyyp18MNGRFcDGzvKpfa4om014ddmf3ZpTzN99p2i3qOgXPsK5m032yyegxJG4439YD87gT221P4/AUA4GgQBAMAlkQDgcvA151C21RDo7j2dw1f0p556IjLmxmXWUOQb47T6rIn0n+zLA1Sc2kPg13vRB819JupRwRHJ9x/fRXyNty8O9L1rPf5OHA+vpr6o+t7t5VvWJn90+Wmi88Nx6OoY/psxu2Smrbo+jz6YPpk3tF+pxAE2zrBFwIAABaP0hAAgKWdUPr6eULtk1d+1/AoqvxbR+D1lYRIZN7Rp/pePjdCSxyns2FwfeTkSvqDZKsp9I0C1tEnxI1G7Pc4Dts9zqdpX/EZjRy3HKNcA8HUo+8/yvxfmOiXSs8O61BVDY7vT2jfzgNeU7D7AwAwP4JgAMChn0zenXD72DAzkv7LsPUmS9uFb1Od1jXY17Apa2mPUOcBrxkjCGZE8OlqCgovR5qPi/WAfSiu/ftW3EPgB2kuUbMd8bNhjG2T9kxzWWmH9QKWeV+ymT7vAQDAQATBAIAl0RPDZ4/pH+X478jedeJsg4e053Vz3GTJ92S/LfC58egDenl60tAeIUJfO7Q0xCO7/Ukb80uA+4DXRCO9t+4/HzymT1qeL2ZoJx8aVvfdxDBtacvtifVj3/6Xs/sDADA/gmAAwNL43Mjs5sTapqj9OxK30cDaTk8LWWZr2zF96jH/6rSrAcsZ8lrf4GPt0T44DW37ZTzivubTJ7eBr/XZZ7uuTtDnnz2OF3Po+5yxo4LXC1rmJbZT1XeOfQAA7Ac3iwMALE0qZW3GvkuM9URyaWUhIgkbYZd7TKvh42Xl/fpOvvc9arrpZP+hJwTQLwNiab8hVZWWc9AAZjuwzeOAdfMN11eO7YPTkXv0lZB9zeWYVbUN6M8rx33V7vt9xyw9rl8vbBtVj7tNmr5s89ke64Bt/iTLCpt1Wd5L/w1Cn+X0vsQFAGAxCIIBAEsUmzDgsuUkMhW/kcNjLdNKdqO+4soJ/NmI7/NgTvCL2iOvnPxbl9Jf0iBd4PZ12XaJlMGCS93ejfgFIo/yNoALCd58QxhGBMOnr9xN3B9V1NAftU63S7Bb1I6FfZ7F7YaQmSwrCLbH0V8df7+S12Ve7ls+O6JKm1+OuHzPle2/lV1I/CTzlmDIzHtmLZ+LDzL/TUsBAEAFQTAAYImezMnzRnaX3BbmhPZu4pPIyLzf2ry3/vt8xnW/6AgJmurJnvWEA9met2UUuExPZvvn0h+024C1cFympPK6bW0ePgrP6c9GmAeOT9No05AvJkKOi+vavqrzuDHHvjPH/u+679yIW1idS/8I3Lm5LNN5rV2LGT87zirL1vYF6rbhMYU705c2lc9S+/mdsbsDALBfBMEAgCW7k+nLP9jA1z7OFtwevqHC7QKWOa79O/V4rQYVGh71XWpsQ7PCYZ566XIuZSBRbZ8oYN0Kz342dB44HSFfTOQBr6ke7+zoUXsszHuOh08e+85n8QsB9Tjxq3Ycyfe8TXT5Lz3adUmfJTYori6/DYfzymMsT0LoCwDAInGzOADAqdHQMJHdJax/vzy+SHlp79mRresSaijnUtZzVjqi+bZn22xqz+l2eu/4Xl0j3PS935n5ZfK2nETIyL2Qeqp1Bbvkycs9+kuf54DX2ND5SXZhsPbLWMpL+fv6f9Qzf93/09pzcc/rdBnuK6/POJ6OzobDn6QM3Z/MOiYS9sUYAAA4AATBAIBToKHKjTnJ/a+UI0yv5fiC3zoNue2o2n2d2BdSBgt/SH8tUVsSJK8tbyZuYbC+vl4+Q//9UcqwK5dd/dG8Mk1o2/hcis+IYPi4CHxdyOX+1dA5N++dm765NvveQ8d75R3z/i6v6wLre+mXQalD39+Y40ayx/0kqn12HDP9PLwyn4//LOCzAwAATIAgGABwrOzIX3sCb0f9npoLs+56Yl9IGcKs97AchbiFOansavemleczaQ+Dq/PNpQyt9FL0P6UMMewoZO0L9iZU1cAsClif+4D+WPfIbgppD1L3USe4qBw3crMMmZlGQ9mPpu8/OLzXV3kdAidm/h/E7aZxT7KfAFiPB7ae8T98drwKhVfsrgAAHDaCYADAsdHAIpMyQPi2pxN4DfjuZReYLIWWP9AQ5m/TPksc7WVvVqWj0z6Z5bTlInS7anmH+uXveeX3xPSBVHZh78pMU+0LReX3aKZ+WVewu8Kzz/QZOiK42idtGBxV/qZfqsS1Zcsb5vne7Mdipt+a47Hu118X2PftVSPV8PdiQcu3788UGwr/1xyHY3ZPAAAOEzeLAwAci8Q8prjTvL2pTlF7iHk+ZBSePZHWAGJde+5yhvY6Nyf2+tBgQQOeu8B1GVsmZZB7bh4/pAxA9LnctJcu60VlG7SxIXA91Ckqv0cBy5h7Tt80kq5gt4WMf5MuX+uO11+Y/Svu2M+qy6+Bpa2/HZnjylXtWJou7HNjI/N/YWhHVG9rP4f0h7hyPIsqny2RhNVAb3NtHo9mW2bswgAAHA6CYADAoUtkFxqO4dGciG8rjynC0erJftNNiOzJfGx+6gn9VCPUdL46Ws+GwbcSNrJwTKlZJkvDcb2hkdYcvZHdqN+kZ1nvHNotmmF9mt6jYPdFh1j8Q8GhI4KbXq+jeDOzPG3Hwnt5Xc5Ff35qmO5W9v9lU2SOIYlMXyf+sfZZUkx4bM0d+tO68hj6eXJe+dy4Xci2BQAAPQiCAQCHSk/iUxkeAD+YE2j7WMqJbGEeecPJvK73FKOGNRSxo73sKOFsT+uftWxfXbaN+Vsqu/q/TVzbKQpYvtxz+qZ+SmgC636kfXroiOA2F2ZfS1r+vjHvvTHTNfX35579dWq6bDcy/RUXj7L7gmpJ+3heO26tzOeJfYQGw7aMz40QCAMAsHjUCAYAHBo9wS6kHIkUGgL/lLKGpd78aC27u8IfwslrPtP72FHC2iap7OcmQWnL8xo8fOlpC1vzc8m27M7oMFeNYNdRsfolTNTx90zKMi5tx+V9BISrymeGLtscZXeeZFlfKnYt553srrD43Xwu6lUXz4H9yNZ1T9l9AQBYJoJgAMChiGV3w6GQANiGv3qyqyPDMjnOS/PtTYXGUj251zaLZlyXzKxPm0vTJ5oCs424B1wh4dDWs+8Cvv1pzi9f7Hv17d83La/NpQyK+/bnOdcnleFfGjZ56DnGXhxoH3wy2ygx7feXhIXCTTf6BAAAC0EQDABYupU5Of0VcIKtJ+wf5XX4ewyXrHaNFNR1jKUc7fxVwkZ2tZ3ca9Dzj8wbCGc9f9eAJ29ok3ji5RqjH+Xs3ujoT6FBcMgXQXb/6duv44ZlzB2OzRooFjN9XqTmvT7JeDWAn806/Cm7m1UeO13HxPSJ9wH9yt7oM5d5v0AEAAAdCIIBAEt2Y07orz1fpyfs78wJ+zHWK3QJNwrTfitzEv8w4vvbQPhWph+1qO/x7NAeeW1Zop7XuI6AbPLoOX3EroweTceoJY4srS9T7ric6cTLNVUArPv6R7MPJ7Ibub11WJ5j6puZ7L5g9B0lfGk+L1J2cwAA9o8gGACwRHrSnUtZB9b1hF5PTL+aE9VETne0ZdN660m8huIajo9ZNuKDTF8PUkMIlxtMncnr0cN95R5cR0A2KQL6c1N/Bawx60UXAa8JCS51v3cJge9l2tHAiWm/MQNgXeb3Zt8N+TJxfaT9tJDdKOHPnsexT9JeygcAAMyEIBgAsDQbc7LoWrf12ZyQ6ompHUGMZrm8HtU1hjnqQWaO013J4dTj5UZxcBEFvCbkGLj2nDYy+72LdKK2iWVY3fgmGgC/M/POeo6lp8zeRDQSv0BYvzj4WxgdDADA3vyHJgAALISOSNORVz5lIL6aE8qlln6IZBfkVH9vk1d+3064XoWUo7pS87geYZ62HuS9mXcx8vLqtv7gMO2NuIU0Q4LYp4C+DYQeQ4qFLdPK7OMu7mX80DTks8JlOVOZPuDVZa+G7uue40NR2f5PsrwvkGwgfGuOvfpwGZWtXyLEUn55+MRuDgDAfAiCAQBLoCfDmbjXxJwibBy6/NVHJGEj1JpG2D1WwgB95CMud2Ha8dY8LkeYp60H+VnGHfWVmmXtCxl0VLBL8PpU2Xa+tgH9A+iSL2Q5CsdlzRzndzPy8t2YY8FYJSAexP3Lo5Bj1mrgZ0KTZ3MM2prttV1A/6kGwq4h/aXsriTJOQQAADAPgmAAwL7pSWDmeGKvoWiygJPGtVnuWMYJT7ucm4d9n08TvMfWrEtstsUYgcUn00aJjDOKzQYNXxym1fW479k2hfm5r9G6Bbs+JpIPOE709UsNISPHY8RXGW8Ea2SOTWMdbx/N8SSbaBv8mnD7npl2qLeFHX2d7/Ez8skc87VdNRC+cFgXbav3E24LAABQQY1gAMA+6UisH+IWAmuosN7jCW5sTlT1RFdrHH6S6UPgEENGn2rbRi+PjzLOzcxsPUgNBMYIXHU+D45tcNfxdzvKep8Kdn8s9LjRdUPJO8djjB4/0hE/J3zqxvf5LLurUEIlC9yGl+Zz6Zf5nNL12+xpWXLTxq71g78JQTAAALMgCAYA7Iue9LmM7tTQ7p0JA+auJRhJGWYU5uRaL3c9W3i7fhkhBLg16/5zpGX6UAkGhtJ1enbsX23T3bH7Aa+savtO17Fh7bifDj1eR+a48WWk464G3H9IeF15Gx7ra78tfHuemc+rH2Z57TF9btrWsbh9gXdttjc11QEAmBBBMABgbnqSdyduNQR/yn5GAUfmhF9r3eoIq/MDa+NqCFDI7u7uPjQ80DBHQ/jHEZZprLvF6/okjst/2/K3W3ZDHNDxcm56fG76EuW7uJV6+DjCMTuR8UYB67po6YFY/Efia/vfmNf9LYfxZWDT58EH83l2Z9phTlvzOf7VYdpLIQwGAGBSBMEAgDmtzEnelcO0euI+9x3FI9kFwNdH0uYaYn8y65QFhAC5x0m8i08jnOjfmf7RRwPfeqClYVbBrjh4P8E8nvb0nk1flqQOy/Rdhn3RsjLHKR1xO0bg+rNyXA/5LPivlCOSz4+kP+ln7y9zDI5nfm8N1P+S/is6LoQwGACAyRAEAwDmYkPgvpvH6EninzJvvUBdtlSOKwBucl0JARKP1z2Zk/h3Mk7tYHu3+HjAPLR/vO9Y3upyW4+1fyNMRBMcnfpIXz0eVi/n/yy7L1DaRgV/l2G1c+3VH2Mcg/U4pSOTfb9MjM0yHPtnwaX5LMhk3sDV1pjuKxVhw2CONQAAjOw/NAEAYAauIfCDORGfcxScBgU6gm3sEV8aRGzNej/JLjzJe14Xm5+ReazNz4sRl83ecT41j8zxdfbEXKe/GrgM9m7xHyV8BKFd7nq9zm1tmhvTfnOPMO+yPsL9PBeEivc8r6eWY6Pdl25b9i/rqwz7kiUx7zFWLWCdX+HZZqlMcwNQe3NK38+CtfnsrH4WrEf+rLo22zmV+UrmFLK7+WrX54ges+/28H8CAACOGkEwAGBqriHwT3Py/jTjct3KeKO+Hs162kcROJ+ucCCuhAFjLLcGCt9MCHAjbjdRs7WDExknuPli1if0ZoAaJmzNsp+3tGFqtve2pb0/eb6nb4irfeGyof8BfbZ7fO/CHHPi2r6pvz+YY/qzORYMuQGjHkc+jLTMn8WvDnks4wfA96Y9tjLsy5Ftx2dXXHkM/ZLwzByHNzLfl2X2cyTr+SyzI4NjIQwGAGAUBMEAgKndOpyoDr2k2NdaXgeHoR7NfDKZJ7DJK8HCRsa7aZG2g95YTgOMVNzCC7vOY7TjtdkmoSf79mZEWcvf70beDr4hbtHwXMShARVxy/NPM/TPvn2r7QuUJ/EfeVtfTt1nr0ZYzmdzTMwdp4/Me08xAnjqMPXJHNPuKutiv5wbEgrbkj0+7ThUYvrXl45pCIMBABgRNYIBAFPSE+2+katzh8CJOakcEl7qMv9pTsBvZD+j9qY4Ubd1I+/ELai0AezDCO99UZlfiOpI5alFntMXDc+dc3hAT58K3a9C9iHfY5h+wRfLsBBYj2FjhMC2pFDu+L667P/INCHwg8wfVhZmnXS7/yFlmY7QWu62ZM+ctdR12ftu/mnLRAAAgIEIggEAU9ETyaWFwKmE341eT6z1suPfZTeKaZ/uetr1T/MzJBC4MuuXOkyroUds3msoDUdzGVY/1zeEyQOX00fR8nx8pPt+xOHPW1Of387Y/k8j9WnXddX9boy65zYEdmmrxCx3SBmKZ3OM+0vKuuZjHk/GVJjP3sgs52PgfHSEbjbjcut79YXBlzMvEwAAR4kgGAAwhY10X+opMn8IrCeQnwJeZwNgPbFOZTmXpuY97b817bsyJ9j3nvM/M+1VSH9gaS8RHyMMPpPdZcBLFo+wreIj3f+jmV936HQfbQpFQ4PgkNHmc32xNWYI/F3cygXY9wz5EvDRHD8j2dVCTgKPy3PSNrmVXSAc8oXgtVnfueqZZ9IfBusy3QgAAAhGEAwAGFsk/aN29hECh9xc7acJEVJZXm3CQtpHe2nYsamtfyy7UcI+NFT6ZUKFvkAgcTiRd2EvT56rj9wHvMZ31HLTZf4xh4s3x45T1NYPtiPOq88cxzcbyI5R29x+hvQttx67/xb/MhB6THhX+Tx7qvTRrhB7ieULbCD8NeC1V2abLSkM/sKxEwCAcATBAICx3fWc6P+U5YfAjyYE0DC1WHhbt9k0PGdHCWsdSd9A+IN5fezQ3noi/zzC+n2bqa+EbGPfIDhveE7DqRWHjJPXtK8+S9jo0jjgNfczrOOYIfBHh+PC2hyvPgW0xTtprzm86flsWyoNsnUkrX4Z6Ft72t6sbc4wuO/zac6RygAAHBWCYADAmG6le7TUgyw/BLajgPMDaO+s42+bjhPlQsICYdfRwbpcsRxOGDzHyMu8YzsdGgKYcW08+kuf9Uz933eZdH3GCIHfm+NPl1TKUcA+5Sf6AmCrqyzBIdzMzN6Q03d08IVDu48p6flsOhNuHgcAQBCCYADAWPQEuusmPM/iVs9xzBNJ3xBYR5ptZHllILpO6l3LQzQpTDtpAOIzKlC3cy7doZMdPfw4wnpOHQaHBGEaikce099JczB+iEHwhWAsG2kOSENDrnim/u9zHP5bhofAuu/oaNasY5qVOS75jAK2V3/E0h++x9Jdf/mQgkkNtP8Svy/rrmXem7XpMnaNXr4U6gUDAOCNIBgAMIaVwwninAGrnrB/85j+2YQBtwfY9l3hQ+o4j9y0mQYDrsGtvVw46ZjGjj57GGE9v024ffLA121G2FZagzM6kuPAVuArCdivu/rj2YK2W+J5HG7zaI5P255jfiHutYD1mP/R7Hv5wG2lfsrhfIFY7WOx5/H5Wua7qufJ9Onnns+4iMMIAADuCIIBAGPQk7GukVIfZb5SCyvxC1HsSOX8QNs+6/jbufiNENR20+D2s+P0Gjp961mGJ7MMY9TP/CDTjUgLqZOaeE5/N9J8luqpsg+G7runJJLyi4C60FAxZHS5Hv+mCIITGScEfpBdvd82Oir0l7iH4N9N2996bqvrgH176eyVGz5h8DcJK0ESoug5Pp7JvKOUAQA4eATBAICh9CSyqyTEvcw70rbvZnVVD+YE/5BHMm57TuJTz/k9mdf8Ie7hqAYkubQHeXZk1/cR1neqy5PzgNdciH95iKYR1zdyXCHoeubXHaqk5fnQ/r2Zqd/30f48Rgh8L/3lhLStvjjOz5aBSMQ/aO86jj7LYdertV/W+YTBc96sTd+rq6bxpRzPl2kAAEyOIBgAMNRtzwnynCdoN+J+afCDzFuzeEpZz0lyHDDPwrzuo7jVkdT3yaU7zEvM/IaaIgwODXI2I2yrMzmcWpenFtZOZdWyzR9l3rIQYweY2r+/jDCf7z3HZ22/rbjXgderHCIJC74j6R8NfOifI75hsF5tks64fKl0ly3qu4EpAAAwCIIBAEMk0n3jKD15K2ZalsjjxFSDzUO6KVyfTPrrKIbSE2wN/1xGB9u6weue+b0fYZ01mNmOePK/Fb8bJ1k3Ae353DKfQwgyVh37FPz6zVnLvhx6LA6Rj3wcuh5hPp971icyy+1y00INNv8ceAxMB/79UPiGwR8k7EvG0GXr6hOH9GUaAAB7RRAMABii6wR47pIQmbiNiLM1gYsj2g56ktw1sk9H624GzL8Q99HBug3ynvfTbfVOhoeHNngeK0ANGR3pW4e5bVsdepBRLa8Suj1OZbRx1LKtnwOPmTq/q4DXPYx0HLQ3Cx0jBH7f87li6wW7hMBfpb++sEvbXvd8zh3bZ0nscWzOZly2XLpLRHwSbhwHAEAvgmAAQCg9We+6QVwy47LoiatrSYgbOeyawF3bo8sYofytuI0Y01DzR08fyMUvcGhjw+AxAoC7idq+afqm9T6EIGM90jRNTuXS7lSav7TS/SvkKoXQY202wrqszP43NATW/eFdzzKtzXv1feFnawGP8cVKNvK+fwh8wuB9lIh4PrHtAQDAqAiCAQChJ/9dJ9k6aqeYcXkyx+m+yvHeYVzb+/sMJ+waoq+le2SWpTeMSnrmFYnfTYqaXFSWawgNgkOC6UvxC3B1W90O7Mv73PebPAlcxNIcmoaOBu47Fvf196F9IRe30bld7FUaecc0G3ELgX/KLjAeaiPdXzDeyzQ321uCrUe/mrOszVPP59i1MCoYAIBOBMEAgNAT5LOOk/p0xmVJpHtksvUgx19DsK/db0Y8SdZ5/SX9wWlfGGxHn90PXB5bkmJoGJxN1PZ1bbWCL2Xe0fS+2vrP1mGaPpdHvn+uOvpXKmFh+o2E3SROA9NiwLr4lGjoOy73lW/Q/eGHw3p+lPFqv6+kP5hPj7y/ZqafuBx75/xs1e3yeMLbBQCAQQiCAQAh0p6TtKeFLEtVcgLbpZDuUcFnMu6IUx1RqCFO34he1zD4+8Dl0fX7f+zd73HbuBr3/Ss7+946FZhbgZWZ+72Z9/eMvRVYqSBOBaEriFNB6Apiz1NA6AKeiVxB6ArWqiC3ryPgiNFK/E8CBL+fGU3OWdsSBYIg8CMI/ui4r9suodF0Jpp+5+uSbYg8rWN1tutUcKytOlQ2zzKt2cB2xm3X/fwo1eu1r0z7UUbL7630uyZ9UvH9Qp4NvF/+de6SGPthl2Xn/UuZzzIzAAA0RhAMAGjqsmSA3Pb25rZiqRdG6FPo1zPZP0nFz8+l39lbuWyDoaoQtyoMFvPzjz1s09cO31G/T9vZyWmL3z/0WX0H9n06NuM6M/9GPRzTIdLv9aGk3rfRdjbwc4f6tZJ6SzRUuTNl8lLxWVUhsF0KYj3SvqrbzobipWZbeiLjXmxN5fis4Kk/eBMAgEG9+fXrF6UAAGhCQ4Bjt3DfjDxA1lltFxW/82yCApdrmMZ7/1+3x85YWhe2Levp83QffCr5uV2Ts+9wXIOAWykPiT5K9cWCSzPQ7yNsWrX4O/38by0/813D/Ri9vn56cjxV0Tr7T8X37lJ26r2Et463llt+pD7rWtttQquyelOlbb3SY+lrD+VRZx/X+awhjo+yfWU9mHreB3suWMjvF1n225C143NY2Xm/eK6NRtymsjoy9rYAADAZBMEAgCZ0YFUWPvwl4z0kriyUKnIdLCVSHsruezKD7lTaB7UL87enFZ8TS//hgoYZ9xWfXWef1HmfOh5MYND0e+YtP7tNAKFB4OcjP2saLA9J68v3Y33KlvV9n2/hdx90/50fqSttL1LVuQh2yMbUz6afqcfrVcdy0M++rFGf9Xj9WvE+K+n+sLs25box+6zteW5ptl2PpSbrK9tlNHw89oveyrh332g9Phba/z1QHQEAYNJYGgIA0MSq5Gd3Ml4IXLUtVpdboF3RcEBvS/5hylO/Z9P1DuvczqufM8QyHhoCaNhRtrxCnWUi7Ps8ddweDXWyFmWYtPy8U2n34Lhj5aVBRuRJ3YxLjjPrcqDPmCqtC+clbdhLyzK6aLk9acPPXJg62DUEtheeshrt+tca7zNEwHddo1xvW57nVqZN+2Ha97OJ1eNM6i2ZM/aSDGXnsEsBAAD/QhAMAGg6mD3m3qNtsZKJl7eGihqK5C0G2Lo/qp74fiXDrOtY5+FvdcJgfZ866w9XOZPdWsZ1pVL+ZPoyuq+iFvX50EOZTsy+9OHhR/GR/25nAEbSPeA6l3BoEHVsdrQuCZG1eM+FdLu41eTiT2S28aJjOTxIvaVoVlIeAtd9n7Z1+3PF7zy1OKesTNvzVaYX/rY5n8Yjb1NacX7joXEAAOwhCAYA1KUh2rFb5TUwGzMIXtQYVG/Ej9tCdfD8xrz0tlm91V9vf2/yQDINAzWkyBsOtK+l+onvX6VZQNrESrbLQJR99qqH96lbhpk0C76TDp+VNvybvGTbhpq93dSxkDYz//Y1Ay+W6VuW1IEnaT9zMpH2y6XcSP3ZrPYBbF3DyxtTL15qHONlIfCXmu/Tdl/VOVesGr5nZr5Tk/31aMrsb3OusOcOH46JTKovjp0OeD451m6WnUuZFQwAwB6CYABAXWWD4LED1zqD4ntx+3CdQ9ZmMJ2Y7/DGDPgfav69DrJ1nUYNBevMdMqlXniRDTh4T2UbaBwLpOuGwfo+GqQ/d9iWE/N5SYNtb/t5Gpq2mcV9c+RnV+I2DL6sqNci/d0WPvXwZmGOqUNrl246fD9tMz60/NtNg/qjx+MP6fawxo1p25Ka36ssBH4vwy05YGdYV33XG6k/Ezkx5Vd3dvuj+Y7/MWWRmLYg87Bu33p4/KYBtyUAAPSOh8UBAOoqmx029gNidDBaFYiMvU1dRbINYDTwqBPAPMlu3ckqdR4sNdTD4yw7Q/JYHar7UD8b3PRxu/qqxvfVMvne8jM20u5W9rL95erhh8e2aWP2iZbl154+q80D93xhQ+Bj9bztw//qPACyTN2H8Gnd6mM9YA3g8prtQibHQ/M2x09f57Xi91nW3Ed6nNQNgO/MPsknVr+rHtI69oPtyrbJtk8AAMBgRjAAoI6oZLD8LOMHrlWDchfb1FVuQgEt6y81fl/3Ryb1ZjytpHpmq32/oQbNaxMOHJv9XHdm8Iv5zjcdt+dCdg+kK6Nl8tjyM9osEWH311PHcur7+D8WTN+bMuxztvKpTHcmXyblFzuylu+bSvsQ+FmqQ2AbNHcNgb+Y+pDXbMczORwCP5l6N2Q7nkq9JYYua34X/c51QmANgP+S3frBU/Ii1XewLB1s02NJGxzThQMAYIcgGABQR9lAKnOwPVWD7WzCZa2DWp0V/FaqA0gd5H6Teg9d0zCjar3gs4HLzm7HxyPb0iTkTGQ7u7LLUhGnUm/d4FWHz2izvm/V/tL3GzNsSUp+lsvxMK+L6wkeu6kcDxbvpP1Mbi2LiwHLMjb7sct6wHYpiLr7rSwEvjE/fxl4X9UJvVdSHdYuax4Dz6bNqvOePqs6R2g5RCNv033L/gsAALNDEAwAqCNuOQAbQp0A7D6AMrczaD/W+N06Aaq+X52QRsOgdODvdmu+29OR71J34J6Z+vDQYVvsusFaZ47Nhs6l2wxkXcak6QzX3JTD5sg2ZzJOGKyfURaY1V3KpKlzmdas4LSknB6k/cUELf/PHbbrsaI9TGS79MlJx8+o+8A1+52yA59pg9LE4b4quqnxneqGwF8Kvzt1db5D5NE2xQIAAP6HIBgAUHfg3mVQ2Kc6SxesAyp7DU3rPCStThicSr1lJ65k+DDYLstwaHvupX7IaWfP6m33mw7bY5eKiEv2Q5fZx6k0D27LwvsxwuBFjXpwMuDn62dHEzhGUzkeLD5J+xDYrjnb1qbks5emfn3q+N1vZDejuO65JDtQbx5knKC0bF/tb0/S8rvs7wM7U/pFwlDn/Bo72KZj7f+5AACA/yEIBgDUUbY+8NiD26ji5zoYzAMrfxuaVi0VUWfJAA0k7mp8pg2Dh37Qjm7Pu71BfJuQM61ZRmV0qYjvphz3v7fW81WH9z5pWZ76N+9L3rNpOTVxK92WC+hKv9+9+P2wp1TKQ+C4Qxt5L+3XBVbJkbZQ//uPjvtWv9tbaTZ7N5J/B6cbU78vRziXlO2r/e+2avFdDr1Pk5nSU/Lo4TaVBdQx3TgAALYIggEAVeKWA6+hRB0Gg1P2YvZFWYhrg8Gq4EzD16can3klwz5AzsrMfn3c+y5NA5RcdstpdJkd/MG81+rAdn7p8L5tl91I5fgSIXaf97mMwsK855UH9f6ssF99YmdLHysj+5CxtuGmhvBdZjI+yr/Xpo5NWfYxC3jZoq29ld+DUxuUpiPsr7RmfX6W6vDeztSuCoFteYdoXeP4GFvWod8AAMBsEAQDAKosOwwGXcgD3x8rqQ6DqwJUGyrXCYPtA+SGHtjbbbJBqw3S2rAzo7vMWrNrB2fyewiZSLclIi6k3RqotyX73T40MOlhP8TmGPLpdmr9ft9ld8HAhzZRt6UsBI47tEV6jH/osH37S0JEZnu1DLvMMG4zC7ioeAHqTsYJSm1oe1Wz3OqE91Uz5bvOBJ+ClxrHyNjK+iORAACA/yIIBgDUGUi3GXi5ks9gn6yk/AFpGuIlNQbysdQPg9cjDe41LHpvtm3dsR7oe+j6nF1mB2tZ2hAylt2axF18knbLTOjf3FW8b9f9pN/xxNN6r/sicrwN16YuHAsCNx3rru672x62MTdllb6+fkq3YF+/00dpNwv40DF5Y+ry0EGpndl+UfM71tlveuyXhcpzCIF9Pc/mFe0aAAAQgmAAQLWyAdQLxePMSspDXA0Fo4r3aBIG60zCbKQBdSr9XWS4N+XwpeP7FANhfb+bju+ns42XLfd7WRisAaWu/5pwiPTKzgL+LMeD8q4hcCTV685W0bqRyy4A7rq8x530E04X25xkpP21lnrrINfdb1UPT+y6HMiU5B5u01oAAEClPykCAEAHGUXgjH14me6DY8FRKtXBrQ2D9X2qQhN7i77ODrydWFldm23W10WH9zo3r+eejh8NrPKGf2f3+9eS37GzjvV7N1lrWbfl0fxb3K61HA+4yn52yELKQ/BDP48L+3JMui2JVC/VoBdTLqV9QFZn3dk6qmas1vVovvcU2/hL0/bVKcsm4f11xXt22f8Y1jlFAADA1ptfv35RCgCAMjqwPbau5BsH25NI+cOObmResyE1nPhc8vN3Ui/MsbdRn9X83Dvz2VOc/RabOuJDONDlVnL9uzrh4ZRDPVcWpn5f1yjfPo6FzJP6+CzNLx74RC/01F1fuUkIHMl2hvUxX0y5zYWW23cPz8Nlx9EbAQAALA0BAKh06tn2cPvn7zT4KFvaoe5g3M4MrvuAtSvZzWidmsx813fS7YFyfbAP42v7PSIpXy9a7a9zjOPsDOBcthecykJgDRJ1PeuVdAuBU3EfAj+Z76L1aYohcGTq94cG3zdqcD5JKupBMsPjBAAATBBBMACgLVcBWlXgMscBatlMNA2Ylg3KNpbyNWiLpr4ebSa7QPjO4XZoOaYdjge9JV2X66h6KB6BcHm7ofU4l+oA2LZ/yw77zdK/v3L4vR9N/e/ju7ii9X8t9cP0R2k2C39RsY+memdEF1XnlNzDbabNAwBACIIBANNTNYNrOcMyyaQ8mG96y/JKtsFiXZ/MfllOuPz0O/8l21uanx1sgwZNaYe/15nhkdQLtAmEd5oGwFo3/jbllnf87FTchMB6weCLqe+xTHfJEK3vOnv5m9RfW/lOmi/FsqqoD6lgX04RAADgJ4JgAMDU6AB+UxEOzFFS8jOdMdd0prQGi++kepapVZwdPNVZ2bnZ/kh2s4Q3I35+1zDYPkCw7pIXcw6EmwbAWg/0IoFe7Ohj6YRUxg2BN6Y+/y279Y/zCe8/3X69+NTkwY92GY82n9Wm3Q1ZVXvBEk4AAHiKIBgAMEVlg0xd03iOy0Nkcnwmq4Zcly3fU4OvpwZ/Y2cHXwZQnitTl97KNgQcYzmUK1N+i47bHss29Kszu3lOgXDTAFjdmeNA/66PJQBSGScE1uP2i+zCX63P9xPff7E5Pj5L/VnAz+YYTlt83lKOr5O/CaA824oqyvtFAACAlwiCAQBDDASHltUYvM/RbcnP2gazuSnPLw3+RoOTbxJOsKjBU2K+iz55XkMlnV2o4bA+qE0D4j5nDtsHyHWtx/fmONVtnXsg3DYA1uUTVtLP7NmFKds+Q+CNqX+Ppj7qki7vTD3V+nMtYYSVWo9TUz/PGvzdgymHtjNULyuOrzkGnpGUP0SW2cAAAHjsza9fvygFAECZshPFG0fbFJtA4BgNLa9nuK90gP7zyM80MFr0UO4afpw0/DsN1BKZz7qRGh596+F9Nua9sp62a2X2w2nN3380v59NeF/YZRCupdk6sn3XV7ukxGkP76XBfjqjNk33xVWLY0f/7rbj52uoeSx4/lvmOSNY25GvntZPbauOPTTw3cTbMgAAekEQDACo4mMQrHQm1rFgR2+Jnuus4LLg4q10n621MIP8ixZ/e2f+dg6D8ZWUhyVN6EzPpOdtu5b6MysfZJpryur3vJV6AfDG1M3bAb5nk+2ocwytZnD8RNIuAFaP0s8sbm3r/impL3NcgkjdV7T/fzlsK8rOf28EAACwNAQAoNJzxWDd5WD0mDOZ70PjspKfxT28vwbwOku17vqzRRrq2KUHVoHvh1SaLadR5pMps6jHbdMLJXUfKqehj840TyZS9ktTXhrE130InJZt32H3wrRTdbajblsc+p0OsSmzn9JuFvBH8x55T/WoTTsbskXFeeRJ3F4wOhMAAFCKIBgAUKVsUBc53K6qW3KvZ7q/spKfLXsu/6ZrB1t6666GYxoq30p4ob1+n0T6fWCeltm653qtdSWW+oHwJ9MexB6XvZb7Dzl+e7hVDIAT6X+t19iU1UWP73lq6sAqwOPFhvDfpf3dBpF0Xwpifx+2aWdDpm1a2UWN1OG2zXWGNgAAjRAEAwCqlAUkLpdf0CCy7AFdq5nur6zkZ9EAdUMDHL0V+LHF32ug8EG2s/9syBlNuOwvZTebUUPT057fX8vrs/Q7O9jWmVjqBcL6nTSs08DNp+AlNnXoU8XvDR0A26VTvks/s4APlb9eRMnN8TLV8Gth2mh7vHxuebw8mXq7GmBflh1jc30gWtV51eWayWX9kUcBAAD/RRAMAKiybjlQHkPZ7K8TmWcYrGHIpsVAuYtcdkHiU8v30Ft6NQwqhsJTWOdZw9/UlLs+IO5ihM+0s4OTnt83K+zHqmU/PphtiB2X/8K0A9+l+rbwLzJcACyyW5f2aoTvfWqOF13D9t58tu+hcGSO63uz3V87HC9aP9/LbhmQoba3zXkxVLGUz7R/ELfLQsz1uQAAADTCw+IAAFVWcvyhV4/iNgjSgfrPirBAB4cvM9tnWcmA/c1IdSaRfmbEaqh9b77TWtwHMLGpU5dSvfzAGDR4v5ZhwrBrsx+rZrZ+keHC1TK6D25r1LO+Hh5W1g6lntSHR1MXMnG/fEFkjhf76qM9eDZ1LR1h+/OSbZ7jg8eyijr+znGd07bgw5Gf9f3ATQAAJosgGABQRQfw38vOJY63TwOBshl4cxwAlg3Yx9xfK/PqMyDTYHhtvmNuXtmAdT+SbfC7FD+CvmPuTD3Pe35fO+O2apbrs9nX2QjfNTLbdFGjrug2DXW7+sKU+QeP68WT7C6g2NcQgf2ycKzE5t8+l8YYMwC2jg2SXF8AdUEvunwr+bkPZaJ1+9hdAe/F7frFAAB4gyAYANBlQKzeittZmpH5/BOPt3FsZTOjXMza0oBAZ5cOuWzCprCPi98vl+PhaCS/3/4dm3/7DLE0wErNew69bMTG7Ht9DfHwM/0eVbM69fbwlQw3O7juLOWht2Nlyvlk4H36aMo9kX7XnLZrphaD4ZeSdnIhv996vzT/LZL+18Le304tZxdrzxIE7/Z9XlHXXc8G1m38hz4AAADVCIIBAHWUzbT5KP0+qb2NRMofEvVkBu5zWSKirDxcDtgj2c0SPg18HzyZ4yI1/1+Dsx+Fn2toO1SIONTsybozYIeYibs05Xle47tfy3DBYSz1AvEu+6743sWZjJemXC8CP3Y25jvr/s4dbgdB8NZ9RZ17MHXTpaoZy3NcygMAgIN4WBwAoI6ymTQ+DIg1MCh7uNWZuA+rsQ11EtkGwn/LdjmDTUDfb2O+k84+0+Ay3TuGit9VL0q8G+j7a5D4Vfp/mJtu87VUP0xOA+5v5pjr4wFm+pkaoleFwF9MuQ8RAtuHkn2XYUJgrQfvD7S193v/WwOvv8x3fQ6sfXgwZbAw+zynyXRO2+uLinp77cF2lrVzj+xGAAB2CIIBAHVkJT+7EPdPq9eAalXxO7rGaTqT/RVVlJUP7s0+07oz9VC4GGDpd1qXfGdLw8Tc7KungbZLL4B8N8dv1HN7oMHol4rf+1D43TYW5u8/V/yelp+G09cD1O/ItBt1gui27B0Lqfweuj0d+T65+a5RQMfOf2Qbcs+ljZ4Cbcs+VfyOL4H9Zcv+CwAAs0MQDACoo2ogdenJNt5U/M5cwuCo5Gc+rpNYDIU10NOA8cnj8tWZmBq+aQj3RuoHWPcHjhsN+pY16m4XGmD+NNvY10WburODz8yxuWr4/ktTV8vC140pt6X0H/bYZTDWUv2gvC60HsVyePZ22vDY+XsCx86T2cb9Y2dKy/ZEMziHaJ36WvE7D56cT/X4P+3QfwEAYFZYIxgAUFfZOsE+rBFYHPRVzdy7k+bBVCj7akprJWq4FZuBvv33xMF2PJl6tTb/5h2+zz8lx41+x1SGXT95qAfKJVI9e1ADwDq3kdtwsGxfP5pjOB+gjFYy/IPgDq2jvF+GXR5wZY8de9ycOzqGHwvHTSbTCnzLziUhrzlb51j2ad39soejhr6vAABojCAYANDXYOs/ngwK7e3kZzUGsho45QHuq2Mnd525GU38u0XmFRf+t766hqcazK1NHV6bepFL/7PJiiH9Rv49Q7fuA9m6GuKBcrFUB9l6EaZsCYeVlM9E3JjtHmLN79i879nAZX8sxNa6Vgwe+w6wluZYWZrXQvq5uPJcOF7ywvGznnhbs78/iv4K8NyxkH8vT3LsGFx69P3zkjbHp4vUAAB4gSAYANAkRPhR8vPi0+19GNDqIL4q0BkyVPJxP83hafc24KrLhr9j2b+gcixQimX42cG2TlxLf6HdwnzHsuUUjs0mXEl5CDzUxZvIbPPFwGVtH6x1rJ385fBYtaFwE1ngbUnZxU9d2uI+oO9aZxa+rcOx+BPy67Z8n0i/BAAAL/xJEQAAatKBn878OhZMXXs04HoxA8RMysNgHfTqg6hWZvuzAPZTWZiTzaSe+izf+/+RHA42M7MvtV5+GnB7dMajXjjQZRsS6R6K2wc36vYfC3XtusFx4fP0u5ZdkBlqOZfElPHQS45UzYSOHNfjF2Et1SZtidbdEIJge9zVWTrEtxBYarQJ91RjAAB+x8PiAABNlA2qzsSv2aY2DK7z4CTddp1VlMn0Z8yWbf+aKuzcusH+0jqcyHat2MeBt0tnPubS323UqdnuTckxp8fbwry0bTkWxuqsvtUAx4nui08ybAisF8/eme0vC9mjvf+fc6h4d6zWPW6nYGmOUb0IdF6zHseenUO03Si78+BBprUmNQAAoyAIBgA0UbWEwsqz7bVh8F3N39cBsQbCuWxn70UT3EdlQR5BsHt5i79Zm3qsgejzgNumgeg32Yayix7ez273sYsxNgxO5fidBn3f2m2Xrvguw64FrAH4R9OGZDV+f8mx6p21lF/ImNr5YSG72foaAF/V/LtHUz99q5NVD55MqcIAAPwbQTAAoIlcymcmXnk4OLa3qr8vGdTv01BKl4z4aQa/iWwDrYXn+0dD4GOzG+0DneD+GCqKG/xtKttA5qZBXW7jQvqbHVwnDD62Nm/fIXBstmfoB/HdyW7d4boWHBpeyiraW58tTJ1PzPf4R7bLtZw3eI8bObyetw/frSwI1vMdy0IAAHAAQTAAoKm04ueJx9utIdpDw7/ToEpvH/9uBtL6QKe1GVjbV+zJdywLJhgUh8EuF6F1+W7Az7Gzg/W4WfSwzXqMNJnNfCf9hsCJOYaHfPieXiTT5TBW0j04Y0awH8razZUn27jcOx9l5jz1j6nzev46b1mXfT2fV63rnVJ1AQA47M2vX78oBQBAU7mUByp/id+zT2MzwD3v6f1uPBgwL8zA/xgd1BMu+aHY+XqUbhcSmjzsqS2dzbvqof7YwOqkxuctezwu7gcun2fZ3XLfVra3jW84TLxQ1a6+E/cP2dP243uPdVnPZann+ySvaEd874MAAOAMM4IBAG2kHX/uWmYGzzqI72NWpQ+3dVfdJksIHKZ1oS4PtX6wXct31cO2rjrW5Sa0XHIZLgTW8tblKyJxHwZiGDqzu+wukpUH2xj38B76Hf82ddn383ci5SHwnRACAwBwFDOCAQBt1JmR48NMqSbf59IMqPVVdfu4XW93bf7NxG3QWrU/fJixjJ0+ZwTvW8l2hvDJQNt+J93DL62Ln4787Iv0EwSXfUZXG1PG+upr7VRtQ5gR7Cet719Lfu569mlkzl/679L8W+ccZpc4upfpBKf63X5W/A6zgQEAKEEQDABoK5HyoOXZDEpfJvr9dNv3Z/quPf0+VfviPxPeDyEaMggW2T1IaaggtI+lIjR82n9InAasUce6OvRSEHembPs+nrK9bSbM8ovui2Phqs6m9fHBcQv59xIrLzLtu0OyimO7jwtVAAAEjSAYANBlkLmW8plHfc3uw3GRlM+QYmDsn6GD4GLdSF5fVwO898Yc22mH9iOT7bITlt6a3uWhhrH5+yFmQz+a7ztUiKb7qRjcT+mOijnY3z/72F/D0+Pvc8XvcAEFAIAKrBEMAGjrRaqXG/ggw4Vc2Eorfp5QRF5ZjvhZuWwvAmhI9dTze2vY+tXUvzZrZL+YbduY/6+zKruEwFrPv0v/IbDe2fC3acdYZ3u+bgt1tU07jG6iGueyGyEEBgCgEkEwAKALHfw+1fidBUU1iJVU3ybLwNgv+8dCNsJn6mdoAP1RysOsNq7M+0ct/tY+PM7OLm5bnhogD7EMxo0pt/sR9tH+UhMxh4pXdP/clvxc74xJKKbBVM3031TsHwAAYBAEAwC6WlX8XAfIKcXUu2XFwFcHxgnF5J147/+PuXaz1pdItrNv+6TLO2io22ad1HtTl/OWx4B+7kXP30eXgfjLHD9j7Z/92cYRh4p39Ph5Lvn5JyHAH6rcz2r0Q1gHHwCAGgiCAQBdaYDxpeJ3NKhJKKre6CzIVMpnSOngOaeovLM8cPyMScMSDWx1uYM+ZwdrXfzW8jhvU09Xsp2JfNrjd9Dy0FnTsYNjZ11RT+CeHjtVM9fvhTtg+qTH+YeK33mUcWbtAwAQBB4WBwDoQ50Hx6n3wuzgPmgZlj0ATJfrIEjyU753nPxH3M1ksxcU+p5Rq0uSXA/4vfQix4ee31OPmUtxe/HEp7qB4+4rjhmtSzH7rjM9h2VSvSTEUrjoCQBAbcwIBgD0wc4yrHIrBJRdpVIeAqsVxeQlrfvFoO9J3IZF9rj92PP72nWDFwPV/75DYLsWcO64fuzPCo45ZLyk7WvZbPoz4YJnH22ltiEnNfZFTnEBAFAfQTAAoC8aYlQFSieye3AVmkulOgT+KOMvN4B64r3/n3myXXqB5q2Ur3/a1Jn0Gwbbuw6uetxGDfN0iYzEk/2wXx8uOWS8pBdQVhW/cyGEwV2O9VSqQ2C984AlIQAAaIilIQAAfau6bVZpABMLgWUTOjCuCsH0IWCER/7S+l586NHf4leQMcRSEX0c67pdmVQ/MKoJH5aC2Be9vn7ulR3rzfqrzhIlGlauKKrej3WW3wAAoCWCYACAq4EcYXB9qVSHwAyM/RbJdEI+Xd/3c4/vtzHfv03dHCIE/iLVD/1yxfeLBfid1s3zit/RC3Qr2uZabeR9zb4D6wIDANASS0MAAPpm1x3dVPyeXSYipsiOWpiBcVUIvDFlTtDgr/3g0edwr++lIuyx3ib4rhMM1WWXgrj2uOzTinoDv2i7+1TxOxcy3JrZodBgdy31LyDnFBkAAO0wIxgAMOTATge/JzV+972wnuK+SOrPjoqFmdU+0wAo3zsW3k5gn/W9VMSTNFsfXD/7qsfP9m0piGNl/s/ef/tLCL5832cZbXVrsTnX0VcAAGAEzAgGAAxlbQZ4mxq/+1W2sxCxdSnNZkdNOViIzHcIebaczuoshhxPE9lndnb/x57eT+tz2qDM+gqBdSmIqdxKrmV+t/ffksDbu9i0A1NlHx5X5y6YH8Is7/3j/LsQAgMAMBpmBAMAhtZkZvBUZu0NpckMzCmHwAsTAOwHpM/m++tFgZeA9mm+9z2nGGg0OY6rVK17q5/1o4fP2Zg6NsWy3v/+byWsmaQrs2/O9vaX1otkoueASOovZfJoyoBzXT2EwAAA9IQgGAAwhiYh0sYMkOf2gKRLM9CtW0axTDMYikxdOC35nWfZzYqeOt2nV3vfLZrod+nrwW1VD49b9/AZU69DWs7Fh5A9ShjrqdepQ1O/yFX3GNHvmcj87oaJTbt4WvP3CYEBAOgRS0MAAMbQZJkIDUK/yTYInsPDdSLZBgffpP6s6UimG3DVCQBOTZksJ75vtc7vL28w5dvCX8w+uev4Pidy/EKPhmJdQ2C7FvGULyTs15Nzmf6SAnZ5jrOa9WMx0WNEj/vHmsfB58L5MXQLc3x/l3ohsH24Y0oXCgCA/hAEAwDGsjZBwFPN378woUGo6ylGZoD7U36f+VfmwQQGU1024bLBd9WQJJPphsH21uciDYdCmOm+ku26u10cCja1bn/o+L53ps5MfWmRtRxeKzia8PFQ94Fg6tTUsymyYXDdY0SD8e+mfCIJ08rU6brHt50VPrc7gwAAGBxBMABgTLkZ3NWdUWhnTOUTDgX26UA/lW0A3ORhWPrALg1SpxxwNQ11pzwzUGe+7c96C+mihn6X9x3fI5Fd8HUoOG/qLqB2wpbx5sDxMEWZ1F8KwIoDOUY2NX//wpwXUgknEI7Nvv/aYP9P/a4XAAC8RhAMABibfcL6xwZ/c2oGkrlMN+iJpV0ArGud6oOiQlhHMmrxN3aZiCm5PrCPbyS8YEPrc5cw+ER24W8izYPCotBC4GJbWXQm07tVPpV2y30sAjlGmtwJI6bt+GnavcsJn+90+3Wm83mDv/siYczoBwDAWwTBAABXNNh823CAXAyErycQFCzMdq7NgPiq4d/bQfHcZ0ZNKfxayXYWe5HW8STQfZNKt2Uizs17dFkSIsQQ2NIZwA97/+1KpjO7/LpFuxea3LTjNy2OjW/m77X9iCbS/tnzXZMAWGdNv5Nwl4ICAMAbb379+kUpAABc00Hup5Z/qyHQvfhzy7SGv5fmddHyPZ7NgDpjP//mRvwOVGPZBiBFGnDYh2SF7L5Dfe/iSaa9bnbdNkXbgv1ZtTobO/V4u7UN+9rh7x9kujNij9G2QC+Cnneo76k53nKPvtO12VcnLffzSpgFDADAKAiCAQA+DZB1gHvW8u83sguEs5EHlUszCI47DPDtd9CQIAl0H2sZfev4Hr6GXyuz7/aDkHcSXqB/iIaVubQLgrrQuwrmMGN+aerRyUSOB20Lv3d8j48SxpI4TdqLJp5Mnbh30MbY853+23ZJl1AveAIA4DWCYABAiAPk4iB5XXj1IZJtKLM0A+Gl9BN+6cxmnVUV+szGf3p4H9/CL62zXyewnUPrI+hvIuSg8BBtbw6Fq1/Er1vqj4XWTf0lYc+kt0sHXfd0Dnk057nMlNtQ57zzju8X+gVPAAC8RhAMAJjDANnScPjFDJBt4Fr838XPX+79b/33bIDvemcGxPlM9m0q/awZ6kvIeuz7+L6MxVA04Pkwwudo6BXPsHxXcviigy+31/cVAj8V2mDOd/2e8/Ij55v4wP8+73l75nDBEwAArxEEAwDmOkB2SWdEpbINzfKZ7VMNd3709F4uw+CypUxCfnhZnWNWQ6fTgY+f5QyPHUvr1qEw2PWt9pfmmDiZ+LHN+a5/c7vgCQCAtwiCAQBTGiCvZNiAaWg6O0vDX13Tcc4zovp8sNjYt8XbunjsoXdzW67gkFi6rw9bhjIufxCbi1mXZdvTpp1csn//+zqf8HewS0CkQgAMAIA3CIIBAFNzaQbIFxPZXjv7V19rdt9/RaYs+pr1NtZt8foZiRy/GDHHWYzHDLVExFyXhDhEy+H+yHFkQ7jbEY6Lvvf1XB4AWLetvJZuD2Ubm73gSVsIAICHCIIBAFMeINtQ+MyzbdNbtDPZhjT37KqD+n6w2FC3xS/M+2oYcyyI0dAtFsKr/XLLej42574kxCFlS5TYMhtqVmbVZ7cx17W165b3SvwMhZ9ld8GT4xMAAI8RBAMAQhCZwXEs7mYK60xFDX0zIRCsK5HjSyx02Q+JdAuEF6Y+XdaoT4/m93j40b/1uR60YkmI4/VVy6XqIYwPsrs41aW+RuYYu+r5e8x5be0u5zx9uVhTmHMeAAATRBAMAAiRHRwvzavv2VNPZuBrXxlF3loq/QdK6tG8t+6bvOJ3F6ae2HpTd11OgslqifQT9rMkRLUmD2t7MsdGZtqwOsdI3YsjbRACd7MstGGR9L+28KZwrrN1hotfAABMEEEwAGAObNAXmZf9/1Vy2QUkmRn4MvOpf7rswucB3/+5sC/t/ozNv1oPms6m01ByJdwCXZceM12WD9iY45bgqV5bl0q7sPaxsL9eCu3kQoZdfocQeBj2fGf3of3/dc979nxH6AsAQEAIggEAgA9iOf7gK19ooHwtrPvcVNclIv6mzFsdT4n0PzO0TxtzPKXsLgAAgHH8QREAAAAPZLKdrXbn4bZpAPzebB+BZHM6o/Cm5d8+UOatj6fY1NsnD7dPZx/bh80BAABgJMwIBgAAvonFj9mMdp3hlF3Si6zhPmVJiH6PKZ19e+F4O5hVDwAA4BBBMAAA8FUs44dXGlRpSKUPgcvZBb2KZDs7uO7yHywJMcw+0Ae+rWTYdX/3cVEFAADAAwTB6GolPOAD8AkDbYRoYc41sfQfCuus08y8NHTMKe5B1X0woC4JcUlxDSoyx9Sl+bfv9bmfzPmI4woh0qVNbikGwBtr08cAUIEgGF1pB+gDxQB44wudIMxAbF6RedVdbkCDqRczWCi+MK6sYp+xJIQbS/OKzPGlF2Dqzhp+LBxbmfmX/YeQ/d/X1/9HMQDe+P9fX/+HYgCqEQSjr8E4AD9k5gXM+bxUlAuzEX0TSfkSEe9ox7yjofBy778R9mLu7diKYgC8oX29lGIAqhEEAwAAYGyr19fXA/+duxoAAACAgRAEAwAAwAVdO7a45rM+qE9nnTLLFAAAABgAQTAAAABc0OUGctktEcGSEAAAAMCA/qAIAAAA4IDO/F2Z/61LQmQUCQAAADAcZgQDAADApeT1dSssCQEAAAAMiiAYAAAAAAAAAALH0hAAAAAAAAAAEDiCYAAAAAAAAAAIHEEwAAAAAAAAAASOIBgAAAAAAAAAAkcQDAAAAAAAAACBIwgGAAAAAAAAgMARBAMAAAAAAABA4AiCAQAAAAAAACBwBMEAAAAAAAAAEDiCYAAAAAAAAAAIHEEwAAAAAAAAAASOIBgAAAAAAAAAAkcQDAAAAAAAAACBIwgGAAAAAAAAgMARBAMAAAAAAABA4AiCAQAAAAAAACBwBMEAAAAAAAAAEDiCYAAAAAAAAAAIHEEwAAAAAAAAAASOIBgAAAAAAAAAAkcQDAAAAAAAAACBIwgGAAAAAAAAgMARBAMAAAAAAABA4AiCAQAAAAAAACBwBMEAAAAAAAAAEDiCYAAAAAAAAAAIHEEwAAAAAAAAAASOIBgAAAAAAAAAAkcQDAAAAAAAAACBIwgGAAAAAAAAgMARBAMAAAAAAABA4AiCAQAAAAAAACBwBMEAAAAAAAAAEDiCYAAAAAAAAAAIHEEwAAAAAAAAAASOIBgAAAAAAAAAAkcQDAAAAAAAAACBIwgGAAAAAAAAgMARBAMAAAAAAABA4AiCAQAAAAAAACBwBMEAAAAAAAAAEDiCYAAAAAAAAAAIHEEwAAAAAAAAAASOIBgAAAAAAAAAAkcQDAAAAAAAAACBIwgGAAAAAAAAgMARBAMAAAAAAABA4AiCAQAAAAAAACBwBMEAAAAAAAAAEDiCYAAAAAAAAAAIHEEwAAAAAAAAAASOIBgAAAAAAAAAAkcQDAAAAAAAAACBIwgGAAAAAAAAgMARBAMAAAAAAABA4AiCAQAAAAAAACBwBMEAAAAAAAAAEDiCYAAAAAAAAAAIHEEwAAAAAAAAAASOIBgAAAAAAAAAAkcQDAAAAAAAAACB+5MiQEAWr6/49bU0/xZlr6+1+feFogIAAAAAAMCcvPn16xelgKmLXl/J6+uq5u/fmd/PKToAAAAAAADMAUtDYOqS19dPqR8Ci/ndn+ZvAQAAAAAAgOAxIxhTpctAZK+vs47v8/j6uhSWiwAAAAAAAEDACIIxRX2FwNaTbNcVBgAAAAAAAILE0hCYokT6C4HFvNctxQoAAAAAAIBQMSMYUxO/vr4P9N7vZDvTGAAAAAAAAAgKM4IxNclE3xsAAAAAAABwhhnBmBJdx/fHwJ/x9vW1pqgBAAAAAAAQEmYEY0pWgXwGAAAAAAAAMKo/KQJMyDKQz5iy+PW12Cun7PX1IsykBgAAgN9jicvX1z39VgDAXLE0BKZkrMr6hqL+jXaYV6+vi4rf25iO9S2dawAAAHgmf32dvr4eZTu5AQCA2SEIxpQQBI9LZ02kr6+zFn979/q6lu1MYQAAAMCl1evra+H/v5PtXW0AAMxKVRC8fws4/KMzL+cSthEEu+sst6EzhGOZz+zg5Mh/D23ZjNy84KYNfD5Q/lX7pM86uDR9g6b9hki2s7Cm3L5G0n4d+Tmdq4cytyWI4gHLqM/+fVWb0HQ7M6o6Buy/nBb+/xPj3Elo217VPe+2acP6bLeP9eFYdm+a5+245d9y7uuOvnYDVUFw8vr6RDF57aNsb8WfA4LgcaykewhszSkMnvvtFfsBZfFkbP/3sc4u5lPHpty+alv2nWrotI2JON5ne17ZP48w4EMffdv3sr37DZx7ffdUaPOyvXYxo3i8kAjZmUtfZHtHMuoMyCqCYG1Uzikmrz3Idg1XBkU9Hhczrk96VfxHz++5Me+bUz+xN7hfF/6lE1tNy+lswts/9TUZGYy691bmM0NK28ZTdnmtdiU3r0wIiNHseJrbBaapjk302D6hKCrHW+sDL4wnEYJg1+NL2vOayoJgvUXiH4poEo3+YibfdYxB0dxvE9OO1hAXf+bwUA498dgH650J2p7A7UA+owN79Nys9ex6IvVML1amsn2QZCgD0tiUPyHd+OY222Nh6pse81fs/kZ9OXse0baHYHjetF9WdqfbnO6upJ89P4+mLbQvDOuy8OLixfjmNGGgk7IgWCvvN4qICu8R7cxfDPwZd9J+DcgQTlxDHvNzuv1OB+6JcEdFVxtz3DOYP17P9JjyMZB8NG1pHnibeSsEwmOa82yPyNS3C6pBY0/mPJIyQJylvKKd3pjjiz4G/Z85eCj0rXOKYzB6IffavAiEx8PyEDX9UdHAYjonwzm4D+QzfLUa+P2TGZVlZo7L92aAgXa046Sz4HQmzz/m+LykWH6rZzpD9c6z7fpo6n/oA4x7Ex7cUBVHcyrzvWsnN+3fHdWgMZ09+EG2S19pOd4Kt4/OqW9bFRaeEBxMtv/zRFE0dmH61T9le2FMj5EFxdK7FzP21XPNI8UxGsaJNZXNCB5zLcLimjbFq7GRaeR9vf3D3mohe9usrzFnAs5lneCF6cAPdVVt7uvKjLHG7d8yv7Ddh7XNnuTwTJf8SJ2PxO9ZFhszkL8VZvBYqfhx6/hcH7yzkv4estn3cV7s1y1KznNTuYOB2R7+HO+H6l1W8++iQl10VfceTVmmglDlNfszzApmbNhXm2K9SP07EGIP+t93pi3MqFbBnrerAulMjk8w1GNtKkuyzDFvaOxYEDzG+sBPsls3MK/RON56VPluagYQut12jZghG/U5rROs5f5hoPee8xphWlfHeAjSXJfe0O/sKiR600Pd0PZlWXj5EhITCP/O9cPk5vTw0rHPT3X6JUnP5wQxx7tdpzby5NjnYSDuwg/tu1/L7uFsfX8nuwa3/Xes78e5hL7XnPuo9LHbj2k0y8gGeO+l7CbEaVs45sUyLpANd97OHPbT3/T8XZYH/l2KHxdlaMvrVIgjQfBQa4VuTKNy27IDOcYasVXaznayD/cZqiF/J/O4gjfU4GfuD4nTuvl5pMHWXG8/ysTNrKc3A7xnZDqmsfjxMAStV4nwsBfdH98dfv5fMu/15lzOTOo7CK6qZ8XAzkU4zMNA3DydfOwHvy7NYO5ypHpGIByWvEW9mft5bE77ekrnXZsljPkQsmfz/VKqVxD99DcjfU4ku1A4djT2nXPeUNsfJZW0751xYyrGdYcT7Ercrrd516ExvDfl+pcMs75bPJM6+yL9zzjbCFeNxmosT2S+gXtIIWVu2sKVqTt/i9t1K7Ve6YWMtcz7gk5mOu4uPDB4/u/56XYm9ezWnIsj06/5aOrAWOZ+zpaZDM7XZtyg9eztCOcZPZd8Mm0Za8ZOm7YRbYLBhKKjj+2p+0K/WyemDb3urB4/X6V8uQA07z+Fvq51bupqYuqNBtA6YfFmxO9+IqwVXOlYENxnwdkAOJHuV9dfHHd8k54OjpU5IPpswOfUQGfmBNiHjSk7niI9nrkGdfcS7oPjbOd0qAtddZ2Z9mE14+Prfmaf65t0ht85l10wrB1+e2FoyPaODv623Of0oKT1iOeZ4sXFmKo2OYsOY7YrmfcF5SmPDefW14gHyBMO0Rmd381nMsuSfmLb4zMxbet/ZJvjDD15gH5ihUNBcCT93FrxaDprutNfAjh4nqTf2U6ZacDf9zRYOp9Z3U3NYHPTcZ/GQgg8toiOarDywkDd1RNydQD/VeZ7K9s9ddv5MfA88zIYY9bSqRDWyEz7L2OeZ/TiogYgt0IAMiXXHceytxQhbeGExhV95gllrkz7S8BGf7kLO7HzUnZ3lA3Rb6aeVjgUBMcd33Njdmgsw9wmug7soNUDIZJ+rorEM6u/96bs7lrU0RsziCQEHt+cB+9zqW+5aY8+OtyGK5nv7Mw5fS77wG+p/L4sVp8D1RXFO+u6VjzPDB2AfBBmB0+FBvZdl/U4Z19P0uOMv3ufeUIZnWzxzYzBuTjGWLCP8/itqbt9z25neYgKfQfB9oFbQ19JfXRUUYdi173tejVvjp2WF9nNDPlo6sahMtyYk+N72S1VAneddAbt86DngrfibkmMOYbBLurYnAdfh2QUwcF6uTLn35ue2gQ6+Fx0sOcZ7f8OPRNfZ5h+p//ovWvp50Fa7GdMcUxs84ShXQjP5ejimSI42HfWc3mfyz/RTyzRZxD8xTQGoXZKx7h6k5ryb7vmWzzjupwXBgMaNL7Zey1MY5AKT4J2Vb8xz0G7nUVFGAxAz7+J9BMIszwEQXDxPLOUcdZM1ofJMRvOT33MBrZ0VvCKIp3c+QXbPu+7Efrdeg7+ITxYk3N3/2WjbW8fgTBBcIn9IDiSdmsqvR+5EcgC78zG0u7WjnOqNDrUO2Dods2VKzqqgHcD9kTaLe9UtKIoUahTep4ZIwy+MGORiGL3Sl+zga2EImUsM1GZjDcJQx+smVLk6Flu+nhdloxgeYgS+0Fw04G6Ni5vZ3Lw5yN3ZrXSfmnxtzHVGi3r9xPFgIE76C7XDNaOKrewAX6xyzu9bXkOWlGE2KtP8Uj9mTPh1mif9Dkb2DqljcHE+91j5QI64SIT7pRA/zLp9kBEguAj/uhQUHY94Dk9/Ghs19J8nZ+Yao2WeEoyxqhjj9RxAAcGrNqnbPrgL2Z7YJ8Ng8eYCXdiBqmEwe71PRu42G8g3MKUz63vR/qscyEMxnBSafdARPqIR7SdEfxkfjenCEep9E2ugMQUGTrUtTFm0WQU9awlDj+bNf8Av2no0nStVzr52PcyYr0gDHZviNnAxf3L0lKY+vjuy0ifdSaEwRj+3P631M/GmDBwRDEIXkq9K6k2BGZB9nEb8LhmhWedYHSxkuFn0dB2zJt2EF3OCk7YBYDXctMnvan5+3TwcexcczPSZxEGuzXUbODi+xNsYerHyFhLABIGY2j6wNaowXiSfuIBxSA4rvH7hMDu2HV+6oR0McWFEepZl8/AvLlcooE1/4BpSKTerA9me6CsDo0VfhAGuzHkbODivmVWMKZuzL6vhsEpRY4B2WWg6sx2p494QJMgmBDYvbohXUxRtRZRBP+tZ1oODwO890ZYGgLbK7kbh59PhwCYTlsRS3WYxzGNY8YM8DQwTIWZcGPv35MRPucTYwQEML77MuLnXQhhMMY5B1Qto8qEgQPqBsGEwH414jEDot4sTQOiJ6r89fWTIvkvuwbPu9fXnfQX2mUULTyoCxcM1IHJ9Xse6feg5bnmYcTP05lw9xT7KMaYDVyUUOSYuETGnYhxJdyFh+GlUj1Zkn7iHhsEl60PTAjs56DofUUnlJDj35bmZHRrBga/Xl8/Xl+fzYnqVNyuXerrAGpl6pOGwjdmQKXl9Nzi/RgcoVi3XIrZBcBk2FsA7478nNkeKDP2bf36vI6EYh/crYwzG9jSsUJEsWPi59JbB8cpS+ZgaGspf9gwfcQ9f1QUzEYIgX2VSnkYHM+4bBbm+9uZvpnsQt+vr68PwkP12sjMwObSlK92hpsG5wTBKJ6wXaJTCkzPSo6HwXTycUxeUm+G8onzzKC0D3rlaPwFTJkGs2POCj7huMGI5/pYDofBTBjYY4Pg+MDPCIH9p43qsbV+4pl0AvV7Xstulq/W139eX99lN9OX0Hc4TQY5d7QnKHAdBMfsAmCSVnI41KODj6o+89j04jd36A0jcfS55/QfMHEvMv7EnDPhLgmMV7+1jX6in1juz8JJ7VBHe00Ree/adDL3r4qH0knR76ZhY2ReS/PfCHf90OSWvFuKC3snatdtC4BpWpl/r/bOR5fCnSc4LDMDw7MRP/PU9NMTir9XkbiZDWwlQhiMaUscHEOfzPmZfAljjDFjc94vnvMJggv+OHIi+0hHelKu5d9XPaa4TrDtLN/LbjkHO7v3qzmB6EOeCIH90GQ28CMnfnjmjCIAJm0l/54ZTCcfZVIHn6l914ii71Xi+PPPaWswcbkcX0t1SEwKwlgOzQxmeYiCQ0HwHQfpZCv6/no/8cS+x2ch7J2SJgObhOICAPRsJb+HwazJijKpo89lXNVv3/PKg+1gn4L2sLlz2d3RAwztUBhMP9HQILiYimshXVMsk63o+1c4YooFA6pbvx5kO8MbAIC+6aDy79fXX3TwUaOv/ODgcy/ok/cm8WQ7ToVAC9N2P/NjGPM571+afuJ/qH87GgTb22M3ppB4mNN0ZbJd1sOaWqfzjTlA9UD9IuM+0bRYhqinbv3i4hIAoOxcoq8uIa4OaHOKEh738xh8dheJH7OB2acIgZ4zXSwPwUUUNLEo9BMXHeq69hPJOQs0CNbQ7c4ckHSip09vVfpiXlNsZO2TTO1D8G7Ypd52xuussXpDuwJPPVIEgHMa/n43L261xhhczYI7F2asd5WU/MzFTO9TIQwG7WHfxzKwX1dsP5FzaI/+MA3ASng4XEiuzWsdyMH/nl3qnToLrT9xokeJyPHnc1UY8KO/Aowpf309U98n2Wc4Nht4Y8ayj4726YLdg4nKHH3uqfDQLtSzogiG8QdFgAlIZTvDGdNqlGm4UTWoc2nNLgCctwFXFAMcyBx9rtZ3QsN2kpKf6d0EL+Jm8oE+hZ6AH7SFw4wlMW8r08ZiAATBmApuGfWH3pZRtSyELglB0IaqejTXzi8A7hiBOy77JyuKv7FIyi8apYXzuqtZwRG7CRPlaqm0C44blFjQTxwWQTCmIhc363/hcIe3qkNBw40qscPP3ghBMOD6+Gc2MFxx2f6vKP7G0pKf3cnvz6JIHWzfCf1e0B62wvIQOEbzhlOKYTgEweBEhSaiisH7hpM6aoodfjZr4gPu6CyPlGKAQy5nBJ8Js+Ca9hXOS36e7P1/bVtcrAF9xX4F7WFjK4ofR9r9TxTDsAiCwYkKfZ2wN6bh5iFcqKIXC1yu+ZSwCwBnUmGWB9x7cnwORPfz9f5sYNfneJaxwxTlDj9bL4yxbjqKtD4wYWcEBMHgRIW6Iim/Oqe3cBDWw/dB8B1tCeBMKtt1AYE59yljir92OTWZDVxsZ1zMCr5g32KCXI/duDAGS0PgTHhA3CgIgkGnHXUlJT/7KNzqi3oicbs2aMIuAJxIhXWB4Q+X4QcXQ7qfrx8rxgWJh9sM+OrZ4WfHFD9kFwKfURTjIAgGUMeyZACvMyy5HQ5TGCTdCBeUABed+3shBIZfXJ8LluyCUrG0mw1saZuzcbDd50KwBdrDpsc65m1p6iAh8IgIgjE1zxSBE8eCXg2BVxQPGnT2XIVBT8JMHcBF515nXjIDEr7JPTgf4riy87XOBs4q/v5F3E1SYHIEpsblHRL6zADWCZ4vXVryh7AcxOgIgkHHHXUa6EOzMgiB0cTC4eBow6AbcHK8/xAeDAc/uX6wLTOCj4ul22xg61bczAo+o38M2kPaQ5SKZHtB7zNF4QZBMIAqhxbxJwRGU7fi5pYfGwK/sAuAUei5IX99faAo4DHXD0iK2AVHJSU/0zsDs5rv43JWcMJuxIS47iPH7ILZ0IkC6evrp5Rf8MPACIIBVNEg+Knw/wmB0ZSe8F0sCWFD4DW7ABi8Y693j+Svr6/CLX5AFQbAh8XSz2zgYv/DhVPTJgJT4LqfzNIQ4YtMe/yP8MwILxAEA6jyYjrmD0IIjOZSIQQGQnVpjvFctrf3sQwEpoTnTvgnqdhfacP3y03f1dV3IeACqrE0RJi0/VvJ9i6On0IA7JU/KQIANbzI4SUigLKTvw7YXDwkSmewx8JyEEDfInNsxeacwMxfTFkubi9exFJ/mYM50PLoczZw8e9cBBDaPl4Ly0RgGuM8oA/LQj+RBwV7jCAYADDEYC51NMD+ItyOCfQhkl3wa/9lxi+AoSQlP2szG9jKZTsr2EUYrP0RXaeYoA0+c333HEvlTNPS9A9t+Kv/MkFgIgiCAQB9icyAx8UVYB0kroTZVZinlemAVw3m9PhYyOHbMJfmZ/o6o0gBjNyGlYVBacf3T8TdrOBbYVk1AG5p/y83r2NeTD/S9geLin1HAt8AEAQDALqKzSDH1VrAt8Ktl5i3U/OqugjziaICvDp3ZhTDfyU1zvNd5K+vR3Ez8/DKfL+c3QzAkXNh5jUKeFgcAKCNpRmY6cDmu7gJgfVWz0gIgQEAzfEwUT+spHzZmb6WVnDZV6CfAgDwBjOCASBs12awu+4wkIpkt0aoXQfK1S1BOjMolV0IDQBAG6zb6oek4px/29PnZOJ2VvCtcPEBAOABgmAACNvnvf//uDco2heZl/JprdAnM4i6Z/AOAEAQVlI+G7jvc34i27uYXNA+TMwuBwC4RhAMAPNyfuR/++jZDAJTYRYNAAChSTr+vKnM9C1OHfW/YmFdaACAYwTBAACfPJlBUiqEv0Bdj1LvidDHxAf+Nw8VATCklZQHsncyzBJQyevrq6PvnAizggGM78b8m7X4W71DdLn3v326axQtEAQDAFzS4HdtOiaZsO4v0IYeO0nHvy/r/Nu1wfXfU4ob6MXclzlKOv68rdS8t6tZwSuzDQDgS3tb5f7If48O9BNPKG7/EQQDAMaioW8uu+C3ywPsAAzvRXYXaW4LnX7t7F++vi4oIqC1Od/1shI3s4GtRNzOCk6p/gACkJtXMShemj6ivpg17CmCYABAX54LA7es8O+LsMwDEFKnPzWvhenoX9PZxwQtKAJnkoqfpwN/firuZgXrZ66EMBhAmNbmpW1sZNq7lXBHmVcIggEgbG/MSTg6MABetni/XH6fpUPIC8yXHv+pecWmo39FsWAilhSBE1WBgF3zfGjabn1yVAZ6h8W9cFcUYI95hEnHjIl5rcTdBTjsIQgGgHmchPMD//2eogHQk0x2S0joi4fNAdXHzBwlHX/eF22n9G4GF+tZnpjPTjgMAMxEal4rIRB2jiAYAAAAfdE7BGLZLhmhHX4eGgLAqhr8b0z7EY+0Pbm4W9ZGg2ANo5kVDPoNmBPtG96b88EHisMNgmAAAAD0TTv5kWxnPbJ+MPC7Od4KrUtSXVf8jl44+jST8mBWMHwRO/58LobMz4tp/+7Ni0kDI/uDIgAAAMBAHX1dh/WOooCHXK4RnM+wvF0tw+AzDb0jigEzl1EEs973ei5+oijGRRAMAACAIa2EMBj+cRlK5jMr6zqzgecqoQgwczlFMPv9Hwth8KgIggEAADC0lRAGA1Y2s+/LbODjroRZwXArdvjZui54zi6YvRchDB4VQTAAAADGsJJ5ro0K/8SOP39OD0diNnC1lCLATPGgOFgaBuuDhjcUxfB4WBwAAADGop38XJgdiPl6lnk9HKlqNrCWx8qT7bxw9Nnnsr04kXF4wAGX66VT51GUm/PBN4piWATBAAAAGMsLnXx4IHb42dmMyrnObODEkzLJxV0QbMsh5tCEo+OU9hC+uH99PThuj4PH0hAAAAAYu5PPEhFwaeG4/s9F1WzgjUflkYvbdcx1VvAlhyYcYEYwfDx3sETEgAiCAQAAMLaEIoBDBB/DqzMb+Fb8WibDdbt0y6EJB8epq6WaHih+HJHTHg6LIBgAAABjy4RZwXDHVRCsT0Sfy/rAdWYD+zbQz8XtrOBT8WO9ZNAWjtUPAI5JKYLhEAQDAACATj7mIhJ3M+DmUuenOBvYSmb++ZiX2OFn31P8KJELs8YHQxAMAAAAV4PAR/NaD/QZS/M5K9kGgMDScZ2fg0Sqw/bU023Pxf2s4ITDFCNxdV58NMcaUHWesP3EoS4cXpvP0TXaF3MpWIJgAAAAuKCd+ti8rgd4f+3Qa/CmT57+KtxyjS1XQbDObMpnUL7R6+tDxe/ceV4WiePPv5YZBRJwKnb0uSlFjxruC/3E9UD1//Pr6+r19U1mNGGAIBgAAAChDiBO9/4/EDusj3OQ9PQ7LuXidlawzqa+5lDFwKK9c+RYNpyP4Un9L9bDZxnu7jTvEAQDAAAgNOnr63yuHXyUOnfwmRuZxww4HVhfVfyO77OBi22IS9fCcjYYVuzoczV8e6H4R8UdBv8uD62HJ3v1cjYIggEAABCSlfw7jMooFsh2DUAXbmdSvklAZaFtxqPDzz8R1grGsGKP2wn0a0kR/CZ9fZ3NuZ9IEAwAAICQBrZfD/x3bkOFrR8u3M6kbKtmAw/5YMghJI4/X8sz4rBFQO0hD4mDa6lsnx1RNLvlSgiCAQAAEIJlSUeeIBjKxYxgXQphDrdBJzV+Z2qBeCZuZwVPscwwnfOli/WBE4oeDq3k8AXLbG4F8Sd1AQAAABO3MB35kwM/e6B4IAQfQ4qleu1lXad7ihdkdP99d/j5F6Z8Mw5h9Gjl4DMfqcdwXOe/HvnZVM5NsXktZLfch95l82L+zaTmhWeCYAAAAExZWQg8pQ6+76IABoFju5F53AZ929Pv+EjbFg2wzh1uQyLuljWZCsqnmUtH9RhwQUPTrxXtvM9tm/Zf9mcyP5l/P+z99wdzvi39TiwNAQAAuuBJxHBd/7SzezbRDv6UnE58+8cOPjYyj9v6VxXHn5VO+Dsmjj//XNw96BDhiR2053eci+HIsqLuaaCae7jdkdluvSPlyhxDf7++/np9vTHfa2n+93/Mz+7M8f3d/O3RhwQSBDPQBTBugw6E5owi8GJQN9e+WVZRB33t4GNcGqKNHXxcS/hrA+sxWCfsnvo6ydrOPDveBtYKRl9WI3+eXhRLKPb/OacIRmND4JOS3/HxrjHts6xNXfki26B3Zbb1UJ/2xfxMfyeS7d1I+rc/jh3vBMHNKhEAdBHN5DMBYGh1QmBfO/gY3/XIn6dLCaQzKdeTGr8XQlkkjj//1EE9RphjkSsHx05O0Tu1nOl3zmqco3zrJ65eX99kG+6+leYXlV/MMad/q5MhdEmMf11IJAgGgPA7fKCjh/HEFMHg6obAPnbw+y4H1DsPjjkDayPzCOy0XD/V+L1nCeOW8FTczwpOOO4513e0Gvnz9KIYs9npL7gYD+l5pyoE1jZ97dnxqcHtk/kOXbZtbdopfa8P+8c+QTCm2OkEANDRwzz3e5MQ2LcO/hADHVRLHHzeegbletvz701B6vjzNdRgVjC6nD/HrD96UWxFscNB3yiTenerZJ5tt54vNbiNpZ/llF5kFwZ/lcIFrD/p+GJiTikCAAB+M5d1miPZzvCt+30zqsbsaZ0Z8zZo+7Tu0Olg8qLm74Y0K1/3bd3lMIais7BT4VZ7NDd23b2mnnpjLvnZyrTTdeu5T+en1Gy3foeXFv3Ztfnu+8ecvpddczg1/aLJzghezOQz4WbAANB2Ac1DAWDoQYx2YpuE3qwPTD8pHfGznmU+s9/qlmtoD2t88aRd4VZ7tGm7x5wNfCfzWCd9KuOyOYwH9fyrs16bXOy492jbtX97I4fvKDqv8dLlH34e6YfoeTiR7aTK//6cpSGaDUDAAAeYWjtC20UbhnnsgzjgMtWZDJk0n8lEEDzvNuZSxlsbeGM+72UGdcAOJud6DPoQwl4IF2D3nVMElfV2rNnAegFoRZF7NS4LfXyQyjYEbuLBs/Pqpsb5RdfcvjnwKn6Xr0fOD/rez+azJrs0BDPcAACYX0dvCk7Z9713jj+1+LsHjvdZW8i4gZ297XIOda7J8RhiELw2g3HXwWNKG+DUlCZaxFJ/KZeu7PqmoG861rn+vmV77Mv5aWn2zxepvpicyfHnHqxkF4Zfy+HlJFJzDl9OdUawi7XwONFywgX6OFlx3CBEdPrdC62fYjv3n1r+/T37fNaSEQe+72U+61GnDX5XZzetKYfBnAoPjnPpZCLbuRixvj5Lfw+5YixIX73OGFfPMW0vyvly3r7s6byif/9UUc9s3/iSpSHobNNw/rtBAYZyFuhxA786RZhn+ceBlaN27rvMYMqojrNtY3RQ9WGkz3ov81kH87rhgDvkY1AH0xsPtiOhn0ffp4K2T2NcFJvT8jhTrashHSd6PvrRoW77tH697b+PceF0bY7VeIpBsKuBTkS75dwYHZ0TihkTrr90juGqnsUUvdPyV6Gsj5h07Nz71sEPbZ/7HjrpOS8d6bPmFAJHcvx21LIBZ6h8eWjcifBALvrYx2lYNsaSEHYm8Jqq6PU5NA6k/LLX1+eO7+PTXWN6fn3s4X1Wspt0llecm6M/OXhqC3VdlSlZjvg5nMgQSmfxxLSbXKGfB1dhYETRe9HJvpTpLolgA7w+7pzIqIqzPNbtciJjXNSfUwgs5rs2LdfQ+9Ja16482A774Lg5t3uu+tg+XxhbSfewrA67JjDjDL/rqm0rpjwmvGx5LjrWfvtCc8a8wTgjOfLf7RhwI9UXbk//5OBpPMBjcBH+4INbrBBa27Wk7aKNHKETwwUH9yHZSqYXBGu90VlLn3p8z3QGdS2eaR0vq0eZDD9xYyPzm/mWSLuLjKGfD+zyED7cTZiavt5cz8Gu2iVfswGtC19H+BxC4GmNB9XlBPtIkdnmvia7PE/4HH5eUQ7PZh/nVW/E0hDTOnBpOMOvY6D+Uqcx1f1MPXPfT7iQac3OXpnOeJ8hcMgPqPKhrvm4BIkNgYdeh3+Otz8vez4+Q+PLhTe9AJJw7h2dj+fbSxknnL4TQuA29cX1haMpPWByYdq1nz33PbIJ1yHthzwWXsW16m9MHavVR5liEEyYMk8LGe9BWwT+CK39oO2ijlHPwjpPlbmdSF3VjrjOWOp7Buf9TOobD5zZbUs2wrH3KPNbOmwxo+MphEDhw4zPw64uUPlW3qvX1zcZPmz8aD6LEHh69UXPlZcTKCsNrHMZ5kKkb+e1R6l/R3pq6pF9rQo/u6z5Pvo7j39M8OBxeRXlQlg2YA4NZ0xxo2eRuF1n/Jy2i3ZyBJeUvxcuPN4XWkbZ6+v7gAP3e+rbbI51O/Nt6BD4RuY58y0VnpFSJfNwn82tv+eyLTwVf2YF60XgoZeD0NmH72QaF5ypq9NrJ+xSYbls17ceKvfzrZ/40qEfo9/li/nfZzWOTTtpJZ9aEHzJNszWmOWujQ6zgkHbhal1Ll0P2E9n3nb6dIylHu0L7XSuTMd+yADYymZQ15aOj/eV40Hkwgx2hp75prdgauiRzLA90/K96OG8FDpt1zYebc+pzC+kc33udd0+2DsVPgz8OQ+yDb0zwdT7iSdmP/oSBkem3dL29PPA/ZsHD+tF1rF+aHj+aP73lZQv/2HPy9mUgmA7kHBtJXCx7y/Zz5iwa7YBM2mz5lrPFuLH0+P3O/kuBx32gSTasR9iCYhjHfwXjvfBnTo81lemTg0deugMm7k+aHXVU/nO5cKgb8uFXMm8Lv5felDeLuq6vSD2Q4a9K0IvdPxtypmlILq1qycebc+ZOb9FDvvNK7MNP805Z4zy8fGusfse2rJL2V2U/FzSJtn+4/2UguBrTw4enckS05aNfoIfe9+vhFvp0Y9Y/Li18oy2K1iR+BNCXsm0HlbWZx/FN3re1BmT6Uj7xF60Tc1g8ZupD2Oev7MZ1DVfJkZ8Gnk79LNy2V5UGLJOPcl2FvC1zDP0iKW/28un9vDKkKQyjyB+5UkfOxuxvO0DtLQ9HPqC2J05hlkrvJ+66hsdG65NfRoj94hMOWh9+seca84dHKu+0WP5oeMY6kV+D5KzA/s0MudlPa5f3vz69Wsqnd5c/LmK8iQsHTCHfX8j834CL/pr3H1ZY+/ZtF1c0Q9LJu4elHLIo8zrokNkOtInnm/ng+l8Z6Zd6mppvnts/rcPdfCvnr6bz7Rf8smj7flitulloGNrJeOEPc/me6QzPpfYGdB9tmUPEv7sVN/OwcXxahx4n8+nPrbOxrsesA25NK8xLvw/mu8yp4djDknPYV8938aN6SPafmIf7YbtHy7Fj4lRPmd4Wj7fS86ZSeF8k1W0E/Y7ruX3izj2XPXfvvJUgmAfT7B3wvIBY0jF3Uy3jTmQcnYDWtJbtj54tk20XWHRjvpnD7drThfSfA0Bqs5va/N6KXQY9zv+i0KH0v5v+6AJ38zhIr1+vx8BDyCjwoAxHqmePZu+5lzaq7K6pftuiAta7yXsgP1F/L0QGHIYrMfsJw+363mvPexyTNr2cKy7Yx9NuWaCvkQyjckCh9qOl726kB35ftHe/9aXjw8a9X1sYnODIc6Zdrz4xfxvmUIQrIVw5em2EagMS8vW9dWzOVxNR7j1d6onQky/js1h8O97H2Vu/te5DZRvd8eVsRcapEagEJt/lyN/N2YAj3cu2Zj9HOLsQv1e3z3fxhDHq773f/bHki/mta5o411dbNVZiLdCADzEeTsTPy+ez9Fbz89DxfrS57ba9vK3XMvnIDgynTPfZ9k8mMIlKOyXT7PcCIPRlI8zgQ8NDK6p19Qx6lnrzpr2US6oinTwR6DhgM4uO2U3d2JnLqdC4GElMt6syo/m3BXSecAO2n0X0lhGx91fOXQ7s3dD6CunODhvz6C+RxOpN1mhreu6PrdtLzfm+//vHODjw+LsAujamZ/CrZYXpvFMhIeL9SEyFd6nW53PzD6+ZPegwqWM8/CGPlyZdnbFbpvcAGgqdczWs1zCegCn3QeEwP4ozkANie0T/2Aw2YlO2ngvuzWHM4rkf7crj3lr/WdT9nEg5ZfJdGb62YdCLSde5jpGJARuT8MwvXvmrSnPRAiBh6inKedt70zlvL8unCP1ocu3LcdPC/O3/5oJbPkyIziS3fo3Ux5Y2ZkGmfmXmXbN6kAi462B1Naj7K6eAlJouy4nfMLfyG6GFE8G9stir46dTPi7DPEgCvYBVGi3Pi9l95A06ltzT7J7oArntH+3Z3qXhuu1VftaR9XVuWDKSwLdmYBgKhfPLgMoc9djV9sW8vC3YUQSRpYVsr8n1h9YmO09N+fLWzNWf6nxd9fmpf3Ho6sX+BIEJ+LnYu9dvBNmHNTl68OOymyEGeDY+hXY93nDLvWKdiy/B/i9pnSODHUf0MH3k88PnvK1P7Y27Yn9l4kY0+uz6ED3UvwMqoZ8kJ7rMo883j5fH47p+z7dbw8xvETCy7JCM9Xx7crULzvZ7LFwfNu+jl1fXMcr54W2YFXWBvw50QJ5dPCZ5xw/g1k4rgdtFuVnkIaptF1L6uusPImbEOSccq9V7lUPixEJ47bpsYU02K3bXmsnP++5L+brbe6bwqCn+O9aCH2H1KWORdLsLqlT8XeCxaLnfpQvx67vd7E1qQ9996997jvb+mNfth0M6Tw45zFgXtE+RDKNtW59MuWZ8Kl52TsjdIzwqaRtuJOaz0HwaWmIqMEgydeT1XKvwtE5rSc2ZbduMWAew/JAZ0S3OWHXQX4PbaZ43O+3XXQk2T+cI49ve1UHHejznJJ5evwe294+BmsvnI8GlRwo17Hb42P96tTT9lXHqKsa7b+v9bU4zt4/pn0eyxTbHZ/ygUPlWdVOtpF5OCZGdZ0gA4KLPlnjeudLEAwAAAAAAAAAGMgfFAEAAAAAAAAAhI0gGAAAAAAAAAACRxAMAAAAAAAAAIEjCAYAAAAAAACAwBEEAwAAAAAAAEDgCIIBAAAAAAAAIHAEwQAAAAAAAAAQOIJgAAAAAAAAAAgcQTAAAAAAAAAABI4gGAAAAAAAAAACRxAMAAAAAAAAAIEjCAYAAAAAAACAwBEEAwAAAAAAAEDgCIIBAAAAAAAAIHAEwQAAAAAAAAAQOIJgAAAAAAAAAAgcQTAAAAAAAAAABI4gGAAAAAAAAAACRxAMAAAAAAAAAIEjCAYAAAAAAACAwBEEAwAAAAAAAEDgCIIBAAAAAAAAIHAEwQAAAAAAAAAQOIJgAAAAAAAAAAgcQTAAAAAAAAAABI4gGAAAAAAAAAACRxAMAAAAAAAAAIEjCAYAAAAAAACAwBEEAwAAAAAAAEDgCIIBAAAAAAAAIHAEwQAAAAAAAAAQOIJgAAAAAAAAAAgcQTAAAAAAAAAABI4gGAAAAAAAAAACRxAMAAAAAAAAAIEjCAYAAAAAAACAwBEEAwAAAAAAAEDgCIIBAAAAAAAAIHAEwQAAAAAAAAAQOIJgAAAAAAAAAAgcQTAAAAAAAAAABI4gGAAAAAAAAAACRxAMAAAAAAAAAIEjCAYAAAAAAACAwBEEAwAAAAAAAEDgCIIBAAAAAAAAIHAEwQAAAAAAAAAQOIJgAAAAAAAAAAgcQTAAAAAAAAAABI4gGAAAAAAAAAACRxAMAAAAAAAAAIH7kyIAAAA9S0p+lr6+cooIAAAAAMZFEAwAAPp0/fr6dORnd0IIDAAAAABOvPn16xelAAAA+hC9vtavr5MDP3t4fV1SRAAAAADgBmsEAwCAvqRyOAR+en2tKB4AAAAAcIcgGAAA9EFn+54f+O8aAsevrxeKCAAAAADcYWkIAADQ1UK2a//uzwbeyDYEXlNEAAAAAOAWM4IBAEBXiRACAwAAAIDX/qQIAABAB9Hra/n6etz777dCCAwAAAAA3mBpCAAAAAAAAAAIHEtDAAAAAAAAAEDgWBoCQFexeUnh36IX2d4enpt/uVUcAKZnaV6R+XdBWw8AAABMC0tDAGhKQ4BL8zpv8ff6AKns9XX/+ko9/Y6ZR9tyLeMHKrq263Lvv72Yfe5yH7goi6pySUeuxz6Wy8q8itZmu1xZmv0lDvfXsW2IJ9TeXxZeJw3/9tnU19SzNnVqDrU7Q7Yva/Pvy4y++9p836zwv6dO25nkSJnee7rNx9rMMeUHzmdjnjtd15kQ+9q2PVtzLgKALYJgAE06iNphverxPTem03/r2cDLp4bxnYOOq37eoZD/bsRByy9PyqKqXG6ODLbnVC76/T/t/bdHcRt46md/P/KztzJOcH5sG95MoL1fmf162tP7PZv3SzmV9tYeD+3BnJuzGX73J1NXNTDNA6s3+t2Wnm5zWbsdYr/r0LnTaS4wk772g/g9GQUABscawQCqRKazpJ3zq57f+8R0gnMZN0xDO7r/VxQDJk4HgAuK4aBL0x5/lf5CYDHv9dW8d0wxT8KFOe/P8Xg5e319fn39NP2fqX1/PcbOS74bx2A5+qPht216PtIJKNcUB4A5IggGUGYl25lzVwN/jg2E9bOWFLvXbtlHmDgNJe8pht8sTJl8k34D4ENlr+FiKoTxU6GhSTbjdl/7P7lMKzxdVfw8oVqXOhfC8jnQscdn075xPgIwKzwsDsAxqTQLgJ9k95CgIu1MR1IvXDgzHbJr4ZYtnzvO9yYUeKE4MOGB/q0wG0jMsXwv9QPgJ9mtp7rf1i9MO17lynzuSnio3BSczbzd1/OeXsB4P4G+SVSj72aDzoyqfdSK8plVf2BNvxbAnBAEA9i3MJ3fqsH8RnZr6NXpLEeyW2f4vGLA9dVsx61H5aLr4+YOPjf3sI6cmn1/yeGCCfsguwdXztXSlEHVg+DuCm191UB5IbsHzF2U/J698BcLYXBTj9J/SGXP0aeet/tDfHd7LMQVx8Ktqas+19ek5u+txL+gU/s7NyN9ltb3Y4G5XuxydZHwWeY9EWKIvvbCHN/nFf3aTLjjDcBMEAQD2O8sZVIeArd96E9u/iY1HfBEymetfJbdjDEfpMLskKILsw8TigITpsd1LPMMIuuEwBrKNH2Y50uDtv5ECIPbyAZse2Pz3udH2v1Y3D9Absjzjgbd10e+/4k5HmJP60Uk9e/kujLlmHu0/flIfYr/x94ZXreNIwF49r39b3VgbQXWVSBeBfZVYLqCKBWYriBKBaErOLuC0BWsXcHKHVgV7O2cwRc/ZYayYooEiO97Tz9COSIIDgYzg8Fg1iHD2/B+XybeBzHPyc0R3/tFGMPWvHeGXQsAuUCNYADYNcDOOozjG/lxeNxHDd1SXk9mftrjkM15LdFyLWQFQ9qcSJ71aufSHQTWrMs/gkP8kYBIq+v/6ND1bTCYGo1xoO+ikNfMPIty4s9/F57/s/N9zPVjK+f6fw78+xxk/MyxcwuhPMBUaRcpdf67d/6GclEAkAUEggHgrUPgbeN9kh9ZQn0b45qVdtNhkG94NVHw7FxvjWqAFNga184kv624GuzygsA3R9C9m6Drvzrft7XHIR5KeV0Q2CWXxT/NGvzc0TexoQspVjbwQxhbVmD/MsP5uxY/4UH1HjsTps9L0GMPzlxEggMATB4CwQAgwUG/dr57GsA4ruQ1O3i742hhkMflPN07RvOdkM0HabASOxh8LvlkAqm+9QIhV3LcLMFVuIfFUtiSGxtrR+cvMnp+K5O9iFS3eeO9ncMP+X9TtWMuO3QfNmdelM71gq4BgKlDIBgAPGdP5EcQeIhtck2411MwyMkOi9NotpziM4nrYD8Aj4342T5fMnAA5+Iv+qnerQdog97Dy7RcCTsMYqJxrue08GfNbaeRtXEmdkD37cF6jdgZkGUm71Of0wsCf5a8D2jL2R6wEhw4MA4AJg+BYABQ49g6FEWz5jRgMmSttMdggGGQx8lLkBcro/JSqK0GadCIfzK9LkDNJ/zsnm69HVjvrsXeqq7ZphUiGpXOt8oCFRn1wca5HlOwaCV2qZdqz7/bMTf1uVvtlm8duo+F7HwhCxwAsoRAMABUHYbzhu4Bw2gune++CJkUkI7e6yp1MkXmYi/6aaBvjEDQSuwgY451S2Mmdzugca7HkkXrZQM/G21vnDG3kulmBatN0hUELhniAACQGwSCAfKmEHuLY3u4CICFysZNh9NMvWBIAQ0AeKVO6gk+b9XRDy8jtOelo03sLoiHOc8fvR47OWC8W9enekDWQvxA/hN6BgAAcoVAMEDelAdeB3jrTHoZlQ3dAwmwr9TJ1PSgFeh5GHm81mJnKHJqezxYi8U5baeeO9djmee8bOD6wDFXTey9zcI7soLkQ55/AfHLCQBAdhAIBsibwgkMbOgaeAelcHgcpE1XqRPdTjyVUieq662ASB1B2yrj2qlQZiYGvIB8TgG0mOWwFDtQX+35f7Uz5sqJvLOuIPBWxtsFAWn4QfhAADB5CAQD5MvCcSBqugbeSVdG5SchsxzSoKvUiX43hYwhy9ndRqLvvTJEZAWPj7d1vsm8Dx4iaVv1i+N67czb1UTemcrnmdM3qgs5IAyUuSMnyAcATB4CwQD5sjjQKQcQx2AuO5xNsvogBSqxS52cTkQnFsa1JpK26YLSwwFzFAxDLfbhgvcZ9cFK7AXzGMZO6bRt/c4xVzv6rpiA3J519BlBPtjn7zR0DQBMnd/pAoBsmRvXnoTtch7rD/ZNMeG+aTMqr3eun4TvFsgVJEApdibZMoz/lA8WsoKqMQVEGvk56DhHJEeTlbXYQeBW3+dAZcxpLXUk7dtlK+8vy6R/98n53VTtFX0vl853V4nIbtcBd+/hUTgE7z3+Ty32gsGzsFgAABlAIBggb0Nol76CdWrI9rmduomgv84Qmb1Oqb73853rp8HgZps3xE5b6kT1zW5tyU/heqpBsJNI9WrLIzr3IArpfxv/PPzuacffPMl0y0fNwhx2ET5eP9zK+DVEvfd0yIL1JjzLbuB0GX6/Sez9leIHgb8mJLcn4i/CwMf8knkY25d7bFkAgMlDIBggX+ZHDAysezZkf+N1JeOIPRoO6nkwrjGwIXbaUif/Nb6rhfqSx4IdA4exlOGDRe0hW2NzLX6m7hB9EMM8VjltO/SQVv37S+f3i8Rsj2/OdxrsJkM2Hb6PeO8H4ZwUAMgEagQDAEBfaDDnQuxDaK6FrGBIA+/wuJPgJM7oIsgM1ekaTMt9EUT7YDNyGwqxFwHu5PAFFX2fVn1u/f15Iu9EMz29ILA+W8nwhXfqOGxUAMgGAsEA8BYCHPBRuurT1ULdT0iDSuwAyZmQMQR5oeUgiszlXoNEVxJvbeCu60P/3pB01dNVuSWwB4foOHamAEA2UBoCIF828nNWCae0+8RQFzAV6iBLuwfRtIfHYXBDCmgQwSt1oosda7oIJsxTkPE68354kHiyoQuxs4G3R3hPWjKiitju0cSFRuz659uE7YznD75L7NTD+BrkHJsUALKCQDBAvmwcwxpsaknv8JQxUcd5YTitmlGpwYWSLoLIaUud6LjfDTZ8kdfAUMo6YRFR+wvj2jMi6PIsxwn4PL6R603G/auBxLsI531v3jzWAWNVpHP1VIPArW1eoeKOrj/vgi26oTsAIEcIBAPki2Uk93VKe/GB//edVzMZvIzKy3CdjEqInbbUiVWDUh3JRSKOpGY1xrwDxGoLDrpPLfkGi27l49mvOqbPjev3oV9jrIU8F/tgt2PP4TOJL6h612GvFkIt65T53MP7U/mwFgm0vEvD3AIAQCAYIGceO4zohu6BHtiXUdngsEEC1NJd6iSFkjo6zpaGro8Fqy3MQ2Cx6UE2NkHmTgw53ET63NUI99T+WUlciw61+NnPV9gUk/BNPjq+vcXbUih1AwDwfzgsDiBvY8uCwzWgbznzDo9TY59yJJACKsMpHx5n6XvN1C8iaJvOOScHzFEAH2Uj9o4UlcMqwvbOZfhs4Le6L5Z5et3RD5+FIB+8onJglRZaSlwLoAAAo0FGMEC+aLamHgazu72uFD9wB/CrRrmXUdlIOocUWttjh3QqCJqPS1epk0biDkJo5rKXIdWM3DZrvmlrtAIci3WQvd1FiE8SX+3Qyrl+0/N9dD5bGvN0KeOXcioNG6LlVig1BT/Ly3dn3HMwNgBkD4FggLyp5XWLvmX013QP9Mi+w+NSWHzQIOD5iPf3nJcN4jUIXaVOvsmPg7Zibfu9Ib8axK5GlKFC7G3eBIFhiDHhbSFfSzy7o2ZiZ8HeSv/Zy/N/Pn858/eYgdbSeU9tP5SIM+yg8/SDY3Pi4wBA9lAaAiBvPGd7LWQfQv+oY21t1/uUiCNnZQQvB7w/geDx6Sp1che53vQCOWM6xFWEbYJ8qJ05SRdMikjauBpwjOhccmtcPx1xjl506K4nYQcbHD52KnwcAMgdAsEAeaNGv1X3MtY6eZA2bUalRQrb9cauq104jjAMS/3P56txXYMlMWeyNuLXTSxHctKthZQH4aA4GA5P9mOwgWZiB7OOOUaqiPpjIfYOjHbuK8ReoAVobTZvYYMFBADIGgLBAOAZ92NkaXJQXR6G+ZVxXR292DMqxwwEa79YZSkaRGoUvMPjlhL3Ipqn04deiNF7fTlwTgI4Bk3HWC5HbptVw/jYY2Tj9MfQWcGzYBNYz78N8y5BYHiPj7N1xhZZwQCQLQSCAcBzgpRvAwYHZo6T8cwrmhy1+FkadcTtbg9Y3EXrN84HCAh44xfG4cJxMK8l3kUtlZd74/qQBzcuOuT2FpmGEYhxC7mXDfw8wBipnOvlgM/eyM8Hc0rQuYVQEgneh8rJ2pnzOGAQALKFQDAAtMb91vnuz4GMfy/zg0ODpitzVlD1XOLOCKwPvN4HGji7dhxixsd46MJA0SEPi4jH3tZxjBs5bm3UQvyt3hrgYrsujEGMW8hLGT4buEXHqJclXQxwf53Xzjr65RGRhQNYO3PeEIv4AABRQiAYAJTNHmfnmxzvALl2+9+yw4CDaVKIn1EZK7XT5qUcJxjclT3J2BifrlInVaRt7qrVre3+fqS2V+G3T5zv2eoNY+LJ/FhbyL1s4Hrk/ji2Xqs77MErYfETfm3OG/LQRQCA6CEQDABvjaHbju+1ZrAGPcoe71mG3zx3vtf2bHg1kzbOiwTb7AVgNbukkf4yTC7Ez57cCoHg2HXnScRtVrm66vj+OujmPsZnEfR41wLPlZDlB+OiMnrjjOOhda3aRlZZhPXAOsLataNB2mPtdliHedTiRgjawcfmae+w1ILuAYDc+J0uAIAd50M6DHF1TDQ7uApGVS2HB2rn8hrgWjmOTkts24TH3Oa9kekGxNuMym8JtXnd4airU/GXvAYG1/Jrwa0yfJZ7/obsybh0p+qIs8QcY+kYe/osmsH7FGT57gCZm73R8/v65EoI8EA8ut06oO0y2D1DzcOVcW07wjhZO/phJf2XDNPf++R8pzqokeEDds0IMjiTcQOTjxO2Lcowp1lyvhAAgIwgEAwAlqH00mGQKxoAuw6f52A4PjoG5CwYWHN5f6AkxhOhv4x47xuJu27uR6mD43OZSHvbrfWN+Fmfl+Gzb3xIGBet87d8x/01yMz22PjQ97eRuDOBrbGndC3EnIXv9fMQ5HgjPy9ytHq+OEDPr4QgMMSl2ytnvm/nqSFsMC8b+GUE/VAZ7ek7ML54hw76PoI8/DbCPcd61pZ/y3QP7GzCHLY0+rxkLgKAnCAQDAAWq2AwqVG0L6hxGj7nPd37STgMJEdKSSujst023+wZI32Pj1sZ7uR2OIy21MmfibW7DvJ8J927NCQ40Mse7qkLJBfoeYiQNiv41JD9Vucfk6pjnI6Btuebc72vuWgjrwtDJ4gfDODf/OmM+0N2vQAAJA01ggHAQw0iDczdD3jPr8HRIjiQJ4XYB7HFShsMfh7ofjdCEDgFmbhKtN2LoIOH0PML9DxEjFeW6tg1enU+sRZjxjwvoXbmOM0K7usQvReh5j0MN9d5Nf1XdA8A5AKBYADoQh0PzdrSrWIPR7yPBpv/FYwwVuPzJcXD49oAmgZpjxXEfgjjo0JEkqCW7oM3Yx5/qyBrx1gA1D75Az0PCXDn2DztFvJjUR14fSi8+/cZOFtLWgvBkC6VI2sqz3O6BwBygEAwALyHRl4DdOrEazZXHxmQz+G39DfZIgwtKgefE2tzW1dyHtrex/hQJ0UDZ/8WsuRTpJTXMjepjsGLnvT9Wz2vfbJBNCARqo7rsyPcT/W8VXblPoJxcyd+4IysYEiNjSNrJ8KCOwBkwm9///03vQAAv8I8OC4L+XFIkFdf8ikY+U0wwJrIAwJFhEbrkP21MJy7odtgvYOUTrPeHR/an1794+2bZ3sM46NJTBfMDad+zOB1e0jlmDLktSGld/tWJxTho8/l1Ql+kB8HyTXCAsZU9HGuz27d/1h6xNKjMb3rIfrC05ljcWxd7b3zMRlyjhzTzvNkbWzbBQBgEP4nwADyxCCrqoFz5QAAAABJRU5ErkJggg=="

    ICONS_PNG = {
        "wheat": "iVBORw0KGgoAAAANSUhEUgAAACwAAAAsCAYAAAAehFoBAAAABmJLR0QA/wD/AP+gvaeTAAAFUUlEQVRYhe2YW2xUVRSGv7XPzLSEglYSU1CBoiESCJpYK50LFoKY8OBDY2piTCgIrVxE4+2ZBx+MMU1UQKDVGosmYoQEVB5ELu3MoHiJUUkMhIs3QGKsBbS1M7OXDz1Mp5XOzC6tvvR/2nvP/6/1nZl9LnOEa1BHS2y+UW0H5hcZ+Rr0kWhj4thIe8pIg4lt4aUqZhcwMWf5UxU+AX4EEKszVGSJQHWO5zKiddHViY//M+D41kgVRjqACTmg62Or419ezZ9sjVVbyybQu/2lv6xKbGFT51djDpxsrplgywLfgN4GgOqOUnpXVjV9mcqXO7Zzbqiru7wN5WF/6fjF3knzl23Y97dLf+MKnCnzNmRh0QNny6c1FIIFmFt/rC/1S2C5wGF/afbkkstrXfs7AetGjMCVJilRbayvfy9TbH7RxkNpC41AGgDRx1XdfmUn4OS08F3A9P6Z7ok0JU+65AFijfHjoB/608pkS/gOl7zbllCvKjtEPnLK5paRgawVU53PO1RuWwJ785WxKMddsoOaZsz32bFyi1PWxawigewYCp5ow9bxMn0506BL1gnYKBcHulLuks2VWDMlW0b5w4nBxayiZ7JNPb3dJTu4jp0zQKCnXbJuJ53HF9mxlcVO2UEayBqxV707Dpt0bRXfHj0FVAJ/S8ZOj6xJXnDJd2yLTTWiPwBBgRORxvhsl7zznU6Vt/xhifXM0655I/ZZ/BPNkq1VfN45YO0W4DKAwBOJ1vC8YrOJ7eE7Qdb700tpDWx17u8aiKxJXhDlBX9aYq3Z+cW2JdcVyiVba25QzLv4366oPL+o6dBvrv2dgQFK6HlR4SiAwJxe6f0g/npk0nD+T1+5Z7LVwAfAbACFZN85r3kkvUcE/KdOnJnWwHLgrL8UJc3hRGt4xlDv4ZaFlenSYCeqNfTT/mStaTAVtnIkvZ2uEp1bouXi6SuIPAKcFOyDKmYfSoVv6VLhyeiqeDtAsjXaoEozcL3/+TmjmfuteLuBWwXabZonYmvjXaMO/Nmm6impUOgwMNdf6ivV0ht7gr1TJM2enHUQvsMiSM4afOtlAg8ETaCrV3ovACEAlO+Ml7k3vOrI78VwFLUldCMmFQrtzoE6Jyp1VU37u2Mr46dSqcACRFsB9SHm5cAqyNZS7VlQs+bQmaqm/d2iUodw3j+4eWq9XbqxOJaivuHEtugGFV72258ioLHoo4mzQ32dLdFao9Km6Mx+q5w2Vhoij3V0DPUmX6u5yXpeJ/03IRTWxRrjW64Z+GBbbWkwlT4FTAXSgg1HGpOfD3twLdEVqrwBIMLKyOp423De/j+nmgQ8hPPmUmZW+KkjPfl4Cv4MgXS63odFlfZ8sACo6FXHV1F4VedRRHb0e6mwZaauEE9BYFHqs/2RTYX8rlL01YGJPFTIH8j3YbK5ZoKF+/zpRU8zJQc311aETJ8lEJin1tapsNggL0YaO9/MVyuxPdZg0edEOQD6vlg91mdDxlNbYrGXgEkIS5PNNRPybYu8wExkBlcuPzBZjUkGTRrFgFqQ/pNA0WeAvMCKPiMwB2EOyDr1hKCXxg62ldgyMwsY9lVW3i3xc/nNJ0DehqF1B6lLkJfy1QHwPfluEBbhnbPXT/s+j6e4y9rBzbUVgVB6kVipBqYiKgKnVfRoKhD8aNGKQ71XvIntsQZF23zIFblb5WBbbWkwnVomKtUKlQpWVM6r0aNBUgcWrP7s10IsI34ZOJzyAY+GRvTw839qHHisNQ481hoHHmuNA4+1Rh9Y+l+y/Gs8Ssr/tDYCdfeU7Z1UeqkZ4GJP2d7Rrj+usdY/96v8Ktz3RG0AAAAASUVORK5CYII=",
        "snowflake": "iVBORw0KGgoAAAANSUhEUgAAACwAAAAsCAYAAAAehFoBAAAABmJLR0QA/wD/AP+gvaeTAAAE/UlEQVRYhe2Z309bZRjHv89pGW1tpxMvJfEPwDiExEjbTaGQjWVmF9Nk3KhLLIsaI+VmkOxqERcjuCxxWWsy9YbtwkSyRcYEN1daFpMOZmL0VsOtiK5dW0bPebzoOadv29P29PSAMdn3hnPO+/C8H16e93neH8Bj7azIbofLFwP7qA2LAMDbGAy+m9i007+twKlojycP108g6gIAMP/iQv6l3tF7Wbv6kOxylIr2ePLkuq7DAgBRV55c11PRHo9d/dgywjosqN/Ygm+5OH/UjpFueYRT0R5PTnJf02GJHgrNmeIP6s+Te+H25694W+2vJWANlhgD6qdNUvhT3YDpEwB/qm/BtrbCfKvQloGNYKHwEIh+12yIsM4KQrAR2hJwLdjAqWSq0jZ4KvGzndCWgPPwfKTDEm0QlH4BdqhkyUc0aIIyCKINHdopn901YF1EG8RyyB9euQ8AiWhwjMEndFzgeCLmnwAAf3jlPrEcEqAtyRLwgy3vaSI6JjuoW4NNxgIREM9UW9NUMhaIaNCyg7qJ6NiDLe9pK32bysOpaI8nR+6zAJDO+yaHP7ixJbYnosExY1hRPBkIJz8Wv8xfONzuc6WnAMDNuTNm8nTDEdaKAgERAiI+V3q4HiwD3+jPhCslS5rSwkOTz5Ue1vyarYh1gasqGNGG4pT0TJCMBWZFWALGJdB3unOm7wkYF6GTscCs9qY4pVQppqnfDHRN4KoKBmxCVg4dPBlfV2EjDOgTjIBxfzhRFRb+cGJGhGbghBbTB0/G11nmAegpr3FFNARulGdV2GkB9ooRbAW0Hh4MTGvQzebpKuBGsMXUVYIFY8IfTozUghWgR0A8JkJrMd0MdBmwGdiybMCYCIwmzjWC1RR4J3lehBYnolloPa0tXwzsQxvHifX1bBbM51miP9T3EDFeL7HiWxAWKqEY1CcxvwkACtHXBF6ptCHwAJjeEHzNgXCj2EbPgnkMgFf199vWoz3B0Ps/bJQBJ6LBVRB3Nxym/0AErPrDiR7Axh3HbsmpPWxnHf17nijEGXgeAAjIshAS9f6NosyEBOqEV2VIAPg196hdX1CVleaVmZfditdxDUBI/fQ3FB7UJ90X/g/B9JmAV1Vuk7HgWwz+Uu38bX94+Suxvd7EXb4UeIEkLAF4pkjH8e1HbUdefe/HjGZeFhJ9kbs5KSO/BmBJ/fQUJFpMXPL3AvVnuRm1ClsFbBa6XrmtJaMy3iysIbAZ6HrltgZszTLeDKz6u7VltPiRHdQtrCdmK0EA+kuMYYCfNijjIwBw5/KBTofMa2DuUP/0hscBddNa7+i9rIvzRwG+VfTHHVJB6dXajcqtti1SNVSvjEsFpbcZ2IbAIjQDMwzMpPO+ebG9ciIycLz0XNouGZXxdN43r/k1e9Bi6eRn/sLh9ifdmUMFB61WLDenjezFmL1z+UCnU+YX/8l5Fyp3LmZkqdLtbc+cY+Y5h8xryVjffqB6IhrBJmN9+x0yrzHz3N72jOlFU8vAupg7mBxL5dB0tQRLV0VYJsdSKWatyRKw9LAwCS3lMXcwpNtaygNwUzC9CRRTF0Na1GGJ49sFxxkrfVs+vaxVxkmSusS0pii81kyebSTLIVGruID5Oc2GGZ12wgI2nA9XjjSBswzSdr4ZaKsuG2CL/m3Qbh5o23bHURvaPljgf3gpsyPXXnAWr71QsP/a67F2Wv8Cd50npzTWc+YAAAAASUVORK5CYII=",
        "chef": "iVBORw0KGgoAAAANSUhEUgAAACwAAAAsCAYAAAAehFoBAAAABmJLR0QA/wD/AP+gvaeTAAAElElEQVRYhe2Yf2iUdRzHX5/b3XYO5lwm5g9yaX8UYUIjle25uQU5tIgIHJYQae62tJ/QP/0wjaJAiiTtx+5MI4poIyOkgUVucM+pswmFREkRFrT5g9jQKepz93z6Y3fe7+25u51h7P3X9/N8fr343vf5Pt/vwZSmdH1LJqNIV9easjkjQ6sEuU/Ru0S5MVb9rMAxFf1msHrugdbW7mixvYoGDgd8axV9A7hl3EDlD0RfMPzhrmL6FQzcu7fJW25F9ig8nOa6AgzFxnMBT1rLzyxP2cbm9X2XCunrKiRpoLPO44lEvkqF1a8RXVkzY7jK8Ju1ht+sPXepqkrUbkHZnxS3zmNF9g101nkyK0+sgmY4HDB2KDwTM0cQXWe0hXvGywkFjPsFPgWqAUR4u6HNfD7f3o6Au7rWlM0/d6rOjnKniC5QeJGxX2cUW5uNjvCAkzrhQP3diqsPqARs0NcV+dOFfbx+8NAx2YZdFHDPu6sqpk8bfRrVZxlbj6lS7TDaw51OYOMKBY3NouzK4vpblXduqBneeUfrz1dy5edcw6E9xsLp3vNHUd2eDVbhl8GaubvzgQUYqp7zIfBrFtc8Ed4aHqnpD++uX5ArP+sMHwk2zbc0clhgfuyRhco+XPZ3GpUzIlRZEffB5s19p/IFBuh9r+kmjztyD6KjCLMEaVHlQRI7yl+2yvLG9tBQem4GsCpyKGj0KqyIPfqRMl1rPB4+UQicUx0ONt4eVfsLYDGACt8bG817RdBxgc1gw0OofDlGzwmr3L28eX3fSClh4+rftXSmVV7eDywCUJsHfB3m/uSYjDWsKv4k64lrBQuw7Mmj/4jam+K2uKQjPSZlhgc66zyXZNo5wItywmg3b7sGnBkyA77fQG8FLlqD7urmbX2RuC9lhi9TcTPgBUDov6aUyVL7SGxU6Z199cUH0oBVxJswOF96suwS5GJ8HC23KpJ9KcC223UmKWthyclySEVqr44vy5lkXwpw44bQWdCTMXNF796mGaWGS1f/rqUzQRvHLPndt8kcTvZnfumU7tio0m1FtpQaMF1XKspfIf4eqZ1xds4EdrMDGAUQeC4UMDaUFjGhUNDwi/JUzDxvRTw702OyfppjicmHmk9cGn2zvv1wtjNA0Rr7yulLoOsSZLrRaAt/lB6b87QWDhivKbyc9vgnt1gty9v6T08G6JHgstkR9RwAlqRACdsa2sxXs+XkPK01+M0tKjwCJL+lSyz1rJwMWIBYrWTY06q05oKFCa5Ivjbzc8tyL0L4IJEgZcWjZtYSeP+Ct3KRr93sHj9nAjVv7hsVlaOTATi+5IeWR7+9MFFUQZfQ/1LufBNs1dpQ0KibjOa2rbWS5zXYGbDqvPh+IsJWlK15smVXMqzqPCcpjpaEojMLAspDTns4mmFxyXGNXVQE6QE9XjhaSuXFiq6O93CS4XBJiJK4WnU3+M2PC8FLVzjgewxYnegxsa67XWIKuNTKex9WdLvZaUzKOVnRqnxznAELkaS/M2YhzMq3kaMeDuRsW5NIj6rrICq1xTDlbqAnRaLj/l07pSn9X/QvmniV37HXZOgAAAAASUVORK5CYII=",
        "coin": "iVBORw0KGgoAAAANSUhEUgAAACwAAAAsCAYAAAAehFoBAAAABmJLR0QA/wD/AP+gvaeTAAAGLElEQVRYhe2YX4xcVR3HP79zZ7ZitkuBgN0lDbyoMYsaia7u3JnZbbQxkTSaoriFpNa6224wRbpa45OMb8aGgrqJOLPBlkTKn4AlSAg2hO3OnVlaQiTBkgAvIMu2Uey2pVqWmTk/H+6fTndndu/MtuWF79PvnPP7873n/s75nXPgY1xayEodqCLT+exNNbFfMfAZoEfRT/rO5X/ArIU3jLFHUz8qHxNBPxLCpQdS/WrMD4BNwLUxzf6F8IRa9md2eEfaidsyYa/gbkDlHsBtJ2C9K2sllx0tPt+KUWzCLxa++qmKJscFvrtgaA6VvwkcRuWYin1La7wPIA6rRc2NiPYqDCD6TWDNBdYqj9okO7Pbiv++aISn8pkBgx4Auuu6X0Z0TyWRfGr9Dyc/iOPnhT8NfqKjUvsO6G6Fm+uG3kV0c3qkVFwx4VI+M6TofqAjdK7KrswO7/E4JJvBK7jfR2Uv0BN0zYNuSW8vPdY24YDsnwETKB+0VbZl7vTmVkI2RHmi/2pbc/YhbAy6LKK3p0dKj7ZMuFhID4ryHNHM6n3uSOmncbal0gOpfgB3tDy9nK7mMOWe9F6FnwRdHyL6jWbp0ZBw6Q+p69Qxfyf6XTKe3l7cuVxwgKmJjGusFgGskUx2uFiKY1fMp+8VGAtYnah8mPjS+h9PnlioZxpaJ8x4SFbgoDtSvCtOUADH8mnfDAnkWEjPertRngZAWZtIVu9rpLeIsFdwN6jyPd+OGVtl20qrUxxIDmuc2lZgFkBgqFhIDy7UWzzDflEIZMYu1gKLg9Tw9EkVfha2jZJbqHMBYa/gfo3zFezldrYui9YayXGRHvYeAV4BUBgoT2T6mhJGZWski+5pNRiATZhJhRmFGZswk63ai6Aq/CbyV9Mt9eOJiKsipQKbguZcJZF8qh3CA9um3gHWtWMbwnm/dtB2OqeBKzHcqsrOcB1FM1ycyHye4NSl8FzccnspkBqbPqdwCJ/M2uIfM58LxyLCovrlqFNl6rIybARhMhQdQ5TH9Tn82VBQY/9xeVgtiWOhoP7FAKjLYQPd0WZr5e0lPT3W23Fy7qpvY7iqHSZG5T+nP+j867fuena+mU6imnir5lT9hkQHpLpFB52RHJxnm2Hu1Jpfi7Cr3XKiKF2rzt4P7Gqqk5w/g3X8htXVYX/j0rxsRLPiu6CItuUjmmGBs+GEicNqoGmFOzPf+YuuK868qMiV7QQV9L3T57qeWVKnsqqLMCWMRH88UaczG0lGbwD+2cxZkHtNz6wXA9VE9UYJZ1DPc4tSwsIboSzW3HQpycREbySpfT0UI8KOmpeicRi4bLSaQFTXh7Kt4xYlvl+a0yeA64BTlWSi+6OqduW9/VfYTucE0AUcd0e868PSfH7RCerl5UnQUWCNf7vlkVaDHX4wu86p2BJALWnc4GzREmyn2RSQBeXJ+vP4Bduaqu6ra+1Wbf2hxVTtIMI6hHWmagdbtfdjys8jf448dIH/+kbwfOQBKNxcmnBva5kw4jSS46JUcG8HvgAgcDg1XDzalDCAtZKLGip7yxP9V7catF0cGe+7BiQ8hyvW/HKhziLC2dHi86iEe2yPrTn7NNdmRWwBmsNUOjr2E70uycPu6NSiU2NDIjbJTuBd346N5Z707+MGrhneBBTQQI6FUnfmXuAWfOOZijp3N9Jruqi8gptB5RCwKlD8bWrWG5Mcdrng5Xw2BZDaPlVeTldzmFKPez9I+O4xb418vdl7xpK7gJd3bwM5QPgnlKeNU9uaGp4+uRyRODgy3ndNkAa3BF1WlaGlLr9L5qb/MKebAf/cKmy01nm1WEhvbmfLC6GKeHn3jkpHx6t1ZOeXI+tTiIEgPQ4A19d1v4Kyx/y39pfU2PS5OH78CmY2gewGvhh9AMyokaE4z1qxZ2nqwcy1UtXfCQwtGDqtcCi4gx2zmLdNxZ4CsEmzxmBvAHr9s4FsIKxgEVd5uKLO3et3TL4Xh0fLv7VYSA8K+itUsq3aLsCkWHNPo61rKbSdh+WJTJ+t6RYMt6KsjWl2HOQJofaQu7380vLqi7Hiqw6Al3d7RaTPv91KN2r9+6GYs6DHReX1muVodrT42sWI9zEuJf4P4ANUwT4V2H4AAAAASUVORK5CYII=",
        "crown": "iVBORw0KGgoAAAANSUhEUgAAACwAAAAsCAYAAAAehFoBAAAABmJLR0QA/wD/AP+gvaeTAAADTklEQVRYhe2XTYgbZRjHf8+badG99CArFKHq0Zu4rkIys3SLB78raIMHQSombUXEg+BFa1bb+gEKgrRuIkFaW7SrPYg3v5Z8bLG62EMRbKvSwm4P4rVbtpP38ZBJdppmnMy6azzMDwLzPO/7zPuf/7zzzARSUlJSUlJS/kfIaopqVW/U+OwR2/o6t3vuZJLaufJE1qL3WYdDE8/U/0y6tklaAJDxtQo6pcZ8W5v27hq0rl5xxyz2G9Ap4/PRatZO7HCznB1XzKlQ6oJ1ZDzOreah7M2aMT8CW4KUqjDuFRrzSdZP7LBiXutJ3Wp8PfHT9NiGqJrvS1sdzZjPWBELIKKyN+n6iQQ3y9lx4KEg/A24GBy7Vxh5J6rOucV/F9gahBdQfm8f6iP1ijuWREMiwWF3VXhdsNuBywCIvtisuDt7a5oV7ylRXgjCKyo8LoZ9QZzY5YEFB048GKxz3l9wjuWKc6dBi92LUA42Pszd3RVbzt6pqtOdWJA9XqExv7Bp82HgbFD1aHDn1lawKFMED6mKvjFZmvUB3GLzqMD7wbQbMPJFreqN/vDBPTcp5gQwAqDwXq5Y/xggn59pifBW90LJvLKmgsN7V+DcpU2bj4bHlxedl4DZINxifPvp1Y0bPwduDyR95y86L19Ts+AcATkfjA+8lwcSHN67VtiXz8+0wuOTpVl/w/LyE8Af7YxsY+Uhu2gd82TnjoRrRPRAEA68l2MFx7nb4d7nT/2lwg5gKZReMsY8FtWjV+NyrOA4d8N4hca8wi5A2z8tZJ+t/Rw1f7I06wvsD0IRS2+PTya4111/wTkWd0Kv2Dgi1ubE2pxbbPa9G2GWFzOfdF0WHo5z+R9fzY2y+1VHsApPe4XG4TgBq6FZcXeqUg0kfekW69uj5kY63K/vrqnKEEn6cqTgqL67HuTzMy1B3uzEqubVqLl9BR8/viMD3A+D791/S89efkBL/bX1TebzMy1EDqCcsUJxPd3t0O7LsgvljCr7pYRd7zVTUoZB98VRq3qjmZY+h3LjMAVdh7DUysjBzveI081f1bdVuO4fw9BRML69DdraVtqa8OuQJMWj0tV2zbfEycrEHT525L9XFI22ZGlid/2XYetISUlJWSf+BhmyTXNfS3MdAAAAAElFTkSuQmCC",
        "globe": "iVBORw0KGgoAAAANSUhEUgAAACwAAAAsCAYAAAAehFoBAAAABmJLR0QA/wD/AP+gvaeTAAAIE0lEQVRYhe2ZfZCVVR3HP79z711eIjWIBFJ0MkFCrBHj5b4sS8FY48gYGkJNhMjuIomDFE39E9s0vUwMrxG6ezcSmiBkVIwZZ0qJde9zdwWhsQEyafJlQGC0JFGC9e5zfv3xvNy7920voFN/+PvrnOd8z+/3eZ7nPOf8znngQ/tgTS7VgSrS3VZ/oyv28wbGiHC1a3UyQMTIPlWOWThqjN0fv7friAj6PwHOPhyfqsZ8E5gNDK+x2xsIj6llS6rZ2XcxcS8Y2EknZqKyEkhcTMBCV9ZKS/3izJ4L6VQz8HPpyVfmNLZR4K6iptOo/FFUT6lhKWCA93zfMVG1itkAjET0VuCKPr1VdtgYS+sXZt5834A721LTDLodGFlw+SCiq3LR2JMxm4vhymFgtAehS1UkIrDOp3o1l4tNYDC9dTn3DtAVCjcX+Hod0XnJxmzmkoGzbam5im4B6gLnqjyYanZ2BhqnNbUW0WUeG7uTzc4sVSTbnnwK5UsAIqxONDrfCfukE3ejsgYY5V/qAZ2fbMo+Wo3H1AD72wBWYJf2MqEP7K8SYxH9ll89Y6x7nw+obsQ0Ae8CqLI02xq/LuiXbMzuMMadgLLbvzQAZLuTTtx9UcCZdLLBf7K+RtfGG53ZqSXO6UKdutICxPzaj+L3db8etE1b2HkM9Cd+tQ4xKwv7xhd1v5U46dwhsD7kUdnqpBOpCwLOPhT/hCjhkwXZmGzKLi+eQ7Ot8esEvgqgcDwXi20sCfCuXQec8DXzuh9quLawXVqwiSZnmcKa8MaQR/f+smFEzcBEzUb8sSWwK9GYeaCczIpZAkQARHT19Hs6zhdr4su7z4GuDTy7JreknK/kCWdFODyUEdFY79pyuhJgJ52YqZp/araXheVWpwOtE2MC3/CrZ6PnejeXCwCQi8XagbN4dzZ/b0tDtFgjLVgTcRfgvw2BuZl0sqFfYH9R8MssLx6zgZ2XgTMIVzjdOeWBfWcqAU+/p+PfqD7hV6+MjuotAQFvTKsQziRGaakK7KQTU8ivYAcLZ4NSM7OCklrZUVnna1TC6UrQWZV0yUXO74AXABSmdbWnJlUERmVBWBZdVQ1A0C/6xbNDh57+U3/Akf+4zwDnPRCZUdGvoCr8PKhbV+eXBVZF8BIZgNO5aOzJSk47N6eGK1zvV7vHzznyXn/A3sdHN4DADfs2ThpWSRt5x90FvO0T3umz9QXOtKcm4I9JhT+U++IDi/bml1UVas66BPYHxZ7ogM9V0sWXd59TeNoLwIhMa2pcCbCo3hJeVOmsFthFbyjQHq4VGMlrJZL3UV5LR1CMGMJxXDiGxwYFNbYqhKDXhGXDyzXBArjuP8J+Ktf2oz4S8sCYoBzOh6JclR8pksikk+OoZCpx/KnZRadn0smKr7cPL/Jxo8GUrslMOtlUSSsqw4IYKFeXAotM0rzgp9XTuPw6Ipaf1QILULT+TBFlSi0xkLyuarb2/2jhE1bV/QifBkD4vsJblTqJykLwNppq+B5K2dWw2NQbEj/2nEi3oo9UiTEszPSU50qARTiWfwmaTVXJ/jNtybECkwFw6ah1Q9nVVh+3wau21kk1Z9sqxkgnGyQAEo4F18MhYeFoCG/NjdUCG+W1fIVP1QIL4GLDBF5FXu1HPj4sqX0pH863iJrnw3aYVs2TiIQOxGrVmys0UwBhxPytagzV6UHZFrCFwFObOg8Db/jqW/f+umFgJWe9Uf4clNXI5FqBNRhGoNGe83+ppOtaM3UQyEy/ejLVnHmxBNjLeeVxv3qFt7stb/6W/CiAKFOf2vDlAf3BHmidOBjC6enFyffv/1clrR1iZgOXAaA8XpiP95nWVAu/Wl1RmHQUm8AzfnHwZQPPfKE/YD9/HgigEvYtMS+mfDcEjMjWwvY+wP7X7gAo3JxtT8yp5NjacLcLYqrudP1QoS915feVVNl04mvATQACz8YXZfYXtpcsHNZKS1hRWdPVPnVoOceD5NwegjGveteB1hmXV4LIbEp+DPQreBSnTg0d0VFO56WcEuThijU/KNaUANcvzuxBwx3EKOtGHtGWUt0tzQdzIvzGr36kR87fWwmYKE3AYO/e2Dpnzk63WKItmFxd3RbC0yXZlljcWZI1ll2abYylgHe+INzeNSr5i7K6CJsAF0Dh297X3dcOtE4cLMIyv9prjN1Uzld2ZGo1cJvv63hOI8vK6coC1y/MvInoPKDHd7Ak25ZcV/ykUwudlxHCt+EOiZQcB/Qw6EGU4IxhW2JR12uF7dqCcdoSG8KjLuhRI3OnN3f8sxxb1aTMaUvMAdke3piy20TcBfFF3WGe8ezD9ddHjD2Cd/rzTlSin5nS2HEcwGlNjkY4AgwBelwx46Y1dr4S9N23cdIwfxjc5l+yqsyttvmtmq15B3P5J41wu7WRQ5l0cl4w5U1b3Pl3hWDIfLRXe9tUEVUEkbQPi8L6AFYVcdoSX8/V1R0qgO3pD9ZDqMGcdCKFynbgkwWXX0BZZc66T9jLTRRXDgHX+HDLDEQUVvvaV3K56E0Denpcb1GQFcBnA0cKx9XI3PpFmWx/LDUfaHduTg2XXt0gMLeo6W2Fp1FOiHA/3lvL+W3egbawHhjtL7eXFfRVkG05jSyrNGYvGjiwTDrZIOgPUam/0L5F1iHWrCw3dVWzi/4p09WemmRdnY/hzoJZoD87CfKY4G5NNHU937+81C75txeA05YY7+0JGSPIVdbaqQBiTDfocVF5ybXsr1+c+ev7Ee9D+yDtvwONJdhC6HLzAAAAAElFTkSuQmCC",
        "cup": "iVBORw0KGgoAAAANSUhEUgAAACwAAAAsCAYAAAAehFoBAAAABmJLR0QA/wD/AP+gvaeTAAAD0UlEQVRYhe2ZX2jbVRTHP+eXpOm/OVkdakFhoA8qPkgLziQtFqWyycYe3BA6HIx1RWVD+7g9uKnbg+BEp2KTUUHoGFMRLbRTUOfyaye1fVN80CFW5oZ0bqPp0qZJjg/5/dJfkzRLlqYNM18IOfeec+795PLj3N+9garKKykkaOi9Td41tdNHAaZn1xzYvH94zuk3g/4uMDa5RI480X3uV6fvfKj9oYTqQUgOB/aODJQKbBQStLYuskWgV6B3bV1ki9M33tfiAekH7Uok9bXM3FSfdoH0p2JXABilMacNXKlrqgFqABBd5Mvoq7FiS1JhwBWk2xRYiOS0gabolRgQA0BlkS+jL2bFlqSCgK9HGwcVjikcux5tHHT6Wnsm5kF3gwy4DDmcmZvqkwHQ3anYqqr6fytraw4HA08LvAJ4yzz3BqBeRH9TNa4Bk0jSrE3ODrb2TNxYKikL2AwGfgEeLiPozfSvwJGLd9777o4dnyYynTnKmtavBFUerVN4u/na5aHvP3gya6vPV4cnNc66cnxEjKcWptHj8+pej9ImShCwarV2ejyJz0+f3u5yQrmX5tVk20sjV4tZmkI1EvRNp9dKme3oOTsFmIA5EvSdUIxB4G7Qzuarl/cD79i5Ffcu4d87+pNhyFYgDoDBwfG+lvRjWnHAAL494TGFjwFQbZoz6jfbvooEBlAkfTpJom22XbHA3tjcz7Ytyn22XbHAS6ligWPemkdtW2DStisWWJSdjlbYtioS2Az5NwK7rOaUROJDtq/igM2QfyMqX2Jtagpv+nrPR21/np1uhSTUnutvWy8JfUSUnSi70lzCmcBF87gzfFWAXWp44+n3RNlnxHVfVpBwZj7m3i6HSDq7VwU4Ltqc55ZsSpQ3fH+b72fCwioBC/qHpoH1T1QmRPgLJCyR+JDzmc1UNrBKEgFFyvZjBLdH04tnnAz0hA8UmptdJYRLqS/u+fqTzoblQVyspCYeXJguNV+hynXiGLMMd8PszLaSyJaQiDy30Er8WExuFrBhGKccrUO5jimlyOzzdyg8m2rJ777u0fFi8rOAfXvCY8BwqqUPuD3xz5YLevRE+2OInMIqEYq+LoIWM0bO2mL2Be5HmADusrouKByOz7u/6Hj5bPaF3030Q6h9g0u1G/RVoNaa+it/d3jbsgADmB/5WzFk2AENkFC4JErht5DCHRljAPrdTG3D1mde+GamGNjUcHlkrfSHpJ+5knVDhLe8yejRW73JLOhPmXAo0CIqz4M+DjSjheUBqDBnwAUV/VbietL/4ug/twJaVVW3i/4DSz5uXng2HCwAAAAASUVORK5CYII=",
        "scroll": "iVBORw0KGgoAAAANSUhEUgAAACwAAAAsCAYAAAAehFoBAAAABmJLR0QA/wD/AP+gvaeTAAADAklEQVRYhe2YT0gUURzHv79ZlxKR1CBahf4dhMJLVFTuKAohQVAX85RLRrt1rSAQopYOXsIIIoLJEMzTdrRLdjBwRk8eBA/hwVvuIRIr08Vd36/DbDrpjLvz3o6y4RcWdua99/t++PGb35s3wJ6CFRWaMPaq7XA4nO0B6ByAat8GRFWC+ZhGGInGzTtSlM542w2aRvQqQMOQAHV344d63HqmFMJrYGKg9bQQYhLAPhUDF8MH0YT5XHa95jUgBCeRh2XGIBhHOYc6vz+AR51xGei3DP2+LLBrhjkJzarXlwBUgjE3Xxtp7Op6vyZjYBr6BwBXXIylMu2a4dETHZUAKu3I9EUWdotI6/37VzbTniWxIWa/Qb30e9/+lyC+tx4Z6DeNaO92azarCODSSo9bL5zQAPX5gd5xYEANeleAAXnoXQMG5KArgoZyqiqzPGIa+r8dx36kZwE02jeoz3wTPaXHrW63GMEDM4Sj27cXt4ZuTKQuJpq7Jlc2DwVeEqRRCsCq33W/MtWubIEDR+Pjw9lwxQGZbdxNO1LD7T2fMwAyheaZhp4tNGdXu4SM9oCDVtkBKz1044beCEJM4+LiCEIOjKGWhDkr66kETMAQGOeLff8ke+IlABdkPZVKguwtNfA1Till+GtNpCeykO5HqMg4a8jN10VmVDyVgPNHp2mVGH5Vdl2i7ICVSiKVuh6KLKSb/NRwui4yo3IKVwJuWEwPsoZuFNvXNKBhMf0OQEzWU6kkeP2UEOwap5QyzEBMZqdT8VQCzm+xj1Ri+FXZdYn/EZgKfqUvnTa8Dq58d+09rsAdc6MrAJbtKz45lmwL/Oxne3BT/nLpTGJqyxEf8HjoKAlhGvwJoGsAjofrc2+tgebHYlX7GQSsCGs1IZF9CtAR+w5/JHLv7p6ZY0FPSMNl2F/hYyy0GAWU5xALZzVkSOOk11zPGm65a04z0AngR2nxttUiMXVGb094voIWfKCs182HoIVuQuOzzKgtKd6GFohpai2MwdZb498C8thTUfoDslghzDeZThgAAAAASUVORK5CYII=",
        "truck": "iVBORw0KGgoAAAANSUhEUgAAACwAAAAsCAYAAAAehFoBAAAABmJLR0QA/wD/AP+gvaeTAAAD90lEQVRYhe2YS2wbVRSGvzOeOAlpSWlh1aqICHVDN6VBLfUkSqQqdAEUJLJBIgVRJwgQFY8NYsOqrPrYQJQJCBVVQNMFCBCi4dFI9oQiFEBQymuH6AYQpG3SBMe+h8WME3tm8rASp6nkX7Jm7jnnnvv5Pn0NNdV0fUsyrrNNkJOoblq9VplS4em2dHak0qq2qB5AuAupAtgCspRDwEil9WygrqR8EeW/lYKKSBDgNgAtb3fJssvzmQdTfaNfrwBarM681dXUNH11Yjk5rJWCWS3VgKutGnC1ZS8eUjXtyrrOpxGrkMfou06fdyKu0qoCJwu50r33ZmBvJEgBkX2e62xK9WaPht2rOiVyiWRyqbEKRzzXeS5sv3ZzWOQvzbMx8lFeKIbEQV+7OaxG2570/o3xHMkOpgqoHAMfOuum6p1e7xVYo7uEk/aOI/rsnEUOZ93Ui7BGgcGHFnh+ziKHvcG2/WsWGCDVmz2KyEvFshqTWtPAAKKFsr16zQOHVQOutmrA1Vb5SSf2+sxrzk1xgUmTs3JW0iynMXN1unm2i8SSYlsL5hZrPTrnKgNWNZ/LPIf1DMll/xOQoIRJ9Rax+Wex3Krl38NCrGXdYldVYk3YxqbfKkgLqpsBBLUMskugKQgzwKQgRtF1QCKwz4jol6qSq6hNdAvINp1bPwVBJhS18Nu0ABQmLeGcKupXlIsmQX9kJDy3La2oGxTHbLEf2J0e+QPg7JsdG+pmCidA7/fdeszp9SK/WReS5zrHFQ758Lxfrw2PtvZ9dgkgO+BsFeE9hTt9SD3opL03SuvH7BJFGDBGeoqwAJ2PjYzb07lHgCt+qOyvBBZAoZj/sia0pwgL4PRlf8cyB2aDTTR/BFihJXj9s/2JzIWwf/czX10W+AYA4dahoe5EOGY+BbFbg+KY87h3JRyTOjh6Hvg7yN8S9sftw/ngmVSNX7wqUrzqFLq7Ty95q+u+cFphdquIvS4FbRbvfvmwPw745+C5YXSgvS3sPDfYsQXVnUHxF5FgUSxB8jIG+DUoto723705HOO5qQ6gGUDhp8WBRU8VX9Uyb2dcZ29x2DODzs685j8i6B1RTkXqLyadzV9vLOvD0dfbd4A/XbyBPV2InJwLlUj+yJCrIt6gM0z5FXwKmAFuLKn420TDDTvu6RmerIT37Ksd6+rqCt+C3l5ivoTfCY0laMOpdGZfeAQjPSyCNmjDQwpnSsyNpbDAd1bB7qoUFqDzqZGJgkgX8H2JubkMVvjEns51x023eU9bv6dT9ylyr4hsF9V6FflB0S8azdQ7rX1jM5XClurHoTuS4+MbH1a0U2C7gWlLOQ/ywZ7ezMeVrI2aarqe9D8y/2qXfGVjfwAAAABJRU5ErkJggg==",
        "map": "iVBORw0KGgoAAAANSUhEUgAAACwAAAAsCAYAAAAehFoBAAAABmJLR0QA/wD/AP+gvaeTAAAEl0lEQVRYhe2YXWyTZRSAn/O1g5YawgWigEESgpIYEw0qrP2qcIMa440GQuIEwa2FYBD5myjKEubfhBEmhrUdRJfhz4hecSFeKK79yo+akHhBxBCJERGjiSFrN1z7HS9ox1e2jna0etPnqn1/9j7vOedtsgM1qouMZ1Oi6+GVZLUFsa2Uxxd+dMWXqUqL5TnRsWByxuPehDDDdhmvlCVsRfxzFKMN4SnH8B+qtKW9kzorKd7bu9Q18+/fVyu6E7jt6qjsKEn4RMeCyUOeulcFXgQmFllWMfF4zFwkyh7gvoIJ1XfHFB79lgAMAN7c5wzgroR4kQyWJpyIBBYj0k7hLVOovqOG+EV5DEAM971qD20CaRiveNEMihxX9JAo+/LCxkhRc5YVNbsR+cohqyIcFsO+xwxbO0XJ5tf3T5jwsxmyVrmy7rmiRLkacYBpIuzyDabPxyNmy4mOBZNHBKwFIx4zV2S8dT8KbHXIXlBhZaAxHjDUPuXcMxzho91LfJMG0ltEaAY8jiXfGsgGf6gvOXypqHkEeAIg5Zl0izOCx/cvmm0bmW0qrKYw4n+q8n7dlaH2hetPXi5Sp2mB9/4ZcrcuXnesH8CK+h9UjFP5CLtVESsafJ7BdCtSUKe/qPCy2Rj/RAS9PjrFqF977DwQtiL+NhXZ7iiVqSLsyHgnrEtEzV/RAlEV6JFsdpt/7fELY/19IxkLNiIa49qjSqH6utGfnRdsSnxcjqyTQDh5zgxZq0TteaAfkC8V1ak4oqqQNAxZGAglVtxIFsCtas9Fhiuj18hmN5aysRxxYJUV8beqIS2oNOSmBgRZHWiKf1pOUAoenWDvqqSsk0A4eS410bcm/12VbwKh8soNrhP+LxG59ktTDv+b8HipCVebmnC1qQlXm5pwtakJV5uacLWpCVebmnC1GZewKq5Ki5RKWcK5lsByER7Jj/mupDqtiH9O5dUgub9+pmJsdo65iy0esbkr+FAipnsE9RdMqDSoyPJENNAjqq25/5JvTrS93mv7jM22SDPgy4+LGD/dUDi5v36mulxv2bY2SGE/+TQid+T6DG6Q51Sk4WbEVZFEl7ncVt4GZjmmLqGy3R+KHyhaEsn2em88ajbbLtcZhWfJySqcUZvHzVDifo+dvlNhA3Axty0nbpy1YmZvPGreVapsPGbOt2JmnygfOWSHVOjwqOduMxzvEkFHjXC803zSNrRDYPbwoMhfqrrz4pTp+5YtO5wFeCD8fRrY+11kfmxAvE0CzcB0wFBlqcDTVsz8zFa2B0OJs6OdlTgQmCEZ2aFKI443JXAEtTeYocJMSSISaENkC4AKLwjyDKr1jjVXFPbWDQ69sXD9yctjRSnXUFwjwlZgmmMqA9ojRt1utTM/5M76QmxNcl2dAqdR3WiGra9HO6NAeATK54K9tdx6HFt8+KE7m+IAlwR57cKU2w/mM1iO8GkVXgo2JY6VI1qGuJOSMwjgRox+rrW3SrplqeT6xruPdi/pHFU8l0GzjAxK38HgrUbWfhPlN/dgZncptxwvR7uX+HyDqQhqBHBJi9nY92G1zqoxXv4FJ98ZTSpxQ/YAAAAASUVORK5CYII=",
        "leaf": "iVBORw0KGgoAAAANSUhEUgAAACwAAAAsCAYAAAAehFoBAAAABmJLR0QA/wD/AP+gvaeTAAAGn0lEQVRYhe2Yf2xbVxXHP+c9u2mTsCYaY8URgm0ahbXSUBO6Jn6OHFFW7UdXRNcWaQJ1HfmhIhhMSMA6jQit0lBBIJYuiZ3+EEhQZVorQKs21jU/bCctSjcJWsooQwzUtBtTm9JucWr7Hf54z477YidOlKwS6leyfM895973effdd9+5F25oYSXXGwDgWPSeWzPq34XyZrA1vnO62OsO/Mov760oT35wVGA1gNrcF2qLv1ws3vjw0KZqpLu2vCI5fjALCyAGT03X5rqN8MDexk+YabsXWOP1GRjBhpbBoULtfAtO5lHfvvBiXyrdKmm7Hahyqy+g2oXIkwAZ7G8CBYE/lBFWReI91ipR+QroV4Fb8wDOGGJsSJ41zvgD6X8CNcCEYWQCDV8fvuDta0FGWBUZjtQvV8OsVyWciLJWIACaH5ZB6VKf/qD+scHLALGotV+UHUBZRs2HgciCAA9EG2/zqa5StVchRl0iqp9HqL6WL6dxVQ74DGNXfcvg6XyHqP4GZIdT5qFCwHOeEvFocI2qbBNhPcqyGcLHFI4Av/MnU79d863j/y3ab7f1FsLtwJXUqK+6qb0/ne+f9QjHuqy7xeDnKGEBiozi2wrDInpcVBNnq2pe37z5hUwp/YvQr3A7UOlflvocMDInYFUkEbG+h/AjwJ/nGgP6EBkxlBMTap5oau1/r9R+p15H/4jINofemBtwb+8mMxEd7UHYOtkzJ0VkZ1XVhYMrNp+6OldAr2zTOGnY6vLqZ7z+koADY+c7QLa6Zhr06dQ5/y7v/JoPZSbMtwy/062tfNLrnxE4FrG2gba5ZhLRjVZz4vD8Yk4q/J/+dxMBKw34JG+9zmraXCK+JxgQ+JlrKqKPLCQsgLRjI3LJNau8/umTn4w8DdzkGNphNScOzjNfQanquFss9/qKAg/uDd0CuZfsPV8yPW0WNZ8SsN3iFL6iwGaKh4EypwfpnG6xXwBVuv8feB1FgVXsdbkgO/3r2V5RFRmKhNbFI8EVs2k30l1bDlQ7DJz3+qeZw1LnFs41tA7/dTYX7dsdrkxEgy/Y6MsgI327wzN9unOakLIVuCmDwN9KAj78i/vKgIBj6elCMcUU7wrW+X3pYZCNs2mXlapp5QxbXvf6C67DlZVXbiKdS4wulnKhvn3hKv/V9DMIbYDpVr8rtrGp6Rv9Ux5tUYk+lC2qX4+WBJwRM2WQ/YgZ/kIxWQ111tfYpu/bpNItSHYJBOAYypZg2+C/SmUdiDbehtqNrvmn0Lb4P0oCDm/tv5SIWhNAGaIBr3+ke+3SCZlYi+gWW9kAuijPPSYqTzaci3VLe255KkmmZr4D4kxT0X2FYgoCi6Dxbs4grES561TvikVjF5bWYhj323BvkmQtYHpSyyuo9PiMq8+uaTn+zmxAARLdDXco0uKaY4vtJaUDO9QkgJVA+cVL1X/G4NNO9RSdReR5TWlnaHuspPnulSqSiBqduOu+KD+paz1yqVBsUWC1zQExMq2O4cC6ygAjAv3YxuGG84Px2T56rxI9oe+CftGx5O9XF5k/LRZbcIs01FlfkzHMARHumLwDeQOxn9W0vBraHp/TSBaEjYY2qOqLOCtLBtEmqzkRKxY/ZYRjz1vVatIn5MECKppaoslDddtPpOYLNhaxHlTVAy4sCjtC08BCgQ+H+OlQuNOh1JMIpwAEVk/Iki7V+TnLiEWsxwUOAYtd2GioJf7jmdpdc/HBrtAXDEOPuOY7PkndnRHzFrWNY0CFU637q6vGWue6LRrqrK+xDbMTYX1e9Z7gaLyllHfhGuB4JPQS6P2ORzdm899YxHpQ4CDZzafqGygtVltihBJ1vGP1zamyssdRfYLczWMr/NBqju8UKbL/Lgbctzu8zO9Pj7p1fwk2x1fmd5LoDj2gogeYTP0U5Pdg73l/ccVr6772h/e9nfd1hz/qI92E8GWBL+E+frf1v22VRxvbYq+VetPXAMejwS2oHAAQ5alCB8vD0cbPZtT+FVDrcaWAM8Coay8BPoVzTubVuMJzYuoz1mOJy7OBhfxVQo267KmIwkCh4PrmwdO9vZvuCVw8/yii32dyJfEDd7m/YhpFdX8q7X9uVslQUWD0Y9lSxm+8XayBe4LTo+3sHappDKut61W0QWA5sNQNS+OM9puIHAP71dGlgaFST39KBa5wZ4j6x1Izntw4b/TgUSCXAva1h30f+fjlRXWtJ6ZsbeZLOWAV6RCb5RhyqOGJ4fHpGhWTe7Ay74crN3RD/0/6HxVIfEy3q2IlAAAAAElFTkSuQmCC",
        "bolt": "iVBORw0KGgoAAAANSUhEUgAAACwAAAAsCAYAAAAehFoBAAAABmJLR0QA/wD/AP+gvaeTAAAEkUlEQVRYhe1ZXWgcVRT+zuyO2ZgYYx9MWynNgxJFhEoEzc6spP49iNrG0ggtitLsLkEq2hexFcWCii8Vf6rsJmlBUehWrVqk4IMJ2Z1N1RQfpKBYMLW0piJtSZNs6u7M8SEzu7OzM8nOzo340A8W5t579pxvD2e+e+5d4CpWFrRSjrUPozdymF5nplOxRO4tUX5XhPDowd6IXCyNAbgbACTWb4smJ34R4VsS4cQOZpBcKo3AJAsARVmeE+VfOOF8Wt0NxjbbVPF8a8c5Uf6FlkQ2rfYR8BmqEkGn1ET2FlExhGVYS0c3EPBxrU+eEhUDtc4bw+j+3tXM0tcAWhZn+Dvb8mkRMSwEJjx6sDciy6UjIKwDABCPE0nvWusMTAWNYUcgwswguagPA7jHnPq9aMhbmI015QBEU0FiOBGIcD6t7gZ4uzmcAfjRjcmxv0G03rJh439Sw9m02seEveZQJ6ZtakI7CQDM6LTsdFkSWsPhRr6kpaMb2K4IzM8rydw31jqhTFioBgMNZNipCAQcUJPa+1VGVCb8R3//YT0oSTt8ZdjsEY4AFUVov/7SoN0mv6+n2WB0AACTWIUAfBBmBmlDFUUg4Dcio+/2/pP/VBm2YD3MHZSYprMfqDc0Quxy6br5h587dsU5X3dJOBThErH+WHRg4oLTTudQZ2XE2ymMC4182iKX53NDsb1O/3URzg0pj9sUoURsPOHVLjJRpB6fdUAC8ybn5LIlkR++907DMD6C+eMItEtJ5r/1sp+90nqsLTLzNkB3+KbI6CrvmIBhML3qNFmyWxvd37taDpd+sJwQcEBJ5Hb4JlIHciNKF3Q6DqB9kRntUePZN5x2niXh1iO0t18c9LIPgvxwzyrS6ahFlgiHlYHsm262roS9eoQaRRCAyVS3rHMow4DVM59oMgpPE4HrJuzZI6wAFqTmd4hxvzk8J+n6pruSJ+a97GtqeDwV2ywRf47ytoszIPzqnwr/PLPQ9pKbllrQ0rGdDLZa0TlJkmLRgfGflvJaoxIS8WuwZ36xhtc57ZYHPdDaNDsO4EtXsqnoQwzeZw4NED+1HFnArSSIvgJg+CdYC2JecJvPp3puZZIOwUwYMV5R49oXdfl0m5xMdV9b0JubGiIZpves+ne7j8gP96xiI3S88pLRJ0o8+6TXS+aE68ZhFr1n4S+FXFpZY+aBMVd9nptMdcsFDmXIpggRnk/USxZYgXsJMHWaT9PRXRMF+5JfRXCDUMKZzNZQZaOpPstp6dhOMKyNp0AwNkcHJ876jSGU8NqLf94EQAYA4spZLjekPGhTBAbxM0oi/2MjMcSWhMSVwycWM5wbUbrAlIH1vhC9rMa1Qw2HCEixClypX4CN0356hHohlDAxd1rPTHTWT49QL8SWhO0+gsAvlBWBcaZYDD/iVxHcILYkKqdlAHSf+VAgMrZsfHZsWkSMhu4lvEBsJwzAUoR4Y4rgBmEZzmS2huBskgIqghuEEe6YPb8WwDXlCcKnQRXBDcIIy8Vii234fTEc3hFUEdwg+i+DF4n4ZirxHmUw/5dI31fxX+Ff6CrKj6MPa8UAAAAASUVORK5CYII=",
        "star": "iVBORw0KGgoAAAANSUhEUgAAACwAAAAsCAYAAAAehFoBAAAABmJLR0QA/wD/AP+gvaeTAAAGQklEQVRYhe2Yb2wU1xHAf7N35zsQBqwkoqbQNmkPpNpSG0wV8O1dzZdWtJFSqOJK5UMg/tsQ5V+lVqQSsqgKraqGCAnK2Y5RpCAluCpJmgRVioSx95y0Namq2Ing0qiViKNIKQFshcN3u9MP984+G/+5s5dKlZhP++a9mfnt6O28eQu35daK+O3wjSPbw5WRsYMAY5nKp7/32JkbfvoP+ukMoDI89qjAU+Z5FPidn/59zfBQsm55RpZ9CKwxqk8iev2ezW3nP/crhuWXI4AMy9uYggVYk7EirX7G8C3DZ080RELZ3IdANeAadQD4OBsK3rNtT1/Gjzi+Zbgi67aQhwXkRYSXzFR1aCLb7FccXzJsKsMHAusATyzvG5YGXVe9YfJJ8S3LvmS4MjLWZGBROBVrHhze2tL/Piq9Zkl1xYS7x49YSwYeStaFBH5mhhpQ62BhzlMOAB6AovtGTtVULDXekoEzRB4Gvgwgwh/q2/rfLcwl2gfeU/gjAML6y1erdi813pKAh5J1IUR+bobqufxq5hpBOzBZFuUXS83yooG1A+u6LGsB7jaql+Ptzj9mrrNbUyPAK2b4pctXqpq1Y/FxF6wS/cl4NcIGS4mqaFQgihJFiALhAr9lWXX1zf1/n83HYHfiXs/zzhfFu4GSRkgrpEUl7QlplIuJtoGPSwJ+48j28KrI+A5PtFY8+ZqIRhWiQGUJb/1yrNXZMd+aVKd9WuEHC/kCxgTSqpJWSz+wVIavZlacLjRRk83PyvD4rxV9QhQQRed3eg0K2dH3gtnssYUoghMTzdlQ6B0V+brkExEFVs6ytFJhE6KbREFRVobHnwWenAYslro6nXICuICSFiGtomlxA+lA4MaFLS1/+WQhwJly36N//Q/wy2Ld2133rXHd8EYNuFFRierUVtsITH6cYqk7+Vx4GHxm6zJ3ReC0wHeN6hrK9+02xykXbiniJG0b4XVM9hX+HBh3d9Q/9db1acAAI6dqKq5cqXqpaK99rvBAvNV5838BmzqeSKjlvcbUd3MmGwruLD7Sp5WXmsaRidWrP/sRaoo9LBd4NZWs/86thh3oshvU8l4vwAq8di1TuWNm/3FTPaxpHJkYrapuRPUFo1qmYv0p1RV/4FbBOp2x7aKcAVYAiNAb1us7Z7tezVrAGxt73dGqtbsRed6oKlT1lNMV2+k37ECnfT/IaSACoPDixEfBH29uO5+dbf28B4d2YA2utbsUHjYqV0R2x1oGXpjPrmTYpP2gCCeBkME5Obr6Cw81Nva6c9kseNJpB1ZqbfwoaHsBGvQhuzV1cimwTmdsF8jz5G8lgByPjQ7slY5837FoYABVxOm2nxXlMaPKWa77lfqfvPXRYmAHf7/1i14g8C/MOaDCEbvZeUJkofOqxOZHBI23OI8j2m9UQSEcWgwsgLHNH1qi/fEW5/FSYKHcbk3lTvM0vqW9799l2RaJsR0zPu8ox7ZkYNPHRgvDUjMym4igCu+b4YZyeuSSga9+dsdGJr9mhsvgmytwwUfoyrVVG8qwK01cy6udHKiMlI42p0z58AK186ybJuX8W6uZfLK8OTM82J241/W8pwEClnVwrqZeRd9F80XKE/UfWFRrJ6ugxU0Zdp6LbcSVA57nPShmoed5P3Q67V4Cut9uSl2YZmAxUvg/lPddmpRRJaTg9LLdlBotaM/1JNY7nXY3rgwDjUyv7QI04sqw02l3n+tJrC9MGB+XZ/j2B3goWbecwmVTdBigvyd+l5OMHw7kvItAE4VDAC4J0ipIq8Il4yIINAVy3kUnGT/c3xO/q9gXcLeJsaCUdNI5x2ObseRvxuQkeP8EeZLp971PBQ5NhILHCi3h2RMNkYps7hGFfcCdRWvHQA+D9VXQXQB4+i27PTW0EEtJe1gsq3bqlqe7ZrznuMDRsEYObW5782rxhAF/5uzRhs5gKLdX8uCr8i8q+yku5UIN4A+wFleIKckoHMtp8NC2tr5P57PftrdvHPjN2WTDc0HJ7RN4BNNOTgFbJe3jEoF1XVFOcwonQhI8sKWl79LcVjeLebGfvt3VcDiruf0Ce5jc+7rON2AR7UHlmwjvuK7V8e32/nQ5oDPFvGjrueOJ3wYCXgfKJhHtWYrP23Jb/l/kv+WDUf9rU5GrAAAAAElFTkSuQmCC",
        "heart": "iVBORw0KGgoAAAANSUhEUgAAACwAAAAsCAYAAAAehFoBAAAABmJLR0QA/wD/AP+gvaeTAAAEX0lEQVRYhe3XTWwUZRzH8e9/dlu2pALVS0GjMR40rgkHKofOtJYoVrh4EYJoTOnbqkFjSIiReGhi0BA9EYl0sRGBaKLGyMHXIDTsDAUTo21s1EQT0FgOJtKqDbTbnb+Hedlladlud7rEpL/D7jPP8+zzfGZ23h5YylL+35FSHU4daKuvqck+CjwERhLVW/xfjgmMKPplw6rxL5JbR6evN87oB8na8YmbN6H6sMJaoBFFEP4UGFXl68m65cfbn/pqckHgUwfa6mvi2ZcQ2QmsKLFfFwV5bXos9taGvsGZq8bpa4vX3JrdCfIiSmOJcSYQ2T+5rG7fXPBZwZlD1jpRPgTuLGqaAr2ISgxhNRAvbFT4RpQtVsr+DcDut25H5CPQ+4vGmUG5iOACjcCyovZfBfcxs/fM9yXBzsHWVjXcT4F6v2pShQFD3WPTY7XfBUfws/2blq2smzRVtRN0GxDz+/9BTB8EICcngTUBUuB9VX0nW1sztGHH4BXw/oH46pl1IvIkaCew3O//D8pmK2Xbc4LtAfNucnIWWOUfs5O5eKzjgc7TvxfvWGEyB621YvAecK9fNeZ/B9hRV2R7a09m5HrjOG8336EqR1Bp9asuGUZufXP30C9BHyMoaB+G5uRIiFU+TuiVR0phAVqetoezNXETOFcADbDnEpowS2EBzO4zFxpWjm8U+MSvanDd2GHV/IENwc4a83GB9f7mcLY2/kRT6ttsqUmCbNgxOJ7QRHsBOsC2N6VOTMx3nOTW0Wn5N7cd5YdgP5xD5pZrwCDPhSXXfSY4x8pJU+rEREIT7QJHBY6Wiw3SvGvoMoY+m6cZL4RFCK5mLng1MmT1ZJrLnWQxkklb5/x/XYnpbVaXM2YAqIEZdBI3PH9ueETkeFDUGTEhOCWUu/KdGK4+bfaIm7cIntHwP4J7LjnV8erTZo+IXAo3DFZ6XwDC5aBeVeqLf3jjojeFJWUSfLC6ej5oiBnhzf+Gx9W8xRA5D+Ep4Yb3ThXdXHXZXCmwiMychYJHs522fgTuAVxDc8nm1NBP1RfmY6fNJMgI3kEdtXrt+6DwwaF6wC8ZrsT2Fz4Oqx3twwB5E98nKoEtD/57asUhkOAlY6OTNvdVl5mPvcZ6HWgDQPl5VcNfA0FbCN78/OdTrkEH4L0/iOy2+8291YQC2P3mXoFd/uY0hnYUrmaMws6t3RlHoQtwARDZU0203W/uRWSPv+mCdlo9ztnCPrOep066ZZuixwhfyuUNqzezezGxTtp6ReHlEKvaaaWcd4v7zXlhVRM9XyyUWDVXA10OFuaxzF9MdLlYmAfYGzh69EKwME+wN0F06IVioQywN1Hl6EqwUCbYm3Dh6EqxsACwN3H56CiwsECwB5g/OiosVAD2IKXRUWKhQrAHmhsdNRYiAAM46ZYORQcIV+H6qjd6/kVGkC6zN3O40rkie0m/Bp1PZFiIEAyzoiPFQsRguApN1NhFi502k94icilLWfT8ByO/LpF7FvClAAAAAElFTkSuQmCC",
        "gift": "iVBORw0KGgoAAAANSUhEUgAAACwAAAAsCAYAAAAehFoBAAAABmJLR0QA/wD/AP+gvaeTAAAEVklEQVRYhe2YXWxTZRjHf8/p6di6SUBgcUpUIiYaERNmonbtHFEhfsRdmCxGCAMyCy7EC6+MmGi880YTgug6o2iGN1zogkrEGGDttsTwESTERIkoiZggU4SNltK+jxc96z56TnsaV5Wk/7v3eZ+P33nPe96PAzVVV1KJcyIeeVSQTtBlIlxR1e8tK/BZuGfouK/4/kirIJ2qerdAA+hPxliD7VsT38wp8PC74WZjW5+I8oiHy36xzAttPSO/uHUe7m9fFlDTBzzmEf+1sWVd++bE7/8Y+OA7HU1BOzuKsKJ0JhmzhCfCPYlvp5tH4u1hg9kH3Fim1HcT9aHw2g0HJko5WeWA64LZV6fBnlHD08FMZrGVyy1FdSPIaQBUFxmj+w+/135nAbbvobsM5stpsD8obCCgtwQzmcWgnaA/O30rG1MTr5TjKTnCB1/vsIM3Z88DC4FxscyK2a/9q4/XNDamr+wFHncSHsucsx+4oeWypKXhCLAy/0Dsqyf17P1bjl6ZHu9Ml5NAI3Dh3IKWm7q69ua8mEqO8LyWq8sdWAQ+dZujazccmLDGc88AJ/JcrKpryfamCW0rwMLRa3V212xYgIefHzqDMOg0F7dc/O2OUkwlgbPYoWnNP7z8wi+NptTQDeQAVNgumO2FNGj36k2H0p6FVMemgEzI068ccI7AWUCdZlTVewpFtyZPIDLgNJtVZBEAorsjseFT3qwISLvTNCZrua40voBXbzl0AUiSp16VjEdeKwVt4K0iW856uxTsSH/kDeA+xzQU7U3+WYqpUHzog+iSQE57URpmpV2qIs8xBXoe5awIWVcIaAWCTvaMKMc8YG2E24AlkxZRaw+iv84iTOUCsmtyjbYL9mv6pgqbXJ9JZxiaEZq12NGNvk7hQde+ovckoqLrXXJgZc3tkGezptxp9sPw38gqsNmu/UpULVJuXTmj8+sCgZasquv8FzQqSgxAhbgiCTc/W8Tk1JwDLrsiiNRbRpNFcW7OEw2h4+W2SC8Nx6O2ojEAS2W0LZYYKBfjpiN9raG0NBTZy27N/zfVgKut6w7Y9aNrSqc2J96PhCy41RhO4bFJuEkhPLluq9CR6I/U+aZRbMviHgNn0zl1/ehdgRXdISa/X4hD4V/TnFW7BborijZOTXE/AUxNCZUfK0n8b0rQAlthhC9dbXp5fv2lURVZoCrzLHSH03VShZ3+k0sY1fyoinyk6IjvWGUbcC+AQV4U0auievGv9PzBgo9boHOLGHeaX0Riyaf8Fh2ORzcq+qEDv6ktltjtNzYZj3wOPAkwUR9qctu8rrtVogZcbdWAq60acLVVA662asDVVg242nL/LzFDEkzsiiz0m9CoNk6evY1qYyWxIMFytwUfwLpGbO9frSXLCzux/Z+l/VxtXKfEotRYBsj4LzTnyjgMRfL8dZqMt61DZD3q5y3MSGkpujyfXE7nb2mVhJNFdSASG95TWd2a5kZ/AwhcmlWt4yGtAAAAAElFTkSuQmCC",
        "calendar": "iVBORw0KGgoAAAANSUhEUgAAACwAAAAsCAYAAAAehFoBAAAABmJLR0QA/wD/AP+gvaeTAAADT0lEQVRYhe2ZzWsUZxyAn9/sZOMmgaZVqIYgCh4qXqRpDyarNaARv1ChLYFgSj9MD4GCelAotBHsn9DqJiJUEMSLCGIwiIa4ayikpYWWgpYeapC0tKkEs9nP+XnYSR1nxji7ZLYI85zmnd/H+wzz7gzzLkSEi9RSNDmybWNZ9XOwRpMDmYvV1KaHu/rA2B0T+WrLkYlfq53bqLYAoGzpl6B9IOenUh0NQesquXIetK/So3pqEka0xT6K/5NYGQ9aZufGXT2qmzpI0lRqxysL5dx/FyemXgLpAcg3xtvj84VskD6F5nhTY74wXRnpmJakdzFWSpja/eH4o5qFJ1Jb34yJnlJ4G3g9iNAykAW5aRil452fTP7ml+ArnB7p+hiVM0Dg9bnMzIH2Jgcyo+6AR/huassblsR+ABL2qXvAH84chc0Cq+zhbaAcUCQGdNs9/hb40aWzDnSDPZgtFs1N3YPjM84M093REnMQNGFfzenOI+kvRFBnTno4eQ3YCzC/omn/rv6x+SC2Ny70NDfnso/t3t8lB9L7nPGpVEdDzmgaQfUD4DWzoXQUOOHM8T4lhA77KFt4aJ5yy4bJW59+X5xvTAwCCxUVutw5XmG11tgHf3UPjZd8O6vcB1CY7jk8FugJAdBzeCyrMO3s4ca+W4vLoP3FwgGYy7ecFJGDlml0VnMHRNBKjRycy7ecrGVuzxoOwp7PRvPA1Vpq3/lo4gHwoJZaqPVN9z/y0gkvsSTEuPNN8tX6qThmBuN5PwyvsEir/TNaKyazIXo9l6ey0uqOeZaEqDaF7BMYsV9gTnzedDIrFqtFWFDI1EftWQS6FBKK/uuOeYRFNYcIiv6ZHMjsrI/is6SHk78D60Fy7thL95SIhMMmEg6bSDhsIuGwiYTDJhIOm0g4bHyEpbJPphKrs4sT20s9n3Y+wvozAEL75Mi2jaFq+XD76+2rgbbKSGbccZ8vDuOGih4ApGxZV+6e29pfLqjvttJyE4sZ7WUpnWVxm1f1usfPfUKHMDJtyQl8NuLqzD3jcXlz57HJBedJ71fzEFaxaL4Leqt+bh5+siw55JaFJf4yuHz5vVjbo4e9iLET1balcpcJFZgBudPaOvvtpvd/KYQ8X4QvTwDshwLNOceL/AAAAABJRU5ErkJggg==",
        "mail": "iVBORw0KGgoAAAANSUhEUgAAACwAAAAsCAYAAAAehFoBAAAABmJLR0QA/wD/AP+gvaeTAAAELElEQVRYhe2YW2gcZRSAvzO722woMfXBVGLr5cUi9uKlqTYzG9L2wUZKVNA+JUJtNltv1adW8KWCSqQgGoSaSb2geVAqiggRsZqQnU0VL40p9QaCjRitD9UkQpLN7hwfsrOZpNltkm62FfZ7Onv++f/5ODtz9t8fSpQoUVQk36DzmlmhkxIslgxASoNT2x7t/TfX+HnCTod1LSLPg1rAdctqlxP9FcUBedqKOUP+kVnCic667aruB8AVRfXLzYgK90aiTq+XyAr3vFG/KjiVOiWwJpMaVuEHUbSYhiqIKDcB1dMJfgsT3rA5dnwEIPt8hpJTTYisARDh2KrKv5tu3n06WUxZj+72hrKK8FiXwP0Ia8fdiX3ACwDGzGUSy4Yu310qWYC79388Kci33mcR9njxTAcQWUnm21fhWceObJF0Omo+3P9XMWUTR2qrNBDoBG2ccWOFF2YrrMpPs6dqowaMwbht7SqCJwBx29qlAWNwliyg8KMXZyssgpsZRNDfQa4BVgt8lLCtt5NTwUfy9ceLof/FreXpikCbKI+TaQQielZVVjOdcL1rjbmTBbgyFNqMapeXU2gOhVKDfUcjZqFlE3ZtjbsycFKU/Z6swnupZKBmvuvPEwYYT46NWrFEsyq7gXOZ9A2Gq70J22r7uuP20MWK9hyqD8Zt66BiOAjrMulRFWKRVueBiYrwufnmzSvsEYk5x1BuBXozqaDCwXEpd+K2deNSZU8cqb8+VJ3qEWiDzAslckLUvS0Sdex8c/MKA1gxZ8iMOtsVngQmAQS2CAzEbesJ1fz7kbnEO60H04HUKcDKpKZUeWa48uqIGev/5ULzF7SxEUHBedmxzeMgXcAtQLnAS4mj1s6+DnmoLhb/I98afa9HrjJSbifKPb7094ZhNNW29J1ciAcsoMJ+rNbE6alQcKtM/+pMv7nKTkN0wLHNxlzz+u3IXUZKB0A8WRXFDut4zWJkYYEV9rNtT+8E8FTfq5FPxdA3M3uPKpAP57Y/r125qtl2hfAn6F6zNdG92HvDIivsp25f/LNyDa/P1f5ytavQZHK9FV2aLCyhwn4yO6jmRGfkE1V9Bagk0/4UAyS7/oiIPGZF4125V1sYS66wHzMa7xLD3YSv/TFTjF4x3E1mAWShQMIAZkv/GXPY2YHIAabb3yQiB8xhZ4fZ0n+mUPcp6P81OYQL8cNfdN7xFsCd0S/PFnJ9KLCwx3KIehTskSgWJeHlpiS83PyfhTW7PUwGVlRfChk/4cmJKi9WGPPirLCofOPFrht4rru9oax4erPpbm8oC2j6sPfZgK+8ON9R1RDwc/E0fSjrENZ6HmENb/SOquY7DHyf6V3X5cA/KtznPwyc9dKZ0b7P00FjAyrvAqPFtvMxqvBOOmhs9MvCZXigLWWasvYmxi58ZYkSJYrCf1pFjw96coWwAAAAAElFTkSuQmCC",
        "phone": "iVBORw0KGgoAAAANSUhEUgAAACwAAAAsCAYAAAAehFoBAAAABmJLR0QA/wD/AP+gvaeTAAAF8UlEQVRYhe2Zf2xV5RnHP897b0sZIj9MVNqQqYNtiNMM2SY959KaoU4FE5dYNUFFS2/rwoyQaMxiZrfFLNniryCztxW2MGtYWbZpxKaEWOw9tyjCFuNqdOqWkIBbHIGBob8477M/7rk/entaL9BT/+H713me83zf93Pfc+/741w4r/MaIylNeFud2WLNTViusKLj7hsYVdH3jsypfqOhYac/PZgFjQFKt7v3CDwHzCvD+w/BrnWS/e9EgxauPHCmPXGXoi8TMuqT6KRR/7u1zfs+mHq0cMUB+p9eMdOim8nBir4I/FlVRkoNInohKj8G6oHZVuJPAbdOF7AAeO3OzSCvB7ntbtK7bzLTQNfSymPH570HfB3wq7TqouXNe/4XMSsABkBFFuYSCnu+yLS0YWBEoDsIY0NmcElEfONkAAT9bz6jXFyO0QrvFyxy9ZSTTaDsCPvyST4jfKMco1o5kLcozpSTTSADcHpG/EPgNIBAWaP173mXvgscD8KbtDXbVtQyANffv3cI8o/4moGupZVfZGxo2OkXfY8v8WrclRExjlF+VER5K7isOnpi/nfKMfsqXQW/TMvUlge2hr35pLU3lmfWvxaFC6YOa7I+A2lM9gDZvYGypky/m79Se2QqwSZSHnjlA+nPgH0AiHz7zbaViycz9qdWfBNhc74hif0lKshijfllq7Ajdx039t6JTJkXai+2xHYB8wEQXq5N9vVHBVmsMcCVwyM7gGEAhcYDqWsrwkwai/0W4YogfMuc9NdHi1nQGODvbdh/FOSPQbhg0My8O9ymcwqX+rvaTfsGowIs1bjJXkWfKQQ81tV1R2x8DY8Dmm3BPOltdaojZByjccCJJu+gQg+AwJLqY0fWhtTsBbYBoHoRvnSGfbAoFL6cKk+QG0GRJ3u31F9QWhIfGt2E8s8grF9w/NNfRQVZrFDgRLP3NoVVrCZecfpnpTXXPfT2CVTvBIYABDZlUu5D0aFmNcmGRR8FPg9gHvbandtKK9yWzAHQ9QRPQ4VnvXanJRrUrCYEdpu9Q6CPFOqksy+VWDauLpnpFJXHglBAfuN1JH4SBWzQweRKp9xtItwfhP8hpnVuY+bDcXXt7uMCvyhq+PczdLBlefPBUxCcGy+IPwLcA7oI+BylV4x0xYeHu7NT6hQAB+e3XcCqIHXYqL8q7KTspZwNiDwL5GaMAcGuFaOHrI3tBq6doBsLDIAeVOQTEX3fnLTdYfN7WUd6b6szG192A9cFqc+MkdW169P7x9V2OLeg0gnMDVKjoIdBLgviIUT3o7IQuHySbj8Vteuc5v7dZwwMcCC1as6QGe5GdUWQOiVIo5NM7yitzaRqv6ZiXir6gDkd8q1ZVdfS9xGA1+YsJ2ZuU9XvCywDqkrqRxQenD/32EtLGwZGzggYoGf7jbNmDZ3aCdxcyOrm0YqKR4NTS169rfXxihp/I6pPALOAQ76Y+rqmvn+Ftd3bWh+vqrGLffzLsdKI8MPcPYHnnKT38BkD50Gq/WdANxSY+buIfSDstdWb21YujI/qao3bV9zGTFl7Zm3FZKrdDIUntMtNeqvPCjindId7ryhbgNwq6Auk4iMjPy33Fz+ZvA73lyi56TIPfNYn3USTt923ZplCbh8cU/jRaGXlx14qsbG3tT5+LsAKoV+dczqa17X0feQe8RIqbKBw5J+L6NMV1acPeh3ODWfb9qkZX+kMy5/zuwRpxSaavC2jGl8M8jzB+w3galR2e+3uG17Kuf5c+8n3N1UN5ZRuc68Rw2YgMeaG6t/USNtMW/WHcl4cBnP/iQDyNSfprYkEOKd0u3u7KD9HuKrk1qBAD8KrarRnopkjWIB2ZSNpc5PpByMFhuz01F+TWIPqRoW68Cr5WNB3LLxrRD5Q3x5WI0sEfg1cErR0i5vMdEcOXKy+tsSVRuw6RBqAr56Btdtp8m4VyW5hpw24WH2pxDJj+AGq9WQXh9mhhcqfiOs6tzFzMpf6UoCLpa0Yr9pdZJBvWXSRqFyqokcNtme6//A5L4D/A4kdIGGKxVAQAAAAAElFTkSuQmCC",
        "bulb": "iVBORw0KGgoAAAANSUhEUgAAACwAAAAsCAYAAAAehFoBAAAABmJLR0QA/wD/AP+gvaeTAAAEPElEQVRYhe2ZXWgcVRTHf2dmN7bRxibaF634hUKSImir2N1MbQTFD1JbMEHxA1PIJqVYJU9SX/pmRRHyoCbbVh9EilvBtgYlVXCbTJKX1Cra6kMRFbVUiwklZZPuzBwfdjeN0WTvTDb0Jf+ncy/3f86P4d47c+/AspZWUokkqsjoAefeQHlAA623hDWqqMJ5ET0t6PFEx8iYCHpFgU9lGqsmxmu7VNgF3F6m1Bk06KmtnUg3tp26FLVmZOCR/c59QaAfAHeG8Sn8gPCc0+GeiFLXimIaTjtPBYEOchnWBzmq8Lylfr161KlHnS1WA6ovCPQDAYBAvSju8D7nySi1Qz/hoXTTNoFDgF1McFzU70p0jv64kG+w12mwLO0DmopdnsI2J+X2Lxnw8d5Nd9hW8DVwTdHck/jD7ZY9hadXTplMq33D+NkehJ3Frgsa425nu/uTKUMsDLBt+b0gBVglnex0Xw7jb2s75KvyoptuqhahHagRn7eBR01zGM9hty/ZDPIgFBbOparYS2FgSxJB7Yv+TpAzFJI94u5LOqb+EIvO6ixFqrK7uT07FYLzX0p0j+ZUdfdMR0DKmMJk0Fjf+jiijxebvztnh46GQ/yvvLOxT4BzAIi0ZDKttonPCHharmrk8kIbMF1kC6l5T9YDBorNa9eOn2sw8RkBq8htM7HK9+Hx5s38XSny0DJvyoLM5nDAqlIolk6E5ppHivw9K2+NiccIWMSaLsWBsjI82nyJL+cSdHqhoSWZASO/zorrw5PNk1dl3azmLyYesykxmT8J5AAU3TrWtz4emm6OCjm0pdi8eCFXc9LEZwSc6B7NKRwBEFg7LdVdETlnlLNWpoAbKSQ98tiuzys3JQAI2Av4AIrudXuTG8JjFjTY59wjyuulzKLBG6ZeY2Cny/0W1beKzWosGRjc59wVBhRgeH9inWVxDLi60KM9ydTIN6b+UN/DK5h6FeRYsVlnBdo/sn9jnal/6J2mWg2sflSvK3Z9Wbt64pUwDKGAN3SeyFuT3lYgC4Bwk/r2a8bFYrwJ3FxsDq3Q3BNhj0uRjkiD7zlrLE9PA9cDXkxit97fkf1tIc/ou5tv8W3vDIUP//PiB43JHSN/hq0d6Yi0afvQXzAzn2O+em3lPL6df5riKUWVvVFgISIwgO3HD5ZiRRNlDSozu0rcin0UtW5k4I07sj8LHAbyglX2XCYinxbGcrjc9FlShXnrVeINuaxlzdGi7ta+en/z6ljea5n9XbtgsYDJfFXss+b2bORDQKh7ibmK570M8JDxnaRAPO99ATwctWbkbe1KaVFPOB+PtUWZEoupuaxlzVFF/nEAuAeSq3Ra/ndN2FP+VKJ7NFeJOhUBHko3HRJY8EZd4WMn5bYuttait7VMptUW2FJunMAW0wu/MnkWLzedfAaRZ9F5t8lp0INNqeEPK1FvWUupfwBnX22Ggvw/rAAAAABJRU5ErkJggg==",
    }

    def png_ico(key, size=22):
        b64 = ICONS_PNG.get(key, "")
        if not b64: return ""
        return f'<img src="data:image/png;base64,{b64}" width="{size}" height="{size}" alt="" style="display:inline-block;vertical-align:middle;border:0">'

    BENS = {'restaurante': [('wheat', 'Receita artesanal desde 1984', 'Ingredientes simples, sem conservantes, feito com o cuidado de quem tem 40 anos de prática'), ('snowflake', 'Dose individual congelada', 'Serve só o que precisa, quando precisa — zero desperdício, total comodidade'), ('chef', 'Sem precisar de pasteleiro', 'Descongela em minutos e vai directo ao prato, mantendo toda a qualidade'), ('coin', 'Quatro sabores à escolha', 'Original, Chocolate, Canela e Café — para combinar com qualquer carta')], 'hotel': [('crown', 'Qualidade constante, em qualquer turno', 'Artesanal e consistente — do primeiro ao último hóspede do dia'), ('snowflake', 'Sem preparação, sem perdas', 'Dose individual congelada — do frigorífico ao prato em minutos'), ('globe', 'Uma história para contar', 'Receita levada pelos portugueses ao Japão no século XVI — genuinamente portuguesa'), ('cup', 'Versátil em qualquer serviço', 'Restaurante, room service, pequeno-almoço ou eventos')], 'pastelaria': [('wheat', 'Complemento, não concorrente', 'Uma referência nacional que os vossos clientes já conhecem e procuram'), ('snowflake', 'Dose individual, controlo total', 'Sem perdas, sem desperdício — vai ao balcão quando é pedido'), ('scroll', 'Receita com 40 anos intacta', 'Ingredientes simples, sem conservantes, origem verificável'), ('coin', 'Margem interessante para revenda', 'Quatro sabores: Original, Chocolate, Canela e Café')], 'catering': [('star', 'Para qualquer dimensão de evento', 'De 20 a 2000 convidados, com a mesma qualidade e simplicidade'), ('snowflake', 'Zero logística no evento', 'Dose individual — sem corte, sem preparação, sem imprevistos'), ('heart', 'Uma sobremesa que fica na memória', 'Produto artesanal com história — os convidados percebem a diferença'), ('wheat', 'Consistência em qualquer volume', 'Quatro sabores com qualidade igual em qualquer quantidade')], 'distribuidor': [('truck', 'Produto com procura crescente', 'Cada vez mais solicitado no canal HORECA em todo o país'), ('coin', 'Condições competitivas por volume', 'Margem de distribuição atrativa, MOQ negociável por região'), ('map', 'Regiões com potencial por explorar', 'Zonas sem cobertura e clientela identificada à vossa espera'), ('leaf', 'Marca com 40 anos e terceira geração', 'Reconhecimento nacional, história real, produto diferenciador')], 'gourmet': [('scroll', 'Uma narrativa que se conta em duas frases', '40 anos de receita intacta — os seus clientes vão querer saber mais'), ('leaf', 'Origem verificável, ingredientes simples', 'Produção familiar portuguesa — transparência total'), ('snowflake', 'Fácil de gerir em prateleira', 'Congelado individual — sem risco de validade, sem perdas'), ('star', 'Posicionamento premium acessível', 'Quatro sabores com valor percebido elevado')], 'cafe': [('cup', 'O acompanhamento natural do café', 'Pão de ló artesanal e café de especialidade — uma combinação com alma'), ('coin', 'Ticket médio mais elevado', 'Produto de impulso com margem interessante ao balcão'), ('scroll', 'Uma história que o barista pode contar', 'Dois minutos de conversa que tornam o produto inesquecível'), ('snowflake', 'Dose individual, zero desperdício', 'Serve só o que precisa — sem bolos inteiros que ficam para trás')], 'cervejaria': [('bolt', 'Do frio ao prato em minutos', 'Zero preparação, zero espera — perfeito para o ritmo de uma cervejaria'), ('snowflake', 'Custo fixo e previsível', 'Dose individual — sem surpresas no final do mês'), ('coin', 'Margem superior ao habitual', 'PVP atrativo com rentabilidade melhor do que pudim ou mousse'), ('wheat', 'Os clientes reconhecem e pedem', 'Uma sobremesa portuguesa de sempre — sem precisar de ser explicada')]}

    INTROS_ALL = {'hotel': ['identificámos o vosso hotel como um espaço que cuida da experiência dos hóspedes em cada detalhe — incluindo à mesa.', 'Na semana passada tivemos o prazer de nos apresentar. Hoje gostaríamos de partilhar algo concreto — um resultado real, de uma unidade hoteleira semelhante à vossa.', 'Há uma tendência que temos observado no setor hoteleiro e que nos parece relevante partilhar.', 'Ao longo das últimas semanas partilhámos a nossa história e alguns casos reais no canal hoteleiro.'], 'pastelaria': ['identificámos a vossa pastelaria como um espaço que valoriza o produto artesanal — e gostaríamos de vos apresentar um complemento que acreditamos ser uma boa adição ao vosso catálogo.', 'Na semana passada tivemos o prazer de nos apresentar. Hoje gostaríamos de partilhar a experiência de uma pastelaria que trabalha connosco.', 'Há algo que temos observado no mercado da pastelaria artesanal que nos parece relevante partilhar convosco.', 'Ao longo das últimas semanas partilhámos a nossa história e casos reais em pastelarias parceiras.'], 'catering': ["identificámos a vossa empresa como uma referência na organização de eventos — e gostaríamos de partilhar como o Ti'Piedade pode simplificar a vossa proposta de sobremesas.", 'Na semana passada tivemos o prazer de nos apresentar. Hoje partilhamos a experiência de uma empresa de catering que trabalha connosco.', 'Há uma tendência nos eventos premium que nos parece relevante partilhar convosco.', 'Ao longo das últimas semanas partilhámos a nossa história. Esta semana gostaríamos de passar para deixar amostras.'], 'distribuidor': ["identificámos a vossa empresa como um parceiro com cobertura regional relevante — e gostaríamos de explorar uma possível colaboração na distribuição do Ti'Piedade.", 'Na semana passada tivemos o prazer de nos apresentar. Hoje partilhamos dados concretos sobre a procura do nosso produto.', 'Há regiões com procura crescente onde ainda não temos representação — e acreditamos que a vossa empresa seria o parceiro certo.', 'Ao longo das últimas semanas partilhámos a nossa história e o potencial de mercado. Esta semana gostaríamos de apresentar as condições comerciais.'], 'gourmet': ["identificámos o vosso espaço como um lugar que cuida da narrativa de cada produto em prateleira — e acreditamos que o Ti'Piedade tem muito a oferecer nesse sentido.", 'Na semana passada tivemos o prazer de nos apresentar. Hoje partilhamos a experiência de uma mercearia gourmet que trabalha connosco.', 'Há algo que os consumidores de produto gourmet estão a valorizar cada vez mais — e que nos parece relevante partilhar convosco.', 'Ao longo das últimas semanas partilhámos a nossa história. Esta semana gostaríamos de passar para mostrar o produto pessoalmente.'], 'cafe': ["identificámos o vosso café como um espaço que cuida da experiência de cada cliente — e gostaríamos de partilhar como o Ti'Piedade pode complementar a vossa oferta.", 'Na semana passada tivemos o prazer de nos apresentar. Hoje partilhamos a experiência de um café que trabalha connosco.', 'Há uma tendência que temos observado nos cafés de especialidade que nos parece relevante partilhar.', 'Ao longo das últimas semanas partilhámos a nossa história. Esta semana gostaríamos de passar com amostras para a vossa equipa provar.'], 'cervejaria': ["identificámos a vossa cervejaria como um espaço com grande volume de refeições — e gostaríamos de partilhar como o Ti'Piedade pode tornar a vossa sobremesa mais simples e mais rentável.", 'Na semana passada tivemos o prazer de nos apresentar. Hoje partilhamos a experiência de uma cervejaria que trabalha connosco.', 'Há algo que estamos a observar no setor das cervejarias que nos parece relevante partilhar.', 'Ao longo das últimas semanas partilhámos a nossa história. Esta semana gostaríamos de passar com amostras.'], 'restaurante': ['identificámos o vosso restaurante como um espaço que valoriza a qualidade à mesa — e gostaríamos, com toda a simplicidade, de vos apresentar o nosso produto.', 'Na semana passada tivemos o prazer de nos apresentar. Hoje gostaríamos de partilhar algo concreto — um resultado real, de um espaço semelhante ao vosso.', 'Há uma tendência que temos observado no setor e que nos parece relevante partilhar — não como argumento de venda, mas como reflexão genuína sobre o que os clientes estão a valorizar.', 'Ao longo das últimas semanas partilhámos a nossa história e alguns resultados reais. Esta semana, se nos permitirem, gostaríamos de passar para deixar amostras dos quatro sabores.']}

    HISTORIA = {'1': {'titulo': 'Uma receita com mais de quatro décadas', 'texto': 'Em 1984, a D.ª Piedade começou a fazer pão de ló numa pequena cozinha familiar, com ingredientes simples e a mesma receita de sempre. Hoje, em terceira geração, mantemos cada ingrediente, cada proporção, cada detalhe intacto. Não por teimosia — mas porque acreditamos que o que é genuíno não precisa de ser melhorado. Só precisa de chegar a mais mesas.'}, '2': {'titulo': 'Feito com mãos, não com máquinas', 'texto': "O Pão de Ló Ti'Piedade é produzido de forma artesanal, em pequenas séries, por uma equipa que conhece cada lote pelo toque e pelo cheiro. Num tempo em que tudo é escalonado e optimizado, escolhemos manter o ritmo de quem faz com cuidado. Os nossos clientes percebem isso — e os seus clientes também."}, '3': {'titulo': 'Uma história que atravessou oceanos', 'texto': "O pão de ló chegou ao Japão no século XVI, trazido pelos navegadores portugueses. Hoje, o kasutera japonês é primo directo da nossa receita — e prova que um produto genuíno resiste ao tempo e à distância. Na Ti'Piedade, orgulhamo-nos de ser guardiões desta tradição, com a humildade de quem sabe que serve algo maior do que si próprio."}, '4': {'titulo': 'Três gerações. A mesma receita.', 'texto': 'Quando a D.ª Piedade criou este pão de ló, não imaginou que um dia chegaria a hotéis, restaurantes e pastelarias por todo o país. Mas sempre soube que o produto era bom. E os bons produtos encontram o seu caminho. A nossa família passou a receita de mão em mão, com o mesmo cuidado com que se passa um nome. Esperamos que a vossa equipa possa sentir isso em cada fatia.'}}

    CTA = {1: ('✉ Responder', '#5C2D0E', 'Ficamos ao dispor para qualquer questão. Será um prazer apresentar o produto pessoalmente.'), 2: ('✉ Saber mais', '#D4682A', 'Se quiser perceber como isto poderia funcionar no vosso espaço, estamos disponíveis para uma conversa sem compromisso.'), 3: ('✉ Vamos conversar?', '#5C2D0E', 'Podemos passar pela vossa zona na próxima semana, sem pressa nem compromisso.'), 4: ('✉ Agendar uma visita', '#1A6B3A', 'Sem apresentações formais, sem pressão. Só o produto, para o descobrirem por si próprios.')}

    cors = {1:"#C49A3C",2:"#D4682A",3:"#5C2D0E",4:"#1A6B3A"}
    cor_barra = cors.get(email_num,"#C49A3C")
    intros = INTROS_ALL.get(grupo, INTROS_ALL["restaurante"])
    frase  = intros[email_num-1] if email_num <= len(intros) else intros[0]
    intro  = frase if frase.lower().startswith("bom dia") else (
        f"Bom dia,<br><br>O meu nome é <strong>Rui Bernardes</strong> e faço parte da equipa "
        f"comercial do <strong>Pão de Ló Ti'Piedade</strong>.<br><br>{frase}"
    )
    cta_btn, cta_cor, cta_sub = CTA.get(email_num, CTA[1])
    h = HISTORIA.get(str(email_num), HISTORIA["1"])
    historia = (
        f'<div style="background:#4a2417;border-radius:10px;padding:18px 20px;margin-top:16px">'
        f'<p style="margin:0;font-size:9px;font-weight:700;color:#C49A3C;letter-spacing:.12em;text-transform:uppercase">A nossa história</p>'
        f'<p style="margin:6px 0 0;font-size:13px;font-weight:700;color:#ffffff;font-family:Georgia,serif">{h["titulo"]}</p>'
        f'<p style="margin:8px 0 0;font-size:12px;color:rgba(255,255,255,.75);line-height:1.7">{h["texto"]}</p>'
        f'</div>'
    )

    def bens_html(tipo):
        rows = ""
        for ico_key, title, desc in BENS.get(tipo, BENS["restaurante"]):
            ico_img = png_ico(ico_key)
            rows += (
                f'<tr><td style="padding:0 0 16px">'
                f'<table width="100%" cellpadding="0" cellspacing="0"><tr>'
                f'<td width="56" style="vertical-align:top;padding-right:14px">'
                f'<div style="width:44px;height:44px;background:#FDF6EF;border:1px solid #E8D5B8;border-radius:8px;text-align:center;line-height:44px">'
                f'{ico_img}</div></td>'
                f'<td style="vertical-align:middle">'
                f'<p style="margin:0;font-size:13px;font-weight:700;color:#4a2417;font-family:Georgia,serif">{title}</p>'
                f'<p style="margin:3px 0 0;font-size:12px;color:#6B5744;line-height:1.5">{desc}</p>'
                f'</td></tr></table></td></tr>'
            )
        return rows

    if email_num == 2:
        central = (
            '<div style="background:#FDF6EF;border:1px solid #E8D5B8;border-radius:10px;padding:16px 18px">'
            '<p style="margin:0;font-size:10px;font-weight:700;color:#C49A3C;letter-spacing:.08em;text-transform:uppercase">Um caso real</p>'
            '<p style="margin:5px 0 0;font-size:13px;font-weight:700;color:#4a2417;font-family:Georgia,serif">Espaço semelhante ao vosso &nbsp;·&nbsp; Três meses de parceria</p>'
            '<p style="margin:3px 0 10px;font-size:12px;color:#6B5744">Canal HORECA &nbsp;·&nbsp; Portugal</p>'
            + f'<p style="margin:4px 0;font-size:12px;color:#2C1A0A">{png_ico("star",14)}&nbsp; Passou a ser uma das sobremesas mais pedidas da carta</p>'
            + f'<p style="margin:4px 0;font-size:12px;color:#2C1A0A">{png_ico("star",14)}&nbsp; Desperdício em sobremesas: praticamente zero</p>'
            + f'<p style="margin:4px 0;font-size:12px;color:#2C1A0A">{png_ico("star",14)}&nbsp; <em>&ldquo;É o produto mais fácil que temos. Sai do frio, vai ao prato.&rdquo;</em></p>'
            + '</div>'
        )
    elif email_num == 3:
        central = (
            '<div style="background:#FDF6EF;border-left:3px solid #C49A3C;border-radius:0 8px 8px 0;padding:14px 16px;margin-bottom:10px">'
            '<p style="margin:0;font-size:13px;font-weight:700;color:#4a2417;font-family:Georgia,serif">O que os clientes estão a valorizar</p>'
            '<p style="margin:6px 0 0;font-size:12px;color:#2C1A0A;line-height:1.7">Há uma preferência crescente por produtos com história verificável, ingredientes simples e produção artesanal. A doçaria industrial tem cada vez menos espaço — e o produto artesanal com narrativa está a ganhar terreno de forma consistente.</p>'
            '</div>'
            '<div style="background:#fff;border:1px solid #E8D5B8;border-radius:8px;padding:14px 16px">'
            '<table width="100%" cellpadding="0" cellspacing="0"><tr>'
            '<td width="50%" style="vertical-align:top;padding-right:10px">'
            '<p style="margin:0 0 4px;font-size:10px;font-weight:700;color:#C49A3C;text-transform:uppercase;letter-spacing:.06em">Para os vossos clientes</p>'
            '<p style="margin:0;font-size:12px;color:#2C1A0A;line-height:1.6">Um produto artesanal com 40 anos de história — genuíno e reconhecível</p>'
            '</td><td width="50%" style="vertical-align:top;padding-left:10px;border-left:1px solid #E8D5B8">'
            '<p style="margin:0 0 4px;font-size:10px;font-weight:700;color:#C49A3C;text-transform:uppercase;letter-spacing:.06em">Para a vossa operação</p>'
            '<p style="margin:0;font-size:12px;color:#2C1A0A;line-height:1.6">Dose individual congelada — zero preparação, zero desperdício, margem previsível</p>'
            '</td></tr></table></div>'
        )
    elif email_num == 4:
        central = (
            f'<div style="background:#FDF6EF;border-radius:10px;padding:22px;text-align:center">'
            f'<div style="margin-bottom:12px">{png_ico("gift",36)}</div>'
            f'<p style="margin:0 0 4px;font-size:15px;font-weight:700;color:#4a2417;font-family:Georgia,serif">Amostras sem compromisso</p>'
            f'<p style="margin:0;font-size:12px;color:#6B5744">Original &nbsp;·&nbsp; Chocolate &nbsp;·&nbsp; Canela &nbsp;·&nbsp; Café</p>'
            f'<p style="margin:12px 0 0;font-size:12px;color:#2C1A0A;line-height:1.7">Sem reuniões formais, sem apresentações longas.<br>Só o produto — para a vossa equipa provar e julgar com toda a liberdade.</p>'
            f'</div>'
        )
    else:
        central = f'<table width="100%" cellpadding="0" cellspacing="0">{bens_html(grupo)}</table>'

    return f"""<!DOCTYPE html>
<html lang="pt">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<!--[if mso]><xml><o:OfficeDocumentSettings><o:PixelsPerInch>96</o:PixelsPerInch></o:OfficeDocumentSettings></xml><![endif]-->
</head>
<body style="margin:0;padding:0;background:#F0E8DC;font-family:Arial,Helvetica,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#F0E8DC;padding:24px 0">
<tr><td align="center">
<table width="580" cellpadding="0" cellspacing="0" style="max-width:580px;width:100%">
  <tr><td style="background:#4a2417;border-radius:16px 16px 0 0;padding:22px 36px 16px;text-align:center">
    <img src="{LOGO}" alt="Ti'Piedade" width="180" style="height:28px;width:auto;display:block;margin:0 auto;border:0">
    <div style="height:2px;background:linear-gradient(90deg,transparent,#C49A3C 30%,#D4682A 70%,transparent);margin-top:14px;border-radius:2px"></div>
  </td></tr>
  <tr><td style="background:{cor_barra};height:4px"></td></tr>
  <tr><td style="background:#ffffff;padding:26px 36px 0;border-left:1px solid #DDD0BB;border-right:1px solid #DDD0BB">
    <p style="margin:0;font-size:13px;color:#2C1A0A;line-height:1.8">{intro}</p>
    <div style="height:1px;background:#E8D5B8;margin:20px 0"></div>
    <p style="margin:0 0 16px;font-size:10px;font-weight:700;color:#C49A3C;letter-spacing:.12em;text-transform:uppercase">Porquê o Ti'Piedade?</p>
  </td></tr>
  <tr><td style="background:#ffffff;padding:0 36px;border-left:1px solid #DDD0BB;border-right:1px solid #DDD0BB">
    {central}
    {historia}
  </td></tr>
  <tr><td style="background:#ffffff;padding:18px 36px 26px;border-left:1px solid #DDD0BB;border-right:1px solid #DDD0BB">
    <div style="height:1px;background:#E8D5B8;margin-bottom:16px"></div>
    <p style="margin:0 0 14px;font-size:12px;color:#6B5744;line-height:1.7">{cta_sub}</p>
    <table cellpadding="0" cellspacing="0"><tr><td style="background:{cta_cor};border-radius:8px">
      <a href="mailto:sales@tipiedade.com" style="display:block;padding:11px 24px;color:#fff;font-size:13px;font-weight:700;text-decoration:none">{cta_btn}</a>
    </td></tr></table>
  </td></tr>
  <tr><td style="background:#FDF6EF;padding:16px 36px;border:1px solid #DDD0BB;border-top:none">
    <p style="margin:0;font-size:10px;font-weight:700;color:#C49A3C;letter-spacing:.08em;text-transform:uppercase">O produto</p>
    <p style="margin:4px 0 0;font-size:13px;font-weight:700;color:#4a2417;font-family:Georgia,serif">Pão de Ló Ti'Piedade · Unidose 85g (congelado)</p>
    <p style="margin:2px 0 0;font-size:11px;color:#6B5744">Original · Chocolate · Canela · Café &nbsp;·&nbsp; Sem conservantes · Desde 1984</p>
  </td></tr>
  <tr><td style="background:#4a2417;border-radius:0 0 16px 16px;padding:14px 36px">
    <table width="100%" cellpadding="0" cellspacing="0"><tr>
      <td><p style="margin:0;font-size:11px;color:#C49A3C;font-weight:600">Rui Bernardes · Departamento Comercial</p>
          <p style="margin:2px 0 0;font-size:10px;color:rgba(255,255,255,.45)">sales@tipiedade.com · www.tipiedade.com</p></td>
      <td align="right"><a href="https://www.tipiedade.com" style="font-size:10px;font-weight:700;color:#C49A3C;text-decoration:none;border:1px solid rgba(196,154,60,.4);padding:5px 12px;border-radius:5px">Ver produto ↗</a></td>
    </tr></table>
  </td></tr>
</table></td></tr></table>
</body></html>"""


def enviar_nurturing_lead(server, smtp_user, lead, email_num, nome_comercial):
    """Envia o email de nurturing em HTML directamente ao email da lead."""
    email_dest = lead.get("email","").strip()
    if not email_dest or email_dest == "—" or "@" not in email_dest:
        return False

    emails = [e.strip() for e in email_dest.replace(";",",").split(",") if "@" in e.strip()]
    if not emails:
        return False

    grupo   = tipologia_grupo(lead.get("t",""))
    assunto = ASSUNTOS.get(grupo, ASSUNTOS["restaurante"]).get(email_num, "")
    texto   = corpo_email(email_num, grupo, lead.get("n",""), lead.get("zona",""), nome_comercial)
    html    = corpo_html(email_num, grupo, lead.get("n",""), lead.get("zona",""), nome_comercial, texto)

    for dest_email in emails:
        msg = MIMEMultipart("alternative")
        msg["From"]    = "Pão de Ló Ti'Piedade <sales@tipiedade.com>"
        msg["To"]      = dest_email
        msg["Subject"] = f"{assunto} | Ti'Piedade"
        # Plain text fallback
        msg.attach(MIMEText(texto, "plain", "utf-8"))
        # HTML principal
        msg.attach(MIMEText(html, "html", "utf-8"))
        # Enviar para destinatário + BCC duplo
        server.sendmail(EMAIL_FROM, [dest_email, EMAIL_CC, EMAIL_BCC2], msg.as_string())

    return True


def enviar(server, de, para, cc, assunto, corpo, ficheiro=None):
    """Envia email com ou sem anexo."""
    msg = MIMEMultipart()
    msg["From"] = f"Ti'Piedade HORECA <{de}>"
    msg["To"]   = para
    msg["CC"]   = cc
    msg["Subject"] = assunto
    msg.attach(MIMEText(corpo, "plain", "utf-8"))
    if ficheiro:
        with open(ficheiro,"rb") as f:
            part = MIMEBase("application","octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{os.path.basename(ficheiro)}"')
        msg.attach(part)
    server.sendmail(de, [para, cc], msg.as_string())


def enviar_emails(ficheiro, sem, modo):
    smtp_host = os.environ["SMTP_HOST"]
    smtp_port = int(os.environ.get("SMTP_PORT","587"))
    smtp_user = os.environ["SMTP_USER"]
    smtp_pass = os.environ["SMTP_PASS"]

    with smtplib.SMTP(smtp_host, smtp_port) as server:
        server.ehlo(); server.starttls(); server.login(smtp_user, smtp_pass)

        if modo == "coordenador":
            total_horeca = sum(len(gerar_leads_horeca(c,sem)) for c in ["nuno","joao","oscar"])
            corpo = f"""Bom dia,

Resumo de leads gerados para revisão — Semana {sem} ({semana_datas()}):

HORECA por comercial:
  Nuno  → {len(gerar_leads_horeca('nuno',sem))} leads
  João  → {len(gerar_leads_horeca('joao',sem))} leads
  Óscar → {len(gerar_leads_horeca('oscar',sem))} leads

Para ti:
  Catering & Eventos → {len(gerar_leads_canal(DB_CATERING,'rui_catering',sem,5))} leads
  Distribuidores     → {len(gerar_leads_canal(DB_DISTRIBUIDORES,'rui_distribuidores',sem,5))} leads

Total: {total_horeca + 10} leads esta semana

Na quarta-feira:
→ Os comerciais recebem as suas leads
→ O Email 1 de nurturing é enviado automaticamente a cada lead com email disponível
→ O histórico é atualizado no CRM

O Excel em anexo tem o detalhe completo.

Ti'Piedade — Sistema de Prospeção HORECA
"""
            enviar(server, smtp_user, EMAIL_RUI, EMAIL_CC,
                   f"[REVISÃO] Leads Semana {sem} — {total_horeca+10} contactos | Ti'Piedade",
                   corpo, ficheiro)
            print(f"✓ Resumo enviado para {EMAIL_RUI}")

        elif modo == "comerciais":
            historico, sha = ler_historico()
            enviados_leads = 0
            sem_email = 0

            for com_id in ["nuno","joao","oscar"]:
                com = COMERCIAIS[com_id]
                leads = gerar_leads_horeca(com_id, sem)
                dest = com["email"]

                # Determinar email_num para esta semana (leads novas = 1, continuação = próximo)
                email_num = next(
                    (calcular_email_num(historico, com_id, l) for l in leads
                     if calcular_email_num(historico, com_id, l)),
                    1
                )

                # ── Enviar nurturing diretamente às leads ──────────
                leads_com_email = 0
                for lead in leads:
                    n = calcular_email_num(historico, com_id, lead)
                    if n is None:
                        continue  # sequência completa
                    ok = enviar_nurturing_lead(server, smtp_user, lead, n, "Rui Bernardes")
                    if ok:
                        leads_com_email += 1
                        enviados_leads += 1
                        historico = registar_envios(historico, com_id, [lead], n, com["nome"])
                    else:
                        sem_email += 1

                # ── Email resumo ao comercial ───────────────────────
                if dest:
                    grupo_exemplo = tipologia_grupo(leads[0]["t"]) if leads else "restaurante"
                    assunto_ex = ASSUNTOS.get(grupo_exemplo, ASSUNTOS["restaurante"]).get(email_num,"")
                    corpo_com = f"""Olá {com['nome']},

Aqui estão os teus leads HORECA para a semana {sem} ({semana_datas()}).

O Rui Bernardes (Departamento Comercial) enviou automaticamente o Email {email_num} da sequência de nurturing a {leads_com_email} leads com email disponível ({sem_email} sem email — para estes contacta por telefone).

A sequência tem 4 emails semanais. Quando aparecer "✓ Pronto p/ visita" no Excel — o contacto já recebeu todos os emails e está preparado para a tua visita.

O teu papel entra aqui: contacto presencial, apresentação do produto e fecho da venda.

Bom trabalho,
Rui Bernardes
Departamento Comercial | Ti'Piedade
sales@tipiedade.com
"""
                    enviar(server, smtp_user, dest, EMAIL_CC,
                           f"Leads Semana {sem} — {len(leads)} contactos | Ti'Piedade",
                           corpo_com, ficheiro)
                    print(f"✓ {com['nome']} — {leads_com_email} emails nurturing enviados às leads")

            # ── Rui — catering ────────────────────────────────────
            leads_cat = gerar_leads_canal(DB_CATERING, "rui_catering", sem, 5)
            for lead in leads_cat:
                n = calcular_email_num(historico, "rui_catering", lead)
                if n:
                    ok = enviar_nurturing_lead(server, smtp_user, lead, n, "Rui Bernardes")
                    if ok:
                        enviados_leads += 1
                        historico = registar_envios(historico, "rui_catering", [lead], n, "Rui")

            corpo_cat = f"""Olá,

Aqui estão os 5 leads de Catering & Eventos para a semana {sem} ({semana_datas()}).

O Email 1 de nurturing foi enviado automaticamente às leads com email disponível.
Para as restantes, o contacto é telefónico direto.

Ti'Piedade — Sistema de Prospeção
"""
            enviar(server, smtp_user, EMAIL_RUI, EMAIL_CC,
                   f"Leads Catering & Eventos — Semana {sem} | Ti'Piedade",
                   corpo_cat, ficheiro)
            print(f"✓ Catering → Rui")

            # ── Rui — distribuidores ──────────────────────────────
            leads_dist = gerar_leads_canal(DB_DISTRIBUIDORES, "rui_distribuidores", sem, 5)
            for lead in leads_dist:
                n = calcular_email_num(historico, "rui_distribuidores", lead)
                if n:
                    ok = enviar_nurturing_lead(server, smtp_user, lead, n, "Rui Bernardes")
                    if ok:
                        enviados_leads += 1
                        historico = registar_envios(historico, "rui_distribuidores", [lead], n, "Rui")

            corpo_dist = f"""Olá,

Aqui estão os 5 leads de Distribuidores para a semana {sem} ({semana_datas()}).

O Email 1 de nurturing foi enviado automaticamente às leads com email disponível.

Ti'Piedade — Sistema de Prospeção
"""
            enviar(server, smtp_user, EMAIL_RUI, EMAIL_CC,
                   f"Leads Distribuidores — Semana {sem} | Ti'Piedade",
                   corpo_dist, ficheiro)
            print(f"✓ Distribuidores → Rui")

            print(f"\n📧 Total nurturing enviado: {enviados_leads} emails às leads | {sem_email} sem email")
            escrever_historico(historico, sha)

# ════════════════════════════════════════════════════════════════

if __name__=="__main__":
    sem=semana_num(); modo=os.environ.get("MODO","coordenador")
    print(f"▶ Semana {sem} — {semana_datas()} — Modo: {modo}")
    historico,sha=ler_historico()
    ficheiro=criar_excel(sem,historico)
    print(f"✓ Excel: {ficheiro}")
    enviar_emails(ficheiro,sem,modo)
    print("✓ Concluído.")
